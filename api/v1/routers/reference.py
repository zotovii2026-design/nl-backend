"""Роутер Справочника — вынесен из api/v1/nl.py

Эндпоинты:
- GET  /api/v1/nl/reference
- POST /api/v1/nl/reference
- GET  /api/v1/nl/fbs-warehouses
- GET  /api/v1/nl/cost-prices
- POST /api/v1/nl/cost-prices
- POST /api/v1/nl/cost-prices/batch
- GET  /api/v1/nl/commission-rate
- POST /api/v1/nl/cost-prices/auto-fill
- POST /api/v1/nl/cost-prices/upload
"""
import uuid
import json
import time
import io
import csv
import logging
from datetime import datetime, date
from typing import Optional

from fastapi import APIRouter, Depends, Request, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, text
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.sql import func

from core.database import get_db
from core.security import decrypt_data
from core.tenant_auth import require_query_organization_access
from models.organization import WbApiKey
from models.reference_book import ReferenceBook
from schemas.reference import RefItem
from services.reference import (
    resolve_org_id, pfloat, pint,
    normalize_product_class, normalize_fulfillment,
    auto_calc_volume, resolve_entity_id,
)
from repositories.reference import fetch_reference, fetch_cost_prices

_log = logging.getLogger(__name__)
router = APIRouter(dependencies=[Depends(require_query_organization_access)])


# ============================================================================
# GET /api/v1/nl/reference — короткий формат
# ============================================================================
@router.get("/api/v1/nl/reference")
async def get_reference(org_id: str, target_date: Optional[str] = None, db: AsyncSession = Depends(get_db)):
    org_id = await resolve_org_id(org_id, db)
    return await fetch_reference(db, org_id, target_date)


# ============================================================================
# POST /api/v1/nl/reference — сохранить строку (RefItem)
# ============================================================================
@router.post("/api/v1/nl/reference")
async def save_reference(item: RefItem, org_id: str, db: AsyncSession = Depends(get_db)):
    """Сохранить строку справочного листа"""
    t_date = datetime.strptime(item.target_date, "%Y-%m-%d").date() if item.target_date else date.today()
    # entity_id lookup
    ent_q = await db.execute(text(
        "SELECT id FROM product_entities WHERE organization_id = :org AND nm_id = :nm LIMIT 1"
    ), {"org": org_id, "nm": item.nm_id})
    ent_row = ent_q.first()
    eid = ent_row[0] if ent_row else None
    ins = pg_insert(ReferenceBook).values(
        organization_id=org_id, nm_id=item.nm_id, vendor_code=item.vendor_code,
        valid_from=t_date, entity_id=eid,
        cost_price=item.cost_price,
        purchase_cost=item.purchase_price, packaging_cost=item.packaging_cost,
        logistics_cost=item.logistics_cost, other_costs=item.other_costs, notes=item.notes,
        product_class=item.product_class, brand=item.brand,
        tax_system=item.tax_system, tax_rate=item.tax_rate, vat_rate=item.vat_rate,
    )
    stmt = ins.on_conflict_do_update(
        constraint="reference_book_org_entity_vf_key",
        set_={
            "vendor_code": ins.excluded.vendor_code,
            "cost_price": ins.excluded.cost_price, "purchase_cost": ins.excluded.purchase_cost,
            "packaging_cost": ins.excluded.packaging_cost, "logistics_cost": ins.excluded.logistics_cost,
            "other_costs": ins.excluded.other_costs, "notes": ins.excluded.notes,
            "product_class": ins.excluded.product_class, "brand": ins.excluded.brand,
            "tax_system": ins.excluded.tax_system, "tax_rate": ins.excluded.tax_rate, "vat_rate": ins.excluded.vat_rate,
            "updated_at": func.now(),
        }
    )
    await db.execute(stmt)
    await db.commit()
    return {"status": "ok"}


# ============================================================================
# GET /api/v1/nl/fbs-warehouses — список складов FBS (кэш 24ч)
# ============================================================================
@router.get("/api/v1/nl/fbs-warehouses")
async def get_fbs_warehouses(org_id: str, db: AsyncSession = Depends(get_db)):
    """Список складов FBS + СЦ(СГТ) + КГТ+ из WB API (кэш 24ч)"""
    cache_key = "fbs_wh_v2_" + org_id
    cache_file = "/tmp/" + cache_key + ".json"
    try:
        with open(cache_file, "r") as f:
            cache = json.load(f)
            if time.time() - cache["ts"] < 86400:
                return cache["data"]
    except (FileNotFoundError, json.JSONDecodeError, KeyError):
        pass
    result = await db.execute(
        select(WbApiKey).where(WbApiKey.organization_id == org_id)
    )
    key_rec = result.scalar_one_or_none()
    if not key_rec or not key_rec.personal_token:
        return []
    token = decrypt_data(key_rec.personal_token)
    try:
        from services.wb_api.client import WBApiClient
        async with WBApiClient(token) as client:
            all_offices = await client.get_all_offices()
    except Exception as e:
        raise HTTPException(502, f"WB API error: {e}")
    type_names = {1: "Склад", 2: "СЦ", 3: "КГТ+"}
    merged = []
    seen_ids = set()
    for ct in [1, 2, 3]:
        items = [o for o in all_offices if o.get("cargoType") == ct]
        for o in items:
            if o["id"] not in seen_ids:
                seen_ids.add(o["id"])
                merged.append({
                    "id": o["id"],
                    "name": o["name"],
                    "address": o.get("address", ""),
                    "city": o.get("city", ""),
                    "type": type_names.get(ct, "?"),
                    "cargoType": ct,
                })
    # Фильтруем: оставляем только склады с ФБС-тарифами
    tarif_rows = await db.execute(text(
        "SELECT DISTINCT warehouse_name FROM wb_box_tariffs "
        "WHERE box_delivery_marketplace_base IS NOT NULL "
        "AND organization_id = :org"
    ), {"org": org_id})
    tarif_names = set(r[0] for r in tarif_rows.fetchall())
    if tarif_names:
        filtered = []
        for wh in merged:
            has_tarif = any(
                wh["name"] in tn or tn in wh["name"]
                for tn in tarif_names
            )
            if has_tarif:
                filtered.append(wh)
        merged = filtered
    with open(cache_file, "w") as f:
        json.dump({"ts": time.time(), "data": merged}, f, ensure_ascii=False)
    return merged


# ============================================================================
# GET /api/v1/nl/cost-prices — полный справочник с entity
# ============================================================================
@router.get("/api/v1/nl/cost-prices")
async def get_cost_prices(org_id: str, db: AsyncSession = Depends(get_db)):
    org_id = await resolve_org_id(org_id, db)
    return await fetch_cost_prices(db, org_id)


# ============================================================================
# POST /api/v1/nl/cost-prices — сохранить одну запись
# ============================================================================
@router.post("/api/v1/nl/cost-prices")
async def save_cost_price(data: dict, org_id: str, db: AsyncSession = Depends(get_db)):
    """Сохранить себестоимость (создать/обновить)"""
    nm_id = data.get("nm_id")
    cost = data.get("cost_price")
    valid_from = data.get("valid_from", date.today().isoformat())
    if isinstance(valid_from, str):
        valid_from = datetime.strptime(valid_from, "%Y-%m-%d").date()
    if not nm_id:
        raise HTTPException(400, "nm_id обязателен")
    entity_id = data.get("entity_id")
    if not entity_id:
        size_name = data.get("size_name", "")
        entity_id = await resolve_entity_id(db, org_id, nm_id, None, size_name)
    if not entity_id:
        raise HTTPException(400, "entity_id обязателен для сохранения справочника")
    await db.execute(text(_SAVE_COST_PRICE_SQL), _build_save_params(data, org_id, nm_id, entity_id, valid_from))
    await db.commit()
    return {"ok": True}


# ============================================================================
# POST /api/v1/nl/cost-prices/batch — batch-сохранение
# ============================================================================
@router.post("/api/v1/nl/cost-prices/batch")
async def save_cost_prices_batch(request: Request, org_id: str, db: AsyncSession = Depends(get_db)):
    """Batch-сохранение справочника — один запрос вместо N отдельных"""
    # Сбрасываем кэш юнит-экономики
    try:
        import redis as _rinv2
        _rinv2.from_url("redis://redis:6379/0").delete(f"ue_cache:{org_id}")
    except Exception:
        pass

    items = await request.json()
    saved = 0
    errors = 0
    for data in items:
        try:
            nm_id = data.get("nm_id")
            if not nm_id:
                errors += 1
                continue
            cost = data.get("cost_price")
            valid_from = data.get("valid_from", date.today().isoformat())
            if isinstance(valid_from, str):
                valid_from = datetime.strptime(valid_from, "%Y-%m-%d").date()
            entity_id = data.get("entity_id")
            if not entity_id:
                size_name = data.get("size_name", "")
                entity_id = await resolve_entity_id(db, org_id, nm_id, None, size_name)
            if not entity_id:
                errors += 1
                print(f"[batch] skip nm={nm_id}: entity_id not resolved")
                continue
            await db.execute(text(_SAVE_COST_PRICE_SQL), _build_save_params(data, org_id, nm_id, entity_id, valid_from))
            saved += 1
        except Exception as e:
            errors += 1
            print(f"[batch] error nm={data.get('nm_id')}: {e}")
    await db.commit()
    return {"ok": True, "saved": saved, "errors": errors}


# ============================================================================
# GET /api/v1/nl/commission-rate — комиссия МП по subject_id
# ============================================================================
@router.get("/api/v1/nl/commission-rate")
async def get_commission_rate(org_id: str, subject_id: int, model: str = "fbo", db: AsyncSession = Depends(get_db)):
    """Получить комиссию МП по subject_id и модели (fbo/fbs)"""
    result = await db.execute(text(
        "SELECT raw_response FROM raw_api_data "
        "WHERE api_method = 'tariffs_commission' "
        "ORDER BY target_date DESC LIMIT 1"
    ))
    row = result.first()
    if not row or not row[0]:
        return {"commission_pct": None, "source": "no_data"}
    cdata = row[0] if isinstance(row[0], dict) else json.loads(row[0])
    for item in cdata.get("report", []):
        if item.get("subjectID") == subject_id:
            if model == "fbs":
                pct = item.get("kgvpMarketplace")
            else:
                pct = item.get("paidStorageKgvp")
            return {"commission_pct": float(pct) if pct else None, "source": "api", "model": model}
    return {"commission_pct": None, "source": "subject_not_found"}


# ============================================================================
# POST /api/v1/nl/cost-prices/auto-fill — автозаполнение из wb_tariff_snapshot
# ============================================================================
@router.post("/api/v1/nl/cost-prices/auto-fill")
async def auto_fill_reference(org_id: str, db: AsyncSession = Depends(get_db)):
    """Автозаполнение справочника из wb_tariff_snapshot (только пустые поля)"""
    snap_result = await db.execute(text("""
        SELECT entity_id, nm_id, commission_pct, logistics_tariff, storage_tariff,
               price_retail, price_with_spp, ad_cost_fact, buyout_pct_fact
        FROM wb_tariff_snapshot
        WHERE organization_id = :org AND target_date = (
            SELECT MAX(target_date) FROM wb_tariff_snapshot WHERE organization_id = :org
        )
    """), {"org": org_id})
    snapshots = snap_result.all()
    if not snapshots:
        return {"ok": False, "error": "Нет данных в wb_tariff_snapshot. Запустите синхронизацию."}

    snap_by_entity = {}
    snap_by_nm = {}
    for s in snapshots:
        eid, nm_id = str(s[0]) if s[0] else None, s[1]
        data = {
            "commission_pct": float(s[2]) if s[2] else None,
            "logistics_tariff": float(s[3]) if s[3] else None,
            "storage_tariff": float(s[4]) if s[4] else None,
            "price_retail": float(s[5]) if s[5] else None,
            "price_with_spp": float(s[6]) if s[6] else None,
            "ad_cost_fact": float(s[7]) if s[7] else None,
            "buyout_pct_fact": float(s[8]) if s[8] else None,
        }
        if eid:
            snap_by_entity[eid] = data
        if nm_id:
            snap_by_nm[nm_id] = data

    subj_result = await db.execute(text(
        "SELECT id::text, subject_id, subject_name FROM product_entities WHERE organization_id = :org AND subject_id IS NOT NULL"
    ), {"org": org_id})
    subj_map = {}
    for sr in subj_result.all():
        subj_map[sr[0]] = {"subject_id": sr[1], "subject_name": sr[2]}

    vc_result = await db.execute(text(
        "SELECT DISTINCT nm_id, vendor_code FROM product_entities WHERE organization_id = :org AND vendor_code IS NOT NULL AND vendor_code != ''"
    ), {"org": org_id})
    vendor_code_by_nm = {}
    for vr in vc_result.all():
        vendor_code_by_nm[vr[0]] = vr[1]

    for eid_str, snap in snap_by_entity.items():
        if eid_str in subj_map:
            snap["subject_id"] = subj_map[eid_str]["subject_id"]
            snap["subject_name"] = subj_map[eid_str]["subject_name"]

    ref_result = await db.execute(text("""
        SELECT id, entity_id, nm_id,
               mp_base_pct, logistics_cost, storage_pct,
               price_before_spp_plan, buyout_niche_pct, ad_plan_rub, vendor_code,
               subject_id, subject_name
        FROM reference_book
        WHERE organization_id = :org AND (valid_to IS NULL OR valid_to >= CURRENT_DATE)
    """), {"org": org_id})
    refs = ref_result.all()

    stats = {"updated": 0, "skipped": 0, "fields_filled": {}}
    field_map = {
        "mp_base_pct": "commission_pct",
        "logistics_cost": "logistics_tariff",
        "storage_pct": "storage_tariff",
        "price_before_spp_plan": "price_retail",
        "buyout_niche_pct": "buyout_pct_fact",
        "ad_plan_rub": "ad_cost_fact",
        "subject_id": "subject_id",
        "subject_name": "subject_name",
    }

    for r in refs:
        rid = str(r[0])
        eid = str(r[1]) if r[1] else None
        nm_id = r[2]
        current = {
            "mp_base_pct": r[3], "logistics_cost": r[4], "storage_pct": r[5],
            "price_before_spp_plan": r[6], "buyout_niche_pct": r[7], "ad_plan_rub": r[8],
            "subject_id": r[9], "subject_name": r[10], "vendor_code": r[11],
        }
        snap = None
        if eid and eid in snap_by_entity:
            snap = snap_by_entity[eid]
        elif nm_id and nm_id in snap_by_nm:
            snap = snap_by_nm[nm_id]

        updates = {}
        vc = vendor_code_by_nm.get(nm_id)
        if vc and (not current.get("vendor_code") or current["vendor_code"] == ""):
            updates["vendor_code"] = vc

        if not snap:
            if updates:
                set_clauses = ", ".join(f"{k} = :{k}" for k in updates)
                updates["rid"] = rid
                await db.execute(text(f"UPDATE reference_book SET {set_clauses} WHERE id = :rid"), updates)
                stats["updated"] += 1
            else:
                stats["skipped"] += 1
            continue

        for ref_field, snap_field in field_map.items():
            snap_val = snap.get(snap_field)
            cur_val = current.get(ref_field)
            if snap_val is not None and (cur_val is None or cur_val == 0):
                updates[ref_field] = snap_val
                stats["fields_filled"][ref_field] = stats["fields_filled"].get(ref_field, 0) + 1

        if not updates:
            stats["skipped"] += 1
            continue

        set_clauses = ", ".join(f"{k} = :{k}" for k in updates)
        updates["rid"] = rid
        await db.execute(text(f"UPDATE reference_book SET {set_clauses} WHERE id = :rid"), updates)
        stats["updated"] += 1

    # Шаг 2: создать записи для nm_id без справочника
    existing_nms = set()
    for r in refs:
        if r[2]:
            existing_nms.add(r[2])
    all_entities = await db.execute(text(
        "SELECT pe.id, pe.nm_id, pe.size_name, pe.brand, pe.subject_id, pe.subject_name, pe.vendor_code FROM product_entities pe WHERE pe.organization_id = :org"
    ), {"org": org_id})
    created_count = 0
    seen_nms = set()
    for ent in all_entities.all():
        nm_id = ent[1]
        if nm_id in existing_nms or nm_id in seen_nms:
            continue
        seen_nms.add(nm_id)
        eid = str(ent[0])
        snap = snap_by_nm.get(nm_id)
        ins = pg_insert(ReferenceBook)
        vals = {
            "id": str(uuid.uuid4()),
            "organization_id": org_id,
            "nm_id": nm_id,
            "entity_id": eid,
            "size_name": ent[2] or "",
            "brand": ent[3] or "",
            "subject_id": ent[4],
            "subject_name": ent[5] or "",
            "vendor_code": ent[6] or "",
            "valid_from": date.today(),
            "mp_base_pct": float(snap["commission_pct"]) if snap and snap.get("commission_pct") else None,
            "logistics_cost": float(snap["logistics_tariff"]) if snap and snap.get("logistics_tariff") else None,
            "storage_pct": float(snap["storage_tariff"]) if snap and snap.get("storage_tariff") else None,
            "price_before_spp_plan": float(snap["price_retail"]) if snap and snap.get("price_retail") else None,
            "price_before_spp_change": float(snap["price_with_spp"]) if snap and snap.get("price_with_spp") else None,
            "buyout_niche_pct": float(snap["buyout_pct_fact"]) if snap and snap.get("buyout_pct_fact") else None,
        }
        stmt = ins.values(**vals).on_conflict_do_nothing(constraint="reference_book_org_nm_vf_key")
        try:
            await db.execute(stmt)
            created_count += 1
        except Exception:
            await db.rollback()

    stats["created"] = created_count
    await db.commit()
    _log.info(f"[auto_fill] org={org_id}: {stats}")
    return {"ok": True, "stats": stats}


# ============================================================================
# POST /api/v1/nl/cost-prices/upload — загрузка из Excel/CSV
# ============================================================================
@router.post("/api/v1/nl/cost-prices/upload")
async def upload_cost_prices_excel(org_id: str, request: Request, db: AsyncSession = Depends(get_db)):
    """Загрузка справочника из Excel/CSV — колонки строго по Tabulator"""
    body = await request.body()
    filename = request.headers.get("x-filename", "upload.csv")
    rows = []
    if filename.endswith(".csv"):
        reader = csv.DictReader(io.StringIO(body.decode("utf-8-sig")), delimiter=";")
        rows = list(reader)
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(body))
            ws = wb.active
            headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        except ImportError:
            raise HTTPException(400, "xlsx не поддерживается")

    def pf(row, *keys):
        for k in keys:
            v = row.get(k)
            if v is not None and str(v).strip():
                try:
                    return float(str(v).replace(',', '.'))
                except:
                    pass
        return None

    def ps(row, *keys):
        for k in keys:
            v = row.get(k)
            if v and str(v).strip():
                return str(v).strip()
        return None

    def pd(row, *keys):
        s = ps(row, *keys)
        if not s:
            return None
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except:
            return None

    def pstr(row, *keys):
        for k in keys:
            v = row.get(k)
            if v is not None and str(v).strip():
                return str(v).strip()
        return None

    updated = 0
    skipped = 0
    warnings = []

    for idx, row in enumerate(rows, start=2):
        nm_raw = row.get("Арт WB") or row.get("nm_id")
        if not nm_raw or not str(nm_raw).strip():
            skipped += 1
            continue
        try:
            nm = int(str(nm_raw).replace(',', '').strip())
        except ValueError:
            warnings.append(f"Строка {idx}: Арт WB '{nm_raw}' не число — пропущено")
            skipped += 1
            continue

        sz_val = ps(row, "Размер", "size_name")
        eid = None
        if sz_val:
            ent_q = await db.execute(text(
                "SELECT pe.id FROM product_entities pe "
                "WHERE pe.organization_id = :org AND pe.nm_id = :nm AND pe.size_name = :sz LIMIT 1"
            ), {"org": org_id, "nm": nm, "sz": sz_val})
            ent_row = ent_q.first()
            eid = str(ent_row[0]) if ent_row else None
            if not eid:
                warnings.append(f"Строка {idx} (Арт WB {nm}): размер '{sz_val}' не найден — пропущено")
                skipped += 1
                continue
        else:
            ent_q = await db.execute(text(
                "SELECT pe.id, pe.size_name FROM product_entities pe "
                "WHERE pe.organization_id = :org AND pe.nm_id = :nm"
            ), {"org": org_id, "nm": nm})
            ent_rows = ent_q.all()
            if len(ent_rows) == 1:
                eid = str(ent_rows[0][0])
                sz_val = ent_rows[0][1] or ""
            elif len(ent_rows) > 1:
                warnings.append(f"Строка {idx} (Арт WB {nm}): несколько размеров, укажите колонку Размер — пропущено")
                skipped += 1
                continue
            else:
                warnings.append(f"Строка {idx} (Арт WB {nm}): товар не найден — пропущено")
                skipped += 1
                continue

        ffm_raw = pstr(row, "Отгрузка", "ФБО/ФБС", "fulfillment_model") or "fbo"
        ffm = normalize_fulfillment(ffm_raw)
        fbs_wh = ps(row, "Склад FBS", "Склад отгрузки FBS", "fbs_warehouse") or ""
        if ffm != "fbs":
            if fbs_wh:
                warnings.append(f"Строка {idx} (Арт WB {nm}): ФБО но указан склад FBS — очищено")
            fbs_wh = ""

        plen = pf(row, "План длина", "Длина", "plan_length")
        pwid = pf(row, "План ширина", "Ширина", "plan_width")
        phei = pf(row, "План высота", "Высота", "plan_height")
        pvol = auto_calc_volume(plen, pwid, phei)

        tax_override = pf(row, "Налог %", "tax_rate")
        vat_rate_raw = pstr(row, "НДС от дохода", "vat_rate")
        vat_rate = 0
        if vat_rate_raw:
            try:
                vat_rate = float(str(vat_rate_raw).replace('%', '').replace(',', '.'))
            except:
                pass

        params = {
            "org": org_id, "nm": nm,
            "bc": ps(row, "Баркод", "barcode"),
            "vc": ps(row, "Арт продавца", "vendor_code"),
            "sz": sz_val or "", "eid": eid,
            "cp": pf(row, "Себестоимость", "Себестоимость ₽", "cost_price"),
            "ec": pf(row, "Доп расходы", "Доп расходы ₽", "extra_costs"),
            "vat": None,
            "minp": pf(row, "Мин. цена", "Мин цена", "min_price"),
            "mpc": pf(row, "Корр. комиссии %", "Корр. % МП", "mp_correction_pct"),
            "ffm": ffm,
            "bnp": pf(row, "% выкупа", "% выкупа по категории", "buyout_niche_pct"),
            "rrc": pf(row, "РРЦ", "rrc_price"),
            "adpr": pf(row, "Рекл. расходы %", "Реклама план", "ad_plan_rub"),
            "pcls": normalize_product_class(ps(row, "Класс", "Класс товара", "product_class")),
            "brand": ps(row, "Бренд", "brand"),
            "pstatus": ps(row, "Статус товара", "product_status"),
            "trate": tax_override,
            "vrate": vat_rate,
            "sjan": pf(row, "Сезон янв", "season_jan"), "sfeb": pf(row, "Сезон фев", "season_feb"),
            "smar": pf(row, "Сезон мар", "season_mar"), "sapr": pf(row, "Сезон апр", "season_apr"),
            "smay": pf(row, "Сезон май", "season_may"), "sjun": pf(row, "Сезон июн", "season_jun"),
            "sjul": pf(row, "Сезон июл", "season_jul"), "saug": pf(row, "Сезон авг", "season_aug"),
            "ssep": pf(row, "Сезон сен", "season_sep"), "soct": pf(row, "Сезон окт", "season_oct"),
            "snov": pf(row, "Сезон ноя", "season_nov"), "sdec": pf(row, "Сезон дек", "season_dec"),
            "plen": plen, "pwid": pwid, "phei": phei, "pvol": pvol,
            "pwgt": pf(row, "План вес", "Вес, гр", "plan_weight"),
            "tq1": ps(row, "ТОП запрос 1", "top_query_1"),
            "tq2": ps(row, "ТОП запрос 2", "top_query_2"),
            "tq3": ps(row, "ТОП запрос 3", "top_query_3"),
            "fbsw": fbs_wh,
            "subn": ps(row, "Категория", "subject_name"),
            "sdays": pf(row, "Скорость достав. дн", "Скорость доставки, дн", "supply_days"),
            "minbat": pf(row, "Мин партия", "Минимальная партия FBO", "min_batch_fbo"),
            "vfrom": pd(row, "Дата начала", "valid_from"),
        }

        await db.execute(text(_UPLOAD_SQL), params)
        updated += 1

    await db.commit()
    return {"updated": updated, "total": len(rows), "skipped": skipped, "warnings": warnings[:20]}


# ============================================================================
# SQL-константы (общие INSERT для single и batch)
# ============================================================================
_SAVE_COST_PRICE_SQL = (
    "INSERT INTO reference_book (organization_id, nm_id, barcode, vendor_code, size_name, entity_id, "
    "subject_id, subject_name, "
    "cost_price, purchase_cost, logistics_cost, packaging_cost, other_costs, extra_costs, vat, min_price, "
    "mp_base_pct, mp_correction_pct, fulfillment_model, storage_pct, "
    "buyout_niche_pct, "
    "price_before_spp_plan, price_before_spp_change, change_date, "
    "wb_club_discount_pct, ad_plan_rub, supply_days, min_batch_fbo, "
    "product_status, product_class, brand, tax_system, tax_rate, "
    "season_jan, season_feb, season_mar, season_apr, season_may, season_jun, season_jul, season_aug, season_sep, season_oct, season_nov, season_dec, "
    "plan_length, plan_width, plan_height, plan_volume, plan_weight, "
    "delivery_days_to_seller, delivery_days_to_mp, "
    "top_query_1, top_query_2, top_query_3, "
    "shipment_method, fbs_warehouse, rrc_price, vat_rate, "
    "valid_from, source, notes) "
    "VALUES (:org, :nm, :bc, :vc, :sz, :eid, "
    ":subid, :subn, "
    ":cp, :pc, :lc, :pk, :oc, :ec, :vat, :minp, "
    ":mpb, :mpc, :ffm, :stp, "
    ":bnp, "
    ":pspp, :psppc, :cdate, "
    ":wbcd, :adpr, :sdays, :minb, "
    ":pstatus, :pcls, :brand, :tsys, :tr, "
    ":sjan, :sfeb, :smar, :sapr, :smay, :sjun, :sjul, :saug, :ssep, :soct, :snov, :sdec, "
    ":pl, :pw, :ph, :pv, :pwg, "
    ":dds, :ddm, "
    ":tq1, :tq2, :tq3, "
    ":sm, :fw, :rrc, :vr, "
    ":vf, :src, :notes) "
    "ON CONFLICT (organization_id, nm_id, entity_id, valid_from) DO UPDATE SET "
    "barcode = COALESCE(EXCLUDED.barcode, reference_book.barcode), "
    "vendor_code = COALESCE(EXCLUDED.vendor_code, reference_book.vendor_code), "
    "cost_price = COALESCE(EXCLUDED.cost_price, reference_book.cost_price), "
    "purchase_cost = COALESCE(EXCLUDED.purchase_cost, reference_book.purchase_cost), "
    "logistics_cost = COALESCE(EXCLUDED.logistics_cost, reference_book.logistics_cost), "
    "packaging_cost = COALESCE(EXCLUDED.packaging_cost, reference_book.packaging_cost), "
    "other_costs = COALESCE(EXCLUDED.other_costs, reference_book.other_costs), "
    "extra_costs = COALESCE(EXCLUDED.extra_costs, reference_book.extra_costs), "
    "vat = COALESCE(EXCLUDED.vat, reference_book.vat), "
    "min_price = COALESCE(EXCLUDED.min_price, reference_book.min_price), "
    "mp_base_pct = COALESCE(EXCLUDED.mp_base_pct, reference_book.mp_base_pct), "
    "mp_correction_pct = COALESCE(EXCLUDED.mp_correction_pct, reference_book.mp_correction_pct), "
    "fulfillment_model = COALESCE(EXCLUDED.fulfillment_model, reference_book.fulfillment_model), "
    "storage_pct = COALESCE(EXCLUDED.storage_pct, reference_book.storage_pct), "
    "buyout_niche_pct = COALESCE(EXCLUDED.buyout_niche_pct, reference_book.buyout_niche_pct), "
    "price_before_spp_plan = COALESCE(EXCLUDED.price_before_spp_plan, reference_book.price_before_spp_plan), "
    "price_before_spp_change = COALESCE(EXCLUDED.price_before_spp_change, reference_book.price_before_spp_change), "
    "change_date = COALESCE(EXCLUDED.change_date, reference_book.change_date), "
    "wb_club_discount_pct = COALESCE(EXCLUDED.wb_club_discount_pct, reference_book.wb_club_discount_pct), "
    "ad_plan_rub = COALESCE(EXCLUDED.ad_plan_rub, reference_book.ad_plan_rub), "
    "supply_days = COALESCE(EXCLUDED.supply_days, reference_book.supply_days), "
    "min_batch_fbo = COALESCE(EXCLUDED.min_batch_fbo, reference_book.min_batch_fbo), "
    "product_status = COALESCE(EXCLUDED.product_status, reference_book.product_status), "
    "product_class = COALESCE(EXCLUDED.product_class, reference_book.product_class), "
    "brand = COALESCE(EXCLUDED.brand, reference_book.brand), "
    "tax_system = COALESCE(EXCLUDED.tax_system, reference_book.tax_system), "
    "tax_rate = COALESCE(EXCLUDED.tax_rate, reference_book.tax_rate), "
    "season_jan = COALESCE(EXCLUDED.season_jan, reference_book.season_jan), "
    "season_feb = COALESCE(EXCLUDED.season_feb, reference_book.season_feb), "
    "season_mar = COALESCE(EXCLUDED.season_mar, reference_book.season_mar), "
    "season_apr = COALESCE(EXCLUDED.season_apr, reference_book.season_apr), "
    "season_may = COALESCE(EXCLUDED.season_may, reference_book.season_may), "
    "season_jun = COALESCE(EXCLUDED.season_jun, reference_book.season_jun), "
    "season_jul = COALESCE(EXCLUDED.season_jul, reference_book.season_jul), "
    "season_aug = COALESCE(EXCLUDED.season_aug, reference_book.season_aug), "
    "season_sep = COALESCE(EXCLUDED.season_sep, reference_book.season_sep), "
    "season_oct = COALESCE(EXCLUDED.season_oct, reference_book.season_oct), "
    "season_nov = COALESCE(EXCLUDED.season_nov, reference_book.season_nov), "
    "season_dec = COALESCE(EXCLUDED.season_dec, reference_book.season_dec), "
    "plan_length = COALESCE(EXCLUDED.plan_length, reference_book.plan_length), "
    "plan_width = COALESCE(EXCLUDED.plan_width, reference_book.plan_width), "
    "plan_height = COALESCE(EXCLUDED.plan_height, reference_book.plan_height), "
    "plan_volume = COALESCE(EXCLUDED.plan_volume, reference_book.plan_volume), "
    "plan_weight = COALESCE(EXCLUDED.plan_weight, reference_book.plan_weight), "
    "delivery_days_to_seller = COALESCE(EXCLUDED.delivery_days_to_seller, reference_book.delivery_days_to_seller), "
    "delivery_days_to_mp = COALESCE(EXCLUDED.delivery_days_to_mp, reference_book.delivery_days_to_mp), "
    "top_query_1 = COALESCE(EXCLUDED.top_query_1, reference_book.top_query_1), "
    "top_query_2 = COALESCE(EXCLUDED.top_query_2, reference_book.top_query_2), "
    "top_query_3 = COALESCE(EXCLUDED.top_query_3, reference_book.top_query_3), "
    "shipment_method = COALESCE(EXCLUDED.shipment_method, reference_book.shipment_method), "
    "fbs_warehouse = COALESCE(EXCLUDED.fbs_warehouse, reference_book.fbs_warehouse), "
    "rrc_price = COALESCE(EXCLUDED.rrc_price, reference_book.rrc_price), "
    "vat_rate = COALESCE(EXCLUDED.vat_rate, reference_book.vat_rate), "
    "subject_id = COALESCE(EXCLUDED.subject_id, reference_book.subject_id), "
    "subject_name = COALESCE(EXCLUDED.subject_name, reference_book.subject_name), "
    "source = EXCLUDED.source, notes = EXCLUDED.notes"
)


def _build_save_params(data: dict, org_id: str, nm_id: int, entity_id: str, valid_from) -> dict:
    """Общие параметры для INSERT в single и batch сохранении"""
    return {
        "org": org_id, "nm": nm_id,
        "bc": data.get("barcode"), "vc": data.get("vendor_code"),
        "sz": data.get("size_name"), "eid": entity_id,
        "subid": int(data["subject_id"]) if data.get("subject_id") is not None and str(data["subject_id"]).strip().lstrip("-").isdigit() else None,
        "subn": data.get("subject_name"),
        "cp": data.get("cost_price"),
        "pc": pfloat(data.get("purchase_cost")),
        "lc": pfloat(data.get("logistics_cost")),
        "pk": pfloat(data.get("packaging_cost")),
        "oc": pfloat(data.get("other_costs")),
        "ec": pfloat(data.get("extra_costs")),
        "vat": pfloat(data.get("vat")) or 0,
        "minp": pfloat(data.get("min_price")),
        "mpb": pfloat(data.get("mp_base_pct")),
        "mpc": pfloat(data.get("mp_correction_pct")),
        "ffm": data.get("fulfillment_model", "fbo"),
        "stp": pfloat(data.get("storage_pct")),
        "bnp": pfloat(data.get("buyout_niche_pct")),
        "pspp": pfloat(data.get("price_before_spp_plan")),
        "psppc": pfloat(data.get("price_before_spp_change")),
        "cdate": date.today(),
        "wbcd": pfloat(data.get("wb_club_discount_pct")),
        "adpr": (
            min(99, max(0, pfloat(data.get("ad_plan_rub"))))
            if pfloat(data.get("ad_plan_rub")) is not None else None
        ),
        "sdays": pint(data.get("supply_days")),
        "minb": pint(data.get("min_batch_fbo")),
        "pstatus": data.get("product_status"),
        "pcls": data.get("product_class"),
        "brand": data.get("brand"),
        "tsys": data.get("tax_system"),
        "tr": pfloat(data.get("tax_rate")),
        "sjan": pfloat(data.get("season_jan")), "sfeb": pfloat(data.get("season_feb")),
        "smar": pfloat(data.get("season_mar")), "sapr": pfloat(data.get("season_apr")),
        "smay": pfloat(data.get("season_may")), "sjun": pfloat(data.get("season_jun")),
        "sjul": pfloat(data.get("season_jul")), "saug": pfloat(data.get("season_aug")),
        "ssep": pfloat(data.get("season_sep")), "soct": pfloat(data.get("season_oct")),
        "snov": pfloat(data.get("season_nov")), "sdec": pfloat(data.get("season_dec")),
        "pl": pfloat(data.get("plan_length")), "pw": pfloat(data.get("plan_width")),
        "ph": pfloat(data.get("plan_height")), "pv": pfloat(data.get("plan_volume")),
        "pwg": pfloat(data.get("plan_weight")),
        "dds": pint(data.get("delivery_days_to_seller")),
        "ddm": pint(data.get("delivery_days_to_mp")),
        "tq1": data.get("top_query_1"), "tq2": data.get("top_query_2"), "tq3": data.get("top_query_3"),
        "sm": data.get("shipment_method"), "fw": data.get("fbs_warehouse"),
        "rrc": pfloat(data.get("rrc_price")),
        "vr": pfloat(data.get("vat_rate")) or 0,
        "vf": valid_from,
        "src": data.get("source", "manual"),
        "notes": data.get("notes"),
    }


_UPLOAD_SQL = (
    "INSERT INTO reference_book ("
    "organization_id, nm_id, barcode, vendor_code, size_name, entity_id, "
    "subject_name, "
    "cost_price, extra_costs, vat, min_price, "
    "mp_correction_pct, fulfillment_model, buyout_niche_pct, "
    "rrc_price, ad_plan_rub, "
    "product_class, brand, product_status, "
    "tax_rate, vat_rate, "
    "season_jan, season_feb, season_mar, season_apr, season_may, season_jun, "
    "season_jul, season_aug, season_sep, season_oct, season_nov, season_dec, "
    "plan_length, plan_width, plan_height, plan_volume, plan_weight, "
    "top_query_1, top_query_2, top_query_3, "
    "fbs_warehouse, "
    "supply_days, min_batch_fbo, "
    "valid_from, change_date, source) "
    "VALUES ("
    ":org, :nm, :bc, :vc, :sz, :eid, "
    ":subn, "
    ":cp, :ec, :vat, :minp, "
    ":mpc, :ffm, :bnp, "
    ":rrc, :adpr, "
    ":pcls, :brand, :pstatus, "
    ":trate, :vrate, "
    ":sjan, :sfeb, :smar, :sapr, :smay, :sjun, "
    ":sjul, :saug, :ssep, :soct, :snov, :sdec, "
    ":plen, :pwid, :phei, :pvol, :pwgt, "
    ":tq1, :tq2, :tq3, "
    ":fbsw, "
    ":sdays, :minbat, "
    ":vfrom, CURRENT_DATE, 'excel') "
    "ON CONFLICT (organization_id, nm_id, entity_id, valid_from) DO UPDATE SET "
    "barcode = COALESCE(EXCLUDED.barcode, reference_book.barcode), "
    "vendor_code = COALESCE(EXCLUDED.vendor_code, reference_book.vendor_code), "
    "cost_price = COALESCE(EXCLUDED.cost_price, reference_book.cost_price), "
    "extra_costs = COALESCE(EXCLUDED.extra_costs, reference_book.extra_costs), "
    "min_price = COALESCE(EXCLUDED.min_price, reference_book.min_price), "
    "mp_correction_pct = COALESCE(EXCLUDED.mp_correction_pct, reference_book.mp_correction_pct), "
    "fulfillment_model = COALESCE(EXCLUDED.fulfillment_model, reference_book.fulfillment_model), "
    "buyout_niche_pct = COALESCE(EXCLUDED.buyout_niche_pct, reference_book.buyout_niche_pct), "
    "rrc_price = COALESCE(EXCLUDED.rrc_price, reference_book.rrc_price), "
    "ad_plan_rub = COALESCE(EXCLUDED.ad_plan_rub, reference_book.ad_plan_rub), "
    "product_class = COALESCE(EXCLUDED.product_class, reference_book.product_class), "
    "brand = COALESCE(EXCLUDED.brand, reference_book.brand), "
    "product_status = COALESCE(EXCLUDED.product_status, reference_book.product_status), "
    "tax_rate = COALESCE(EXCLUDED.tax_rate, reference_book.tax_rate), "
    "vat_rate = COALESCE(EXCLUDED.vat_rate, reference_book.vat_rate), "
    "season_jan = COALESCE(EXCLUDED.season_jan, reference_book.season_jan), "
    "season_feb = COALESCE(EXCLUDED.season_feb, reference_book.season_feb), "
    "season_mar = COALESCE(EXCLUDED.season_mar, reference_book.season_mar), "
    "season_apr = COALESCE(EXCLUDED.season_apr, reference_book.season_apr), "
    "season_may = COALESCE(EXCLUDED.season_may, reference_book.season_may), "
    "season_jun = COALESCE(EXCLUDED.season_jun, reference_book.season_jun), "
    "season_jul = COALESCE(EXCLUDED.season_jul, reference_book.season_jul), "
    "season_aug = COALESCE(EXCLUDED.season_aug, reference_book.season_aug), "
    "season_sep = COALESCE(EXCLUDED.season_sep, reference_book.season_sep), "
    "season_oct = COALESCE(EXCLUDED.season_oct, reference_book.season_oct), "
    "season_nov = COALESCE(EXCLUDED.season_nov, reference_book.season_nov), "
    "season_dec = COALESCE(EXCLUDED.season_dec, reference_book.season_dec), "
    "plan_length = COALESCE(EXCLUDED.plan_length, reference_book.plan_length), "
    "plan_width = COALESCE(EXCLUDED.plan_width, reference_book.plan_width), "
    "plan_height = COALESCE(EXCLUDED.plan_height, reference_book.plan_height), "
    "plan_volume = COALESCE(EXCLUDED.plan_volume, reference_book.plan_volume), "
    "plan_weight = COALESCE(EXCLUDED.plan_weight, reference_book.plan_weight), "
    "top_query_1 = COALESCE(EXCLUDED.top_query_1, reference_book.top_query_1), "
    "top_query_2 = COALESCE(EXCLUDED.top_query_2, reference_book.top_query_2), "
    "top_query_3 = COALESCE(EXCLUDED.top_query_3, reference_book.top_query_3), "
    "fbs_warehouse = CASE WHEN :ffm = 'fbs' THEN COALESCE(EXCLUDED.fbs_warehouse, reference_book.fbs_warehouse) ELSE '' END, "
    "subject_name = COALESCE(EXCLUDED.subject_name, reference_book.subject_name), "
    "supply_days = COALESCE(EXCLUDED.supply_days, reference_book.supply_days), "
    "min_batch_fbo = COALESCE(EXCLUDED.min_batch_fbo, reference_book.min_batch_fbo), "
    "change_date = CURRENT_DATE"
)
