import io
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, File, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from core.database import get_db
from core.tenant_auth import require_query_organization_access
from models.promotion import WbPromotion, WbPromotionProduct, WbPromotionSnapshot


router = APIRouter(
    tags=["nl"],
    dependencies=[Depends(require_query_organization_access)],
)


@router.get("/api/v1/nl/promotions/summary")
async def get_promotions_summary(
    org_id: str,
    snapshot_date: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    """Возвращает сводку по акциям: % товаров в акциях, разбивка по акциям."""
    params = {"org": org_id}
    
    # Сначала определяем latest snapshot date если не задан
    if not snapshot_date:
        latest_date_query = text(
            "SELECT MAX(snapshot_date) FROM wb_promotion_snapshots WHERE organization_id = :org"
        )
        latest_result = await db.execute(latest_date_query, params)
        latest_date = latest_result.scalar()
        if not latest_date:
            # Нет снимков — возвращаем нули
            return {
                "total_products": 0,
                "in_promotion": 0,
                "in_promotion_pct": 0.0,
                "by_promotion": [],
            }
        snapshot_date = latest_date
    
    params["date"] = snapshot_date
    
    # total_products — только товары, доступные к покупке на витрине WB.
    total_query = text(
        """
        SELECT COUNT(DISTINCT nm_id) as total
        FROM wb_promotion_snapshots
        WHERE organization_id = :org AND snapshot_date = :date
          AND available_to_buy = true
        """
    )
    total_result = await db.execute(total_query, params)
    total_products = total_result.scalar() or 0

    in_promo_query = text(
        """
        SELECT COUNT(DISTINCT nm_id) as count
        FROM wb_promotion_snapshots
        WHERE organization_id = :org
          AND snapshot_date = :date
          AND available_to_buy = true
          AND in_any_promo = true
        """
    )
    in_promo_result = await db.execute(in_promo_query, params)
    in_promotion = in_promo_result.scalar() or 0
    
    # in_promotion_pct
    in_promotion_pct = (
        round(in_promotion * 100.0 / total_products, 1) if total_products > 0 else 0.0
    )
    
    # by_promotion — разбивка по каждой акции
    # Для каждой акции из wb_promotions считаем сколько товаров в ней есть в snapshot
    by_promo_query = text(
        """
        SELECT
            wp.id,
            wp.promotion_id,
            wp.title,
            COUNT(DISTINCT snp.nm_id) as count
        FROM wb_promotions wp
        LEFT JOIN LATERAL (
            SELECT snp2.nm_id
            FROM wb_promotion_snapshots snp2
            WHERE snp2.organization_id = :org
              AND snp2.snapshot_date = :date
              AND snp2.promotions IS NOT NULL
              AND snp2.promotions::text != '[]'
              AND snp2.promotions::text != 'null'
              AND EXISTS (
                  SELECT 1 FROM jsonb_array_elements(snp2.promotions) p
                  WHERE (p->>'id')::text = wp.promotion_id::text
              )
        ) snp ON true
        WHERE wp.organization_id = :org
        GROUP BY wp.id, wp.promotion_id, wp.title
        ORDER BY count DESC
        """
    )
    by_promo_result = await db.execute(by_promo_query, params)
    
    by_promotion = []
    for row in by_promo_result.all():
        count = row.count or 0
        pct = round(count * 100.0 / total_products, 1) if total_products > 0 else 0.0
        by_promotion.append({
            "promotion_id": row.promotion_id,
            "title": row.title,
            "count": count,
            "pct": pct,
        })
    
    return {
        "total_products": total_products,
        "in_promotion": in_promotion,
        "in_promotion_pct": in_promotion_pct,
        "by_promotion": by_promotion,
    }


@router.get("/api/v1/nl/promotions")
async def get_promotions(
    org_id: str,
    is_active: Optional[bool] = None,
    db: AsyncSession = Depends(get_db),
):
    """Return WB promotions for an organization with participating counts."""
    # Базовый запрос с LATERAL JOIN для подсчёта участников из wb_promotion_products
    query = text(
        """
        SELECT
            wp.id,
            wp.promotion_id,
            wp.title,
            wp.promo_type,
            wp.start_date,
            wp.end_date,
            wp.max_price,
            wp.min_discount,
            wp.has_boost,
            wp.boost_value,
            wp.is_active,
            wp.importance,
            wp.source,
            COALESCE(pp_count.count, 0) as participating_count
        FROM wb_promotions wp
        LEFT JOIN LATERAL (
            SELECT COUNT(DISTINCT nm_id) as count
            FROM wb_promotion_products pp
            WHERE pp.organization_id = :org
              AND pp.wb_promotion_ext_id = wp.promotion_id
              AND pp.in_action = true
        ) pp_count ON true
        WHERE wp.organization_id = :org
        """
    )
    
    params = {"org": org_id}
    if is_active is not None:
        query = text(str(query) + " AND wp.is_active = :is_active")
        params["is_active"] = is_active
    
    query = text(str(query) + " ORDER BY wp.start_date DESC")
    
    result = await db.execute(query, params)
    
    items = []
    for row in result.all():
        items.append({
            "id": str(row.id),
            "promotion_id": row.promotion_id,
            "title": row.title,
            "promo_type": row.promo_type,
            "start_date": row.start_date.isoformat() if row.start_date else None,
            "end_date": row.end_date.isoformat() if row.end_date else None,
            "max_price": float(row.max_price) if row.max_price else None,
            "min_discount": row.min_discount,
            "has_boost": row.has_boost,
            "boost_value": float(row.boost_value) if row.boost_value else None,
            "is_active": row.is_active,
            "importance": row.importance,
            "source": row.source,
            "participating_count": row.participating_count,
        })
    
    return items


@router.get("/api/v1/nl/promotions/products")
async def get_promotion_products(
    org_id: str,
    promotion_id: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    """Build promotion product rows for the Tabulator frontend."""
    params = {"org": org_id}
    promo_product_filter = ""
    promo_where = ""
    if promotion_id:
        promo_product_filter = " AND pp2.wb_promotion_ext_id = :promo_id"
        promo_where = " AND pp.id IS NOT NULL"
        params["promo_id"] = int(promotion_id)

    query = text(
        """
        WITH latest_snapshot AS (
            SELECT MAX(snapshot_date) AS snapshot_date
            FROM wb_promotion_snapshots
            WHERE organization_id = :org
        )
        SELECT
            pp.id,
            pp.wb_promotion_ext_id,
            pe.nm_id,
            pp.in_action,
            pp.auto_matched,
            pp.current_price,
            pp.required_price,
            pp.price_in_promo,
            pp.profit_in_promo,
            pp.margin_delta,
            pp.plan,
            pp.status_text,
            pe.id as entity_id,
            pp.promotion_id_col,
            pe.vendor_code,
            pe.product_name,
            pe.photo_main as photo,
            pe.size_name,
            pe.brand,
            pe.subject_name,
            wp.title as promo_title,
            wp.promo_type,
            wp.start_date as promo_start,
            wp.end_date as promo_end,
            wp.importance as promo_importance,
            COALESCE(ts.price, pp.current_price) as price_before_spp,
            ts.stock_qty,
            rb.cost_price,
            snp.snapshot_date,
            snp.promotions as snapshot_promotions_raw,
            snp.sale_conditions as sale_conditions_raw,
            snp.available_qty,
            snp.available_to_buy,
            snp.regular_in_promo,
            snp.auto_in_promo,
            snp.in_any_promo,
            snp.regular_promotion_ids,
            snp.auto_promotion_ids,
            snp.price_basic,
            snp.price_product
        FROM product_entities pe
        LEFT JOIN LATERAL (
            SELECT pp2.*
            FROM wb_promotion_products pp2
            WHERE pp2.organization_id = :org
              AND pp2.nm_id = pe.nm_id
        """
        + promo_product_filter
        + """
            ORDER BY pp2.in_action DESC, pp2.synced_at DESC NULLS LAST, pp2.updated_at DESC NULLS LAST
            LIMIT 1
        ) pp ON true
        LEFT JOIN wb_promotions wp ON wp.id = pp.promotion_id_col
        LEFT JOIN LATERAL (
            SELECT price, stock_qty
            FROM tech_status ts2
            WHERE ts2.organization_id = :org AND ts2.nm_id = pe.nm_id
            ORDER BY target_date DESC LIMIT 1
        ) ts ON true
        LEFT JOIN LATERAL (
            SELECT cost_price FROM reference_book rb2
            WHERE rb2.organization_id = :org AND rb2.entity_id = pe.id
            ORDER BY valid_from DESC LIMIT 1
        ) rb ON true
        LEFT JOIN LATERAL (
            SELECT
                snapshot_date,
                promotions,
                sale_conditions,
                available_qty,
                available_to_buy,
                regular_in_promo,
                auto_in_promo,
                in_any_promo,
                regular_promotion_ids,
                auto_promotion_ids,
                price_basic,
                price_product
            FROM wb_promotion_snapshots snp2
            WHERE snp2.organization_id = :org
              AND snp2.nm_id = pe.nm_id
              AND snp2.snapshot_date = (SELECT snapshot_date FROM latest_snapshot)
            ORDER BY snp2.snapshot_date DESC LIMIT 1
        ) snp ON true
        WHERE pe.organization_id = :org
          AND pe.nm_id IS NOT NULL
          AND COALESCE(snp.available_to_buy, false) = true
        """
        + promo_where
        + """
        ORDER BY pe.nm_id, pe.size_name
        """
    )
    result = await db.execute(query, params)

    items = []
    for row in result.all():
        price_before = (
            float(row.price_before_spp) if row.price_before_spp else None
        )
        cost = float(row.cost_price) if row.cost_price else None
        margin_pct = None
        if price_before and cost and price_before > 0:
            margin_pct = round((price_before - cost) / price_before * 100, 1)
        
        # Обработка snapshot_promotions (JSONB массив)
        snapshot_promotions = []
        if row.snapshot_promotions_raw:
            try:
                promo_list = row.snapshot_promotions_raw
                if isinstance(promo_list, list):
                    for promo in promo_list:
                        if not promo:
                            continue
                        if isinstance(promo, dict):
                            title = promo.get("title") or f"Акция {promo.get('id')}"
                            snapshot_promotions.append(title)
                        else:
                            snapshot_promotions.append(str(promo))
            except AttributeError:
                pass
        regular_in_promo = bool(row.regular_in_promo)
        auto_in_promo = bool(row.auto_in_promo)
        in_any_promo = bool(row.in_any_promo)
        snapshot_in_promo = auto_in_promo
        
        # Обработка available_promotions — enrich из wb_promotions
        available_promotions = []
        promo_sources = row.regular_promotion_ids or row.auto_promotion_ids
        if promo_sources:
            try:
                promo_list = promo_sources
                if isinstance(promo_list, list):
                    # Для каждой акции из snapshot ищем детали в wb_promotions
                    for p in promo_list:
                        if not p:
                            continue
                        if isinstance(p, dict):
                            promo_id = p.get("id")
                            title = p.get("title", "")
                            active = p.get("active", False)
                            start_dt = p.get("startDateTime") or p.get("start_date")
                            end_dt = p.get("endDateTime") or p.get("end_date")
                        else:
                            # Фолбэк если структура другая
                            promo_id = None
                            title = str(p)
                            active = False
                            start_dt = None
                            end_dt = None
                        
                        # Форматируем даты
                        start_date = None
                        end_date = None
                        if start_dt:
                            try:
                                from datetime import datetime
                                start_date = datetime.fromisoformat(start_dt.replace("Z", "+00:00")).strftime("%d.%m.%Y")
                            except (ValueError, AttributeError):
                                pass
                        if end_dt:
                            try:
                                from datetime import datetime
                                end_date = datetime.fromisoformat(end_dt.replace("Z", "+00:00")).strftime("%d.%m.%Y")
                            except (ValueError, AttributeError):
                                pass
                        
                        available_promotions.append({
                            "id": promo_id,
                            "title": title or (f"Акция {promo_id}" if promo_id else ""),
                            "active": active,
                            "start_date": start_date,
                            "end_date": end_date,
                        })
            except AttributeError:
                pass
        
        # current_action_id — если товар в wb_promotion_products и in_action
        current_action_id = None
        if row.in_action and row.promotion_id_col:
            current_action_id = str(row.promotion_id_col)
        promo_title = None
        if regular_in_promo:
            promo_title = row.promo_title
        if not promo_title and in_any_promo:
            promo_title = (
                ", ".join([p["title"] for p in available_promotions[:2] if p.get("title")])
                or (", ".join(snapshot_promotions[:2]) if snapshot_promotions else None)
            )
        
        items.append(
            {
                "id": str(row.id) if row.id else None,
                "wb_promotion_ext_id": row.wb_promotion_ext_id,
                "nm_id": row.nm_id,
                "in_action": row.in_action or False,
                "regular_in_promo": regular_in_promo,
                "auto_in_promo": auto_in_promo,
                "in_any_promo": in_any_promo,
                "snapshot_in_promo": snapshot_in_promo,
                "snapshot_date": row.snapshot_date.isoformat() if row.snapshot_date else None,
                "available_qty": row.available_qty,
                "available_to_buy": bool(row.available_to_buy),
                "auto_matched": row.auto_matched or False,
                "current_price": (
                    float(row.current_price) if row.current_price else None
                ),
                "required_price": (
                    float(row.required_price) if row.required_price else None
                ),
                "price_in_promo": (
                    float(row.price_in_promo) if row.price_in_promo else None
                ),
                "profit_in_promo": (
                    float(row.profit_in_promo) if row.profit_in_promo else None
                ),
                "margin_delta": (
                    float(row.margin_delta) if row.margin_delta else None
                ),
                "plan": row.plan or False,
                "status_text": row.status_text,
                "entity_id": str(row.entity_id) if row.entity_id else None,
                "vendor_code": row.vendor_code,
                "product_name": row.product_name,
                "photo": row.photo,
                "size_name": row.size_name,
                "brand": row.brand,
                "subject_name": row.subject_name,
                "promo_title": promo_title,
                "promo_type": (row.promo_type if regular_in_promo else None) or ("auto" if auto_in_promo else None),
                "promo_start": (
                    row.promo_start.strftime("%d.%m.%Y")
                    if row.promo_start
                    else None
                ),
                "promo_end": (
                    row.promo_end.strftime("%d.%m.%Y") if row.promo_end else None
                ),
                "promo_importance": row.promo_importance,
                "price_before_spp": price_before,
                "stock_qty": row.available_qty if row.available_qty is not None else row.stock_qty,
                "margin_pct": margin_pct,
                # Новые поля из snapshot
                "snapshot_promotions": snapshot_promotions,
                "price_basic": float(row.price_basic) if row.price_basic else None,
                "price_product": float(row.price_product) if row.price_product else None,
                # Доступные акции для товара
                "available_promotions": available_promotions,
                "current_action_id": current_action_id,
            }
        )
    return {"items": items}


class PromoProductSave(BaseModel):
    items: list


@router.post("/api/v1/nl/promotions/products/save")
async def save_promotion_products(
    data: PromoProductSave,
    org_id: str,
    db: AsyncSession = Depends(get_db),
):
    """Save manual promotion fields."""
    saved = 0
    for item in data.items:
        row_id = item.get("id")
        if not row_id:
            continue
        query = select(WbPromotionProduct).where(
            WbPromotionProduct.id == row_id,
            WbPromotionProduct.organization_id == org_id,
        )
        result = await db.execute(query)
        promotion_product = result.scalar_one_or_none()
        if not promotion_product:
            continue
        if "plan" in item:
            promotion_product.plan = item["plan"]
        if "price_in_promo" in item:
            promotion_product.price_in_promo = item["price_in_promo"]
        saved += 1
    await db.commit()
    return {"ok": True, "saved": saved}


@router.post("/api/v1/nl/promotions/upload-excel")
async def upload_promo_excel(
    org_id: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Upload a WB promotion Excel template."""
    content = await file.read()
    workbook = openpyxl.load_workbook(io.BytesIO(content))
    worksheet = workbook.active

    col_map = {}
    header_row = 1
    for col_idx in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=header_row, column=col_idx).value
        if value:
            value_lower = str(value).strip().lower()
            if "уже участв" in value_lower:
                col_map["in_action"] = col_idx
            elif "бренд" in value_lower:
                col_map["brand"] = col_idx
            elif "предмет" in value_lower:
                col_map["subject"] = col_idx
            elif "наименован" in value_lower:
                col_map["name"] = col_idx
            elif "артикул постав" in value_lower:
                col_map["vendor_code"] = col_idx
            elif "артикул wb" in value_lower or (
                "артикул продавца" in value_lower and "wb" in value_lower
            ):
                col_map["nm_id"] = col_idx
            elif "баркод" in value_lower or "штрих" in value_lower:
                col_map["barcode"] = col_idx
            elif "оборач" in value_lower:
                col_map["turnover"] = col_idx
            elif "остаток" in value_lower:
                col_map["stock"] = col_idx
            elif "плановая цена" in value_lower or "план. цена" in value_lower:
                col_map["planned_price"] = col_idx
            elif "текущая цена" in value_lower:
                col_map["current_price"] = col_idx
            elif "минимальная цена" in value_lower or "мин. цена" in value_lower:
                col_map["min_price"] = col_idx
            elif "текущая скидка" in value_lower:
                col_map["current_discount"] = col_idx
            elif "загружаемая скидка" in value_lower or "загружаем" in value_lower:
                col_map["upload_discount"] = col_idx
            elif "статус" in value_lower:
                col_map["status"] = col_idx

    nm_col = col_map.get("nm_id")
    if not nm_col:
        for col_idx in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row=header_row, column=col_idx).value
            if value and (
                "артикул" in str(value).lower() and "wb" in str(value).lower()
            ):
                nm_col = col_idx
                break

    filename = file.filename or "upload"
    promo_title = filename.replace(".xlsx", "").replace(".xls", "")
    new_promotion = WbPromotion(
        organization_id=org_id,
        promotion_id=abs(hash(promo_title)) % 1000000,
        title=promo_title,
        promo_type="excel",
        is_active=True,
        source="excel",
    )
    db.add(new_promotion)
    await db.flush()

    count = 0
    for row_idx in range(2, worksheet.max_row + 1):
        nm_value = (
            worksheet.cell(row=row_idx, column=nm_col).value if nm_col else None
        )
        if not nm_value:
            continue
        try:
            nm_id = int(nm_value)
        except (ValueError, TypeError):
            continue

        entity_result = await db.execute(
            text(
                "SELECT id FROM product_entities "
                "WHERE organization_id = :org AND nm_id = :nm LIMIT 1"
            ),
            {"org": org_id, "nm": nm_id},
        )
        entity_row = entity_result.first()
        entity_id = entity_row[0] if entity_row else None

        def cell_value(column_key):
            column = col_map.get(column_key)
            return (
                worksheet.cell(row=row_idx, column=column).value
                if column
                else None
            )

        in_action = cell_value("in_action")
        if isinstance(in_action, str):
            in_action = in_action.strip().upper() in ("ДА", "YES", "1", "+")
        elif isinstance(in_action, (int, float)):
            in_action = bool(in_action)

        existing = await db.execute(
            text(
                "SELECT id FROM wb_promotion_products "
                "WHERE organization_id = :org "
                "AND wb_promotion_ext_id = :ext_id AND nm_id = :nm"
            ),
            {
                "org": org_id,
                "ext_id": new_promotion.promotion_id,
                "nm": nm_id,
            },
        )
        existing_row = existing.first()

        if existing_row:
            await db.execute(
                text(
                    "UPDATE wb_promotion_products "
                    "SET in_action = :ia, status_text = :st, "
                    "current_price = :cp, entity_id = :eid, "
                    "promotion_id_col = :pid, updated_at = now() "
                    "WHERE id = :rid"
                ),
                {
                    "ia": in_action,
                    "st": str(cell_value("status") or "")[:200],
                    "cp": cell_value("current_price"),
                    "eid": entity_id,
                    "pid": str(new_promotion.id),
                    "rid": existing_row[0],
                },
            )
        else:
            await db.execute(
                text(
                    "INSERT INTO wb_promotion_products "
                    "(id, organization_id, promotion_id_col, "
                    "wb_promotion_ext_id, nm_id, entity_id, in_action, "
                    "current_price, status_text, created_at) "
                    "VALUES (gen_random_uuid(), :org, :pid, :ext_id, :nm, "
                    ":eid, :ia, :cp, :st, now())"
                ),
                {
                    "org": org_id,
                    "pid": str(new_promotion.id),
                    "ext_id": new_promotion.promotion_id,
                    "nm": nm_id,
                    "eid": entity_id,
                    "ia": in_action,
                    "cp": cell_value("current_price"),
                    "st": str(cell_value("status") or "")[:200],
                },
            )
        count += 1

    await db.commit()
    return {
        "ok": True,
        "count": count,
        "promotion_id": new_promotion.promotion_id,
    }


@router.post("/api/v1/nl/promotions/sync-api")
async def sync_promo_api(
    org_id: str,
    db: AsyncSession = Depends(get_db),
):
    """Start the existing WB promotion synchronization task."""
    from tasks.promo_sync import do_promo_sync

    result = do_promo_sync.delay()
    return {"status": "started", "task_id": result.id}


@router.get("/api/v1/nl/promotions/download-excel")
async def download_promo_excel(
    org_id: str,
    promotion_id: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    """
    Сборный xlsx для выбранных товаров из раздела Акции (заглушка).
    Формат: шаблон WB — nm_id, бренд, предмет, текущая цена, цена участия,
    скидка, статус. Валидация: цена участия < минимальной цены справочника → ошибка.
    """
    # Получаем товары с plan=true (отмечены для участия)
    params = {"org": org_id}
    promo_filter = ""
    if promotion_id:
        promo_filter = " AND pp.wb_promotion_ext_id = :promo_id"
        params["promo_id"] = int(promotion_id)

    query = text(
        """
        SELECT
            pp.nm_id,
            pp.in_action,
            pp.current_price,
            pp.price_in_promo,
            pp.plan,
            pp.status_text,
            pe.vendor_code,
            pe.product_name,
            pe.brand,
            pe.subject_name,
            COALESCE(ts.price, pp.current_price) as price_before_spp,
            ts.stock_qty,
            rb.cost_price,
            rb.min_price
        FROM wb_promotion_products pp
        LEFT JOIN product_entities pe ON pe.id = pp.entity_id
        LEFT JOIN LATERAL (
            SELECT price, stock_qty
            FROM tech_status ts2
            WHERE ts2.organization_id = :org AND ts2.nm_id = pp.nm_id
            ORDER BY target_date DESC LIMIT 1
        ) ts ON true
        LEFT JOIN LATERAL (
            SELECT cost_price, min_price FROM reference_book rb2
            WHERE rb2.organization_id = :org AND rb2.entity_id = pp.entity_id
            ORDER BY valid_from DESC LIMIT 1
        ) rb ON true
        WHERE pp.organization_id = :org
        """
        + promo_filter
        + """
        ORDER BY pp.nm_id
        """
    )
    result = await db.execute(query, params)
    rows = result.all()

    # Создаём xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Акции"

    # Заголовки (шаблон WB)
    headers = [
        "Артикул WB",
        "Бренд",
        "Предмет",
        "Наименование",
        "Артикул продавца",
        "Текущая цена",
        "Цена участия",
        "Скидка %",
        "Остаток",
        "Себестоимость",
        "Мин. цена справочника",
        "Статус",
    ]
    ws.append(headers)

    # Стилизация заголовков
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6c5ce7", end_color="6c5ce7", fill_type="solid")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Ширины колонок
    col_widths = [12, 16, 18, 30, 14, 12, 12, 10, 10, 12, 16, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    warnings = []

    for row in rows:
        current_price = float(row.current_price) if row.current_price else None
        price_in_promo = float(row.price_in_promo) if row.price_in_promo else None
        min_price = float(row.min_price) if row.min_price else None
        cost_price = float(row.cost_price) if row.cost_price else None

        # Скидка %
        discount_pct = None
        if current_price and price_in_promo and current_price > 0:
            discount_pct = round((current_price - price_in_promo) / current_price * 100, 1)

        # Валидация: цена участия < мин. цены справочника
        status_val = row.status_text or ""
        if price_in_promo and min_price and price_in_promo < min_price:
            status_val = f"⚠ Цена участия {price_in_promo}₽ < мин. цены {min_price}₽"
            warnings.append(
                f"Артикул {row.nm_id}: цена участия {price_in_promo}₽ "
                f"ниже минимальной цены справочника {min_price}₽"
            )

        ws.append([
            row.nm_id,
            row.brand or "",
            row.subject_name or "",
            row.product_name or "",
            row.vendor_code or "",
            current_price,
            price_in_promo,
            discount_pct,
            row.stock_qty or 0,
            cost_price,
            min_price,
            status_val,
        ])

    # Лист с предупреждениями
    if warnings:
        ws_warn = wb.create_sheet("Предупреждения")
        ws_warn.append(["Внимание", "Детали"])
        for col_idx in range(1, 3):
            cell = ws_warn.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
        ws_warn.column_dimensions["A"].width = 20
        ws_warn.column_dimensions["B"].width = 80
        for w in warnings:
            ws_warn.append(["Валидация", w])

    # Экспорт в память
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"promotions_export_{org_id[:8]}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
