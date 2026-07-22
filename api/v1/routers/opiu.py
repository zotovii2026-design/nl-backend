"""API endpoints for the rebuilt OPIU report."""

import io
from datetime import date, timedelta
from decimal import Decimal

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from core.database import get_db
from core.dependencies import get_current_user
from core.role_deps import require_organization_role
from domain.opiu import build_opiu_report, serialize_report
from domain.unit_economics import normalize_tax_system
from models.organization import Role
from models.user import User
from models.wb_finance import WbFinanceSync
from tasks.celery_app import celery_app


router = APIRouter(tags=["nl"])

ZERO = Decimal("0")


def _validate_period(date_from: date, date_to: date):
    if date_to < date_from:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="date_to must not be earlier than date_from",
        )
    if (date_to - date_from).days > 366:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="The maximum report period is 366 days",
        )


async def _load_report_rows(
    db: AsyncSession,
    organization_id: str,
    date_from: date,
    date_to: date,
):
    result = await db.execute(
        text(
            """
            SELECT
                fr.entity_id,
                fr.nm_id,
                fr.vendor_code,
                fr.barcode,
                fr.size_name,
                fr.doc_type_name,
                fr.seller_oper_name,
                fr.quantity,
                fr.return_amount,
                fr.retail_price,
                fr.retail_amount,
                fr.for_pay,
                fr.acquiring_fee,
                fr.delivery_service,
                fr.penalty,
                fr.paid_storage,
                fr.deduction,
                fr.paid_acceptance,
                fr.cashback_amount,
                fr.cashback_discount,
                fr.cashback_commission_change,
                pe.product_name,
                pe.photo_main,
                COALESCE(rb.brand, pe.brand, '') AS brand,
                COALESCE(rb.product_class, '') AS product_class,
                COALESCE(rb.product_status, '') AS product_status,
                COALESCE(rb.subject_name, pe.subject_name, '') AS subject_name
            FROM wb_finance_rows fr
            LEFT JOIN product_entities pe ON pe.id = fr.entity_id
            LEFT JOIN LATERAL (
                SELECT
                    ref.brand,
                    ref.product_class,
                    ref.product_status,
                    ref.subject_name
                FROM reference_book ref
                WHERE ref.organization_id = fr.organization_id
                  AND (
                      ref.entity_id = fr.entity_id
                      OR (
                          fr.entity_id IS NULL
                          AND ref.nm_id = fr.nm_id
                      )
                  )
                  AND ref.valid_from <= :date_to
                  AND (ref.valid_to IS NULL OR ref.valid_to >= :date_from)
                ORDER BY
                    CASE WHEN ref.entity_id = fr.entity_id THEN 0 ELSE 1 END,
                    ref.valid_from DESC
                LIMIT 1
            ) rb ON TRUE
            WHERE fr.organization_id = :org
              AND (
                  (
                      LOWER(COALESCE(fr.seller_oper_name, '')) IN (
                          'хранение',
                          'коррекция хранения'
                      )
                      AND fr.operation_date::date BETWEEN :date_from AND :date_to
                  )
                  OR (
                      LOWER(COALESCE(fr.seller_oper_name, '')) NOT IN (
                          'хранение',
                          'коррекция хранения'
                      )
                      AND fr.report_date_from BETWEEN :date_from AND :date_to
                  )
              )
            ORDER BY fr.rrd_id
            """
        ),
        {
            "org": organization_id,
            "date_from": date_from,
            "date_to": date_to,
        },
    )

    names = [
        "entity_id",
        "nm_id",
        "vendor_code",
        "barcode",
        "size_name",
        "doc_type_name",
        "seller_oper_name",
        "quantity",
        "return_amount",
        "retail_price",
        "retail_amount",
        "for_pay",
        "acquiring_fee",
        "delivery_service",
        "penalty",
        "paid_storage",
        "deduction",
        "paid_acceptance",
        "cashback_amount",
        "cashback_discount",
        "cashback_commission_change",
        "product_name",
        "photo_main",
        "brand",
        "product_class",
        "product_status",
        "subject_name",
    ]
    return [dict(zip(names, row)) for row in result.all()]


async def _load_paid_storage_report_rows(
    db: AsyncSession,
    organization_id: str,
    date_from: date,
    date_to: date,
):
    storage_from = date_from + timedelta(days=1)
    storage_to = date_to + timedelta(days=1)
    result = await db.execute(
        text(
            """
            SELECT
                ps.entity_id,
                ps.nm_id,
                COALESCE(ps.vendor_code, pe.vendor_code, '') AS vendor_code,
                '' AS barcode,
                '' AS size_name,
                'Продажа' AS doc_type_name,
                'Хранение' AS seller_oper_name,
                0 AS quantity,
                0 AS return_amount,
                0 AS retail_price,
                0 AS retail_amount,
                0 AS for_pay,
                0 AS acquiring_fee,
                0 AS delivery_service,
                0 AS penalty,
                SUM(ps.storage_amount) AS paid_storage,
                0 AS deduction,
                0 AS paid_acceptance,
                0 AS cashback_amount,
                0 AS cashback_discount,
                0 AS cashback_commission_change,
                pe.product_name,
                pe.photo_main,
                COALESCE(rb.brand, ps.brand, pe.brand, '') AS brand,
                COALESCE(rb.product_class, '') AS product_class,
                COALESCE(rb.product_status, '') AS product_status,
                COALESCE(rb.subject_name, ps.subject_name, pe.subject_name, '') AS subject_name
            FROM wb_paid_storage_rows ps
            LEFT JOIN product_entities pe ON pe.id = ps.entity_id
            LEFT JOIN LATERAL (
                SELECT
                    ref.brand,
                    ref.product_class,
                    ref.product_status,
                    ref.subject_name
                FROM reference_book ref
                WHERE ref.organization_id = ps.organization_id
                  AND (
                      ref.entity_id = ps.entity_id
                      OR (
                          ps.entity_id IS NULL
                          AND ref.nm_id = ps.nm_id
                      )
                  )
                  AND ref.valid_from <= :date_to
                  AND (ref.valid_to IS NULL OR ref.valid_to >= :date_from)
                ORDER BY
                    CASE WHEN ref.entity_id = ps.entity_id THEN 0 ELSE 1 END,
                    ref.valid_from DESC
                LIMIT 1
            ) rb ON TRUE
            WHERE ps.organization_id = :org
              AND ps.storage_date BETWEEN :storage_from AND :storage_to
            GROUP BY
                ps.entity_id,
                ps.nm_id,
                ps.vendor_code,
                ps.brand,
                ps.subject_name,
                pe.vendor_code,
                pe.product_name,
                pe.photo_main,
                pe.brand,
                pe.subject_name,
                rb.brand,
                rb.product_class,
                rb.product_status,
                rb.subject_name
            """
        ),
        {
            "org": organization_id,
            "date_from": date_from,
            "date_to": date_to,
            "storage_from": storage_from,
            "storage_to": storage_to,
        },
    )
    names = [
        "entity_id",
        "nm_id",
        "vendor_code",
        "barcode",
        "size_name",
        "doc_type_name",
        "seller_oper_name",
        "quantity",
        "return_amount",
        "retail_price",
        "retail_amount",
        "for_pay",
        "acquiring_fee",
        "delivery_service",
        "penalty",
        "paid_storage",
        "deduction",
        "paid_acceptance",
        "cashback_amount",
        "cashback_discount",
        "cashback_commission_change",
        "product_name",
        "photo_main",
        "brand",
        "product_class",
        "product_status",
        "subject_name",
    ]
    return [dict(zip(names, row)) for row in result.all()]


def _as_decimal(value) -> Decimal:
    if value in (None, ""):
        return ZERO
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _money(value) -> float:
    return float(_as_decimal(value).quantize(Decimal("0.01")))


async def _load_ad_spend_by_nm(
    db: AsyncSession,
    organization_id: str,
    date_from: date,
    date_to: date,
):
    result = await db.execute(
        text(
            """
            SELECT sn.nm_id, COALESCE(SUM(sn.spent), 0) AS spent
            FROM ad_stats_nm sn
            WHERE sn.organization_id = :org
              AND sn.stat_date BETWEEN :date_from AND :date_to
              AND sn.nm_id IS NOT NULL
            GROUP BY sn.nm_id
            """
        ),
        {
            "org": organization_id,
            "date_from": date_from,
            "date_to": date_to,
        },
    )
    return {int(nm_id): _as_decimal(spent) for nm_id, spent in result.all()}


async def _load_control_totals(
    db: AsyncSession,
    organization_id: str,
    date_from: date,
    date_to: date,
):
    result = await db.execute(
        text(
            """
            WITH finance AS (
                SELECT
                    COALESCE(SUM(
                        CASE
                            WHEN fr.operation_date::date BETWEEN :date_from AND :date_to
                            THEN fr.paid_storage
                            ELSE 0
                        END
                    ), 0) AS finance_storage,
                    COALESCE(SUM(
                        CASE
                            WHEN fr.report_date_from BETWEEN :date_from AND :date_to
                             AND COALESCE(fr.deduction, 0) <> 0
                             AND (
                                 LOWER(COALESCE(fr.seller_oper_name, '')) LIKE '%wb продвиж%'
                                 OR LOWER(COALESCE(fr.seller_oper_name, '')) LIKE '%вб продвиж%'
                                 OR (
                                     LOWER(COALESCE(fr.seller_oper_name, '')) = 'удержание'
                                     AND COALESCE(fr.nm_id, 0) = 0
                                     AND COALESCE(fr.vendor_code, '') = ''
                                     AND COALESCE(fr.barcode, '') = ''
                                     AND fr.entity_id IS NULL
                                 )
                             )
                            THEN fr.deduction
                            ELSE 0
                        END
                    ), 0) AS wb_promotion_deduction
                FROM wb_finance_rows fr
                WHERE fr.organization_id = :org
                  AND (
                      fr.operation_date::date BETWEEN :date_from AND :date_to
                      OR fr.report_date_from BETWEEN :date_from AND :date_to
                  )
            ),
            bank_control AS (
                SELECT bank_payment_sum
                FROM wb_finance_syncs
                WHERE organization_id = :org
                  AND date_from = :date_from
                  AND date_to = :date_to
                  AND status IN ('success', 'warning')
                  AND bank_payment_sum IS NOT NULL
                ORDER BY started_at DESC
                LIMIT 1
            ),
            ads AS (
                SELECT COALESCE(SUM(sn.spent), 0) AS advertising_total
                FROM ad_stats_nm sn
                WHERE sn.organization_id = :org
                  AND sn.stat_date BETWEEN :date_from AND :date_to
                  AND sn.nm_id IS NOT NULL
            )
            SELECT
                finance.finance_storage,
                finance.wb_promotion_deduction,
                ads.advertising_total,
                bank_control.bank_payment_sum
            FROM finance
            CROSS JOIN ads
            LEFT JOIN bank_control ON TRUE
            """
        ),
        {
            "org": organization_id,
            "date_from": date_from,
            "date_to": date_to,
        },
    )
    row = result.one()
    return {
        "finance_storage": _as_decimal(row.finance_storage),
        "wb_promotion_deduction": _as_decimal(row.wb_promotion_deduction),
        "advertising_total": _as_decimal(row.advertising_total),
        "bank_payment_sum": (
            _as_decimal(row.bank_payment_sum)
            if row.bank_payment_sum is not None
            else None
        ),
    }


async def _load_sales_funnel_orders_by_nm(
    db: AsyncSession,
    organization_id: str,
    date_from: date,
    date_to: date,
):
    result = await db.execute(
        text(
            """
            SELECT target_date, raw_response
            FROM raw_api_data
            WHERE organization_id = :org
              AND api_method = 'sales_funnel'
              AND status = 'ok'
              AND target_date BETWEEN :date_from AND :date_to
            """
        ),
        {
            "org": organization_id,
            "date_from": date_from,
            "date_to": date_to,
        },
    )
    orders_by_nm = {}
    for target_date, raw_response in result.all():
        if not isinstance(raw_response, list):
            continue
        for item in raw_response:
            if not isinstance(item, dict):
                continue
            stat = (item.get("statistic") or {}).get("selected") or {}
            period = stat.get("period") or {}
            if period.get("start") and period.get("end"):
                if (
                    period.get("start") != str(target_date)
                    or period.get("end") != str(target_date)
                ):
                    continue
            nm_id = (item.get("product") or {}).get("nmId")
            if not nm_id:
                continue
            bucket = orders_by_nm.setdefault(
                int(nm_id),
                {"orders_qty": ZERO, "orders_sum": ZERO},
            )
            bucket["orders_qty"] += _as_decimal(stat.get("orderCount"))
            bucket["orders_sum"] += _as_decimal(stat.get("orderSum"))
    return orders_by_nm


async def _load_raw_orders_by_nm(
    db: AsyncSession,
    organization_id: str,
    date_from: date,
    date_to: date,
):
    result = await db.execute(
        text(
            """
            WITH raw_orders AS (
                SELECT
                    r.target_date,
                    COALESCE(o.elem->>'srid', '') AS srid,
                    COALESCE(
                        NULLIF(o.elem->>'nmId', '')::bigint,
                        NULLIF(o.elem->>'nm_id', '')::bigint
                    ) AS nm_id,
                    LEFT(o.elem->>'date', 10)::date AS order_date,
                    COALESCE(
                        NULLIF(o.elem->>'priceWithDisc', '')::numeric,
                        NULLIF(o.elem->>'totalPrice', '')::numeric,
                        NULLIF(o.elem->>'price', '')::numeric,
                        0
                    ) AS order_sum
                FROM raw_api_data r
                CROSS JOIN LATERAL jsonb_array_elements(r.raw_response) AS o(elem)
                WHERE r.organization_id = :org
                  AND r.api_method = 'orders'
                  AND r.status = 'ok'
                  AND r.target_date BETWEEN :date_from AND :date_to
            ),
            dedup_non_empty_orders AS (
                SELECT DISTINCT ON (srid) *
                FROM raw_orders
                WHERE srid <> ''
                ORDER BY srid, target_date
            ),
            dedup_orders AS (
                SELECT *
                FROM dedup_non_empty_orders
                UNION ALL
                SELECT *
                FROM raw_orders
                WHERE srid = ''
            )
            SELECT nm_id,
                   COUNT(*) AS orders_qty,
                   COALESCE(SUM(order_sum), 0) AS orders_sum
            FROM dedup_orders
            WHERE order_date BETWEEN :date_from AND :date_to
              AND nm_id IS NOT NULL
            GROUP BY nm_id
            """
        ),
        {
            "org": organization_id,
            "date_from": date_from,
            "date_to": date_to,
        },
    )
    return {
        int(nm_id): {
            "orders_qty": _as_decimal(orders_qty),
            "orders_sum": _as_decimal(orders_sum),
        }
        for nm_id, orders_qty, orders_sum in result.all()
    }


async def _load_cost_by_item(
    db: AsyncSession,
    organization_id: str,
    date_from: date,
    date_to: date,
):
    result = await db.execute(
        text(
            """
            SELECT DISTINCT ON (entity_id, nm_id)
                   entity_id,
                   nm_id,
                   COALESCE(cost_price, 0) AS cost_price,
                   COALESCE(extra_costs, 0) AS extra_costs,
                   COALESCE(purchase_cost, 0) AS purchase_cost,
                   COALESCE(tax_system, '') AS tax_system,
                   COALESCE(tax_rate, 0) AS tax_rate,
                   COALESCE(vat_rate, 0) AS vat_rate
            FROM reference_book
            WHERE organization_id = :org
              AND valid_from <= :date_to
              AND (valid_to IS NULL OR valid_to >= :date_from)
            ORDER BY entity_id, nm_id, valid_from DESC, created_at DESC NULLS LAST
            """
        ),
        {"org": organization_id, "date_from": date_from, "date_to": date_to},
    )
    by_entity = {}
    by_nm = {}
    for (
        entity_id,
        nm_id,
        cost_price,
        extra_costs,
        purchase_cost,
        tax_system,
        tax_rate,
        vat_rate,
    ) in result.all():
        value = {
            "unit_cost": _as_decimal(cost_price) + _as_decimal(extra_costs),
            "purchase_cost": _as_decimal(purchase_cost),
            "tax_system": normalize_tax_system(tax_system),
            "tax_rate": _as_decimal(tax_rate),
            "vat_rate": _as_decimal(vat_rate),
        }
        if entity_id:
            by_entity[str(entity_id)] = value
        if nm_id:
            by_nm[int(nm_id)] = value
    return by_entity, by_nm


async def _load_cost_totals_by_item(
    db: AsyncSession,
    organization_id: str,
    date_from: date,
    date_to: date,
):
    result = await db.execute(
        text(
            """
            SELECT
                fr.entity_id,
                fr.nm_id,
                SUM(
                    COALESCE(fr.quantity, 0)
                    * (
                        COALESCE(ref.cost_price, 0)
                        + COALESCE(ref.extra_costs, 0)
                    )
                ) AS cost_total
            FROM wb_finance_rows fr
            LEFT JOIN LATERAL (
                SELECT cost_price, extra_costs
                FROM reference_book rb
                WHERE rb.organization_id = fr.organization_id
                  AND (
                      rb.entity_id = fr.entity_id
                      OR (
                          fr.entity_id IS NULL
                          AND rb.nm_id = fr.nm_id
                      )
                  )
                  AND rb.valid_from <= COALESCE(fr.operation_date::date, :date_to)
                  AND (
                      rb.valid_to IS NULL
                      OR rb.valid_to >= COALESCE(fr.operation_date::date, :date_to)
                  )
                ORDER BY
                    CASE WHEN rb.entity_id = fr.entity_id THEN 0 ELSE 1 END,
                    rb.valid_from DESC,
                    rb.created_at DESC NULLS LAST
                LIMIT 1
            ) ref ON TRUE
            WHERE fr.organization_id = :org
              AND fr.report_date_from BETWEEN :date_from AND :date_to
              AND LOWER(COALESCE(fr.seller_oper_name, '')) = 'продажа'
              AND LOWER(COALESCE(fr.doc_type_name, '')) NOT LIKE '%возврат%'
              AND (
                  ref.cost_price IS NOT NULL
                  OR ref.extra_costs IS NOT NULL
              )
            GROUP BY fr.entity_id, fr.nm_id
            """
        ),
        {
            "org": organization_id,
            "date_from": date_from,
            "date_to": date_to,
        },
    )
    by_entity = {}
    by_nm = {}
    for entity_id, nm_id, cost_total in result.all():
        if entity_id:
            by_entity[str(entity_id)] = _as_decimal(cost_total)
        if nm_id:
            by_nm[int(nm_id)] = _as_decimal(cost_total)
    return by_entity, by_nm


def _item_nm_id(item):
    nm_id = item.get("nm_id")
    if nm_id in (None, "", 0, "0"):
        return None
    return int(nm_id)


def _calculate_selected_tax(
    tax_system: str | None,
    tax_rate: Decimal,
    revenue: Decimal,
    expenses_before_tax: Decimal,
) -> Decimal:
    normalized = normalize_tax_system(tax_system)
    if normalized == "usn":
        return revenue * tax_rate / Decimal("100")
    if normalized == "usn_dr":
        return max(revenue - expenses_before_tax, ZERO) * tax_rate / Decimal("100")
    if normalized == "osn" and tax_rate:
        return max(revenue - expenses_before_tax, ZERO) * tax_rate / Decimal("100")
    return ZERO


def _enrich_serialized_report(
    data: dict,
    ad_spend_by_nm: dict[int, Decimal],
    orders_by_nm: dict[int, dict],
    cost_by_entity: dict[str, dict],
    cost_by_nm: dict[int, dict],
    cost_total_by_entity: dict[str, Decimal] | None = None,
    cost_total_by_nm: dict[int, Decimal] | None = None,
    control_totals: dict[str, Decimal] | None = None,
):
    cost_total_by_entity = cost_total_by_entity or {}
    cost_total_by_nm = cost_total_by_nm or {}
    control_totals = control_totals or {}
    enriched_items = []
    totals = {
        "marketplace_commission_sum": ZERO,
        "acquiring_sum": ZERO,
        "realized_sum": ZERO,
        "advertising_api_spend": ZERO,
        "external_ad_spend": ZERO,
        "orders_qty": ZERO,
        "orders_sum": ZERO,
        "cost_total": ZERO,
        "vat_tax": ZERO,
        "selected_tax": ZERO,
        "other_expenses": ZERO,
        "net_profit": ZERO,
        "gross_profit_after_ads": ZERO,
    }

    for item in data["items"]:
        nm_id = _item_nm_id(item)
        sales_qty = _as_decimal(item.get("sales_qty"))
        mp_sum = _as_decimal(item.get("marketplace_commission_unit")) * sales_qty
        acquiring_sum = _as_decimal(item.get("acquiring_unit")) * sales_qty
        realized_sum = _as_decimal(item.get("realized_unit")) * sales_qty
        ad_spend = ad_spend_by_nm.get(nm_id, ZERO) if nm_id else ZERO
        external_ad_spend = ZERO
        orders = orders_by_nm.get(nm_id, {}) if nm_id else {}
        orders_qty = _as_decimal(orders.get("orders_qty"))
        orders_sum = _as_decimal(orders.get("orders_sum"))
        cost_ref = cost_by_entity.get(
            str(item.get("entity_id") or ""),
            cost_by_nm.get(nm_id, {}) if nm_id else {},
        )
        unit_cost = _as_decimal(cost_ref.get("unit_cost"))
        historical_cost_total = cost_total_by_entity.get(
            str(item.get("entity_id") or ""),
            cost_total_by_nm.get(nm_id) if nm_id else None,
        )
        cost_total = (
            historical_cost_total
            if historical_cost_total is not None
            else unit_cost * sales_qty
        )
        if sales_qty and historical_cost_total is not None:
            unit_cost = cost_total / sales_qty
        revenue_tax_base = _as_decimal(item.get("retail_net_sum")) or realized_sum
        vat_rate = _as_decimal(cost_ref.get("vat_rate"))
        tax_rate = _as_decimal(cost_ref.get("tax_rate"))
        vat_tax = revenue_tax_base * vat_rate / Decimal("100")
        other_expenses = (
            _as_decimal(item.get("deduction"))
            + _as_decimal(item.get("acceptance"))
            + _as_decimal(item.get("distributed_other_expenses"))
            + _as_decimal(item.get("loyalty_points"))
            + _as_decimal(item.get("loyalty_participation"))
        )
        expenses_before_tax = (
            mp_sum
            + acquiring_sum
            + _as_decimal(item.get("delivery_total"))
            + _as_decimal(item.get("penalty"))
            + _as_decimal(item.get("storage"))
            + ad_spend
            + external_ad_spend
            + cost_total
            + other_expenses
        )
        selected_tax = _calculate_selected_tax(
            cost_ref.get("tax_system"),
            tax_rate,
            revenue_tax_base,
            expenses_before_tax,
        )
        gross_profit_after_ads = _as_decimal(item.get("gross_profit")) - ad_spend
        net_profit = (
            gross_profit_after_ads
            - external_ad_spend
            - cost_total
            - vat_tax
            - selected_tax
        )
        drr = ad_spend / orders_sum * 100 if orders_sum else ZERO
        gross_margin = (
            gross_profit_after_ads / realized_sum * 100 if realized_sum else ZERO
        )
        roi = net_profit / cost_total * 100 if cost_total else ZERO
        net_margin = net_profit / realized_sum * 100 if realized_sum else ZERO
        markup = realized_sum / cost_total * 100 if cost_total else ZERO
        loyalty_pct = (
            (
                _as_decimal(item.get("loyalty_points"))
                + _as_decimal(item.get("loyalty_participation"))
            )
            / _as_decimal(item.get("retail_net_sum"))
            * 100
            if _as_decimal(item.get("retail_net_sum"))
            else ZERO
        )

        item.update(
            {
                "retail_sum": _money(item.get("retail_sum")),
                "returns_retail_sum": _money(item.get("returns_retail_sum")),
                "retail_net_sum": _money(item.get("retail_net_sum")),
                "marketplace_commission_sum": _money(mp_sum),
                "acquiring_sum": _money(acquiring_sum),
                "realized_sum": _money(realized_sum),
                "advertising_api_spend": _money(ad_spend),
                "external_ad_spend": _money(external_ad_spend),
                "storage_difference_info": 0,
                "advertising_difference_info": 0,
                "orders_qty": float(orders_qty),
                "orders_sum": _money(orders_sum),
                "drr": _money(drr),
                "cost_unit": _money(unit_cost),
                "cost_total": _money(cost_total),
                "tax_system": cost_ref.get("tax_system") or "",
                "tax_rate": _money(tax_rate),
                "vat_rate": _money(vat_rate),
                "vat_tax": _money(vat_tax),
                "selected_tax": _money(selected_tax),
                "other_expenses": _money(other_expenses),
                "gross_profit_after_ads": _money(gross_profit_after_ads),
                "net_profit": _money(net_profit),
                "gross_margin": _money(gross_margin),
                "net_margin": _money(net_margin),
                "roi": _money(roi),
                "markup": _money(markup),
                "loyalty_pct": _money(loyalty_pct),
                "is_unassigned": item.get("vendor_code") == "(без артикула)",
            }
        )
        for key in totals:
            totals[key] += {
                "marketplace_commission_sum": mp_sum,
                "acquiring_sum": acquiring_sum,
                "realized_sum": realized_sum,
                "advertising_api_spend": ad_spend,
                "external_ad_spend": external_ad_spend,
                "orders_qty": orders_qty,
                "orders_sum": orders_sum,
                "cost_total": cost_total,
                "vat_tax": vat_tax,
                "selected_tax": selected_tax,
                "other_expenses": other_expenses,
                "net_profit": net_profit,
                "gross_profit_after_ads": gross_profit_after_ads,
            }[key]
        enriched_items.append(item)

    total = data["total"]
    finance_storage = _as_decimal(control_totals.get("finance_storage"))
    wb_promotion_deduction = _as_decimal(
        control_totals.get("wb_promotion_deduction")
    )
    advertising_total = _as_decimal(control_totals.get("advertising_total"))
    bank_payment_sum = control_totals.get("bank_payment_sum")
    product_storage = sum(
        (
            _as_decimal(item.get("storage"))
            for item in enriched_items
            if not item.get("is_unassigned")
        ),
        ZERO,
    )
    if advertising_total:
        totals["advertising_api_spend"] = advertising_total
    advertising_difference = (
        wb_promotion_deduction - totals["advertising_api_spend"]
        if wb_promotion_deduction
        else ZERO
    )
    storage_difference = (
        finance_storage - product_storage if finance_storage else ZERO
    )
    if wb_promotion_deduction:
        totals["gross_profit_after_ads"] = (
            _as_decimal(total.get("gross_profit"))
            - totals["advertising_api_spend"]
            - advertising_difference
        )
        totals["net_profit"] = (
            totals["gross_profit_after_ads"]
            - totals["external_ad_spend"]
            - totals["cost_total"]
            - totals["vat_tax"]
            - totals["selected_tax"]
        )
    total.update({key: _money(value) for key, value in totals.items()})
    total["wb_promotion_deduction"] = _money(wb_promotion_deduction)
    total["orders_qty"] = float(totals["orders_qty"])
    total["drr"] = _money(
        totals["advertising_api_spend"] / totals["orders_sum"] * 100
        if totals["orders_sum"]
        else ZERO
    )
    realized_sum = _as_decimal(total.get("realized_sum"))
    cost_total = _as_decimal(total.get("cost_total"))
    total["gross_margin"] = _money(
        totals["gross_profit_after_ads"] / realized_sum * 100
        if realized_sum
        else ZERO
    )
    total["roi"] = _money(
        totals["net_profit"] / cost_total * 100 if cost_total else ZERO
    )
    total["net_margin"] = _money(
        totals["net_profit"] / realized_sum * 100 if realized_sum else ZERO
    )
    total["markup"] = _money(
        realized_sum / cost_total * 100 if cost_total else ZERO
    )
    total["cost_unit"] = None
    if bank_payment_sum is not None:
        total["net_for_pay"] = _money(bank_payment_sum)
    total["storage_difference_info"] = _money(storage_difference)
    total["advertising_difference_info"] = _money(advertising_difference)
    total["is_unassigned"] = False
    data["items"] = enriched_items
    data["unassigned_items"] = [
        item for item in enriched_items if item.get("is_unassigned")
    ]
    product_items = [
        item for item in enriched_items if not item.get("is_unassigned")
    ]
    data["product_total"] = _product_total_row(product_items)
    return data


async def _enrichment_context(
    db: AsyncSession,
    organization_id: str,
    date_from: date,
    date_to: date,
):
    ad_spend_by_nm = await _load_ad_spend_by_nm(
        db, organization_id, date_from, date_to
    )
    orders_by_nm = await _load_sales_funnel_orders_by_nm(
        db, organization_id, date_from, date_to
    )
    raw_orders_by_nm = await _load_raw_orders_by_nm(
        db, organization_id, date_from, date_to
    )
    for nm_id, orders in raw_orders_by_nm.items():
        if nm_id not in orders_by_nm or not orders_by_nm[nm_id]["orders_sum"]:
            orders_by_nm[nm_id] = orders
    cost_by_entity, cost_by_nm = await _load_cost_by_item(
        db, organization_id, date_from, date_to
    )
    cost_total_by_entity, cost_total_by_nm = await _load_cost_totals_by_item(
        db, organization_id, date_from, date_to
    )
    control_totals = await _load_control_totals(
        db, organization_id, date_from, date_to
    )
    return (
        ad_spend_by_nm,
        orders_by_nm,
        cost_by_entity,
        cost_by_nm,
        cost_total_by_entity,
        cost_total_by_nm,
        control_totals,
    )


async def _sync_info(
    db: AsyncSession,
    organization_id: str,
    date_from: date,
    date_to: date,
):
    result = await db.execute(
        select(WbFinanceSync)
        .where(
            WbFinanceSync.organization_id == organization_id,
            WbFinanceSync.date_from == date_from,
            WbFinanceSync.date_to == date_to,
        )
        .order_by(WbFinanceSync.started_at.desc())
        .limit(1)
    )
    sync = result.scalar_one_or_none()
    if not sync:
        return None
    return {
        "status": sync.status,
        "rows_count": sync.rows_count,
        "bank_payment_sum": (
            float(sync.bank_payment_sum)
            if sync.bank_payment_sum is not None
            else None
        ),
        "calculated_payment_sum": (
            float(sync.calculated_payment_sum)
            if sync.calculated_payment_sum is not None
            else None
        ),
        "difference": (
            float(sync.difference) if sync.difference is not None else None
        ),
        "started_at": sync.started_at.isoformat(),
        "finished_at": (
            sync.finished_at.isoformat() if sync.finished_at else None
        ),
        "error_message": sync.error_message,
    }


async def _authorized_report(
    organization_id: str,
    date_from: date,
    date_to: date,
    current_user: User,
    db: AsyncSession,
):
    _validate_period(date_from, date_to)
    await require_organization_role(
        organization_id, Role.VIEWER, current_user, db
    )
    rows = await _load_report_rows(
        db, organization_id, date_from, date_to
    )
    paid_storage_rows = await _load_paid_storage_report_rows(
        db, organization_id, date_from, date_to
    )
    if paid_storage_rows:
        rows = [{**row, "paid_storage": ZERO} for row in rows]
        rows.extend(paid_storage_rows)
    return build_opiu_report(rows)


@router.get("/api/v1/nl/opiu/report")
async def get_opiu_report(
    org_id: str,
    date_from: date,
    date_to: date,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    report = await _authorized_report(
        org_id, date_from, date_to, current_user, db
    )
    serialized = serialize_report(report)
    serialized = _enrich_serialized_report(
        serialized,
        *(await _enrichment_context(db, org_id, date_from, date_to)),
    )
    serialized["period"] = {
        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
    }
    serialized["sync"] = await _sync_info(
        db, org_id, date_from, date_to
    )
    return serialized


@router.post("/api/v1/nl/opiu/sync")
async def trigger_opiu_sync(
    org_id: str,
    date_from: date,
    date_to: date,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _validate_period(date_from, date_to)
    await require_organization_role(
        org_id, Role.ADMIN, current_user, db
    )
    task = celery_app.send_task(
        "wb.opiu.sync_org",
        args=[org_id, date_from.isoformat(), date_to.isoformat()],
    )
    return {"status": "queued", "task_id": task.id}


EXPORT_COLUMNS = [
    ("Артикул поставщика", "vendor_code"),
    ("Артикул WB", "nm_id"),
    ("Название", "product_name"),
    ("Кол-во реализовано, шт", "sales_qty"),
    ("Цена розничная с учетом согласованной скидки, руб", "retail_sum"),
    ("Возвраты, сумма, руб", "returns_retail_sum"),
    (
        "Цена розн. с учетом скидки за вычетом возвратов, руб",
        "retail_net_sum",
    ),
    ("Вайлдберриз реализовал Товар (Пр), руб", "realized_sum"),
    ("Комиссия ВБ с учетом НДС, %", "marketplace_commission_pct"),
    ("Комиссия ВБ с учетом НДС, сумма, руб", "marketplace_commission_sum"),
    ("Эквайринг, %", "acquiring_pct"),
    ("Эквайринг, сумма, руб", "acquiring_sum"),
    ("Услуги по доставке товара покупателю, руб", "delivery_total"),
    ("Общая сумма штрафов, руб", "penalty"),
    ("Хранение, руб", "storage"),
    ("Расхождение по хранению (инфо, не распределяется), руб", "storage_difference_info"),
    ("Внутренняя реклама WB, руб", "advertising_api_spend"),
    ("Внешняя реклама (заглушка), руб", "external_ad_spend"),
    ("Расхождение по рекламе (инфо, не распределяется), руб", "advertising_difference_info"),
    ("ДРР, %", "drr"),
    ("Заказы, шт", "orders_qty"),
    ("Заказы, сумма, руб", "orders_sum"),
    ("Стоимость участия в программе лояльности, руб", "loyalty_participation"),
    ("Сумма баллов, удержанных по программе лояльности, руб", "loyalty_points"),
    ("% программы лояльности", "loyalty_pct"),
    ("Платная приемка, руб", "acceptance"),
    ("Прочие затраты, руб", "other_expenses"),
    ("К перечислению на р/с", "net_for_pay"),
    ("Валовая прибыль, руб", "gross_profit_after_ads"),
    ("Валовая рентабельность, %", "gross_margin"),
    ("Себестоимость, руб/ед", "cost_unit"),
    ("Себестоимость, сумма руб", "cost_total"),
    ("Налоговые удержания НДС, руб", "vat_tax"),
    ("Налоговые удержания (выбранный режим), руб", "selected_tax"),
    ("Чистая прибыль", "net_profit"),
    ("Рентабельность, %", "net_margin"),
    ("ROI, %", "roi"),
    ("Наценка, %", "markup"),
    ("Баркод", "barcode"),
    ("Размер", "size_name"),
    ("Бренд", "brand"),
    ("Категория", "subject_name"),
]

CONTROL_COLUMNS = [
    ("Группа", "control_group"),
    ("Артикул поставщика", "vendor_code"),
    ("Артикул WB", "nm_id"),
    ("Название", "product_name"),
    ("Штрафы", "penalty"),
    ("Хранение", "storage"),
    ("Удержания WB по фин. отчету, руб", "deduction"),
    ("Прочие затраты", "other_expenses"),
    ("К перечислению на р/с", "net_for_pay"),
    ("Валовая прибыль finance", "gross_profit"),
]

OTHER_EXPENSE_COLUMNS = [
    ("Операция / группа", "operation"),
    ("Поле WB", "source_field"),
    ("Куда попало в отчете", "target_field"),
    ("Сумма, руб", "amount"),
    ("Правило распределения", "allocation"),
    ("Кол-во артикулов", "items_count"),
]

RAW_FINANCE_COLUMNS = [
    ("Дата операции", "operation_date"),
    ("Тип документа", "doc_type_name"),
    ("Операция WB", "seller_oper_name"),
    ("Артикул поставщика", "vendor_code"),
    ("Артикул WB", "nm_id"),
    ("Баркод", "barcode"),
    ("Размер", "size_name"),
    ("Количество", "quantity"),
    ("Вайлдберриз реализовал, руб", "retail_amount"),
    ("К перечислению, руб", "for_pay"),
    ("Эквайринг, руб", "acquiring_fee"),
    ("Доставка, руб", "delivery_service"),
    ("Штрафы", "penalty"),
    ("Хранение", "paid_storage"),
    ("Удержания", "deduction"),
    ("Приёмка", "paid_acceptance"),
    ("Компенсация скидки", "cashback_amount"),
    ("Баллы/скидка лояльности", "cashback_discount"),
    ("Участие в лояльности", "cashback_commission_change"),
]


def _style_worksheet(worksheet, money_from_column: int = 4):
    header_fill = PatternFill("solid", fgColor="5B4B8A")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    if worksheet.max_row > 1:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = (
            f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
        )
    for column in range(1, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(column)].width = 18
    for row in range(2, worksheet.max_row + 1):
        for column in range(money_from_column, worksheet.max_column + 1):
            worksheet.cell(row, column).number_format = "#,##0.00"


def _append_sheet(workbook, title: str, columns: list[tuple[str, str]], rows):
    worksheet = workbook.create_sheet(title=title)
    worksheet.append([column_title for column_title, _ in columns])
    for item in rows:
        worksheet.append([item.get(field) for _, field in columns])
    _style_worksheet(worksheet)
    return worksheet


def _sum_rows(rows, field: str) -> Decimal:
    return sum((_as_decimal(row.get(field)) for row in rows), ZERO)


def _product_total_row(rows):
    realized_sum = _sum_rows(rows, "realized_sum")
    mp_sum = _sum_rows(rows, "marketplace_commission_sum")
    acquiring_sum = _sum_rows(rows, "acquiring_sum")
    ad_spend = _sum_rows(rows, "advertising_api_spend")
    orders_sum = _sum_rows(rows, "orders_sum")
    gross_profit = _sum_rows(rows, "gross_profit_after_ads")
    net_profit = _sum_rows(rows, "net_profit")
    cost_total = _sum_rows(rows, "cost_total")

    total = {
        "vendor_code": "ИТОГО ПО АРТИКУЛАМ",
        "nm_id": "",
        "product_name": "",
        "barcode": "",
        "size_name": "",
        "brand": "",
        "subject_name": "",
    }
    for field in {
        key for _, key in EXPORT_COLUMNS
    } - set(total):
        total[field] = _money(_sum_rows(rows, field))

    total["orders_qty"] = float(_sum_rows(rows, "orders_qty"))
    total["marketplace_commission_pct"] = _money(
        mp_sum / realized_sum * 100 if realized_sum else ZERO
    )
    total["acquiring_pct"] = _money(
        acquiring_sum / realized_sum * 100 if realized_sum else ZERO
    )
    total["drr"] = _money(ad_spend / orders_sum * 100 if orders_sum else ZERO)
    total["cost_unit"] = None
    total["storage_difference_info"] = 0
    total["advertising_difference_info"] = 0
    total["gross_margin"] = _money(
        gross_profit / realized_sum * 100 if realized_sum else ZERO
    )
    total["roi"] = _money(net_profit / cost_total * 100 if cost_total else ZERO)
    total["net_margin"] = _money(
        net_profit / realized_sum * 100 if realized_sum else ZERO
    )
    total["markup"] = _money(realized_sum / cost_total * 100 if cost_total else ZERO)
    return total


@router.get("/api/v1/nl/opiu/export")
async def export_opiu_report(
    org_id: str,
    date_from: date,
    date_to: date,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    report = await _authorized_report(
        org_id, date_from, date_to, current_user, db
    )
    data = serialize_report(report)
    data = _enrich_serialized_report(
        data,
        *(await _enrichment_context(db, org_id, date_from, date_to)),
    )
    product_rows = [
        item for item in data["items"] if not item.get("is_unassigned")
    ]
    product_rows = [data["total"]] + product_rows
    control_rows = [
        {
            **item,
            "control_group": "Нераспределено: нет артикула/nm_id/barcode",
        }
        for item in data.get("unassigned_items", [])
    ]
    raw_rows = await _load_report_rows(db, org_id, date_from, date_to)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "ОПиУ по артикулам"
    worksheet.append([title for title, _ in EXPORT_COLUMNS])
    for item in product_rows:
        worksheet.append([item.get(field) for _, field in EXPORT_COLUMNS])
    _style_worksheet(worksheet)

    total_fill = PatternFill("solid", fgColor="D9EAD3")
    for cell in worksheet[2]:
        cell.fill = total_fill
        cell.font = Font(bold=True)

    _append_sheet(
        workbook,
        "Расшифровка прочих затрат",
        OTHER_EXPENSE_COLUMNS,
        data.get("allocations", []),
    )
    _append_sheet(
        workbook,
        "Контроль нераспределено",
        CONTROL_COLUMNS,
        control_rows,
    )
    _append_sheet(
        workbook,
        "Raw finance rows",
        RAW_FINANCE_COLUMNS,
        raw_rows,
    )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"opiu_{date_from.isoformat()}_{date_to.isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
