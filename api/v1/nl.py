import uuid
"""API для справочного листа, авторизации и фронтенд НЛ"""
from fastapi import APIRouter, Depends, Query, Request, HTTPException, status
from fastapi.responses import HTMLResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, text
from sqlalchemy.dialects.postgresql import insert as pg_insert
from pydantic import BaseModel
from typing import Optional
from datetime import datetime, date, timedelta

from core.database import get_db
from core.security import verify_password, get_password_hash, create_access_token, decode_token, encrypt_data, decrypt_data
from core.dependencies import get_current_user
from core.role_deps import require_organization_role
from models.organization import Role
from models.user import User
from models.reference_book import ReferenceBook
from models.raw_data import TechStatus
from models.sales_plan import SalesPlan, PlanType, Seasonality

router = APIRouter(tags=["nl"])


# ─── AUTH HELPERS ──────────────────────────────────────────

def get_org_from_token(token: str) -> str:
    """Извлечь org_id из JWT"""
    payload = decode_token(token)
    if not payload:
        return None
    return payload.get("org_id")


# ─── ORG ID RESOLVER ─────────────────────────────────────
async def resolve_org_id(org_id: str, db) -> str:
    """Если org_id — числовой (wb_seller_id), найти UUID организации"""
    try:
        uuid.UUID(org_id)
        return org_id  # Уже UUID
    except ValueError:
        pass
    # Попробовать найти по wb_seller_id
    from sqlalchemy import text as sql_text
    result = await db.execute(
        sql_text("SELECT id FROM organizations WHERE wb_seller_id = :sid"),
        {"sid": int(org_id)}
    )
    row = result.first()
    if row:
        return str(row[0])
    raise HTTPException(status_code=400, detail=f"Организация не найдена: {org_id}")


# ─── API ENDPOINTS ─────────────────────────────────────────

class RegisterData(BaseModel):
    email: str
    password: str
    org_name: str = "Моя организация"


class LoginData(BaseModel):
    email: str
    password: str


class RefItem(BaseModel):
    nm_id: int
    vendor_code: Optional[str] = None
    product_name: Optional[str] = None
    target_date: Optional[str] = None  # YYYY-MM-DD
    cost_price: Optional[float] = None
    purchase_price: Optional[float] = None
    packaging_cost: Optional[float] = None
    logistics_cost: Optional[float] = None
    other_costs: Optional[float] = None
    notes: Optional[str] = None
    product_class: Optional[str] = None
    brand: Optional[str] = None
    tax_system: Optional[str] = None  # usn / osn / usn_dr
    tax_rate: Optional[float] = None
    vat_rate: Optional[float] = None


class SalesPlanItem(BaseModel):
    nm_id: int
    vendor_code: Optional[str] = None
    size_name: Optional[str] = None
    period: str  # YYYY-MM-DD (первый день месяца)
    plan_type: str = "quantity"  # quantity / revenue
    plan_value: float = 0
    actual_value: float = 0
    sales_temp: Optional[float] = None
    seasonality: str = "medium"  # low / medium / high / peak
    entity_id: Optional[str] = None


@router.post("/api/v1/nl/register")
async def nl_register(data: RegisterData, db: AsyncSession = Depends(get_db)):
    """Регистрация нового пользователя + создание организации"""
    from models.organization import Organization, Membership, WbApiKey, Role
    from models.organization import SubscriptionTier, SubscriptionStatus

    # Проверка email
    result = await db.execute(select(User).where(User.email == data.email))
    if result.scalar_one_or_none():
        raise HTTPException(400, "Email уже зарегистрирован")

    # Создаём пользователя
    user = User(email=data.email, password_hash=get_password_hash(data.password))
    db.add(user)
    await db.flush()

    # Создаём организацию
    org = Organization(name=data.org_name, subscription_tier=SubscriptionTier.TRIAL, subscription_status=SubscriptionStatus.ACTIVE)
    db.add(org)
    await db.flush()

    # Привязываем пользователя к организации
    membership = Membership(user_id=user.id, organization_id=org.id, role=Role.OWNER)
    db.add(membership)
    await db.commit()

    # Токен с org_id
    token = create_access_token(data={"sub": str(user.id), "org_id": str(org.id)})
    return {"access_token": token, "org_id": str(org.id)}


@router.post("/api/v1/nl/login")
async def nl_login(data: LoginData, db: AsyncSession = Depends(get_db)):
    """Логин"""
    from models.organization import Membership

    result = await db.execute(select(User).where(User.email == data.email))
    user = result.scalar_one_or_none()
    if not user or not verify_password(data.password, user.password_hash):
        raise HTTPException(401, "Неверный email или пароль")

    # Получаем организацию
    result = await db.execute(
        select(Membership).where(Membership.user_id == user.id)
    )
    membership = result.scalars().first()
    org_id = str(membership.organization_id) if membership else None

    token = create_access_token(data={"sub": str(user.id), "org_id": org_id})
    return {"access_token": token, "org_id": org_id}


@router.get("/api/v1/nl/me")
async def nl_me(token: str = Query(""), db: AsyncSession = Depends(get_db)):
    """Текущий пользователь"""
    payload = decode_token(token)
    if not payload:
        raise HTTPException(401, "Не авторизован")
    user_id = payload.get("sub")
    org_id = payload.get("org_id")
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(401, "Пользователь не найден")
    return {"email": user.email, "org_id": org_id}


@router.get("/api/v1/nl/reference")
async def get_reference(org_id: str, target_date: Optional[str] = None, db: AsyncSession = Depends(get_db)):
    org_id = await resolve_org_id(org_id, db)
    """Справочник — все актуальные записи"""
    from datetime import datetime as dt
    from sqlalchemy import text
    sql = (
        "SELECT nm_id, vendor_code, subject_id, subject_name, cost_price, purchase_cost as purchase_price, packaging_cost, "
        "logistics_cost, other_costs, notes, product_class, brand, tax_system, tax_rate, vat_rate, "
        "valid_from FROM reference_book "
        "WHERE organization_id = :org AND (valid_to IS NULL OR valid_to >= CURRENT_DATE)"
    )
    params = {"org": org_id}
    if target_date:
        sql += " AND valid_from <= :td"
        params["td"] = dt.strptime(target_date, "%Y-%m-%d").date()
    sql += " ORDER BY nm_id, valid_from DESC"
    result = await db.execute(text(sql), params)
    return [{
        "nm_id": r[0],
        "vendor_code": r[1],
        "target_date": str(r[13]),
        "cost_price": float(r[2]) if r[2] else None,
        "purchase_price": float(r[3]) if r[3] else None,
        "packaging_cost": float(r[4]) if r[4] else None,
        "logistics_cost": float(r[5]) if r[5] else None,
        "other_costs": float(r[6]) if r[6] else None,
        "notes": r[7],
        "product_class": r[8],
        "brand": r[9],
        "tax_system": r[10],
        "tax_rate": float(r[11]) if r[11] else None,
        "vat_rate": float(r[12]) if r[12] else None,
    } for r in result.all()]


@router.post("/api/v1/nl/reference")
async def save_reference(item: RefItem, org_id: str, db: AsyncSession = Depends(get_db)):
    """Сохранить строку справочного листа"""
    from datetime import datetime as dt_mod
    t_date = dt_mod.strptime(item.target_date, "%Y-%m-%d").date() if item.target_date else date.today()
    # entity_id lookup
    from sqlalchemy import text as _sql_text
    ent_q = await db.execute(_sql_text(
        "SELECT id FROM product_entities WHERE organization_id = :org AND nm_id = :nm LIMIT 1"
    ), {"org": org_id, "nm": item.nm_id})
    ent_row = ent_q.first()
    eid = ent_row[0] if ent_row else None
    ins = pg_insert(ReferenceBook).values(
        organization_id=org_id, nm_id=item.nm_id, vendor_code=item.vendor_code,
        valid_from=t_date, entity_id=eid,
        cost_price=item.cost_price,
        purchase_cost=item.purchase_price, packaging_cost=item.packaging_cost,
        logistics_cost=item.logistics_cost, other_costs=item.other_costs, notes=item.notes,
        product_class=item.product_class, brand=item.brand,
        tax_system=item.tax_system, tax_rate=item.tax_rate, vat_rate=item.vat_rate,
    )
    stmt = ins.on_conflict_do_update(
        constraint="reference_book_org_entity_vf_key",
        set_={
            "vendor_code": ins.excluded.vendor_code,
            "cost_price": ins.excluded.cost_price, "purchase_cost": ins.excluded.purchase_cost,
            "packaging_cost": ins.excluded.packaging_cost, "logistics_cost": ins.excluded.logistics_cost,
            "other_costs": ins.excluded.other_costs, "notes": ins.excluded.notes,
            "product_class": ins.excluded.product_class, "brand": ins.excluded.brand,
            "tax_system": ins.excluded.tax_system, "tax_rate": ins.excluded.tax_rate, "vat_rate": ins.excluded.vat_rate,
            "updated_at": func.now(),
        }
    )
    await db.execute(stmt)
    await db.commit()
    return {"status": "ok"}


@router.get("/api/v1/nl/products")
async def get_products(org_id: str, target_date: Optional[str] = None, db: AsyncSession = Depends(get_db)):
    org_id = await resolve_org_id(org_id, db)
    """Список уникальных карточек из ТС на дату с entity_id и size_name"""
    from datetime import datetime as dt_mod
    from models.product_entity import ProductEntity, EntityBarcode
    q = select(
        TechStatus.entity_id, TechStatus.nm_id, TechStatus.vendor_code, TechStatus.product_name,
        TechStatus.photo_main, TechStatus.barcode, TechStatus.sku
    ).where(TechStatus.organization_id == org_id, TechStatus.nm_id.isnot(None))
    if target_date:
        q = q.where(TechStatus.target_date == dt_mod.strptime(target_date, "%Y-%m-%d").date())
    q = q.distinct()
    result = await db.execute(q)

    # Получаем маппинг entity_id → size_name
    ent_result = await db.execute(
        select(ProductEntity.id, ProductEntity.size_name, ProductEntity.subject_name).where(
            ProductEntity.organization_id == org_id
        )
    )
    size_map = {str(r[0]): r[1] for r in ent_result.all()}

    # Получаем все активные ШК по сущностям
    bc_result = await db.execute(
        select(EntityBarcode.entity_id, EntityBarcode.barcode).where(
            EntityBarcode.organization_id == org_id,
            EntityBarcode.is_active == True,
        )
    )
    barcode_map = {}
    for r in bc_result.all():
        eid = str(r[0])
        if eid not in barcode_map:
            barcode_map[eid] = []
        barcode_map[eid].append(r[1])

    items = []
    seen = set()
    for r in result.all():
        eid = str(r[0]) if r[0] else None
        key = (r[1], eid)  # уникальность по nm_id + entity_id
        if key in seen:
            continue
        seen.add(key)
        items.append({
            "entity_id": eid,
            "nm_id": r[1],
            "vendor_code": r[2],
            "product_name": r[3],
            "photo_main": r[4],
            "barcode": r[5],
            "sku": r[6],
            "size_name": size_map.get(eid, "") if eid else "",
            "barcodes": barcode_map.get(eid, []) if eid else ([r[5]] if r[5] else []),
        })
    return items




@router.get("/api/v1/nl/organizations")
async def nl_organizations(token: str = Query(""), db: AsyncSession = Depends(get_db)):
    """Список организаций пользователя"""
    from models.organization import Membership, Organization, WbApiKey
    payload = decode_token(token)
    if not payload:
        raise HTTPException(401, "Не авторизован")
    user_id = payload.get("sub")
    result = await db.execute(
        select(Membership, Organization)
        .join(Organization, Membership.organization_id == Organization.id)
        .where(Membership.user_id == user_id)
    )
    rows = result.all()
    orgs = []
    for m, o in rows:
        # Count WB keys
        keys_result = await db.execute(
            select(WbApiKey).where(WbApiKey.organization_id == o.id)
        )
        keys = keys_result.scalars().all()
        orgs.append({
            "id": str(o.id),
            "name": o.name,
            "wb_seller_id": o.wb_seller_id,
            "wb_keys_count": len(keys),
            "role": m.role.value if hasattr(m.role, "value") else str(m.role),
        })
    return orgs


@router.post("/api/v1/nl/organizations")
async def nl_create_org(data: dict, token: str = Query(""), db: AsyncSession = Depends(get_db)):
    """Создать новую организацию"""
    from models.organization import Organization, Membership, Role, SubscriptionTier, SubscriptionStatus
    payload = decode_token(token)
    if not payload:
        raise HTTPException(401, "Не авторизован")
    user_id = payload.get("sub")
    org = Organization(name=data.get("name", "Новый магазин"), subscription_tier=SubscriptionTier.TRIAL, subscription_status=SubscriptionStatus.ACTIVE)
    db.add(org)
    await db.flush()
    membership = Membership(user_id=user_id, organization_id=org.id, role=Role.OWNER)
    db.add(membership)
    await db.commit()
    return {"id": str(org.id), "name": org.name}


@router.post("/api/v1/nl/connect-wb")
async def nl_connect_wb(data: dict, token: str = Query(""), db: AsyncSession = Depends(get_db)):
    """Подключить новый магазин WB: создать организацию + ключ + membership"""
    import base64, json as _json
    from models.organization import Organization, Membership, WbApiKey, Role, SubscriptionTier, SubscriptionStatus
    from core.security import encrypt_data
    
    payload = decode_token(token)
    if not payload:
        raise HTTPException(401, "Не авторизован")
    user_id = payload.get("sub")
    
    name = data.get("name", "").strip()
    api_key = data.get("api_key", "").strip()
    
    if not api_key:
        raise HTTPException(400, "API ключ обязателен")
    if not name:
        name = "Новый магазин"
    
    # Parse JWT to extract seller_id (oid) as integer
    oid = None
    try:
        parts = api_key.split(".")
        if len(parts) >= 2:
            payload_b64 = parts[1] + "=" * (4 - len(parts[1]) % 4)
            payload_json = base64.urlsafe_b64decode(payload_b64)
            payload_data = _json.loads(payload_json)
            oid_val = payload_data.get("oid")
            if oid_val is not None:
                oid = int(oid_val)
    except Exception:
        pass
    
    # Check if org with this seller_id already exists for this user
    if oid:
        result = await db.execute(
            select(Organization, Membership)
            .join(Membership, Membership.organization_id == Organization.id)
            .where(Organization.wb_seller_id == oid, Membership.user_id == user_id)
        )
        existing = result.first()
        if existing:
            raise HTTPException(400, f"Магазин с seller_id {oid} уже подключен ({existing[0].name})")
    
    # Create organization
    org = Organization(
        name=name,
        wb_seller_id=oid,
        subscription_tier=SubscriptionTier.TRIAL,
        subscription_status=SubscriptionStatus.ACTIVE
    )
    db.add(org)
    await db.flush()
    
    # Create membership (OWNER)
    membership = Membership(user_id=user_id, organization_id=org.id, role=Role.OWNER)
    db.add(membership)
    
    # Create WB API key (encrypted)
    encrypted = encrypt_data(api_key)
    wb_key = WbApiKey(organization_id=org.id, name=name, personal_token=encrypted, api_key="unused")
    db.add(wb_key)
    
    await db.commit()
    
    return {
        "status": "ok",
        "org_id": str(org.id),
        "name": org.name,
        "wb_seller_id": oid,
        "key_id": str(wb_key.id)
    }


@router.get("/api/v1/nl/tax-settings")
async def get_tax_settings(org_id: str, db: AsyncSession = Depends(get_db)):
    """Налоговые настройки кабинета"""
    from models.organization import Organization
    result = await db.execute(
        select(Organization).where(Organization.id == org_id)
    )
    org = result.scalar_one_or_none()
    if not org:
        raise HTTPException(404, "Организация не найдена")
    return {
        "tax_system": org.tax_system or "",
        "tax_rate": float(org.tax_rate) if org.tax_rate else None,
        "vat_type": org.vat_type or "нет"
    }


@router.post("/api/v1/nl/tax-settings")
async def save_tax_settings(data: dict, org_id: str, db: AsyncSession = Depends(get_db)):
    """Сохранить налоговые настройки кабинета"""
    from models.organization import Organization
    result = await db.execute(
        select(Organization).where(Organization.id == org_id)
    )
    org = result.scalar_one_or_none()
    if not org:
        raise HTTPException(404, "Организация не найдена")
    
    allowed_systems = ["УСН Доходы", "УСН Доходы-Расходы", "ОСНО", "АУСН Доходы", "АУСН Доходы-Расходы"]
    ts = data.get("tax_system", "")
    if ts and ts not in allowed_systems:
        raise HTTPException(400, f"Недопустимый вид налогообложения. Допустимые: {', '.join(allowed_systems)}")
    
    allowed_vat = ["нет", "5%", "7%"]
    vt = data.get("vat_type", "нет")
    if vt and vt not in allowed_vat:
        raise HTTPException(400, f"Недопустимый НДС. Допустимые: {', '.join(allowed_vat)}")
    
    org.tax_system = ts if ts else None
    org.tax_rate = float(data["tax_rate"]) if data.get("tax_rate") else None
    org.vat_type = vt if vt else "нет"
    await db.commit()
    
    return {
        "tax_system": org.tax_system or "",
        "tax_rate": float(org.tax_rate) if org.tax_rate else None,
        "vat_type": org.vat_type or "нет"
    }


@router.get("/api/v1/nl/wb-keys")
async def nl_wb_keys(org_id: str, db: AsyncSession = Depends(get_db)):
    """Список WB API ключей организации"""
    from models.organization import WbApiKey
    result = await db.execute(
        select(WbApiKey).where(WbApiKey.organization_id == org_id)
    )
    keys = result.scalars().all()
    return [{"id": str(k.id), "name": k.name, "created_at": str(k.created_at)} for k in keys]


@router.post("/api/v1/nl/wb-keys")
async def nl_add_wb_key(data: dict, org_id: str, db: AsyncSession = Depends(get_db)):
    """Добавить WB API ключ"""
    import base64, json as _json
    from models.organization import WbApiKey, Organization
    name = data.get("name", "WB Key")
    api_key = data.get("api_key", "")
    if not api_key:
        raise HTTPException(400, "API ключ обязателен")
    encrypted = encrypt_data(api_key)
    key = WbApiKey(organization_id=org_id, name=name, personal_token=encrypted, api_key="unused")
    db.add(key)
    # Извлекаем wb_seller_id (oid) из JWT payload
    try:
        parts = api_key.split(".")
        if len(parts) >= 2:
            payload_b64 = parts[1] + "=" * (4 - len(parts[1]) % 4)
            payload_json = base64.urlsafe_b64decode(payload_b64)
            payload_data = _json.loads(payload_json)
            oid = payload_data.get("oid")
            if oid:
                result = await db.execute(select(Organization).where(Organization.id == org_id))
                org = result.scalar_one_or_none()
                if org:
                    org.wb_seller_id = oid
    except Exception:
        pass  # Не критично если не удалось распарсить
    await db.commit()
    return {"status": "ok", "id": str(key.id)}


@router.delete("/api/v1/nl/wb-keys/{key_id}")
async def nl_delete_wb_key(key_id: str, org_id: str, db: AsyncSession = Depends(get_db)):
    """Удалить WB API ключ"""
    from models.organization import WbApiKey
    result = await db.execute(
        select(WbApiKey).where(WbApiKey.id == key_id, WbApiKey.organization_id == org_id)
    )
    key = result.scalar_one_or_none()
    if not key:
        raise HTTPException(404, "Ключ не найден")
    await db.delete(key)
    await db.commit()
    return {"status": "ok"}




@router.get("/api/v1/nl/fbs-warehouses")
async def get_fbs_warehouses(org_id: str, db: AsyncSession = Depends(get_db)):
    """Список складов FBS + СЦ(СГТ) + КГТ+ из WB API (кэш 24ч)"""
    import json, time
    from models.organization import WbApiKey
    cache_key = "fbs_wh_v2_" + org_id
    cache_file = "/tmp/" + cache_key + ".json"
    try:
        with open(cache_file, "r") as f:
            cache = json.load(f)
            if time.time() - cache["ts"] < 86400:
                return cache["data"]
    except (FileNotFoundError, json.JSONDecodeError, KeyError):
        pass
    result = await db.execute(
        select(WbApiKey).where(WbApiKey.organization_id == org_id)
    )
    key_rec = result.scalar_one_or_none()
    if not key_rec or not key_rec.personal_token:
        return []
    token = decrypt_data(key_rec.personal_token)
    try:
        from services.wb_api.client import WBApiClient
        async with WBApiClient(token) as client:
            all_offices = await client.get_all_offices()
    except Exception as e:
        raise HTTPException(502, f"WB API error: {e}")
    # Группируем по cargoType: 1=склад, 2=СЦ(СГТ), 3=КГТ+
    type_names = {1: "Склад", 2: "СЦ", 3: "КГТ+"}
    merged = []
    seen_ids = set()
    for ct in [1, 2, 3]:
        items = [o for o in all_offices if o.get("cargoType") == ct]
        for o in items:
            if o["id"] not in seen_ids:
                seen_ids.add(o["id"])
                merged.append({
                    "id": o["id"],
                    "name": o["name"],
                    "address": o.get("address", ""),
                    "city": o.get("city", ""),
                    "type": type_names.get(ct, "?"),
                    "cargoType": ct,
                })
    # Фильтруем: оставляем только склады с ФБС-тарифами в wb_box_tariffs
    from sqlalchemy import text as sa_text
    tarif_rows = await db.execute(sa_text(
        "SELECT DISTINCT warehouse_name FROM wb_box_tariffs "
        "WHERE box_delivery_marketplace_base IS NOT NULL "
        "AND organization_id = :org"
    ), {"org": org_id})
    tarif_names = set(r[0] for r in tarif_rows.fetchall())
    
    if tarif_names:
        filtered = []
        for wh in merged:
            has_tarif = any(
                wh["name"] in tn or tn in wh["name"]
                for tn in tarif_names
            )
            if has_tarif:
                filtered.append(wh)
        merged = filtered
    
    with open(cache_file, "w") as f:
        json.dump({"ts": time.time(), "data": merged}, f, ensure_ascii=False)
    return merged

@router.get("/api/v1/nl/dates")
async def get_available_dates(org_id: str, db: AsyncSession = Depends(get_db)):
    """Доступные даты в ТС"""
    result = await db.execute(
        select(TechStatus.target_date)
        .where(TechStatus.organization_id == org_id)
        .distinct()
        .order_by(TechStatus.target_date.desc())
        .limit(30)
    )
    return [str(r[0]) for r in result.all()]


@router.get("/api/v1/nl/control")
async def get_control_metrics(org_id: str, target_date: Optional[str] = None, db: AsyncSession = Depends(get_db)):
    org_id = await resolve_org_id(org_id, db)
    """Оперативный контроль — метрики на дату"""
    from sqlalchemy import func, case, and_
    from datetime import datetime as dt_mod
    import decimal

    d = dt_mod.strptime(target_date, "%Y-%m-%d").date() if target_date else date.today()

    # Основные метрики за день
    result = await db.execute(
        select(
            func.count(TechStatus.id).label("total_products"),
            func.sum(TechStatus.stock_qty).label("total_stock"),
            func.sum(TechStatus.stock_fbo_qty).label("total_stock_fbo"),
            func.sum(TechStatus.orders_count).label("total_orders"),
            func.sum(TechStatus.buyouts_count).label("total_buyouts"),
            func.sum(TechStatus.returns_count).label("total_returns"),
            func.sum(TechStatus.impressions).label("total_impressions"),
            func.sum(TechStatus.clicks).label("total_clicks"),
            func.sum(TechStatus.ad_cost).label("total_ad_cost"),
            func.sum(TechStatus.price_discount * TechStatus.buyouts_count).label("total_revenue"),
            func.sum(TechStatus.price_discount * TechStatus.buyouts_count).label("total_revenue_gross"),
            func.avg(TechStatus.rating).label("avg_rating"),
        ).where(TechStatus.organization_id == org_id, TechStatus.target_date == d)
    )
    row = result.one()

    # Товары с нулевым остатком
    zero_stock = await db.execute(
        select(func.count(TechStatus.id)).where(
            TechStatus.organization_id == org_id, TechStatus.target_date == d,
            TechStatus.stock_qty <= 0
        )
    )

    # Товары с низким остатком (<=5)
    low_stock = await db.execute(
        select(func.count(TechStatus.id)).where(
            TechStatus.organization_id == org_id, TechStatus.target_date == d,
            TechStatus.stock_qty > 0, TechStatus.stock_qty <= 5
        )
    )

    # Товары по рейтингу (< 4)
    low_rating = await db.execute(
        select(func.count(TechStatus.id)).where(
            TechStatus.organization_id == org_id, TechStatus.target_date == d,
            TechStatus.rating < 4.0
        )
    )

    # Детализация по товарам (с entity_id)
    products_detail = await db.execute(
        select(
            TechStatus.entity_id, TechStatus.nm_id, TechStatus.vendor_code, TechStatus.product_name,
            TechStatus.photo_main, TechStatus.stock_qty, TechStatus.stock_fbo_qty, TechStatus.orders_count,
            TechStatus.buyouts_count, TechStatus.returns_count, TechStatus.rating,
            TechStatus.impressions, TechStatus.clicks, TechStatus.ad_cost,
            TechStatus.price, TechStatus.price_discount, TechStatus.tariff,
            TechStatus.barcode,
        ).where(TechStatus.organization_id == org_id, TechStatus.target_date == d)
        .order_by(TechStatus.orders_count.desc().nullslast())
    )

    # Маппинг entity_id -> size_name + Д×Ш×В, вес, объём (факт)
    from models.product_entity import ProductEntity
    ent_result = await db.execute(
        select(ProductEntity.id, ProductEntity.size_name,
               ProductEntity.length, ProductEntity.width, ProductEntity.height,
               ProductEntity.weight).where(
            ProductEntity.organization_id == org_id
        )
    )
    _ent_rows = ent_result.all()
    size_map = {str(r[0]): r[1] for r in _ent_rows}
    dims_map = {}
    for r in _ent_rows:
        eid = str(r[0])
        l, w, h = r[2], r[3], r[4]
        wt = r[5]
        vol = round((l * w * h) / 1000, 2) if l and w and h else None
        dims_map[eid] = {"length": l, "width": w, "height": h, "weight": wt, "volume": vol}

    # Маппинг entity_id -> все ШК (для поиска)
    from models.product_entity import EntityBarcode
    bc_result = await db.execute(
        select(EntityBarcode.entity_id, EntityBarcode.barcode).where(
            EntityBarcode.is_active == True
        )
    )
    barcodes_map = {}
    for r in bc_result.all():
        eid = str(r[0])
        if eid not in barcodes_map:
            barcodes_map[eid] = []
        barcodes_map[eid].append(r[1])

    # --- Юнит Экономика для ТС ---
    from sqlalchemy import text

    # Себестоимость и справочник
    ref_result = await db.execute(text(
        "SELECT entity_id, nm_id, cost_price, product_class, brand, tax_system, tax_rate, vat_rate, wb_price_fact "
        "FROM reference_book WHERE organization_id = :org AND (valid_to IS NULL OR valid_to >= CURRENT_DATE)"
    ), {"org": org_id})
    ref_by_entity = {}
    ref_by_nm = {}
    for r in ref_result.all():
        d_item = {
            "cost_price": float(r[2]) if r[2] else 0,
            "product_class": r[3] or "",
            "brand": r[4] or "",
            "tax_system": r[5] or "",
            "tax_rate": float(r[6]) if r[6] else 0,
            "vat_rate": float(r[7]) if r[7] else 0,
            "wb_price_fact": float(r[8]) if r[8] else None,
        }
        if r[0]:
            ref_by_entity[str(r[0])] = d_item
        if r[1]:
            ref_by_nm[r[1]] = d_item

    # WB тарифы (из wb_tariff_snapshot)
    snap_result = await db.execute(text(
        "SELECT nm_id, logistics_tariff, storage_tariff, commission_pct, buyout_pct_fact "
        "FROM wb_tariff_snapshot WHERE organization_id = :org ORDER BY target_date DESC"
    ), {"org": org_id})
    snap_by_nm_ts = {}
    for r in snap_result.all():
        if r[0] not in snap_by_nm_ts:
            snap_by_nm_ts[r[0]] = {
                "logistics_tariff": float(r[1]) if r[1] else 0,
                "storage_tariff": float(r[2]) if r[2] else 0,
                "commission_pct": float(r[3]) if r[3] else 0,
                "buyout_pct_fact": float(r[4]) if r[4] else 0,
            }

    def _get_ref(eid, nm):
        return ref_by_nm.get(nm, ref_by_entity.get(eid, {"cost_price":0,"product_class":"","brand":"","tax_system":"","tax_rate":0,"vat_rate":0,"wb_price_fact":None}))

    def _get_snap(nm):
        return snap_by_nm_ts.get(nm, {"logistics_tariff":0,"storage_tariff":0,"commission_pct":0,"buyout_pct_fact":0})

    # subject_map for product display
    _subj_result = await db.execute(
        select(ProductEntity.id, ProductEntity.subject_name)
        .where(ProductEntity.organization_id == org_id)
    )
    subject_map = {str(r[0]): r[1] or "" for r in _subj_result.all()}

    def _calc_unit(price_disc, ad_cost, ref, snap):
        """Упрощённый расчёт юнитки для ТС"""
        p = float(price_disc or 0)
        a = float(ad_cost or 0)
        cp = ref["cost_price"]
        comm = p * snap["commission_pct"] / 100 if p and snap["commission_pct"] else 0
        logist = snap["logistics_tariff"]
        expenses = round(cp + comm + logist + a, 2)
        profit = round(p - expenses, 2)
        margin = round(profit / p * 100, 1) if p else 0
        roi = round(profit / cp * 100, 1) if cp else 0
        return {"unit_expenses": expenses, "unit_profit": profit, "unit_margin": margin, "unit_roi": roi}

    def safe_float(v):
        return float(v) if v is not None and not isinstance(v, decimal.Decimal) else (float(v) if isinstance(v, decimal.Decimal) else None)
    def safe_int(v):
        return int(v) if v is not None else None

    total_clicks = safe_int(row.total_clicks) or 0
    total_impressions = safe_int(row.total_impressions) or 0

    return {
        "date": str(d),
        "summary": {
            "total_products": safe_int(row.total_products) or 0,
            "total_stock": (safe_int(row.total_stock) or 0) + (safe_int(row.total_stock_fbo) or 0),
            "total_stock_fbo": safe_int(row.total_stock_fbo) or 0,
            "total_stock_fbs": safe_int(row.total_stock) or 0,
            "total_orders": safe_int(row.total_orders) or 0,
            "total_buyouts": safe_int(row.total_buyouts) or 0,
            "total_returns": safe_int(row.total_returns) or 0,
            "total_impressions": total_impressions,
            "total_clicks": total_clicks,
            "ctr": round(total_clicks / total_impressions * 100, 2) if total_impressions > 0 else 0,
            "total_ad_cost": safe_float(row.total_ad_cost) or 0,
            "total_revenue": safe_float(row.total_revenue) or 0,
            "avg_rating": round(float(row.avg_rating), 2) if row.avg_rating else None,
            "zero_stock_count": safe_int(zero_stock.scalar()) or 0,
            "low_stock_count": safe_int(low_stock.scalar()) or 0,
            "low_rating_count": safe_int(low_rating.scalar()) or 0,
        },
        "products": (lambda rows: [
            {**{
                "entity_id": str(r[0]) if r[0] else None,
                "nm_id": r[1],
                "vendor_code": r[2],
                "product_name": r[3],
                "photo_main": r[4],
                "stock_qty": safe_int(r[5]),
                "stock_fbo_qty": safe_int(r[6]),
                "orders_count": safe_int(r[7]),
                "buyouts_count": safe_int(r[8]),
                "returns_count": safe_int(r[9]),
                "rating": safe_float(r[10]),
                "impressions": safe_int(r[11]),
                "clicks": safe_int(r[12]),
                "ad_cost": safe_float(r[13]),
                "price": safe_float(r[14]),
                "price_discount": safe_float(r[15]),
                "tariff": safe_float(r[16]),
                "barcode": r[17] or "",
                "barcodes": ", ".join(barcodes_map.get(str(r[0]), [])) or (r[17] or ""),
                "size_name": size_map.get(str(r[0]), "") if r[0] else "",
                "subject_name": subject_map.get(str(r[0]), "") if r[0] else "",
                **(dims_map.get(str(r[0]), {}) if r[0] else {}),
            },
            **{k: v for k, v in _get_ref(str(r[0]) if r[0] else "", r[1]).items()},
            **{f"snap_{k}": v for k, v in _get_snap(r[1]).items()},
            **_calc_unit(safe_float(r[14]), safe_float(r[12]), _get_ref(str(r[0]) if r[0] else "", r[1]), _get_snap(r[1])),
            }
            for r in rows
        ])(products_detail.all())
    }



# ─── FRONTEND ──────────────────────────────────────────────




# === РНП — API эндпоинт ===
# Вставить в api/v1/nl.py перед @router.get("/nl/register"...)

@router.get("/api/v1/nl/rnp")
async def get_rnp(
    org_id: str,
    month: Optional[str] = None,  # YYYY-MM (например 2026-05)
    days: Optional[int] = None,   # Количество дней назад (30, 60, 90)
    sort_by: Optional[str] = "orders_revenue",  # orders_revenue, roi, buyout_pct
    filter_status: Optional[str] = None,
    search: Optional[str] = None,
    use_buyout_pct: bool = False,  # чекбокс "Учесть % выкупа"
    db: AsyncSession = Depends(get_db),
):
    """
    РНП — Раздел Нормативных Показателей.
    Данные за месяц: каждая карточка = строка, дни = столбцы.
    """
    from datetime import datetime as dt_mod
    import calendar
    import decimal

    from models.product_entity import ProductEntity
    # Период: N дней назад до сегодня, или выбранный месяц
    today = date.today()
    if month:
        year, mon = month.split("-")
        year, mon = int(year), int(mon)
        first_day = date(year, mon, 1)
        last_day = date(year, mon, calendar.monthrange(year, mon)[1])
        days_in_month = calendar.monthrange(year, mon)[1]
    else:
        num_days = days if days else 90
        last_day = today
        first_day = today - timedelta(days=num_days - 1)
        days_in_month = num_days

    # 1. Список дней (по убыванию)
    day_list = []
    d = min(last_day, today)
    while d >= first_day:
        day_list.append(d)
        d -= timedelta(days=1)

    from models.product_entity import ProductEntity

    # 2. Получаем entity_id -> размер
    ent_result = await db.execute(
        select(ProductEntity.id, ProductEntity.size_name, ProductEntity.nm_id, ProductEntity.subject_name)
        .where(ProductEntity.organization_id == org_id)
    )
    _ent_rows = ent_result.all()
    size_map = {str(r[0]): r[1] for r in _ent_rows}

    # 3. Справочник (reference_book) — последние записи по entity
    ref_result = await db.execute(text(
        "SELECT DISTINCT ON (entity_id) entity_id, nm_id, cost_price, purchase_cost, "
        "packaging_cost, logistics_cost, other_costs, extra_costs, vat, "
        "mp_base_pct, mp_correction_pct, tax_system, tax_rate, vat_rate, "
        "product_class, brand, product_status, subject_id, subject_name, "
        "in_promo, ad_shows_organic, ad_shows_paid, ad_strategy, tags, rating_reviews, localization_pct "
        "FROM reference_book "
        "WHERE organization_id = :org AND entity_id IS NOT NULL "
        "AND (valid_to IS NULL OR valid_to >= :fd) "
        "ORDER BY entity_id, valid_from DESC"
    ), {"org": org_id, "fd": first_day})
    ref_map = {}  # entity_id -> dict
    ref_map_nm = {}  # nm_id -> dict (fallback)
    import logging as _log2
    _l = _log2.getLogger(__name__)
    for r in ref_result.all():
        d_item = {
            "cost_price": float(r[2]) if r[2] else 0,
            "purchase_cost": float(r[3]) if r[3] else 0,
            "packaging_cost": float(r[4]) if r[4] else 0,
            "logistics_cost": float(r[5]) if r[5] else 0,
            "other_costs": float(r[6]) if r[6] else 0,
            "extra_costs": float(r[7]) if r[7] else 0,
            "vat": float(r[8]) if r[8] else 0,
            "mp_base_pct": float(r[9]) if r[9] else 0,
            "mp_correction_pct": float(r[10]) if r[10] else 0,
            "tax_system": r[11] or "",
            "tax_rate": float(r[12]) if r[12] else 0,
            "vat_rate": float(r[13]) if r[13] else 0,
            "product_class": r[14] or "",
            "brand": r[15] or "",
            "product_status": r[16] or "",
            "subject_id": r[17],
            "subject_name": r[18] or "",
            "in_promo": bool(r[19]) if r[19] is not None else False,
            "ad_shows_organic": int(r[20]) if r[20] else None,
            "ad_shows_paid": int(r[21]) if r[21] else None,
            "ad_strategy": r[22] or "",
            "tags": r[23] or "",
            "rating_reviews": float(r[24]) if r[24] else None,
            "localization_pct": r[25] or "",
        }
        if r[0]:
            ref_map[str(r[0])] = d_item
        if r[1]:
            ref_map_nm[r[1]] = d_item
    print(f"[CONTROL-DEBUG] ref_map size={len(ref_map)}, sample={list(ref_map.values())[0] if ref_map else None}")

    def get_ref(eid, nm):
        return ref_map_nm.get(nm, ref_map.get(eid, {
            "cost_price": 0, "purchase_cost": 0, "packaging_cost": 0,
            "logistics_cost": 0, "other_costs": 0, "extra_costs": 0, "vat": 0,
            "mp_base_pct": 0, "mp_correction_pct": 0, "tax_system": "",
            "tax_rate": 0, "vat_rate": 0, "product_class": "", "brand": "",
            "product_status": "", "in_promo": False, "ad_shows_organic": None,
            "ad_shows_paid": None, "ad_strategy": "", "tags": "",
            "rating_reviews": None, "localization_pct": "",
        }))

    # 4. WB тарифы — последние по nm_id
    snap_result = await db.execute(text(
        "SELECT DISTINCT ON (nm_id) nm_id, logistics_tariff, storage_tariff, "
        "commission_pct, buyout_pct_fact, price_retail, price_with_spp, spp_pct "
        "FROM wb_tariff_snapshot "
        "WHERE organization_id = :org "
        "ORDER BY nm_id, target_date DESC"
    ), {"org": org_id})
    snap_map = {}
    for r in snap_result.all():
        snap_map[r[0]] = {
            "logistics_tariff": float(r[1]) if r[1] else 0,
            "storage_tariff": float(r[2]) if r[2] else 0,
            "commission_pct": float(r[3]) if r[3] else 0,
            "buyout_pct_fact": float(r[4]) if r[4] else 0,
            "price_retail": float(r[5]) if r[5] else 0,
            "price_with_spp": float(r[6]) if r[6] else 0,
            "spp_pct": float(r[7]) if r[7] else 0,
        }

    def get_snap(nm):
        return snap_map.get(nm, {
            "logistics_tariff": 0, "storage_tariff": 0,
            "commission_pct": 0, "buyout_pct_fact": 0,
            "price_retail": 0, "price_with_spp": 0, "spp_pct": 0,
        })

    # 5. План продаж
    plan_result = await db.execute(text(
        "SELECT entity_id, plan_type, plan_value "
        "FROM sales_plans "
        "WHERE organization_id = :org AND period = :period"
    ), {"org": org_id, "period": first_day})
    plan_map = {}  # entity_id -> {quantity: X, revenue: Y}
    for r in plan_result.all():
        eid = str(r[0]) if r[0] else None
        if not eid:
            continue
        if eid not in plan_map:
            plan_map[eid] = {"quantity": 0, "revenue": 0}
        ptype = str(r[1]) if r[1] else "quantity"
        plan_map[eid][ptype] = float(r[2]) if r[2] else 0

    # 6. tech_status за весь месяц — агрегация по entity_id + target_date
    ts_result = await db.execute(text(
        "SELECT entity_id, target_date, nm_id, vendor_code, product_name, photo_main, barcode, "
        "orders_count, buyouts_count, returns_count, stock_qty, "
        "price, price_discount, price_spp, ad_cost, impressions, clicks, tariff "
        "FROM tech_status "
        "WHERE organization_id = :org AND target_date BETWEEN :fd AND :ld "
        "AND entity_id IS NOT NULL "
        "ORDER BY entity_id, target_date DESC"
    ), {"org": org_id, "fd": first_day, "ld": last_day})
    ts_rows = ts_result.all()

    # Группируем по entity_id
    from collections import defaultdict
    entities_data = defaultdict(lambda: {"days": {}, "last_row": None})
    for r in ts_rows:
        eid = str(r[0])
        tdate = r[1]
        entities_data[eid]["days"][str(tdate)] = {
            "date": str(tdate),
            "orders_count": int(r[7]) if r[7] else 0,
            "buyouts_count": int(r[8]) if r[8] else 0,
            "returns_count": int(r[9]) if r[9] else 0,
            "stock_qty": int(r[10]) if r[10] else 0,
            "price": float(r[11]) if r[11] else 0,
            "price_discount": float(r[12]) if r[12] else 0,
            "price_spp": float(r[13]) if r[13] else 0,
            "ad_cost": float(r[14]) if r[14] else 0,
            "impressions": int(r[15]) if r[15] else 0,
            "clicks": int(r[16]) if r[16] else 0,
            "tariff": float(r[17]) if r[17] else 0,
        }
        # Сохраняем последнюю строку для идентификации
        if entities_data[eid]["last_row"] is None:
            entities_data[eid]["last_row"] = r

    # 7. Собираем результат
    def safe_f(v): return float(v) if v is not None else 0
    def safe_i(v): return int(v) if v is not None else 0

    products = []
    for eid, edata in entities_data.items():
        last = edata["last_row"]
        if not last:
            continue
        nm = last[2]
        ref = get_ref(eid, nm)
        snap = get_snap(nm)
        plan = plan_map.get(eid, {"quantity": 0, "revenue": 0})

        # Фильтр поиска
        if search:
            vc = str(last[3] or "")
            pn = str(last[4] or "")
            if not (search.lower() in vc.lower() or search.lower() in pn.lower() or (search.isdigit() and int(search) == nm)):
                continue

        # Фильтр по статусу
        if filter_status and ref["product_status"] != filter_status:
            continue

        # Последний сток
        last_day_key = max(edata["days"].keys()) if edata["days"] else None
        last_day_data = edata["days"].get(last_day_key, {})
        current_stock = last_day_data.get("stock_qty", 0)

        # Сумма за месяц
        total_orders = sum(d["orders_count"] for d in edata["days"].values())
        total_buyouts = sum(d["buyouts_count"] for d in edata["days"].values())
        total_ad_cost = sum(d["ad_cost"] for d in edata["days"].values())
        total_orders_revenue = sum(d["orders_count"] * d["price_discount"] for d in edata["days"].values())
        total_buyouts_revenue = sum(d["buyouts_count"] * d["price_discount"] for d in edata["days"].values())

        # Себестоимость единицы
        total_cost = ref["cost_price"] + ref["purchase_cost"] + ref["packaging_cost"] + ref["logistics_cost"] + ref["other_costs"] + ref["extra_costs"] + ref["vat"]

        # Комиссия МП
        mp_pct = (ref["mp_base_pct"] or snap["commission_pct"]) + ref["mp_correction_pct"]

        # Маржа до ДРР (на выкуп)
        commission = total_buyouts_revenue * mp_pct / 100 if total_buyouts_revenue else 0
        logistics = snap["logistics_tariff"] * total_buyouts
        margin_before_drr = total_buyouts_revenue - (total_cost * total_buyouts) - commission - logistics

        # Прибыль расчёт
        profit_calc = margin_before_drr - total_ad_cost

        # Маржа с ДРР
        margin_with_drr = profit_calc

        # ДРР
        drr = round(total_ad_cost / total_orders_revenue * 100, 2) if total_orders_revenue else 0

        # КРРР
        krrr = round(margin_with_drr / margin_before_drr * 100, 1) if margin_before_drr else 0

        # Себестоимость остатков
        cost_of_stock = round(total_cost * current_stock, 2)

        # % выкупа
        buyout_pct = snap["buyout_pct_fact"] or 0

        # План / факт / % выполнения
        plan_val = plan.get("revenue", 0) or plan.get("quantity", 0)
        days_passed = len(edata["days"])
        daily_norm = round(plan_val / days_in_month, 2) if days_in_month else 0
        pct_complete = round(total_orders_revenue / (daily_norm * days_passed) * 100, 1) if daily_norm and days_passed else 0

        # «Хватит на» дней
        avg_orders_day = total_orders / days_passed if days_passed else 0
        if use_buyout_pct and buyout_pct > 0:
            effective_demand = avg_orders_day * buyout_pct / 100
        else:
            effective_demand = avg_orders_day
        enough_days = round(current_stock / effective_demand, 1) if effective_demand > 0 else 999

        # ROI
        total_invested = total_cost * (total_orders if not use_buyout_pct else total_buyouts)
        roi = round(total_buyouts_revenue / total_invested * 100 - 100, 1) if total_invested else 0

        # CPL
        total_clicks = sum(d["clicks"] for d in edata["days"].values())
        cpl = round(total_clicks / total_orders, 2) if total_orders else 0

        # CTR
        total_impressions = sum(d["impressions"] for d in edata["days"].values())
        ctr = round(total_clicks / total_impressions * 100, 2) if total_impressions else 0

        # Дни для столбцов (по убыванию)
        day_columns = []
        for day in day_list:
            dk = str(day)
            dd = edata["days"].get(dk, {
                "date": dk, "orders_count": 0, "buyouts_count": 0,
                "returns_count": 0, "stock_qty": 0, "price": 0,
                "price_discount": 0, "price_spp": 0, "ad_cost": 0,
                "impressions": 0, "clicks": 0, "tariff": 0,
            })
            o_rev = dd["orders_count"] * dd["price_discount"]
            b_rev = dd["buyouts_count"] * dd["price_discount"]
            dd_comm = b_rev * mp_pct / 100 if b_rev else 0
            dd_logist = snap["logistics_tariff"] * dd["buyouts_count"]
            dd_margin_before = b_rev - (total_cost * dd["buyouts_count"]) - dd_comm - dd_logist
            dd_profit = dd_margin_before - dd["ad_cost"]
            dd_margin_with = dd_profit
            dd_drr = round(dd["ad_cost"] / o_rev * 100, 2) if o_rev else 0

            day_columns.append({
                "date": dk,
                "orders_count": dd["orders_count"],
                "orders_revenue": round(o_rev, 2),
                "buyouts_count": dd["buyouts_count"],
                "buyouts_revenue": round(b_rev, 2),
                "ad_cost": round(dd["ad_cost"], 2),
                "drr": dd_drr,
                "margin_before_drr": round(dd_margin_before, 2),
                "profit_calc": round(dd_profit, 2),
                "margin_with_drr": round(dd_margin_with, 2),
            })

        products.append({
            "entity_id": eid,
            "nm_id": nm,
            "vendor_code": last[3] or "",
            "product_name": last[4] or "",
            "photo_main": last[5] or "",
            "barcode": last[6] or "",
            "size_name": size_map.get(eid, ""),
            # Справочник
            "brand": ref["brand"],
            "product_status": ref["product_status"],
            "product_class": ref["product_class"],
            "in_promo": ref["in_promo"],
            "ad_strategy": ref["ad_strategy"],
            "tags": ref["tags"],
            "rating_reviews": ref["rating_reviews"],
            "localization_pct": ref["localization_pct"],
            "ad_shows_organic": ref["ad_shows_organic"],
            "ad_shows_paid": ref["ad_shows_paid"],
            # Цены
            "price_retail": snap["price_retail"],
            "price_with_spp": snap["price_with_spp"],
            "spp_pct": snap["spp_pct"],
            "buyout_pct": buyout_pct,
            # Себестоимость
            "cost_price": ref["cost_price"],
            "cost_of_stock": cost_of_stock,
            # План
            "plan_value": plan_val,
            "daily_norm": daily_norm,
            "pct_complete": pct_complete,
            # Итоги за месяц
            "total_orders": total_orders,
            "total_orders_revenue": round(total_orders_revenue, 2),
            "total_buyouts": total_buyouts,
            "total_buyouts_revenue": round(total_buyouts_revenue, 2),
            "total_ad_cost": round(total_ad_cost, 2),
            "drr": drr,
            "margin_before_drr": round(margin_before_drr, 2),
            "profit_calc": round(profit_calc, 2),
            "margin_with_drr": round(margin_with_drr, 2),
            "krrr": krrr,
            "roi": roi,
            "cpl": cpl,
            "ctr": ctr,
            "enough_days": enough_days,
            "current_stock": current_stock,
            # Дни
            "days": day_columns,
        })

    # Сортировка
    sort_key_map = {
        "orders_revenue": lambda x: x["total_orders_revenue"],
        "roi": lambda x: x["roi"],
        "buyout_pct": lambda x: x["buyout_pct"],
    }
    sort_fn = sort_key_map.get(sort_by, sort_key_map["orders_revenue"])
    products.sort(key=sort_fn, reverse=True)

    # Сводка по всем карточкам
    summary = {
        "total_orders": sum(p["total_orders"] for p in products),
        "total_orders_revenue": round(sum(p["total_orders_revenue"] for p in products), 2),
        "total_buyouts": sum(p["total_buyouts"] for p in products),
        "total_buyouts_revenue": round(sum(p["total_buyouts_revenue"] for p in products), 2),
        "total_ad_cost": round(sum(p["total_ad_cost"] for p in products), 2),
        "total_drr": round(
            sum(p["total_ad_cost"] for p in products) / sum(p["total_orders_revenue"] for p in products) * 100, 2
        ) if sum(p["total_orders_revenue"] for p in products) else 0,
        "total_margin_before_drr": round(sum(p["margin_before_drr"] for p in products), 2),
        "total_profit_calc": round(sum(p["profit_calc"] for p in products), 2),
        "total_margin_with_drr": round(sum(p["margin_with_drr"] for p in products), 2),
        "total_products": len(products),
        "total_stock": sum(p["current_stock"] for p in products),
    }

    period_label = f"{year}-{mon:02d}" if month else f"{first_day} — {last_day}"
    return {
        "month": period_label,
        "days_in_month": days_in_month,
        "day_list": [str(d) for d in day_list],
        "summary": summary,
        "products": products,
    }
@router.get("/nl/register", response_class=HTMLResponse)
async def nl_register_page():
    html = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>НЛ — Регистрация</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f5f7fa;color:#1a1a2e}
.auth-container{max-width:400px;margin:80px auto;background:#fff;padding:32px;border-radius:12px;box-shadow:0 2px 12px rgba(0,0,0,.08)}
h2{color:#6c5ce7;margin-bottom:8px;font-size:1.3em}
p{color:#999;font-size:.85em;margin-bottom:20px}
.field{margin-bottom:14px}
label{display:block;font-size:.8em;color:#666;margin-bottom:4px}
input{width:100%;border:1px solid #e0e0e0;border-radius:4px;padding:8px;font-size:.9em}
input:focus{outline:none;border-color:#6c5ce7;box-shadow:0 0 0 2px rgba(108,92,231,.15)}
.btn{background:#6c5ce7;color:#fff;border:none;padding:10px;width:100%;border-radius:6px;cursor:pointer;font-size:.95em;font-weight:500}
.btn:hover{background:#5a4bd1}
.link{text-align:center;margin-top:16px;font-size:.85em}
.link a{color:#6c5ce7;text-decoration:none}
.link a:hover{text-decoration:underline}
.error{color:#e74c3c;font-size:.85em;margin-bottom:10px;display:none}
</style>
</head>
<body>
<div class="auth-container">
<h2>📝 Регистрация</h2>
<p>Создайте аккаунт для доступа к аналитике</p>
<div id="err" class="error"></div>
<div class="field"><label>Email</label><input type="email" id="email"></div>
<div class="field"><label>Пароль</label><input type="password" id="password"></div>
<div class="field"><label>Название магазина</label><input type="text" id="org" value="Мой магазин"></div>
<button class="btn" id="submitBtn">Зарегистрироваться</button>
<div class="link"><a href="/nl/v2">Уже есть аккаунт? Войти</a></div>
</div>
<script>
document.getElementById('submitBtn').addEventListener('click', async function() {
    var email = document.getElementById('email').value;
    var password = document.getElementById('password').value;
    var org_name = document.getElementById('org').value;
    var err = document.getElementById('err');
    err.style.display = 'none';
    try {
        var res = await fetch('/api/v1/nl/register', {
            method: 'POST',
            headers: {'Content-Type': 'application/json'},
            body: JSON.stringify({email: email, password: password, org_name: org_name})
        });
        if (!res.ok) {
            var d = await res.json();
            throw new Error(d.detail || 'Ошибка');
        }
        var data = await res.json();
        localStorage.setItem('nl_token', data.access_token);
        localStorage.setItem('nl_org_id', data.org_id);
        window.location.href = '/nl/v2';
    } catch(e) {
        err.textContent = e.message;
        err.style.display = 'block';
    }
});
</script>
</body>
</html>"""
    from fastapi.responses import HTMLResponse
    resp = HTMLResponse(html)
    resp.headers["Cache-Control"] = "no-cache, no-store"
    return resp



@router.get("/nl/login", response_class=HTMLResponse)
async def nl_login_page():
    html = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>НЛ — Вход</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f5f7fa;color:#1a1a2e}
.auth-container{max-width:400px;margin:80px auto;background:#fff;padding:32px;border-radius:12px;box-shadow:0 2px 12px rgba(0,0,0,.08)}
h2{color:#6c5ce7;margin-bottom:8px;font-size:1.3em}
p{color:#999;font-size:.85em;margin-bottom:20px}
.field{margin-bottom:14px}
label{display:block;font-size:.8em;color:#666;margin-bottom:4px}
input{width:100%;border:1px solid #e0e0e0;border-radius:4px;padding:8px;font-size:.9em}
input:focus{outline:none;border-color:#6c5ce7;box-shadow:0 0 0 2px rgba(108,92,231,.15)}
.btn{background:#6c5ce7;color:#fff;border:none;padding:10px;width:100%;border-radius:6px;cursor:pointer;font-size:.95em;font-weight:500}
.btn:hover{background:#5a4bd1}
.link{text-align:center;margin-top:16px;font-size:.85em}
.link a{color:#6c5ce7;text-decoration:none}
.link a:hover{text-decoration:underline}
.error{color:#e74c3c;font-size:.85em;margin-bottom:10px;display:none}
</style>
</head>
<body>
<div class="auth-container">
<h2>🔑 Вход</h2>
<p>Войдите в свой аккаунт НЛ</p>
<div id="err" class="error"></div>
<div class="field"><label>Email</label><input type="email" id="email"></div>
<div class="field"><label>Пароль</label><input type="password" id="password"></div>
<button class="btn" id="submitBtn">Войти</button>
<div class="link"><a href="/nl/register">Нет аккаунта? Зарегистрироваться</a></div>
</div>
<script>
document.getElementById('submitBtn').addEventListener('click', async function() {
    var email = document.getElementById('email').value;
    var password = document.getElementById('password').value;
    var err = document.getElementById('err');
    err.style.display = 'none';
    try {
        var res = await fetch('/api/v1/nl/login', {
            method: 'POST',
            headers: {'Content-Type': 'application/json'},
            body: JSON.stringify({email: email, password: password})
        });
        if (!res.ok) {
            var d = await res.json();
            throw new Error(d.detail || 'Неверный email или пароль');
        }
        var data = await res.json();
        localStorage.setItem('nl_token', data.access_token);
        localStorage.setItem('nl_org_id', data.org_id);
        window.location.href = '/nl/v2';
    } catch(e) {
        err.textContent = e.message;
        err.style.display = 'block';
    }
});
</script>
</body>
</html>"""
    from fastapi.responses import HTMLResponse
    resp = HTMLResponse(html)
    resp.headers["Cache-Control"] = "no-cache, no-store"
    return resp

@router.get("/api/v1/nl/analytics")
async def get_analytics(org_id: str, target_date: Optional[str] = None, search: Optional[str] = None, db: AsyncSession = Depends(get_db)):
    org_id = await resolve_org_id(org_id, db)
    """Аналитика по товарам — детальная таблица"""
    from datetime import datetime as dt_mod
    import decimal

    d = dt_mod.strptime(target_date, "%Y-%m-%d").date() if target_date else date.today()

    query = select(
        TechStatus.nm_id, TechStatus.vendor_code, TechStatus.product_name,
        TechStatus.photo_main, TechStatus.stock_qty, TechStatus.orders_count,
        TechStatus.buyouts_count, TechStatus.returns_count,
        TechStatus.price, TechStatus.price_discount, TechStatus.tariff,
        TechStatus.ad_cost, TechStatus.rating,
        TechStatus.impressions, TechStatus.clicks,
        TechStatus.warehouse_name, TechStatus.barcode,
    ).where(TechStatus.organization_id == org_id, TechStatus.target_date == d)

    if search:
        query = query.where(
            (TechStatus.vendor_code.ilike(f"%{search}%")) |
            (TechStatus.product_name.ilike(f"%{search}%")) |
            (TechStatus.nm_id == int(search) if search.isdigit() else False)
        )

    query = query.order_by(TechStatus.orders_count.desc().nullslast())
    result = await db.execute(query)
    rows = result.all()

    def sf(v): return float(v) if v and not isinstance(v, decimal.Decimal) else (float(v) if isinstance(v, decimal.Decimal) else None)
    def si(v): return int(v) if v else None

    products = []
    for r in rows:
        price = sf(r[8])
        price_disc = sf(r[9])
        tariff = sf(r[10])
        ad_cost = sf(r[11]) or 0
        orders = si(r[5]) or 0
        buyouts = si(r[6]) or 0
        revenue = price_disc * buyouts if price_disc and buyouts else 0
        commission = revenue * (tariff / 100) if revenue and tariff else 0
        payout = revenue - commission - ad_cost

        products.append({
            "nm_id": r[0], "vendor_code": r[1], "product_name": r[2],
            "photo_main": r[3], "stock_qty": si(r[4]),
            "orders_count": orders, "buyouts_count": buyouts,
            "returns_count": si(r[7]),
            "buyout_percent": round(buyouts / orders * 100, 1) if orders else 0,
            "price": price, "price_discount": price_disc,
            "tariff_percent": tariff,
            "commission": round(commission, 2),
            "logistics": 0, "ad_cost": round(ad_cost, 2),
            "drr": round(ad_cost / revenue * 100, 1) if revenue else 0,
            "fines": 0, "storage": 0, "reception": 0, "other_deductions": 0,
            "avg_check": round(revenue / buyouts, 2) if buyouts else 0,
            "revenue": round(revenue, 2), "payout": round(payout, 2),
            "cost_price": 0, "margin": round(payout, 2),
            "margin_per_unit": round(payout / buyouts, 2) if buyouts else 0,
            "profitability": round(payout / revenue * 100, 1) if revenue else 0,
            "roi": 0, "rating": sf(r[12]),
            "impressions": si(r[13]), "clicks": si(r[14]),
            "ctr": round((r[14] or 0) / (r[13] or 1) * 100, 2) if r[13] else 0,
            "turnover": 0, "in_transit": 0,
        })

    return {"date": str(d), "count": len(products), "products": products}


@router.get("/api/v1/nl/warehouses")
async def get_warehouse_stock(org_id: str, target_date: Optional[str] = None, db: AsyncSession = Depends(get_db)):
    """Остатки на складах WB"""
    from datetime import datetime as dt_mod
    d = dt_mod.strptime(target_date, "%Y-%m-%d").date() if target_date else date.today()
    result = await db.execute(
        select(TechStatus.nm_id, TechStatus.vendor_code, TechStatus.product_name,
               TechStatus.warehouse_name, TechStatus.stock_qty, TechStatus.barcode)
        .where(TechStatus.organization_id == org_id, TechStatus.target_date == d)
        .order_by(TechStatus.stock_qty.desc().nullslast())
    )
    return [{"nm_id": r[0], "vendor_code": r[1], "product_name": r[2],
             "warehouse": r[3], "qty": int(r[4]) if r[4] else 0, "barcode": r[5]} for r in result.all()]


@router.get("/api/v1/nl/operating-expenses")
async def get_operating_expenses(org_id: str, db: AsyncSession = Depends(get_db)):
    """Операционные расходы"""
    # TODO: добавить модель OperatingExpense
    return []


@router.post("/api/v1/nl/operating-expenses")
async def add_operating_expense(data: dict, org_id: str, db: AsyncSession = Depends(get_db)):
    """Добавить операционный расход"""
    # TODO: сохранить в БД
    return {"ok": True}


@router.get("/api/v1/nl/opiu")
async def get_opiu(org_id: str, period: str = "4", db: AsyncSession = Depends(get_db)):
    """ОПиУ — отчёт о прибылях и убытках по неделям"""
    from sqlalchemy import func
    from datetime import timedelta
    import decimal

    weeks = int(period) if period.isdigit() else 4
    today = date.today()
    
    # Получаем данные за N недель
    start_date = today - timedelta(weeks=weeks)
    
    result = await db.execute(
        select(
            TechStatus.target_date,
            func.sum(TechStatus.orders_count).label("orders"),
            func.sum(TechStatus.buyouts_count).label("buyouts"),
            func.sum(TechStatus.returns_count).label("returns"),
            func.sum(TechStatus.stock_qty).label("stock"),
            func.sum(TechStatus.ad_cost).label("ad_cost"),
            func.sum(TechStatus.impressions).label("impressions"),
            func.sum(TechStatus.clicks).label("clicks"),
            func.avg(TechStatus.tariff).label("avg_tariff"),
            func.sum(TechStatus.price_discount).label("revenue"),
        ).where(
            TechStatus.organization_id == org_id,
            TechStatus.target_date >= start_date
        ).group_by(TechStatus.target_date)
        .order_by(TechStatus.target_date.desc())
    )
    rows = result.all()
    
    # Группируем по неделям
    from collections import OrderedDict
    weeks_data = OrderedDict()
    for r in rows:
        d = r[0]
        # ISO week
        week_start = d - timedelta(days=d.weekday())
        week_key = week_start.isoformat()
        week_label = week_start.strftime("%d.%m") + " - " + (week_start + timedelta(days=6)).strftime("%d.%m.%Y")
        
        if week_key not in weeks_data:
            weeks_data[week_key] = {"label": week_label, "orders": 0, "buyouts": 0, "returns": 0, "revenue": 0, "ad_cost": 0, "stock": 0}
        
        def safe(v): return float(v) if v and not isinstance(v, decimal.Decimal) else (float(v) if isinstance(v, decimal.Decimal) else 0)
        
        w = weeks_data[week_key]
        w["orders"] += safe(r[1])
        w["buyouts"] += safe(r[2])
        w["returns"] += safe(r[3])
        w["stock"] += safe(r[4]) or 0
        w["ad_cost"] += safe(r[5])
        w["revenue"] += safe(r[9])
    
    # Считаем итоги
    total = {"orders": 0, "buyouts": 0, "returns": 0, "revenue": 0, "ad_cost": 0}
    for w in weeks_data.values():
        for k in total:
            total[k] += w.get(k, 0)
    
    return {
        "total": total,
        "weeks": [{"key": k, **v} for k, v in weeks_data.items()]
    }


@router.get("/api/v1/nl/opiu")
async def get_opiu(org_id: str, period: str = "4", db: AsyncSession = Depends(get_db)):
    """ОПиУ по неделям"""
    from sqlalchemy import func
    from datetime import timedelta
    import decimal
    weeks = int(period) if period.isdigit() else 4
    start_date = date.today() - timedelta(weeks=weeks)
    result = await db.execute(
        select(TechStatus.target_date,
            func.sum(TechStatus.orders_count).label("orders"),
            func.sum(TechStatus.buyouts_count).label("buyouts"),
            func.sum(TechStatus.returns_count).label("returns"),
            func.sum(TechStatus.ad_cost).label("ad_cost"),
            func.sum(TechStatus.price_discount).label("revenue"),
        ).where(TechStatus.organization_id == org_id, TechStatus.target_date >= start_date)
        .group_by(TechStatus.target_date).order_by(TechStatus.target_date.desc())
    )
    from collections import OrderedDict
    weeks_data = OrderedDict()
    total = {"orders": 0, "buyouts": 0, "returns": 0, "revenue": 0, "ad_cost": 0}
    for r in result.all():
        d = r[0]
        ws = d - timedelta(days=d.weekday())
        wk = ws.isoformat()
        if wk not in weeks_data:
            weeks_data[wk] = {"label": ws.strftime("%d.%m") + " - " + (ws + timedelta(days=6)).strftime("%d.%m"), "orders": 0, "buyouts": 0, "returns": 0, "revenue": 0, "ad_cost": 0}
        def sf(v): return float(v) if v else 0
        w = weeks_data[wk]
        w["orders"] += sf(r[1]); w["buyouts"] += sf(r[2]); w["returns"] += sf(r[3]); w["ad_cost"] += sf(r[4]); w["revenue"] += sf(r[5])
        total["orders"] += sf(r[1]); total["buyouts"] += sf(r[2]); total["returns"] += sf(r[3]); total["ad_cost"] += sf(r[4]); total["revenue"] += sf(r[5])
    return {"total": total, "weeks": [{"key": k, **v} for k, v in weeks_data.items()]}



# ==================== USER DATA APIs ====================

@router.get("/api/v1/nl/cost-prices")
async def get_cost_prices(org_id: str, db: AsyncSession = Depends(get_db)):
    org_id = await resolve_org_id(org_id, db)
    """Справочник товаров — одна запись на entity (nm_id + размер)"""
    from sqlalchemy import text
    # JOIN product_entities с reference_book: каждая entity = своя строка
    # Сначала получаем все entity, потом подтягиваем справочник по nm_id
    result = await db.execute(text(
        "SELECT pe.id as entity_id, pe.nm_id, pe.size_name, pe.vendor_code, pe.brand, "
        "pe.subject_id, pe.subject_name, pe.length, pe.width, pe.height, pe.weight, "
        "(SELECT string_agg(eb.barcode, ', ') FROM entity_barcodes eb WHERE eb.entity_id = pe.id AND eb.is_active = true) as barcodes, "
        "cp.id as ref_id, cp.cost_price, cp.purchase_cost, cp.logistics_cost, cp.packaging_cost, "
        "cp.other_costs, cp.extra_costs, cp.vat, cp.min_price, "
        "cp.mp_base_pct, cp.mp_correction_pct, cp.fulfillment_model, cp.storage_pct, "
        "cp.buyout_niche_pct, cp.price_before_spp_plan, cp.price_before_spp_change, "
        "cp.change_date, cp.wb_club_discount_pct, cp.ad_plan_rub, cp.supply_days, "
        "cp.min_batch_fbo, cp.product_status, cp.valid_from, cp.notes, "
        "cp.product_class, cp.tax_system, cp.tax_rate, "
        "cp.season_jan, cp.season_feb, cp.season_mar, cp.season_apr, cp.season_may, cp.season_jun, "
        "cp.season_jul, cp.season_aug, cp.season_sep, cp.season_oct, cp.season_nov, cp.season_dec, "
        "cp.plan_length, cp.plan_width, cp.plan_height, cp.plan_volume, cp.plan_weight, "
        "cp.delivery_days_to_seller, cp.delivery_days_to_mp, "
        "cp.top_query_1, cp.top_query_2, cp.top_query_3, "
        "cp.shipment_method, cp.fbs_warehouse, cp.rrc_price, cp.vat_rate, "
        "ts.product_name "
        "FROM product_entities pe "
        "LEFT JOIN LATERAL (SELECT * FROM reference_book WHERE organization_id = :org AND nm_id = pe.nm_id AND entity_id = pe.id "
        "  AND (valid_to IS NULL OR valid_to >= CURRENT_DATE) ORDER BY valid_from DESC LIMIT 1) cp ON true "
        "LEFT JOIN (SELECT DISTINCT nm_id, product_name FROM tech_status WHERE organization_id = :org) ts ON pe.nm_id = ts.nm_id "
        "WHERE pe.organization_id = :org "
        "ORDER BY pe.nm_id, pe.size_name"
    ), {"org": org_id})

    def fval(v): return float(v) if v else None
    def ival(v): return int(v) if v else None
    def sval(v): return str(v) if v else None

    return [{
        "id": sval(r[12]) or str(r[0]),  # ref_id или entity_id
        "entity_id": str(r[0]),
        "nm_id": r[1],
        "size_name": r[2] or "",
        "vendor_code": r[3] or "",
        "brand": r[4] or "",
        "subject_id": r[5], "subject_name": r[6] or "",
        "length": fval(r[7]), "width": fval(r[8]), "height": fval(r[9]),
        "weight": fval(r[10]),
        "barcode": r[11] or "",
        "barcodes": r[11] or "",
        "cost_price": fval(r[13]) or 0,
        "purchase_cost": fval(r[14]),
        "logistics_cost": fval(r[15]),
        "packaging_cost": fval(r[16]),
        "other_costs": fval(r[17]),
        "extra_costs": fval(r[18]),
        "vat": fval(r[19]) or 0,
        "min_price": fval(r[20]),
        "mp_base_pct": fval(r[21]),
        "mp_correction_pct": fval(r[22]),
        "fulfillment_model": r[23] or "fbo",
        "storage_pct": fval(r[24]),
        "buyout_niche_pct": fval(r[25]),
        "price_before_spp_plan": fval(r[26]),
        "price_before_spp_change": fval(r[27]),
        "change_date": sval(r[28]),
        "wb_club_discount_pct": fval(r[29]),
        "ad_plan_rub": fval(r[30]),
        "supply_days": r[31],
        "min_batch_fbo": r[32],
        "product_status": r[33] or "",
        "valid_from": sval(r[34]) or "",
        "notes": r[35] or "",
        "product_class": r[36] or "",
        "tax_system": r[37] or "",
        "tax_rate": fval(r[38]) or 0,
        "season_jan": fval(r[39]), "season_feb": fval(r[40]),
        "season_mar": fval(r[41]), "season_apr": fval(r[42]),
        "season_may": fval(r[43]), "season_jun": fval(r[44]),
        "season_jul": fval(r[45]), "season_aug": fval(r[46]),
        "season_sep": fval(r[47]), "season_oct": fval(r[48]),
        "season_nov": fval(r[49]), "season_dec": fval(r[50]),
        "plan_length": fval(r[51]), "plan_width": fval(r[52]),
        "plan_height": fval(r[53]), "plan_volume": fval(r[54]),
        "plan_weight": fval(r[55]),
        "delivery_days_to_seller": ival(r[56]), "delivery_days_to_mp": ival(r[57]),
        "top_query_1": r[58] or "", "top_query_2": r[59] or "", "top_query_3": r[60] or "",
        "shipment_method": r[61] or "", "fbs_warehouse": r[62] or "",
        "rrc_price": fval(r[63]), "vat_rate": fval(r[64]) or 0,
        "product_name": r[65] or "",
        "sizes": [],  # больше не нужен массив — каждая entity = своя строка
    } for r in result.all()]


@router.post("/api/v1/nl/cost-prices")
async def save_cost_price(data: dict, org_id: str, db: AsyncSession = Depends(get_db)):
    """Сохранить себестоимость (создать/обновить)"""
    from sqlalchemy import text
    nm_id = data.get("nm_id")
    cost = data.get("cost_price", 0)
    valid_from = data.get("valid_from", date.today().isoformat())
    if isinstance(valid_from, str):
        from datetime import datetime as _dt
        valid_from = _dt.strptime(valid_from, "%Y-%m-%d").date()
    if not nm_id:
        raise HTTPException(400, "nm_id обязателен")
    # entity_id из запроса — каждая сущность (размер) сохраняется отдельно
    entity_id = data.get("entity_id")
    if not entity_id:
        # Fallback: если entity_id не передан, ищем по nm_id + size_name
        size_name = data.get("size_name", "")
        ent_q = await db.execute(text(
            "SELECT pe.id FROM product_entities pe "
            "WHERE pe.organization_id = :org AND pe.nm_id = :nm AND pe.size_name = :sz LIMIT 1"
        ), {"org": org_id, "nm": nm_id, "sz": size_name})
        ent_row = ent_q.first()
        entity_id = str(ent_row[0]) if ent_row else None
    await db.execute(text(
        "INSERT INTO reference_book (organization_id, nm_id, barcode, vendor_code, size_name, entity_id, "
        "subject_id, subject_name, "
        "cost_price, purchase_cost, logistics_cost, packaging_cost, other_costs, extra_costs, vat, min_price, "
        "mp_base_pct, mp_correction_pct, fulfillment_model, storage_pct, "
        "buyout_niche_pct, "
        "price_before_spp_plan, price_before_spp_change, change_date, "
        "wb_club_discount_pct, ad_plan_rub, supply_days, min_batch_fbo, "
        "product_status, product_class, brand, tax_system, tax_rate, "
        "season_jan, season_feb, season_mar, season_apr, season_may, season_jun, season_jul, season_aug, season_sep, season_oct, season_nov, season_dec, "
        "plan_length, plan_width, plan_height, plan_volume, plan_weight, "
        "delivery_days_to_seller, delivery_days_to_mp, "
        "top_query_1, top_query_2, top_query_3, "
        "shipment_method, fbs_warehouse, rrc_price, vat_rate, "
        "valid_from, source, notes) "
        "VALUES (:org, :nm, :bc, :vc, :sz, :eid, "
        ":subid, :subn, "
        ":cp, :pc, :lc, :pk, :oc, :ec, :vat, :minp, "
        ":mpb, :mpc, :ffm, :stp, "
        ":bnp, "
        ":pspp, :psppc, :cdate, "
        ":wbcd, :adpr, :sdays, :minb, "
        ":pstatus, :pcls, :brand, :tsys, :tr, "
        ":sjan, :sfeb, :smar, :sapr, :smay, :sjun, :sjul, :saug, :ssep, :soct, :snov, :sdec, "
        ":pl, :pw, :ph, :pv, :pwg, "
        ":dds, :ddm, "
        ":tq1, :tq2, :tq3, "
        ":sm, :fw, :rrc, :vr, "
        ":vf, :src, :notes) "
        "ON CONFLICT (organization_id, nm_id, entity_id, valid_from) DO UPDATE SET "
        "barcode = COALESCE(EXCLUDED.barcode, reference_book.barcode), vendor_code = COALESCE(EXCLUDED.vendor_code, reference_book.vendor_code), "
        "cost_price = COALESCE(EXCLUDED.cost_price, reference_book.cost_price), purchase_cost = COALESCE(EXCLUDED.purchase_cost, reference_book.purchase_cost), "
        "logistics_cost = COALESCE(EXCLUDED.logistics_cost, reference_book.logistics_cost), packaging_cost = COALESCE(EXCLUDED.packaging_cost, reference_book.packaging_cost), "
        "other_costs = COALESCE(EXCLUDED.other_costs, reference_book.other_costs), extra_costs = COALESCE(EXCLUDED.extra_costs, reference_book.extra_costs), vat = COALESCE(EXCLUDED.vat, reference_book.vat), "
        "min_price = COALESCE(EXCLUDED.min_price, reference_book.min_price), "
        "mp_base_pct = COALESCE(EXCLUDED.mp_base_pct, reference_book.mp_base_pct), mp_correction_pct = COALESCE(EXCLUDED.mp_correction_pct, reference_book.mp_correction_pct), "
        "fulfillment_model = COALESCE(EXCLUDED.fulfillment_model, reference_book.fulfillment_model), storage_pct = COALESCE(EXCLUDED.storage_pct, reference_book.storage_pct), "
        "buyout_niche_pct = COALESCE(EXCLUDED.buyout_niche_pct, reference_book.buyout_niche_pct), "
        "price_before_spp_plan = COALESCE(EXCLUDED.price_before_spp_plan, reference_book.price_before_spp_plan), "
        "price_before_spp_change = COALESCE(EXCLUDED.price_before_spp_change, reference_book.price_before_spp_change), change_date = COALESCE(EXCLUDED.change_date, reference_book.change_date), "
        "wb_club_discount_pct = COALESCE(EXCLUDED.wb_club_discount_pct, reference_book.wb_club_discount_pct), ad_plan_rub = COALESCE(EXCLUDED.ad_plan_rub, reference_book.ad_plan_rub), "
        "supply_days = COALESCE(EXCLUDED.supply_days, reference_book.supply_days), min_batch_fbo = COALESCE(EXCLUDED.min_batch_fbo, reference_book.min_batch_fbo), "
        "product_status = COALESCE(EXCLUDED.product_status, reference_book.product_status), "
        "product_class = COALESCE(EXCLUDED.product_class, reference_book.product_class), brand = COALESCE(EXCLUDED.brand, reference_book.brand), "
        "tax_system = COALESCE(EXCLUDED.tax_system, reference_book.tax_system), " +
        "season_jan = COALESCE(EXCLUDED.season_jan, reference_book.season_jan), season_feb = COALESCE(EXCLUDED.season_feb, reference_book.season_feb), season_mar = COALESCE(EXCLUDED.season_mar, reference_book.season_mar), season_apr = COALESCE(EXCLUDED.season_apr, reference_book.season_apr), " +
        "season_may = COALESCE(EXCLUDED.season_may, reference_book.season_may), season_jun = COALESCE(EXCLUDED.season_jun, reference_book.season_jun), season_jul = COALESCE(EXCLUDED.season_jul, reference_book.season_jul), season_aug = COALESCE(EXCLUDED.season_aug, reference_book.season_aug), " +
        "season_sep = COALESCE(EXCLUDED.season_sep, reference_book.season_sep), season_oct = COALESCE(EXCLUDED.season_oct, reference_book.season_oct), season_nov = COALESCE(EXCLUDED.season_nov, reference_book.season_nov), season_dec = COALESCE(EXCLUDED.season_dec, reference_book.season_dec), " +
        "plan_length = COALESCE(EXCLUDED.plan_length, reference_book.plan_length), plan_width = COALESCE(EXCLUDED.plan_width, reference_book.plan_width), plan_height = COALESCE(EXCLUDED.plan_height, reference_book.plan_height), " +
        "plan_volume = COALESCE(EXCLUDED.plan_volume, reference_book.plan_volume), plan_weight = COALESCE(EXCLUDED.plan_weight, reference_book.plan_weight), " +
        "delivery_days_to_seller = COALESCE(EXCLUDED.delivery_days_to_seller, reference_book.delivery_days_to_seller), delivery_days_to_mp = COALESCE(EXCLUDED.delivery_days_to_mp, reference_book.delivery_days_to_mp), " +
        "top_query_1 = COALESCE(EXCLUDED.top_query_1, reference_book.top_query_1), top_query_2 = COALESCE(EXCLUDED.top_query_2, reference_book.top_query_2), top_query_3 = COALESCE(EXCLUDED.top_query_3, reference_book.top_query_3), " +
        "shipment_method = COALESCE(EXCLUDED.shipment_method, reference_book.shipment_method), fbs_warehouse = COALESCE(EXCLUDED.fbs_warehouse, reference_book.fbs_warehouse), rrc_price = COALESCE(EXCLUDED.rrc_price, reference_book.rrc_price), vat_rate = COALESCE(EXCLUDED.vat_rate, reference_book.vat_rate), " 
        "subject_id = COALESCE(EXCLUDED.subject_id, reference_book.subject_id), subject_name = COALESCE(EXCLUDED.subject_name, reference_book.subject_name), "
        "source = EXCLUDED.source, notes = EXCLUDED.notes"
    ), {"org": org_id, "nm": nm_id, "bc": data.get("barcode"), "vc": data.get("vendor_code"),
        "sz": data.get("size_name"), "eid": entity_id,
        "subid": int(data["subject_id"]) if data.get("subject_id") is not None and str(data["subject_id"]).strip().lstrip("-").isdigit() else None, "subn": data.get("subject_name"),
        "cp": cost, "pc": float(data["purchase_cost"]) if data.get("purchase_cost") is not None and str(data["purchase_cost"]) not in ("", "None") else None,
        "lc": float(data["logistics_cost"]) if data.get("logistics_cost") is not None and str(data["logistics_cost"]) not in ("", "None") else None, "pk": float(data["packaging_cost"]) if data.get("packaging_cost") is not None and str(data["packaging_cost"]) not in ("", "None") else None,
        "oc": float(data["other_costs"]) if data.get("other_costs") is not None and str(data["other_costs"]) not in ("", "None") else None, "ec": float(data["extra_costs"]) if data.get("extra_costs") is not None and str(data["extra_costs"]) not in ("", "None") else None, "vat": float(data.get("vat", 0)) if data.get("vat") is not None and str(data.get("vat", 0)) not in ("", "None") else 0, "minp": float(data["min_price"]) if data.get("min_price") is not None and str(data["min_price"]) not in ("", "None") else None,
        "mpb": float(data["mp_base_pct"]) if data.get("mp_base_pct") is not None and str(data["mp_base_pct"]) not in ("", "None") else None, "mpc": float(data["mp_correction_pct"]) if data.get("mp_correction_pct") is not None and str(data["mp_correction_pct"]) not in ("", "None") else None,
        "ffm": data.get("fulfillment_model", "fbo"), "stp": float(data["storage_pct"]) if data.get("storage_pct") is not None and str(data["storage_pct"]) not in ("", "None") else None,
        "bnp": float(data["buyout_niche_pct"]) if data.get("buyout_niche_pct") is not None and str(data["buyout_niche_pct"]) not in ("", "None") else None,
        "pspp": float(data["price_before_spp_plan"]) if data.get("price_before_spp_plan") is not None and str(data["price_before_spp_plan"]) not in ("", "None") else None, "psppc": float(data["price_before_spp_change"]) if data.get("price_before_spp_change") is not None and str(data["price_before_spp_change"]) not in ("", "None") else None,
        "cdate": date.today(),
        "wbcd": float(data["wb_club_discount_pct"]) if data.get("wb_club_discount_pct") is not None and str(data["wb_club_discount_pct"]) not in ("", "None") else None, "adpr": min(99, max(0, float(data["ad_plan_rub"]) if data.get("ad_plan_rub") is not None and str(data["ad_plan_rub"]) not in ("", "None") else 5)),
        "sdays": int(data["supply_days"]) if data.get("supply_days") and str(data["supply_days"]).isdigit() else None, "minb": int(data["min_batch_fbo"]) if data.get("min_batch_fbo") and str(data["min_batch_fbo"]).isdigit() else None,
        "pstatus": data.get("product_status"),
        "pcls": data.get("product_class"), "brand": data.get("brand"),
        "tsys": data.get("tax_system"), 
        "tr": float(data["tax_rate"]) if data.get("tax_rate") is not None and str(data.get("tax_rate")) not in ("", "None") else None,
        "sjan": float(data["season_jan"]) if data.get("season_jan") is not None and str(data["season_jan"]) not in ("", "None") else None, "sfeb": float(data["season_feb"]) if data.get("season_feb") is not None and str(data["season_feb"]) not in ("", "None") else None, "smar": float(data["season_mar"]) if data.get("season_mar") is not None and str(data["season_mar"]) not in ("", "None") else None, "sapr": float(data["season_apr"]) if data.get("season_apr") is not None and str(data["season_apr"]) not in ("", "None") else None,
        "smay": float(data["season_may"]) if data.get("season_may") is not None and str(data["season_may"]) not in ("", "None") else None, "sjun": float(data["season_jun"]) if data.get("season_jun") is not None and str(data["season_jun"]) not in ("", "None") else None, "sjul": float(data["season_jul"]) if data.get("season_jul") is not None and str(data["season_jul"]) not in ("", "None") else None, "saug": float(data["season_aug"]) if data.get("season_aug") is not None and str(data["season_aug"]) not in ("", "None") else None,
        "ssep": float(data["season_sep"]) if data.get("season_sep") is not None and str(data["season_sep"]) not in ("", "None") else None, "soct": float(data["season_oct"]) if data.get("season_oct") is not None and str(data["season_oct"]) not in ("", "None") else None, "snov": float(data["season_nov"]) if data.get("season_nov") is not None and str(data["season_nov"]) not in ("", "None") else None, "sdec": float(data["season_dec"]) if data.get("season_dec") is not None and str(data["season_dec"]) not in ("", "None") else None,
        "pl": float(data["plan_length"]) if data.get("plan_length") is not None and str(data["plan_length"]) not in ("", "None") else None, "pw": float(data["plan_width"]) if data.get("plan_width") is not None and str(data["plan_width"]) not in ("", "None") else None, "ph": float(data["plan_height"]) if data.get("plan_height") is not None and str(data["plan_height"]) not in ("", "None") else None,
        "pv": float(data["plan_volume"]) if data.get("plan_volume") is not None and str(data["plan_volume"]) not in ("", "None") else None, "pwg": float(data["plan_weight"]) if data.get("plan_weight") is not None and str(data["plan_weight"]) not in ("", "None") else None,
        "dds": int(data["delivery_days_to_seller"]) if data.get("delivery_days_to_seller") and str(data["delivery_days_to_seller"]).lstrip('-').isdigit() else None, "ddm": int(data["delivery_days_to_mp"]) if data.get("delivery_days_to_mp") and str(data["delivery_days_to_mp"]).lstrip('-').isdigit() else None,
        "tq1": data.get("top_query_1"), "tq2": data.get("top_query_2"), "tq3": data.get("top_query_3"),
        "sm": data.get("shipment_method"), "fw": data.get("fbs_warehouse"), "rrc": float(data["rrc_price"]) if data.get("rrc_price") is not None and str(data["rrc_price"]) not in ("", "None") else None,
        "vr": float(data["vat_rate"]) if data.get("vat_rate") is not None and str(data.get("vat_rate")) not in ("", "None") else None,
        "vf": valid_from, "src": data.get("source", "manual"), "notes": data.get("notes")})
    await db.commit()
    return {"ok": True}

@router.post("/api/v1/nl/cost-prices/batch")
async def save_cost_prices_batch(request: Request, org_id: str, db: AsyncSession = Depends(get_db)):
    """Batch-сохранение справочника — один запрос вместо N отдельных"""

    # Сбрасываем кэш юнит-экономики (данные справочника влияют на ЮЭ)
    try:
        import redis as _rinv2
        _rinv2.from_url("redis://redis:6379/0").delete(f"ue_cache:{org_id}")
    except Exception:
        pass

    items = await request.json()
    from sqlalchemy import text
    from datetime import datetime as _dt
    saved = 0
    errors = 0
    for data in items:
        try:
            nm_id = data.get("nm_id")
            if not nm_id:
                errors += 1
                continue
            cost = data.get("cost_price", 0)
            valid_from = data.get("valid_from", date.today().isoformat())
            if isinstance(valid_from, str):
                valid_from = _dt.strptime(valid_from, "%Y-%m-%d").date()
            entity_id = data.get("entity_id")
            if not entity_id:
                size_name = data.get("size_name", "")
                ent_q = await db.execute(text(
                    "SELECT pe.id FROM product_entities pe "
                    "WHERE pe.organization_id = :org AND pe.nm_id = :nm AND pe.size_name = :sz LIMIT 1"
                ), {"org": org_id, "nm": nm_id, "sz": size_name})
                ent_row = ent_q.first()
                entity_id = str(ent_row[0]) if ent_row else None
            def pfloat(v):
                if v is not None and str(v) not in ("", "None"):
                    try: return float(v)
                    except: pass
                return None
            def pint(v):
                if v and str(v).lstrip("-").isdigit():
                    return int(v)
                return None
            await db.execute(text(
                "INSERT INTO reference_book (organization_id, nm_id, barcode, vendor_code, size_name, entity_id, "
                "subject_id, subject_name, "
                "cost_price, purchase_cost, logistics_cost, packaging_cost, other_costs, extra_costs, vat, min_price, "
                "mp_base_pct, mp_correction_pct, fulfillment_model, storage_pct, "
                "buyout_niche_pct, "
                "price_before_spp_plan, price_before_spp_change, change_date, "
                "wb_club_discount_pct, ad_plan_rub, supply_days, min_batch_fbo, "
                "product_status, product_class, brand, tax_system, tax_rate, "
                "season_jan, season_feb, season_mar, season_apr, season_may, season_jun, season_jul, season_aug, season_sep, season_oct, season_nov, season_dec, "
                "plan_length, plan_width, plan_height, plan_volume, plan_weight, "
                "delivery_days_to_seller, delivery_days_to_mp, "
                "top_query_1, top_query_2, top_query_3, "
                "shipment_method, fbs_warehouse, rrc_price, vat_rate, "
                "valid_from, source, notes) "
                "VALUES (:org, :nm, :bc, :vc, :sz, :eid, "
                ":subid, :subn, "
                ":cp, :pc, :lc, :pk, :oc, :ec, :vat, :minp, "
                ":mpb, :mpc, :ffm, :stp, "
                ":bnp, "
                ":pspp, :psppc, :cdate, "
                ":wbcd, :adpr, :sdays, :minb, "
                ":pstatus, :pcls, :brand, :tsys, :tr, "
                ":sjan, :sfeb, :smar, :sapr, :smay, :sjun, :sjul, :saug, :ssep, :soct, :snov, :sdec, "
                ":pl, :pw, :ph, :pv, :pwg, "
                ":dds, :ddm, "
                ":tq1, :tq2, :tq3, "
                ":sm, :fw, :rrc, :vr, "
                ":vf, :src, :notes) "
                "ON CONFLICT (organization_id, nm_id, entity_id, valid_from) DO UPDATE SET "
                "barcode = COALESCE(EXCLUDED.barcode, reference_book.barcode), "
                "vendor_code = COALESCE(EXCLUDED.vendor_code, reference_book.vendor_code), "
                "cost_price = COALESCE(EXCLUDED.cost_price, reference_book.cost_price), "
                "purchase_cost = COALESCE(EXCLUDED.purchase_cost, reference_book.purchase_cost), "
                "logistics_cost = COALESCE(EXCLUDED.logistics_cost, reference_book.logistics_cost), "
                "packaging_cost = COALESCE(EXCLUDED.packaging_cost, reference_book.packaging_cost), "
                "other_costs = COALESCE(EXCLUDED.other_costs, reference_book.other_costs), "
                "extra_costs = COALESCE(EXCLUDED.extra_costs, reference_book.extra_costs), "
                "vat = COALESCE(EXCLUDED.vat, reference_book.vat), "
                "min_price = COALESCE(EXCLUDED.min_price, reference_book.min_price), "
                "mp_base_pct = COALESCE(EXCLUDED.mp_base_pct, reference_book.mp_base_pct), "
                "mp_correction_pct = COALESCE(EXCLUDED.mp_correction_pct, reference_book.mp_correction_pct), "
                "fulfillment_model = COALESCE(EXCLUDED.fulfillment_model, reference_book.fulfillment_model), "
                "storage_pct = COALESCE(EXCLUDED.storage_pct, reference_book.storage_pct), "
                "buyout_niche_pct = COALESCE(EXCLUDED.buyout_niche_pct, reference_book.buyout_niche_pct), "
                "price_before_spp_plan = COALESCE(EXCLUDED.price_before_spp_plan, reference_book.price_before_spp_plan), "
                "price_before_spp_change = COALESCE(EXCLUDED.price_before_spp_change, reference_book.price_before_spp_change), "
                "change_date = COALESCE(EXCLUDED.change_date, reference_book.change_date), "
                "wb_club_discount_pct = COALESCE(EXCLUDED.wb_club_discount_pct, reference_book.wb_club_discount_pct), "
                "ad_plan_rub = COALESCE(EXCLUDED.ad_plan_rub, reference_book.ad_plan_rub), "
                "supply_days = COALESCE(EXCLUDED.supply_days, reference_book.supply_days), "
                "min_batch_fbo = COALESCE(EXCLUDED.min_batch_fbo, reference_book.min_batch_fbo), "
                "product_status = COALESCE(EXCLUDED.product_status, reference_book.product_status), "
                "product_class = COALESCE(EXCLUDED.product_class, reference_book.product_class), "
                "brand = COALESCE(EXCLUDED.brand, reference_book.brand), "
                "tax_system = COALESCE(EXCLUDED.tax_system, reference_book.tax_system), "
        "tax_rate = COALESCE(EXCLUDED.tax_rate, reference_book.tax_rate), "
                "season_jan = COALESCE(EXCLUDED.season_jan, reference_book.season_jan), "
                "season_feb = COALESCE(EXCLUDED.season_feb, reference_book.season_feb), "
                "season_mar = COALESCE(EXCLUDED.season_mar, reference_book.season_mar), "
                "season_apr = COALESCE(EXCLUDED.season_apr, reference_book.season_apr), "
                "season_may = COALESCE(EXCLUDED.season_may, reference_book.season_may), "
                "season_jun = COALESCE(EXCLUDED.season_jun, reference_book.season_jun), "
                "season_jul = COALESCE(EXCLUDED.season_jul, reference_book.season_jul), "
                "season_aug = COALESCE(EXCLUDED.season_aug, reference_book.season_aug), "
                "season_sep = COALESCE(EXCLUDED.season_sep, reference_book.season_sep), "
                "season_oct = COALESCE(EXCLUDED.season_oct, reference_book.season_oct), "
                "season_nov = COALESCE(EXCLUDED.season_nov, reference_book.season_nov), "
                "season_dec = COALESCE(EXCLUDED.season_dec, reference_book.season_dec), "
                "plan_length = COALESCE(EXCLUDED.plan_length, reference_book.plan_length), "
                "plan_width = COALESCE(EXCLUDED.plan_width, reference_book.plan_width), "
                "plan_height = COALESCE(EXCLUDED.plan_height, reference_book.plan_height), "
                "plan_volume = COALESCE(EXCLUDED.plan_volume, reference_book.plan_volume), "
                "plan_weight = COALESCE(EXCLUDED.plan_weight, reference_book.plan_weight), "
                "delivery_days_to_seller = COALESCE(EXCLUDED.delivery_days_to_seller, reference_book.delivery_days_to_seller), "
                "delivery_days_to_mp = COALESCE(EXCLUDED.delivery_days_to_mp, reference_book.delivery_days_to_mp), "
                "top_query_1 = COALESCE(EXCLUDED.top_query_1, reference_book.top_query_1), "
                "top_query_2 = COALESCE(EXCLUDED.top_query_2, reference_book.top_query_2), "
                "top_query_3 = COALESCE(EXCLUDED.top_query_3, reference_book.top_query_3), "
                "shipment_method = COALESCE(EXCLUDED.shipment_method, reference_book.shipment_method), "
                "fbs_warehouse = COALESCE(EXCLUDED.fbs_warehouse, reference_book.fbs_warehouse), "
                "rrc_price = COALESCE(EXCLUDED.rrc_price, reference_book.rrc_price), "
                "vat_rate = COALESCE(EXCLUDED.vat_rate, reference_book.vat_rate), "
                "subject_id = COALESCE(EXCLUDED.subject_id, reference_book.subject_id), "
                "subject_name = COALESCE(EXCLUDED.subject_name, reference_book.subject_name), "
                "source = EXCLUDED.source, notes = COALESCE(EXCLUDED.notes, reference_book.notes)"
            ), {"org": org_id, "nm": nm_id, "bc": data.get("barcode"), "vc": data.get("vendor_code"),
                "sz": data.get("size_name"), "eid": entity_id,
                "subid": pint(data.get("subject_id")), "subn": data.get("subject_name"),
                "cp": cost, "pc": pfloat(data.get("purchase_cost")),
                "lc": pfloat(data.get("logistics_cost")), "pk": pfloat(data.get("packaging_cost")),
                "oc": pfloat(data.get("other_costs")), "ec": pfloat(data.get("extra_costs")),
                "vat": pfloat(data.get("vat")) or 0, "minp": pfloat(data.get("min_price")),
                "mpb": pfloat(data.get("mp_base_pct")), "mpc": pfloat(data.get("mp_correction_pct")),
                "ffm": data.get("fulfillment_model", "fbo"), "stp": pfloat(data.get("storage_pct")),
                "bnp": pfloat(data.get("buyout_niche_pct")),
                "pspp": pfloat(data.get("price_before_spp_plan")), "psppc": pfloat(data.get("price_before_spp_change")),
                "cdate": date.today(),
                "wbcd": pfloat(data.get("wb_club_discount_pct")), "adpr": min(99, max(0, pfloat(data.get("ad_plan_rub")) if pfloat(data.get("ad_plan_rub")) is not None else 5)),
                "sdays": pint(data.get("supply_days")), "minb": pint(data.get("min_batch_fbo")),
                "pstatus": data.get("product_status"),
                "pcls": data.get("product_class"), "brand": data.get("brand"),
                "tsys": data.get("tax_system"),
                "tr": pfloat(data.get("tax_rate")),
                "sjan": pfloat(data.get("season_jan")), "sfeb": pfloat(data.get("season_feb")),
                "smar": pfloat(data.get("season_mar")), "sapr": pfloat(data.get("season_apr")),
                "smay": pfloat(data.get("season_may")), "sjun": pfloat(data.get("season_jun")),
                "sjul": pfloat(data.get("season_jul")), "saug": pfloat(data.get("season_aug")),
                "ssep": pfloat(data.get("season_sep")), "soct": pfloat(data.get("season_oct")),
                "snov": pfloat(data.get("season_nov")), "sdec": pfloat(data.get("season_dec")),
                "pl": pfloat(data.get("plan_length")), "pw": pfloat(data.get("plan_width")),
                "ph": pfloat(data.get("plan_height")), "pv": pfloat(data.get("plan_volume")),
                "pwg": pfloat(data.get("plan_weight")),
                "dds": pint(data.get("delivery_days_to_seller")), "ddm": pint(data.get("delivery_days_to_mp")),
                "tq1": data.get("top_query_1"), "tq2": data.get("top_query_2"), "tq3": data.get("top_query_3"),
                "sm": data.get("shipment_method"), "fw": data.get("fbs_warehouse"),
                "rrc": pfloat(data.get("rrc_price")), "vr": pfloat(data.get("vat_rate")),
                "vf": valid_from, "src": data.get("source", "manual"), "notes": data.get("notes")})
            saved += 1
        except Exception as e:
            errors += 1
            print(f"[batch] error nm={data.get("nm_id")}: {e}")
    await db.commit()
    return {"ok": True, "saved": saved, "errors": errors}




@router.get("/api/v1/nl/commission-rate")
async def get_commission_rate(org_id: str, subject_id: int, model: str = "fbo", db: AsyncSession = Depends(get_db)):
    """Получить комиссию МП по subject_id и модели (fbo/fbs)"""
    from sqlalchemy import text
    import json as _json, logging
    _log = logging.getLogger(__name__)
    
    # Ищем комиссии по любому org_id (WB отдаёт единый справочник)
    result = await db.execute(text(
        "SELECT raw_response FROM raw_api_data "
        "WHERE api_method = 'tariffs_commission' "
        "ORDER BY target_date DESC LIMIT 1"
    ))
    row = result.first()
    if not row or not row[0]:
        return {"commission_pct": None, "source": "no_data"}
    
    cdata = row[0] if isinstance(row[0], dict) else _json.loads(row[0])
    for item in cdata.get("report", []):
        if item.get("subjectID") == subject_id:
            if model == "fbs":
                pct = item.get("kgvpMarketplace")  # FBS = Маркетплейс
            else:
                pct = item.get("paidStorageKgvp")  # FBO = Склад WB
            return {"commission_pct": float(pct) if pct else None, "source": "api", "model": model}
    
    return {"commission_pct": None, "source": "subject_not_found"}


@router.post("/api/v1/nl/cost-prices/auto-fill")
async def auto_fill_reference(org_id: str, db: AsyncSession = Depends(get_db)):
    """Автозаполнение справочника из wb_tariff_snapshot (только пустые поля)"""
    from sqlalchemy import text
    import logging
    _log = logging.getLogger(__name__)

    # Берём свежий snapshot за последнюю дату
    snap_result = await db.execute(text("""
        SELECT entity_id, nm_id, commission_pct, logistics_tariff, storage_tariff,
               price_retail, price_with_spp, ad_cost_fact, buyout_pct_fact
        FROM wb_tariff_snapshot
        WHERE organization_id = :org AND target_date = (
            SELECT MAX(target_date) FROM wb_tariff_snapshot WHERE organization_id = :org
        )
    """), {"org": org_id})
    snapshots = snap_result.all()
    if not snapshots:
        return {"ok": False, "error": "Нет данных в wb_tariff_snapshot. Запустите синхронизацию."}

    # Строим маппинг: entity_id -> snapshot, nm_id -> snapshot
    snap_by_entity = {}
    snap_by_nm = {}
    for s in snapshots:
        eid, nm_id = str(s[0]) if s[0] else None, s[1]
        data = {
            "commission_pct": float(s[2]) if s[2] else None,
            "logistics_tariff": float(s[3]) if s[3] else None,
            "storage_tariff": float(s[4]) if s[4] else None,
            "price_retail": float(s[5]) if s[5] else None,
            "price_with_spp": float(s[6]) if s[6] else None,
            "ad_cost_fact": float(s[7]) if s[7] else None,
            "buyout_pct_fact": float(s[8]) if s[8] else None,
        }
        if eid:
            snap_by_entity[eid] = data
        if nm_id:
            snap_by_nm[nm_id] = data

    # Загружаем subject_id/subject_name из product_entities для auto_fill
    subj_result = await db.execute(text(
        "SELECT id::text, subject_id, subject_name FROM product_entities WHERE organization_id = :org AND subject_id IS NOT NULL"
    ), {"org": org_id})
    subj_map = {}
    for sr in subj_result.all():
        subj_map[sr[0]] = {"subject_id": sr[1], "subject_name": sr[2]}

    # Загружаем vendor_code из product_entities по nm_id
    vc_result = await db.execute(text(
        "SELECT DISTINCT nm_id, vendor_code FROM product_entities WHERE organization_id = :org AND vendor_code IS NOT NULL AND vendor_code != ''"
    ), {"org": org_id})
    vendor_code_by_nm = {}
    for vr in vc_result.all():
        vendor_code_by_nm[vr[0]] = vr[1]

    # Добавляем subject данные в snapshots
    for eid_str, snap in snap_by_entity.items():
        if eid_str in subj_map:
            snap["subject_id"] = subj_map[eid_str]["subject_id"]
            snap["subject_name"] = subj_map[eid_str]["subject_name"]
    for nm, snap in snap_by_nm.items():
        # Попробуем найти entity по nm
        for eid_str, sm in subj_map.items():
            pass  # entity map более точный

    # Загружаем текущий справочник
    ref_result = await db.execute(text("""
        SELECT id, entity_id, nm_id,
               mp_base_pct, logistics_cost, storage_pct,
               price_before_spp_plan, buyout_niche_pct, ad_plan_rub, vendor_code,
               subject_id, subject_name
        FROM reference_book
        WHERE organization_id = :org AND (valid_to IS NULL OR valid_to >= CURRENT_DATE)
    """), {"org": org_id})
    refs = ref_result.all()

    stats = {"updated": 0, "skipped": 0, "fields_filled": {}}
    field_map = {
        "mp_base_pct": "commission_pct",
        "logistics_cost": "logistics_tariff",
        "storage_pct": "storage_tariff",
        "price_before_spp_plan": "price_retail",
        "buyout_niche_pct": "buyout_pct_fact",
        "ad_plan_rub": "ad_cost_fact",
        "subject_id": "subject_id",
        "subject_name": "subject_name",
    }

    # vendor_code: заполняем из product_entities (не из snapshot)
    # Будет обработан отдельно ниже

    for r in refs:
        rid = str(r[0])
        eid = str(r[1]) if r[1] else None
        nm_id = r[2]
        # Текущие значения в справочнике
        current = {
            "mp_base_pct": r[3],
            "logistics_cost": r[4],
            "storage_pct": r[5],
            "price_before_spp_plan": r[6],
            "buyout_niche_pct": r[7],
            "ad_plan_rub": r[8],
            "subject_id": r[9],
            "subject_name": r[10],
            "vendor_code": r[11],
        }

        # Ищем snapshot: сначала по entity_id, потом по nm_id
        snap = None
        if eid and eid in snap_by_entity:
            snap = snap_by_entity[eid]
        elif nm_id and nm_id in snap_by_nm:
            snap = snap_by_nm[nm_id]

        # vendor_code: подтягиваем из product_entities если пустой (даже без snapshot)
        updates = {}
        vc = vendor_code_by_nm.get(nm_id)
        if vc and (not current.get("vendor_code") or current["vendor_code"] == ""):
            updates["vendor_code"] = vc

        if not snap:
            if updates:
                set_clauses = ", ".join(f"{k} = :{k}" for k in updates)
                updates["rid"] = rid
                await db.execute(text(f"UPDATE reference_book SET {set_clauses} WHERE id = :rid"), updates)
                stats["updated"] += 1
            else:
                stats["skipped"] += 1
            continue

        # Собираем обновления (только пустые поля)
        for ref_field, snap_field in field_map.items():
            snap_val = snap.get(snap_field)
            cur_val = current.get(ref_field)
            if snap_val is not None and (cur_val is None or cur_val == 0):
                updates[ref_field] = snap_val
                stats["fields_filled"][ref_field] = stats["fields_filled"].get(ref_field, 0) + 1

        if not updates:
            stats["skipped"] += 1
            continue

        # Обновляем
        set_clauses = ", ".join(f"{k} = :{k}" for k in updates)
        updates["rid"] = rid
        await db.execute(text(f"UPDATE reference_book SET {set_clauses} WHERE id = :rid"), updates)
        stats["updated"] += 1

    # --- Шаг 2: Создать записи для nm_id без справочника ---
    existing_nms = set()
    for r in refs:
        if r[2]:
            existing_nms.add(r[2])
    
    all_entities = await db.execute(text(
        "SELECT pe.id, pe.nm_id, pe.size_name, pe.brand, pe.subject_id, pe.subject_name, pe.vendor_code FROM product_entities pe WHERE pe.organization_id = :org"
    ), {"org": org_id})
    
    created_count = 0
    seen_nms = set()
    for ent in all_entities.all():
        nm_id = ent[1]
        if nm_id in existing_nms or nm_id in seen_nms:
            continue
        seen_nms.add(nm_id)
        eid = str(ent[0])
        snap = snap_by_nm.get(nm_id)
        
        ins = pg_insert(ReferenceBook)
        vals = {
            "id": str(uuid.uuid4()),
            "organization_id": org_id,
            "nm_id": nm_id,
            "entity_id": eid,
            "size_name": ent[2] or "",
            "brand": ent[3] or "",
            "subject_id": ent[4],
            "subject_name": ent[5] or "",
            "vendor_code": ent[6] or "",
            "valid_from": date.today(),
            "mp_base_pct": float(snap["commission_pct"]) if snap and snap.get("commission_pct") else None,
            "logistics_cost": float(snap["logistics_tariff"]) if snap and snap.get("logistics_tariff") else None,
            "storage_pct": float(snap["storage_tariff"]) if snap and snap.get("storage_tariff") else None,
            "price_before_spp_plan": float(snap["price_retail"]) if snap and snap.get("price_retail") else None,
            "price_before_spp_change": float(snap["price_with_spp"]) if snap and snap.get("price_with_spp") else None,
            "buyout_niche_pct": float(snap["buyout_pct_fact"]) if snap and snap.get("buyout_pct_fact") else None,
        }
        stmt = ins.values(**vals).on_conflict_do_nothing(
            constraint="reference_book_org_nm_vf_key"
        )
        try:
            await db.execute(stmt)
            created_count += 1
        except Exception:
            await db.rollback()
    
    stats["created"] = created_count
    
    await db.commit()
    _log.info(f"[auto_fill] org={org_id}: {stats}")
    return {"ok": True, "stats": stats}


@router.get("/api/v1/nl/fbo-needs")
async def get_fbo_needs(org_id: str, days: int = 14, db: AsyncSession = Depends(get_db)):
    """Расчёт потребности FBO: остатки + темп заказов по складам"""
    from sqlalchemy import text

    # 1) Остатки по складам (из tech_status — последний snapshot)
    stocks_result = await db.execute(text("""
        SELECT ts.nm_id, ts.warehouse_name, 0 as warehouse_id,
               ts.stock_qty as qty, ts.stock_qty as qty_full
        FROM tech_status ts
        WHERE ts.organization_id = :org
          AND ts.target_date = (SELECT MAX(target_date) FROM tech_status WHERE organization_id = :org)
          AND ts.entity_id IS NOT NULL
    """), {"org": org_id})
    stocks = stocks_result.all()

    # Маппинг: (nm_id, warehouse_name) -> {qty, qty_full, warehouse_id}
    stock_map = {}
    warehouses = {}  # warehouse_name -> warehouse_id
    for s in stocks:
        key = (s[0], s[1])
        stock_map[key] = {"qty": s[3] or 0, "qty_full": s[4] or 0, "warehouse_id": s[2]}
        if s[1] not in warehouses:
            warehouses[s[1]] = s[2]

    # 2) Темп заказов по складам за N дней (из raw_api_data JSONB)
    import json as _json
    orders_result = await db.execute(text("""
        SELECT raw_response FROM raw_api_data
        WHERE organization_id = :org AND api_method = 'orders'
          AND target_date >= CURRENT_DATE - make_interval(days => :days_back)
    """), {"org": org_id, "days_back": days})
    raw_orders = orders_result.all()

    # Парсим JSONB: подсчёт заказов по (nmId, warehouseName)
    from collections import defaultdict
    order_agg = defaultdict(lambda: {"total_qty": 0, "days": set()})
    for row in raw_orders:
        raw = row[0]
        items = raw if isinstance(raw, list) else []
        for item in items:
            if not isinstance(item, dict):
                continue
            if item.get("isCancel"):
                continue
            nm = item.get("nmId")
            wh = item.get("warehouseName")
            if not nm or not wh:
                continue
            key = (nm, wh)
            order_agg[key]["total_qty"] += 1
            d = item.get("date", "")[:10]
            if d:
                order_agg[key]["days"].add(d)

    order_map = {}
    for key, v in order_agg.items():
        active = max(len(v["days"]), 1)
        order_map[key] = {"total_qty": v["total_qty"], "rate_per_day": round(v["total_qty"] / active, 2), "active_days": len(v["days"])}

    # 3) Все nm_id из entities (чтобы показать даже без остатков)
    entities_result = await db.execute(text("""
        SELECT pe.id, pe.nm_id, pe.size_name, pe.product_name, pe.photo_main
        FROM product_entities pe
        WHERE pe.organization_id = :org
    """), {"org": org_id})
    entities = entities_result.all()

    # Маппинг nm_id -> entity info
    entity_by_nm = {}
    for e in entities:
        if e[1] not in entity_by_nm:
            entity_by_nm[e[1]] = {"entity_id": str(e[0]), "size_name": e[2], "product_name": e[3], "photo_main": e[4]}

    # 4) Справочник: supply_days, min_batch_fbo по entity_id
    ref_result = await db.execute(text("""
        SELECT entity_id, nm_id, supply_days, min_batch_fbo
        FROM reference_book
        WHERE organization_id = :org AND (valid_to IS NULL OR valid_to >= CURRENT_DATE)
    """), {"org": org_id})
    refs = ref_result.all()

    # entity_id -> supply_days, min_batch_fbo; fallback nm_id
    ref_by_entity = {}
    ref_by_nm = {}
    for r in refs:
        d = {"supply_days": r[2], "min_batch_fbo": r[3]}
        if r[0]:
            ref_by_entity[str(r[0])] = d
        ref_by_nm[r[1]] = d

    # 5) Собираем результат — только комбинации с остатками или заказами
    all_keys = set(stock_map.keys()) | set(order_map.keys())
    rows = []
    for key in all_keys:
        nm_id, wname = key
        if nm_id not in entity_by_nm:
            continue
        einfo = entity_by_nm[nm_id]
        eid = einfo["entity_id"]
        wid = warehouses.get(wname, 0)

        qty = stock_map.get(key, {}).get("qty", 0)
        qty_full = stock_map.get(key, {}).get("qty_full", 0)
        ref = ref_by_nm.get(nm_id, ref_by_entity.get(eid, {}))
        supply_days = ref.get("supply_days") or 5
        min_batch = ref.get("min_batch_fbo") or 1

        order_info = order_map.get(key, {})
        rate = order_info.get("rate_per_day", 0)
        total_orders = order_info.get("total_qty", 0)
        active_days = order_info.get("active_days", 0)

        # Расчёт потребности
        need = round(rate * supply_days) - qty
        if need <= 0:
            need = 0
        elif need < min_batch:
            need = min_batch
        else:
            import math
            need = math.ceil(need / min_batch) * min_batch

        # Дней до нуля
        days_to_zero = round(qty / rate, 1) if rate > 0 else 999

        rows.append({
            "entity_id": eid,
            "nm_id": nm_id,
            "product_name": einfo["product_name"],
            "size_name": einfo["size_name"],
            "photo_main": einfo["photo_main"],
            "warehouse_name": wname,
            "warehouse_id": wid,
            "stock_qty": qty,
            "stock_qty_full": qty_full,
            "order_rate": rate,
            "orders_total": total_orders,
            "active_days": active_days,
            "supply_days": supply_days,
            "min_batch": min_batch,
            "need": need,
            "days_to_zero": days_to_zero,
        })

    # Сортировка: сначала критичные (days_to_zero меньше)
    rows.sort(key=lambda x: x["days_to_zero"])

    return {"warehouses": list(warehouses.keys()), "rows": rows, "days": days}


@router.post("/api/v1/nl/cost-prices/upload")
async def upload_cost_prices_excel(org_id: str, request: Request, db: AsyncSession = Depends(get_db)):
    """Загрузка справочника из Excel/CSV — полная версия со всеми колонками"""
    import io, csv
    from sqlalchemy import text
    body = await request.body()
    filename = request.headers.get("x-filename", "upload.csv")
    rows = []
    if filename.endswith(".csv"):
        reader = csv.DictReader(io.StringIO(body.decode("utf-8-sig")), delimiter=";")
        rows = list(reader)
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(body))
            ws = wb.active
            headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        except ImportError:
            raise HTTPException(400, "xlsx не поддерживается")
    
    def pf(row, *keys):
        """Parse float from row by multiple possible key names"""
        for k in keys:
            v = row.get(k)
            if v is not None and str(v).strip():
                try: return float(str(v).replace(',','.'))
                except: pass
        return None
    def ps(row, *keys):
        """Parse string from row"""
        for k in keys:
            v = row.get(k)
            if v and str(v).strip(): return str(v).strip()
        return None
    def pd(row, *keys):
        """Parse date from row (YYYY-MM-DD string to date object)"""
        s = ps(row, *keys)
        if not s: return None
        try:
            from datetime import datetime as _dt
            return _dt.strptime(s, "%Y-%m-%d").date()
        except: return None
    
    updated = 0
    for row in rows:
        nm = row.get("Арт WB") or row.get("nm_id")
        if not nm: continue
        nm = int(nm)
        
        # entity_id: ищем по nm_id + size_name, fallback на nm_id
        sz_val = ps(row, "Размер", "size_name")
        eid = None
        if sz_val:
            ent_q = await db.execute(text(
                "SELECT pe.id FROM product_entities pe "
                "WHERE pe.organization_id = :org AND pe.nm_id = :nm AND pe.size_name = :sz LIMIT 1"
            ), {"org": org_id, "nm": nm, "sz": sz_val})
            ent_row = ent_q.first()
            eid = str(ent_row[0]) if ent_row else None
        if not eid:
            ent_q = await db.execute(text(
                "SELECT pe.id FROM product_entities pe "
                "WHERE pe.organization_id = :org AND pe.nm_id = :nm LIMIT 1"
            ), {"org": org_id, "nm": nm})
            ent_row = ent_q.first()
            eid = str(ent_row[0]) if ent_row else None
        
        # Собираем все поля — полная версия справочника
        params = {
            "org": org_id, "nm": nm,
            "bc": ps(row, "Баркод", "barcode"),
            "vc": ps(row, "Арт продавца", "vendor_code"),
            "sz": ps(row, "Размер", "size_name"), "eid": eid,
            # Себестоимость
            # Себестоимость
            "cp": pf(row, "Себестоимость", "cost_price"),
            "pc": pf(row, "Закупка", "purchase_cost"),
            "lc": pf(row, "Логистика", "logistics_cost"),
            "pk": pf(row, "Упаковка", "packaging_cost"),
            "oc": pf(row, "Прочее", "other_costs"),
            "ec": pf(row, "Доп расходы", "extra_costs"),
            "vat": pf(row, "НДС руб", "vat"),
            "minp": pf(row, "Мин. цена", "min_price"),
            # МП/Комиссия
            "mpb": pf(row, "Баз. % МП", "mp_base_pct"),
            "mpc": pf(row, "Корр. % МП", "mp_correction_pct"),
            "ffm": ps(row, "ФБО/ФБС", "fulfillment_model") or "fbo",
            "stp": pf(row, "% хранения", "storage_pct"),
            "bnp": pf(row, "% выкупа по категории", "buyout_niche_pct"),
            # Цены
            "pspp": pf(row, "Цена до СПП план", "price_before_spp_plan"),
            "psppc": pf(row, "Цена до СПП к изм.", "price_before_spp_change"),
            "cdate": pd(row, "Дата правок", "change_date"),
            "wbcd": pf(row, "Скидка WB Клуб %", "wb_club_discount_pct"),
            # Реклама
            "adpr": pf(row, "Реклама план", "ad_plan_rub"),
            # Классификация
            "pcls": ps(row, "Класс товара", "product_class"),
            "brand": ps(row, "Бренд", "brand"),
            "pstatus": ps(row, "Статус товара", "product_status"),
            # Налоги
            "tsys": ps(row, "Налог. система", "tax_system"),
            # Сезонность (12 месяцев)
            "sjan": pf(row, "Сезон янв", "season_jan"),
            "sfeb": pf(row, "Сезон фев", "season_feb"),
            "smar": pf(row, "Сезон мар", "season_mar"),
            "sapr": pf(row, "Сезон апр", "season_apr"),
            "smay": pf(row, "Сезон май", "season_may"),
            "sjun": pf(row, "Сезон июн", "season_jun"),
            "sjul": pf(row, "Сезон июл", "season_jul"),
            "saug": pf(row, "Сезон авг", "season_aug"),
            "ssep": pf(row, "Сезон сен", "season_sep"),
            "soct": pf(row, "Сезон окт", "season_oct"),
            "snov": pf(row, "Сезон ноя", "season_nov"),
            "sdec": pf(row, "Сезон дек", "season_dec"),
            # Габариты ПЛАН
            "plen": pf(row, "План длина", "plan_length"),
            "pwid": pf(row, "План ширина", "plan_width"),
            "phei": pf(row, "План высота", "plan_height"),
            "pvol": pf(row, "План объём", "plan_volume"),
            "pwgt": pf(row, "План вес", "plan_weight"),
            # Доставка
            "ddts": pf(row, "Доставка до склада (дни)", "delivery_days_to_seller"),
            "ddmp": pf(row, "Доставка до МП (дни)", "delivery_days_to_mp"),
            # ТОП запросы
            "tq1": ps(row, "ТОП запрос 1", "top_query_1"),
            "tq2": ps(row, "ТОП запрос 2", "top_query_2"),
            "tq3": ps(row, "ТОП запрос 3", "top_query_3"),
            # Отгрузка
            "shm": ps(row, "Способ отгрузки", "shipment_method"),
            "fbsw": ps(row, "Склад отгрузки FBS", "fbs_warehouse"), "rrc": pf(row, "РРЦ", "rrc_price"),
            # subject_id/name (из product_entities, но можно передать в файле)
            "subid": int(row.get("subject_id")) if row.get("subject_id") and str(row.get("subject_id")).strip().isdigit() else None,
            "subn": ps(row, "Категория", "subject_name"),
            # Прочее
            "notes": ps(row, "Заметки", "notes"),
        }
        
        await db.execute(text(
            "INSERT INTO reference_book ("
            "organization_id, nm_id, barcode, vendor_code, size_name, entity_id, "
            "subject_id, subject_name, "
            "cost_price, purchase_cost, logistics_cost, packaging_cost, other_costs, extra_costs, vat, min_price, "
            "mp_base_pct, mp_correction_pct, fulfillment_model, storage_pct, buyout_niche_pct, "
            "price_before_spp_plan, price_before_spp_change, change_date, wb_club_discount_pct, ad_plan_rub, "
            "product_class, brand, product_status, tax_system, "
            "season_jan, season_feb, season_mar, season_apr, season_may, season_jun, "
            "season_jul, season_aug, season_sep, season_oct, season_nov, season_dec, "
            "plan_length, plan_width, plan_height, plan_volume, plan_weight, "
            "delivery_days_to_seller, delivery_days_to_mp, "
            "top_query_1, top_query_2, top_query_3, "
            "shipment_method, fbs_warehouse, rrc_price, "
            "notes, valid_from, source) "
            "VALUES ("
            ":org, :nm, :bc, :vc, :sz, :eid, "
            ":subid, :subn, "
            ":cp, :pc, :lc, :pk, :oc, :ec, :vat, :minp, "
            ":mpb, :mpc, :ffm, :stp, :bnp, "
            ":pspp, :psppc, :cdate, :wbcd, :adpr, "
            ":pcls, :brand, :pstatus, :tsys, "
            ":sjan, :sfeb, :smar, :sapr, :smay, :sjun, "
            ":sjul, :saug, :ssep, :soct, :snov, :sdec, "
            ":plen, :pwid, :phei, :pvol, :pwgt, "
            ":ddts, :ddmp, "
            ":tq1, :tq2, :tq3, "
            ":shm, :fbsw, :rrc, "
            ":notes, CURRENT_DATE, 'excel') "
            "ON CONFLICT (organization_id, nm_id, entity_id, valid_from) DO UPDATE SET "
            "barcode = COALESCE(EXCLUDED.barcode, reference_book.barcode), "
            "vendor_code = COALESCE(EXCLUDED.vendor_code, reference_book.vendor_code), "
            "cost_price = COALESCE(EXCLUDED.cost_price, reference_book.cost_price), "
            "purchase_cost = COALESCE(EXCLUDED.purchase_cost, reference_book.purchase_cost), "
            "logistics_cost = COALESCE(EXCLUDED.logistics_cost, reference_book.logistics_cost), "
            "packaging_cost = COALESCE(EXCLUDED.packaging_cost, reference_book.packaging_cost), "
            "other_costs = COALESCE(EXCLUDED.other_costs, reference_book.other_costs), "
            "extra_costs = COALESCE(EXCLUDED.extra_costs, reference_book.extra_costs), "
            "vat = COALESCE(EXCLUDED.vat, reference_book.vat), "
            "min_price = COALESCE(EXCLUDED.min_price, reference_book.min_price), "
            "mp_base_pct = COALESCE(EXCLUDED.mp_base_pct, reference_book.mp_base_pct), "
            "mp_correction_pct = COALESCE(EXCLUDED.mp_correction_pct, reference_book.mp_correction_pct), "
            "fulfillment_model = COALESCE(EXCLUDED.fulfillment_model, reference_book.fulfillment_model), "
            "storage_pct = COALESCE(EXCLUDED.storage_pct, reference_book.storage_pct), "
            "buyout_niche_pct = COALESCE(EXCLUDED.buyout_niche_pct, reference_book.buyout_niche_pct), "
            "price_before_spp_plan = COALESCE(EXCLUDED.price_before_spp_plan, reference_book.price_before_spp_plan), "
            "price_before_spp_change = COALESCE(EXCLUDED.price_before_spp_change, reference_book.price_before_spp_change), "
            "change_date = COALESCE(EXCLUDED.change_date, reference_book.change_date), "
            "wb_club_discount_pct = COALESCE(EXCLUDED.wb_club_discount_pct, reference_book.wb_club_discount_pct), "
            "ad_plan_rub = COALESCE(EXCLUDED.ad_plan_rub, reference_book.ad_plan_rub), "
            "product_class = COALESCE(EXCLUDED.product_class, reference_book.product_class), "
            "brand = COALESCE(EXCLUDED.brand, reference_book.brand), "
            "product_status = COALESCE(EXCLUDED.product_status, reference_book.product_status), "
            "tax_system = COALESCE(EXCLUDED.tax_system, reference_book.tax_system), "
        "tax_rate = COALESCE(EXCLUDED.tax_rate, reference_book.tax_rate), "
            "season_jan = COALESCE(EXCLUDED.season_jan, reference_book.season_jan), "
            "season_feb = COALESCE(EXCLUDED.season_feb, reference_book.season_feb), "
            "season_mar = COALESCE(EXCLUDED.season_mar, reference_book.season_mar), "
            "season_apr = COALESCE(EXCLUDED.season_apr, reference_book.season_apr), "
            "season_may = COALESCE(EXCLUDED.season_may, reference_book.season_may), "
            "season_jun = COALESCE(EXCLUDED.season_jun, reference_book.season_jun), "
            "season_jul = COALESCE(EXCLUDED.season_jul, reference_book.season_jul), "
            "season_aug = COALESCE(EXCLUDED.season_aug, reference_book.season_aug), "
            "season_sep = COALESCE(EXCLUDED.season_sep, reference_book.season_sep), "
            "season_oct = COALESCE(EXCLUDED.season_oct, reference_book.season_oct), "
            "season_nov = COALESCE(EXCLUDED.season_nov, reference_book.season_nov), "
            "season_dec = COALESCE(EXCLUDED.season_dec, reference_book.season_dec), "
            "plan_length = COALESCE(EXCLUDED.plan_length, reference_book.plan_length), "
            "plan_width = COALESCE(EXCLUDED.plan_width, reference_book.plan_width), "
            "plan_height = COALESCE(EXCLUDED.plan_height, reference_book.plan_height), "
            "plan_volume = COALESCE(EXCLUDED.plan_volume, reference_book.plan_volume), "
            "plan_weight = COALESCE(EXCLUDED.plan_weight, reference_book.plan_weight), "
            "delivery_days_to_seller = COALESCE(EXCLUDED.delivery_days_to_seller, reference_book.delivery_days_to_seller), "
            "delivery_days_to_mp = COALESCE(EXCLUDED.delivery_days_to_mp, reference_book.delivery_days_to_mp), "
            "top_query_1 = COALESCE(EXCLUDED.top_query_1, reference_book.top_query_1), "
            "top_query_2 = COALESCE(EXCLUDED.top_query_2, reference_book.top_query_2), "
            "top_query_3 = COALESCE(EXCLUDED.top_query_3, reference_book.top_query_3), "
            "shipment_method = COALESCE(EXCLUDED.shipment_method, reference_book.shipment_method), "
            "fbs_warehouse = COALESCE(EXCLUDED.fbs_warehouse, reference_book.fbs_warehouse), rrc_price = COALESCE(EXCLUDED.rrc_price, reference_book.rrc_price), "
            "notes = COALESCE(EXCLUDED.notes, reference_book.notes), "
            "subject_id = COALESCE(EXCLUDED.subject_id, reference_book.subject_id), "
            "subject_name = COALESCE(EXCLUDED.subject_name, reference_book.subject_name), "
            "source = EXCLUDED.source"
        ), params)
        updated += 1
    await db.commit()
    return {"updated": updated, "total": len(rows)}


@router.get("/api/v1/nl/sellers")
async def get_sellers(org_id: str, db: AsyncSession = Depends(get_db)):
    """Список продавцов"""
    from sqlalchemy import text
    result = await db.execute(text(
        "SELECT id, seller_id, seller_name, inn, seller_type, contact_name, "
        "contact_email, contact_phone, role, is_active, notes, created_at "
        "FROM sellers WHERE organization_id = :org ORDER BY created_at DESC"
    ), {"org": org_id})
    return [{"id": str(r[0]), "seller_id": r[1], "seller_name": r[2], "inn": r[3],
             "seller_type": r[4], "contact_name": r[5], "contact_email": r[6],
             "contact_phone": r[7], "role": r[8], "is_active": r[9],
             "notes": r[10], "created_at": str(r[11])} for r in result.all()]


@router.post("/api/v1/nl/sellers")
async def add_seller(data: dict, org_id: str, db: AsyncSession = Depends(get_db)):
    """Добавить продавца"""
    from sqlalchemy import text
    await db.execute(text(
        "INSERT INTO sellers (organization_id, seller_id, seller_name, inn, seller_type, "
        "contact_name, contact_email, contact_phone, role, notes) "
        "VALUES (:org, :sid, :name, :inn, :type, :cname, :email, :phone, :role, :notes)"
    ), {"org": org_id, "sid": data.get("seller_id"), "name": data.get("seller_name"),
        "inn": data.get("inn"), "type": data.get("seller_type", "fbo"),
        "cname": data.get("contact_name"), "email": data.get("contact_email"),
        "phone": data.get("contact_phone"), "role": data.get("role", "seller"),
        "notes": data.get("notes")})
    await db.commit()
    return {"ok": True}


@router.get("/api/v1/nl/seo-keywords")
async def get_seo_keywords(org_id: str, nm_id: Optional[str] = None, db: AsyncSession = Depends(get_db)):
    """SEO ключевые запросы"""
    from sqlalchemy import text
    q = "SELECT id, nm_id, vendor_code, keyword, position, frequency_monthly, frequency_weekly, "         "season_start, season_end, season_multiplier, trend, trend_value, competition, "         "target_date, source, notes FROM seo_keywords WHERE organization_id = :org"
    params = {"org": org_id}
    if nm_id:
        q += " AND nm_id = :nm"
        params["nm"] = int(nm_id)
    q += " ORDER BY target_date DESC NULLS LAST, frequency_monthly DESC NULLS LAST"
    result = await db.execute(text(q), params)
    return [{"id": str(r[0]), "nm_id": r[1], "vendor_code": r[2], "keyword": r[3],
             "position": r[4], "frequency_monthly": r[5], "frequency_weekly": r[6],
             "season_start": r[7], "season_end": r[8],
             "season_multiplier": float(r[9]) if r[9] else 1.0,
             "trend": r[10], "trend_value": float(r[11]) if r[11] else None,
             "competition": r[12], "target_date": str(r[13]) if r[13] else None,
             "source": r[14], "notes": r[15]} for r in result.all()]


@router.post("/api/v1/nl/seo-keywords")
async def add_seo_keyword(data: dict, org_id: str, db: AsyncSession = Depends(get_db)):
    """Добавить SEO запрос"""
    from sqlalchemy import text
    await db.execute(text(
        "INSERT INTO seo_keywords (organization_id, nm_id, vendor_code, keyword, position, "
        "frequency_monthly, frequency_weekly, season_start, season_end, season_multiplier, "
        "trend, trend_value, competition, target_date, source, notes) "
        "VALUES (:org, :nm, :vc, :kw, :pos, :fm, :fw, :ss, :se, :sm, :trend, :tv, :comp, :td, :src, :notes)"
    ), {"org": org_id, "nm": data.get("nm_id"), "vc": data.get("vendor_code"),
        "kw": data.get("keyword"), "pos": data.get("position"),
        "fm": data.get("frequency_monthly"), "fw": data.get("frequency_weekly"),
        "ss": data.get("season_start"), "se": data.get("season_end"),
        "sm": data.get("season_multiplier", 1.0), "trend": data.get("trend"),
        "tv": data.get("trend_value"), "comp": data.get("competition"),
        "td": data.get("target_date"), "src": data.get("source", "manual"),
        "notes": data.get("notes")})
    await db.commit()
    return {"ok": True}


@router.post("/api/v1/nl/seo-keywords/upload")
async def upload_seo_keywords(org_id: str, request: Request, db: AsyncSession = Depends(get_db)):
    """Загрузка SEO запросов из Excel/CSV"""
    import io, csv
    from sqlalchemy import text
    body = await request.body()
    filename = request.headers.get("x-filename", "upload.csv")
    rows = []
    if filename.endswith(".csv"):
        rows = list(csv.DictReader(io.StringIO(body.decode("utf-8-sig")), delimiter=";"))
    updated = 0
    for row in rows:
        nm = row.get("Арт WB") or row.get("nm_id")
        kw = row.get("Запрос") or row.get("keyword")
        if nm and kw:
            await db.execute(text(
                "INSERT INTO seo_keywords (organization_id, nm_id, vendor_code, keyword, "
                "position, frequency_monthly, season_start, season_end, trend, competition, target_date, source) "
                "VALUES (:org, :nm, :vc, :kw, :pos, :fm, :ss, :se, :trend, :comp, CURRENT_DATE, 'excel')"
            ), {"org": org_id, "nm": int(nm), "vc": row.get("Арт продавца",""),
                "kw": kw, "pos": row.get("Позиция"), "fm": row.get("Частотность"),
                "ss": row.get("Сезон начало"), "se": row.get("Сезон конец"),
                "trend": row.get("Тренд"), "comp": row.get("Конкуренция")})
            updated += 1
    await db.commit()
    return {"updated": updated, "total": len(rows)}



# ─── ПЛАН ПРОДАЖ ────────────────────────────────────────────

@router.get("/api/v1/nl/sales-plans")
async def get_sales_plans(org_id: str, period: Optional[str] = None, db: AsyncSession = Depends(get_db)):
    """План продаж — список по организации с фильтром по периоду"""
    from sqlalchemy import text
    sql = (
        "SELECT sp.id, sp.entity_id, sp.nm_id, sp.vendor_code, sp.size_name, "
        "sp.period, sp.plan_type, sp.plan_value, sp.actual_value, "
        "sp.sales_temp, sp.seasonality, sp.created_at, sp.updated_at, "
        "pe.product_name, pe.photo_main "
        "FROM sales_plans sp "
        "LEFT JOIN product_entities pe ON sp.entity_id = pe.id "
        "WHERE sp.organization_id = :org "
    )
    params = {"org": org_id}
    if period:
        sql += " AND sp.period = :period"
        params["period"] = datetime.strptime(period, "%Y-%m-%d").date()
    sql += " ORDER BY sp.nm_id, sp.period DESC"
    result = await db.execute(text(sql), params)
    return [{
        "id": str(r[0]),
        "entity_id": str(r[1]) if r[1] else None,
        "nm_id": r[2],
        "vendor_code": r[3],
        "size_name": r[4],
        "period": str(r[5]),
        "plan_type": r[6],
        "plan_value": float(r[7]) if r[7] else 0,
        "actual_value": float(r[8]) if r[8] else 0,
        "sales_temp": float(r[9]) if r[9] else None,
        "seasonality": r[10],
        "created_at": str(r[11]) if r[11] else None,
        "updated_at": str(r[12]) if r[12] else None,
        "product_name": r[13],
        "photo_main": r[14],
        "pct_complete": round(float(r[8]) / float(r[7]) * 100, 1) if r[7] and float(r[7]) > 0 else 0,
    } for r in result.all()]


@router.post("/api/v1/nl/sales-plans")
async def save_sales_plan(data: SalesPlanItem, org_id: str, db: AsyncSession = Depends(get_db)):
    """Создать/обновить план продаж"""
    period_date = datetime.strptime(data.period, "%Y-%m-%d").date()
    # Приводим к первому дню месяца
    period_date = period_date.replace(day=1)

    # entity_id lookup
    eid = None
    if data.entity_id:
        eid = data.entity_id
    else:
        ent_q = await db.execute(text(
            "SELECT id FROM product_entities WHERE organization_id = :org AND nm_id = :nm LIMIT 1"
        ), {"org": org_id, "nm": data.nm_id})
        ent_row = ent_q.first()
        eid = str(ent_row[0]) if ent_row else None

    # vendor_code / size_name из entity если не заданы
    vc = data.vendor_code
    sn = data.size_name
    if eid and (not vc or not sn):
        ent_info = await db.execute(text(
            "SELECT vendor_code, size_name FROM product_entities WHERE id = :eid"
        ), {"eid": eid})
        ent_row = ent_info.first()
        if ent_row:
            vc = vc or ent_row[0]
            sn = sn or ent_row[1]

    ins = pg_insert(SalesPlan).values(
        organization_id=org_id,
        entity_id=eid,
        nm_id=data.nm_id,
        vendor_code=vc,
        size_name=sn,
        period=period_date,
        plan_type=data.plan_type,
        plan_value=data.plan_value,
        actual_value=data.actual_value,
        sales_temp=data.sales_temp,
        seasonality=data.seasonality,
    )
    stmt = ins.on_conflict_do_update(
        constraint="sales_plans_org_entity_period_type_key",
        set_={
            "vendor_code": ins.excluded.vendor_code,
            "size_name": ins.excluded.size_name,
            "plan_value": ins.excluded.plan_value,
            "actual_value": ins.excluded.actual_value,
            "sales_temp": ins.excluded.sales_temp,
            "seasonality": ins.excluded.seasonality,
            "updated_at": datetime.utcnow(),
        }
    )
    await db.execute(stmt)
    await db.commit()
    return {"status": "ok"}


@router.put("/api/v1/nl/sales-plans/{plan_id}")
async def update_sales_plan(plan_id: str, data: dict, org_id: str, db: AsyncSession = Depends(get_db)):
    """Обновить отдельные поля плана продаж"""
    from sqlalchemy import text as _t
    fields = []
    params = {"pid": plan_id, "org": org_id}
    for key in ["plan_value", "actual_value", "sales_temp", "plan_type", "seasonality", "vendor_code", "size_name"]:
        if key in data:
            fields.append(f"{key} = :{key}")
            params[key] = data[key]
    if not fields:
        return {"status": "noop"}
    fields.append("updated_at = NOW()")
    sql = f"UPDATE sales_plans SET {', '.join(fields)} WHERE id = :pid AND organization_id = :org"
    await db.execute(_t(sql), params)
    await db.commit()
    return {"status": "ok"}


@router.delete("/api/v1/nl/sales-plans/{plan_id}")
async def delete_sales_plan(plan_id: str, org_id: str, db: AsyncSession = Depends(get_db)):
    """Удалить план продаж"""
    from sqlalchemy import text as _t
    await db.execute(_t(
        "DELETE FROM sales_plans WHERE id = :pid AND organization_id = :org"
    ), {"pid": plan_id, "org": org_id})
    await db.commit()
    return {"status": "ok"}


@router.post("/api/v1/nl/sales-plans/batch")
async def batch_sales_plans(items: list[SalesPlanItem], org_id: str, db: AsyncSession = Depends(get_db)):
    """Массовое создание/обновление планов продаж"""
    updated = 0
    for item in items:
        period_date = datetime.strptime(item.period, "%Y-%m-%d").date().replace(day=1)

        ent_q = await db.execute(text(
            "SELECT id, vendor_code, size_name FROM product_entities "
            "WHERE organization_id = :org AND nm_id = :nm LIMIT 1"
        ), {"org": org_id, "nm": item.nm_id})
        ent_row = ent_q.first()
        eid = str(ent_row[0]) if ent_row else None
        vc = item.vendor_code or (ent_row[1] if ent_row else None)
        sn = item.size_name or (ent_row[2] if ent_row else None)

        ins = pg_insert(SalesPlan).values(
            organization_id=org_id,
            entity_id=eid,
            nm_id=item.nm_id,
            vendor_code=vc,
            size_name=sn,
            period=period_date,
            plan_type=item.plan_type,
            plan_value=item.plan_value,
            actual_value=item.actual_value,
            sales_temp=item.sales_temp,
            seasonality=item.seasonality,
        )
        stmt = ins.on_conflict_do_update(
            constraint="sales_plans_org_entity_period_type_key",
            set_={
                "plan_value": ins.excluded.plan_value,
                "actual_value": ins.excluded.actual_value,
                "sales_temp": ins.excluded.sales_temp,
                "seasonality": ins.excluded.seasonality,
                "updated_at": datetime.utcnow(),
            }
        )
        await db.execute(stmt)
        updated += 1
    await db.commit()
    return {"status": "ok", "updated": updated}


@router.get("/api/v1/nl/sales-plans/summary")
async def sales_plans_summary(org_id: str, period: Optional[str] = None, db: AsyncSession = Depends(get_db)):
    """Сводка по плану продаж"""
    from sqlalchemy import text
    params = {"org": org_id}
    where = "WHERE organization_id = :org"
    if period:
        where += " AND period = :period"
        params["period"] = datetime.strptime(period, "%Y-%m-%d").date()
    result = await db.execute(text(
        f"SELECT plan_type, "
        f"SUM(plan_value) as total_plan, "
        f"SUM(actual_value) as total_actual, "
        f"COUNT(*) as items_count, "
        f"COUNT(*) FILTER (WHERE actual_value / NULLIF(plan_value,0) >= 0.9) as green_count, "
        f"COUNT(*) FILTER (WHERE actual_value / NULLIF(plan_value,0) >= 0.7 AND actual_value / NULLIF(plan_value,0) < 0.9) as yellow_count, "
        f"COUNT(*) FILTER (WHERE actual_value / NULLIF(plan_value,0) < 0.7) as red_count "
        f"FROM sales_plans {where} GROUP BY plan_type"
    ), params)
    rows = result.all()
    return [{
        "plan_type": r[0],
        "total_plan": float(r[1]) if r[1] else 0,
        "total_actual": float(r[2]) if r[2] else 0,
        "items_count": r[3],
        "pct_complete": round(float(r[2]) / float(r[1]) * 100, 1) if r[1] and float(r[1]) > 0 else 0,
        "green_count": r[4],
        "yellow_count": r[5],
        "red_count": r[6],
    } for r in rows]


# ─── РЕКЛАМА ────────────────────────────────────────────

@router.get("/api/v1/nl/ad-stats")
async def get_ad_stats(org_id: str, days: str = "7", date_from: Optional[str] = None, date_to: Optional[str] = None, db: AsyncSession = Depends(get_db)):
    """Рекламная статистика — из ad_stats_nm (те же данные что по артикулам)"""
    import decimal as _dec, json as _json

    if date_from and date_to:
        d_from = date_from
        d_to = date_to
    else:
        try:
            days_int = int(days)
        except:
            days_int = 7
        if days_int == 1:
            d_from = date.today().isoformat()
            d_to = date.today().isoformat()
        elif days_int == 2:
            d = date.today() - timedelta(days=1)
            d_from = d.isoformat()
            d_to = d.isoformat()
        else:
            d_from = (date.today() - timedelta(days=days_int)).isoformat()
            d_to = date.today().isoformat()

    params = {
        "org": org_id,
        "d_from": datetime.strptime(d_from, "%Y-%m-%d").date(),
        "d_to": datetime.strptime(d_to, "%Y-%m-%d").date(),
    }

    def sf(v):
        if v is None: return 0
        return float(v) if not isinstance(v, _dec.Decimal) else float(v)

    # ═══ Статистика по дням (из ad_stats_nm) ═══
    daily_rows = await db.execute(text("""
        SELECT sn.stat_date,
               SUM(sn.views) as views,
               SUM(sn.clicks) as clicks,
               SUM(sn.spent) as spent,
               SUM(sn.orders) as orders,
               SUM(sn.atbs) as atbs
        FROM ad_stats_nm sn
        WHERE sn.organization_id = :org
            AND sn.stat_date >= :d_from AND sn.stat_date <= :d_to
            AND sn.spent > 0
        GROUP BY sn.stat_date
        ORDER BY sn.stat_date DESC
    """), params)

    # ═══ ДРР по дням: sum_price из ad_stats_nm по составу РК ═══
    sum_price_by_day = await db.execute(text("""
        SELECT sn.stat_date, COALESCE(SUM(sn.sum_price), 0) as sum_price
        FROM ad_stats_nm sn
        WHERE sn.organization_id = :org
            AND sn.stat_date >= :d_from AND sn.stat_date <= :d_to
            AND sn.spent > 0
        GROUP BY sn.stat_date
    """), params)
    sp_by_date = {}
    for r in sum_price_by_day:
        sp_by_date[str(r[0])] = round(sf(r[1]), 2)

    daily = []
    for r in daily_rows:
        views = int(r[1] or 0)
        clicks = int(r[2] or 0)
        spent = round(sf(r[3]), 2)
        orders = int(r[4] or 0)
        atbs = int(r[5] or 0)
        date_str = str(r[0])
        sum_price_day = sp_by_date.get(date_str, 0)
        drr_day = round(spent / sum_price_day * 100, 1) if sum_price_day else 0
        daily.append({
            "date": date_str,
            "views": views,
            "clicks": clicks,
            "spent": spent,
            "ctr": round(clicks / views * 100, 2) if views else 0,
            "cpc": round(spent / clicks, 2) if clicks else 0,
            "orders": orders,
            "atbs": atbs,
            "cr": round(orders / clicks * 100, 2) if clicks else 0,
            "sum_price": sum_price_day,
            "drr": drr_day,
        })

    # ═══ Список кампаний (из ad_stats_nm, агрегировано по РК) ═══
    camp_rows = await db.execute(text("""
        SELECT sn.wb_campaign_id, c.name, c.status, c.type,
               SUM(sn.views) as views,
               SUM(sn.clicks) as clicks,
               SUM(sn.spent) as spent,
               SUM(sn.orders) as orders,
               SUM(sn.atbs) as atbs,
               (SELECT COUNT(DISTINCT sn2.nm_id) FROM ad_stats_nm sn2
                    WHERE sn2.organization_id = :org
                    AND sn2.wb_campaign_id = sn.wb_campaign_id
                    AND sn2.spent > 0
                    AND sn2.stat_date >= :d_from AND sn2.stat_date <= :d_to
               ) as nm_count
        FROM ad_stats_nm sn
        JOIN ad_campaigns c ON c.wb_campaign_id = sn.wb_campaign_id
            AND c.organization_id = sn.organization_id
        WHERE sn.organization_id = :org
            AND sn.stat_date >= :d_from AND sn.stat_date <= :d_to
            AND sn.spent > 0
        GROUP BY sn.wb_campaign_id, c.name, c.status, c.type
        ORDER BY SUM(sn.spent) DESC
    """), params)

    campaigns = []
    for r in camp_rows:
        views = int(r[4] or 0)
        clicks = int(r[5] or 0)
        spent = round(sf(r[6]), 2)
        orders = int(r[7] or 0)
        atbs = int(r[8] or 0)
        nm_count = int(r[9] or 0)

        # Состав РК — nm_id из ad_stats_nm
        nm_ids_row = await db.execute(text("""
            SELECT DISTINCT nm_id FROM ad_stats_nm
            WHERE organization_id = :org AND wb_campaign_id = :cid
                AND stat_date >= :d_from AND stat_date <= :d_to AND spent > 0
            ORDER BY nm_id
        """), {**params, "cid": r[0]})
        nm_ids = [int(n[0]) for n in nm_ids_row]

        # Инфо о товарах
        products = []
        if nm_ids:
            prod_row = await db.execute(text("""
                SELECT raw_response FROM raw_api_data
                WHERE api_method = 'products' AND organization_id = :org
                ORDER BY fetched_at DESC LIMIT 1
            """), {"org": org_id})
            pr = prod_row.first()
            if pr and pr[0]:
                cards_data = pr[0] if isinstance(pr[0], list) else (pr[0].get("cards", []) if isinstance(pr[0], dict) else [])
                nm_set = set(nm_ids)
                for cd in cards_data:
                    if not isinstance(cd, dict): continue
                    nm = cd.get("nmID")
                    if nm and int(nm) in nm_set:
                        photos = cd.get("photos") or []
                        photo_url = ""
                        if photos:
                            photo_url = photos[0].get("c246x328", "") or photos[0].get("big", "") or photos[0].get("hq", "")
                        products.append({
                            "nm_id": int(nm),
                            "vendor_code": cd.get("vendorCode", ""),
                            "name": cd.get("title", ""),
                            "photo": photo_url,
                        })

        # Сумма заказов для ДРР из ad_stats_nm (только состав РК)
        sum_price_row = await db.execute(text("""
            SELECT COALESCE(SUM(sn.sum_price), 0) as sum_price
            FROM ad_stats_nm sn
            WHERE sn.organization_id = :org AND sn.wb_campaign_id = :cid
                AND sn.stat_date >= :d_from AND sn.stat_date <= :d_to
                AND sn.spent > 0
        """), {**params, "cid": r[0]})
        sum_price_val = round(sf(sum_price_row.scalar()), 2)

        # Общие заказы и выручка по товарам этой РК из tech_status
        camp_nm_ids_for_drr = nm_ids  # уже есть nm_ids
        total_orders_rk = 0
        total_revenue_rk = 0
        if camp_nm_ids_for_drr:
            rk_totals_row = await db.execute(text("""
                SELECT COALESCE(SUM(ts.orders_count), 0),
                       COALESCE(SUM(ts.orders_count * ts.price_discount), 0)
                FROM tech_status ts
                WHERE ts.organization_id = :org
                    AND ts.target_date >= :d_from AND ts.target_date <= :d_to
                    AND ts.nm_id = ANY(:nm_ids)
            """), {**params, "nm_ids": camp_nm_ids_for_drr})
            rk_totals = rk_totals_row.first()
            if rk_totals:
                total_orders_rk = int(rk_totals[0] or 0)
                total_revenue_rk = round(sf(rk_totals[1]), 2)

        # ДРР = расход / sum_price (из ad_stats_nm по составу РК)
        drr_rk = round(spent / sum_price_val * 100, 1) if sum_price_val else 0

        campaigns.append({
            "campaign_id": r[0],
            "name": r[1] or "Без названия",
            "status": str(r[2]) if r[2] else "",
            "type": str(r[3]) if r[3] else "",
            "views": views,
            "clicks": clicks,
            "spent": spent,
            "ctr": round(clicks / views * 100, 2) if views else 0,
            "cpc": round(spent / clicks, 2) if clicks else 0,
            "orders": orders,
            "atbs": atbs,
            "nm_count": nm_count,
            "products": products,
            "sum_price": sum_price_val,
            "total_orders": total_orders_rk,
            "total_revenue": total_revenue_rk,
            "drr": drr_rk,
        })

    # ═══ Баланс ═══
    balance = None
    bal_row = await db.execute(text("""
        SELECT raw_response FROM raw_api_data
        WHERE api_method = 'ad_balance' AND status = 'ok' AND organization_id = :org
        ORDER BY fetched_at DESC LIMIT 1
    """), {"org": org_id})
    br = bal_row.first()
    if br and br[0]:
        balance = br[0]

    # ═══ Итого ═══
    totals = {"views": 0, "clicks": 0, "spent": 0, "orders": 0, "atbs": 0}
    for d in daily:
        for k in totals:
            totals[k] += d.get(k, 0)
    totals["ctr"] = round(totals["clicks"] / totals["views"] * 100, 2) if totals["views"] else 0
    totals["cpc"] = round(totals["spent"] / totals["clicks"], 2) if totals["clicks"] else 0
    totals["cr"] = round(totals["orders"] / totals["clicks"] * 100, 2) if totals["clicks"] else 0
    # ДРР общий = расход / sum_price из ad_stats_nm
    all_sum_price = sum(d.get("sum_price", 0) for d in daily)
    totals["drr"] = round(totals["spent"] / all_sum_price * 100, 1) if all_sum_price else 0
    totals["sum_price"] = round(all_sum_price, 2)

    return {
        "daily": daily,
        "campaigns": campaigns,
        "top_campaigns": campaigns[:20],  # Топ 20 для совместимости
        "totals": totals,
        "balance": balance,
    }
@router.get("/api/v1/nl/ad-stats/by-art")
async def get_ad_stats_by_art(org_id: str, days: str = "30", date_from: Optional[str] = None, date_to: Optional[str] = None, statuses: Optional[str] = None, db: AsyncSession = Depends(get_db)):
    """Рекламная статистика по артикулам — данные из ad_stats_nm (разбивка WB по nm_id)"""
    import decimal as _dec, json as _json

    if date_from and date_to:
        d_from = date_from
        d_to = date_to
    else:
        try:
            days_int = int(days)
        except:
            days_int = 7
        if days_int == 1:
            d_from = date.today().isoformat()
            d_to = date.today().isoformat()
        elif days_int == 2:
            d = date.today() - timedelta(days=1)
            d_from = d.isoformat()
            d_to = d.isoformat()
        else:
            d_from = (date.today() - timedelta(days=days_int)).isoformat()
            d_to = date.today().isoformat()

    params = {
        "org": org_id,
        "d_from": datetime.strptime(d_from, "%Y-%m-%d").date(),
        "d_to": datetime.strptime(d_to, "%Y-%m-%d").date(),
    }

    # Parse status filter — только реальные статусы WB: 7, 9, 11
    status_list = []
    if statuses:
        status_list = [s.strip() for s in statuses.split(",") if s.strip() and s.strip() in ("7", "9", "11")]

    status_cond = ""
    if status_list:
        status_cond = "AND c.status = ANY(:statuses)"
        params["statuses"] = status_list

    def sf(v):
        if v is None: return 0
        return float(v) if not isinstance(v, _dec.Decimal) else float(v)

    # ═══ Основной запрос: агрегация по nm_id из ad_stats_nm ═══
    rows = await db.execute(text("""
        SELECT
            sn.nm_id,
            SUM(sn.spent) as total_spent,
            SUM(sn.views) as total_views,
            SUM(sn.clicks) as total_clicks,
            SUM(sn.orders) as total_orders,
            SUM(sn.atbs) as total_atbs
        FROM ad_stats_nm sn
        JOIN ad_campaigns c ON c.wb_campaign_id = sn.wb_campaign_id
            AND c.organization_id = sn.organization_id
        WHERE sn.organization_id = :org
            AND sn.stat_date >= :d_from AND sn.stat_date <= :d_to
            """ + status_cond + """
        GROUP BY sn.nm_id
        HAVING SUM(sn.spent) > 0
        ORDER BY SUM(sn.spent) DESC
    """), params)

    art_data = {}
    for r in rows:
        nm_id = int(r[0])
        spent = round(sf(r[1]), 2)
        views = int(r[2] or 0)
        clicks = int(r[3] or 0)
        orders = int(r[4] or 0)
        atbs = int(r[5] or 0)
        art_data[nm_id] = {
            "spent": spent,
            "views": views,
            "clicks": clicks,
            "orders": orders,
            "atbs": atbs,
        }

    all_nm_ids = list(art_data.keys())

    # ═══ Для каждого артикула — список РК с данными по этому nm_id ═══
    nm_campaigns = {}  # nm_id -> [campaigns]
    if all_nm_ids:
        camp_rows = await db.execute(text("""
            SELECT
                sn.nm_id,
                sn.wb_campaign_id,
                c.name,
                c.status,
                c.type,
                SUM(sn.spent) as camp_spent,
                SUM(sn.views) as camp_views,
                SUM(sn.clicks) as camp_clicks,
                AVG(sn.ctr) as camp_ctr,
                SUM(sn.orders) as camp_orders,
                SUM(sn.atbs) as camp_atbs,
                SUM(sn.sum_price) as camp_sum_price
            FROM ad_stats_nm sn
            JOIN ad_campaigns c ON c.wb_campaign_id = sn.wb_campaign_id
                AND c.organization_id = sn.organization_id
            WHERE sn.organization_id = :org
                AND sn.stat_date >= :d_from AND sn.stat_date <= :d_to
                AND sn.nm_id = ANY(:nm_ids)
                """ + status_cond + """
            GROUP BY sn.nm_id, sn.wb_campaign_id, c.name, c.status, c.type
            HAVING SUM(sn.spent) > 0
            ORDER BY SUM(sn.spent) DESC
        """), {**params, "nm_ids": all_nm_ids})

        for r in camp_rows:
            nm_id = int(r[0])
            if nm_id not in nm_campaigns:
                nm_campaigns[nm_id] = []
            nm_campaigns[nm_id].append({
                "campaign_id": int(r[1]),
                "name": r[2] or "Без названия",
                "status": str(r[3]) if r[3] else "",
                "type": str(r[4]) if r[4] else "",
                "spent_share": round(sf(r[5]), 2),
                "views": int(r[6] or 0),
                "clicks": int(r[7] or 0),
                "ctr": round(sf(r[8]), 2),
                "orders": int(r[9] or 0),
                "atbs": int(r[10] or 0),
                "sum_price": round(sf(r[11]), 2),
            })

    # ═══ Собираем items ═══
    # ═══ Общие заказы и цена по nm_id из tech_status (для ДРР) ═══
    nm_orders_price = {}
    if all_nm_ids:
        ts_rows = await db.execute(text("""
            SELECT ts.nm_id,
                   SUM(ts.orders_count) as total_orders,
                   SUM(ts.orders_count * ts.price_discount) as total_revenue
            FROM tech_status ts
            WHERE ts.organization_id = :org
                AND ts.target_date >= :d_from AND ts.target_date <= :d_to
                AND ts.nm_id = ANY(:nm_ids)
            GROUP BY ts.nm_id
        """), {**params, "nm_ids": all_nm_ids})
        for r in ts_rows:
            nm_orders_price[int(r[0])] = {
                "total_orders": int(r[1] or 0),
                "total_revenue": round(sf(r[2]), 2),
            }

    items = []
    for nm_id in all_nm_ids:
        d = art_data[nm_id]
        spent = d["spent"]
        views = d["views"]
        clicks = d["clicks"]
        orders = d["orders"]
        campaigns = nm_campaigns.get(nm_id, [])
        # ДРР = расход / sum_price (из ad_stats_nm, только состав РК)
        sum_price_art = sum(c.get("sum_price", 0) for c in campaigns) if campaigns else 0
        op = nm_orders_price.get(nm_id, {"total_orders": 0, "total_revenue": 0})
        total_orders_art = op["total_orders"]
        total_revenue_art = op["total_revenue"]
        drr_art = round(spent / sum_price_art * 100, 1) if sum_price_art else 0
        items.append({
            "nm_id": nm_id,
            "spent": spent,
            "views": views,
            "clicks": clicks,
            "ctr": round(clicks / views * 100, 2) if views else 0,
            "cpc": round(spent / clicks, 2) if clicks else 0,
            "orders": orders,
            "cr": round(orders / clicks * 100, 2) if clicks else 0,
            "campaigns_count": len(campaigns),
            "campaigns": campaigns,
            "total_orders": total_orders_art,
            "total_revenue": total_revenue_art,
            "drr": drr_art,
        })

    # ═══ Информация о товарах (название, фото, vendor_code) ═══
    nm_to_info = {}
    if all_nm_ids:
        prod_row = await db.execute(text("""
            SELECT raw_response FROM raw_api_data
            WHERE api_method = 'products' AND organization_id = :org
            ORDER BY fetched_at DESC LIMIT 1
        """), {"org": org_id})
        pr = prod_row.first()
        if pr and pr[0]:
            cards_data = pr[0] if isinstance(pr[0], list) else (pr[0].get("cards", []) if isinstance(pr[0], dict) else [])
            nm_set = set(all_nm_ids)
            for c in cards_data:
                if not isinstance(c, dict): continue
                nm = c.get("nmID")
                if nm and int(nm) in nm_set:
                    photos = c.get("photos") or []
                    photo_url = ""
                    if photos:
                        photo_url = photos[0].get("c246x328", "") or photos[0].get("big", "") or photos[0].get("hq", "")
                    nm_to_info[int(nm)] = {
                        "name": c.get("title", ""),
                        "brand": c.get("brand", ""),
                        "vendor_code": c.get("vendorCode", ""),
                        "photo": photo_url,
                    }

    for item in items:
        info = nm_to_info.get(item["nm_id"], {})
        item["name"] = info.get("name", "")
        item["brand"] = info.get("brand", "")
        item["vendor_code"] = info.get("vendor_code", "")
        item["photo"] = info.get("photo", "")

    totals = {
        "spent": round(sum(i["spent"] for i in items), 2),
        "views": sum(i["views"] for i in items),
        "clicks": sum(i["clicks"] for i in items),
        "orders": sum(i["orders"] for i in items),
        "ctr": round(sum(i["clicks"] for i in items) / max(sum(i["views"] for i in items), 1) * 100, 2),
        "cpc": round(sum(i["spent"] for i in items) / max(sum(i["clicks"] for i in items), 1), 2),
        "cr": round(sum(i["orders"] for i in items) / max(sum(i["clicks"] for i in items), 1) * 100, 2),
        "items_count": len(items),
        "campaigns_count": sum(i["campaigns_count"] for i in items),
        "total_orders": sum(i["total_orders"] for i in items),
        "total_revenue": round(sum(i["total_revenue"] for i in items), 2),
        "drr": round(sum(i["spent"] for i in items) / max(sum(i["total_revenue"] for i in items), 1) * 100, 1),
    }

    return {"items": items, "totals": totals}

@router.get("/api/v1/nl/marketer/products")
async def get_marketer_products(
    org_id: str,
    days: str = "30",
    search: Optional[str] = None,
    status: Optional[str] = None,
    abc_class: Optional[str] = None,
    brand: Optional[str] = None,
    db: AsyncSession = Depends(get_db)
):
    """Стол маркетолога — список товаров с рекламными данными"""
    import decimal as _dec, json as _json

    try:
        days_int = int(days)
    except:
        days_int = 30

    def sf(v):
        if v is None: return 0
        return float(v) if isinstance(v, (_dec.Decimal, int)) else (float(v) if v else 0)

    date_from = f"CURRENT_DATE - make_interval(days => {days_int})"

    # 1) Получаем уникальные nm_id из активных/приостановленных кампаний
    active_statuses = ('7', '9', '11')  # WB status codes: 7=активна, 9=приостановлена, 11=завершена

    # Все РК с их составом (из ad_stats_nm, только spent > 0) за период
    camp_rows = await db.execute(text(f"""
        SELECT c.wb_campaign_id, c.name, c.type, c.status,
               COALESCE(SUM(sn.views),0), COALESCE(SUM(sn.clicks),0),
               COALESCE(SUM(sn.spent),0),
               COALESCE(SUM(sn.orders),0), COALESCE(SUM(sn.atbs),0),
               COALESCE(SUM(sn.sum_price),0)
        FROM ad_campaigns c
        JOIN ad_stats_nm sn ON sn.wb_campaign_id = c.wb_campaign_id
            AND sn.organization_id = c.organization_id
            AND sn.stat_date >= {date_from}
            AND sn.spent > 0
        WHERE c.organization_id = :org
        GROUP BY c.wb_campaign_id, c.name, c.type, c.status
        ORDER BY COALESCE(SUM(sn.spent),0) DESC
    """), {"org": org_id})

    # nm_id → какие РК к ним относятся (из ad_stats_nm, только spent > 0)
    nm_camp_rows = await db.execute(text(f"""
        SELECT sn.wb_campaign_id, sn.nm_id, c.name, c.type, c.status,
               COALESCE(SUM(sn.views),0), COALESCE(SUM(sn.clicks),0),
               COALESCE(SUM(sn.spent),0),
               COALESCE(SUM(sn.orders),0), COALESCE(SUM(sn.atbs),0),
               COALESCE(SUM(sn.sum_price),0)
        FROM ad_stats_nm sn
        JOIN ad_campaigns c ON c.wb_campaign_id = sn.wb_campaign_id
            AND c.organization_id = sn.organization_id
        WHERE sn.organization_id = :org
            AND sn.stat_date >= {date_from}
            AND sn.spent > 0
        GROUP BY sn.wb_campaign_id, sn.nm_id, c.name, c.type, c.status
    """), {"org": org_id})

    nm_to_campaigns = {}  # nm_id -> [{campaign_info}]
    all_campaigns = []
    all_nm_ids = set()
    camp_ids_seen = set()

    for r in nm_camp_rows:
        nm_id = int(r[1])
        camp_info = {
            "campaign_id": r[0],
            "name": r[2] or "Без названия",
            "type": str(r[3]) if r[3] else "",
            "status": str(r[4]) if r[4] else "",
            "views": int(sf(r[5])),
            "clicks": int(sf(r[6])),
            "spent": round(sf(r[7]), 2),
            "orders": int(sf(r[8])),
            "atbs": int(sf(r[9])),
            "sum_price": round(sf(r[10]), 2),
        }
        if nm_id not in nm_to_campaigns:
            nm_to_campaigns[nm_id] = []
        nm_to_campaigns[nm_id].append(camp_info)
        all_nm_ids.add(nm_id)
        if r[0] not in camp_ids_seen:
            camp_ids_seen.add(r[0])
            all_campaigns.append(camp_info)

    # 2) Получаем инфо о товарах (фото, название, бренд, статус, класс)
    product_info = {}
    if all_nm_ids:
        pe_rows = await db.execute(text("""
            SELECT DISTINCT ON (pe.nm_id) pe.nm_id, pe.brand, pe.subject_name, pe.photo_main,
                pe.vendor_code, pe.subject_name
            FROM product_entities pe
            WHERE pe.organization_id = :org AND pe.nm_id = ANY(:nms)
            ORDER BY pe.nm_id, pe.created_at DESC
        """), {"org": org_id, "nms": list(all_nm_ids)})
        for r in pe_rows:
            product_info[r[0]] = {
                "nm_id": r[0],
                "brand": r[1] or "",
                "category": r[2] or "",
                "photo": r[3] or "",
                "vendor_code": r[4] or "",
            }

        # Дополняем из tech_status (цены)
        ts_rows = await db.execute(text(f"""
            SELECT DISTINCT ON (ts.nm_id) ts.nm_id,
                ts.price, ts.price_spp, ts.price_discount
            FROM tech_status ts
            WHERE ts.organization_id = :org AND ts.nm_id = ANY(:nms)
              AND ts.target_date >= {date_from}
            ORDER BY ts.nm_id, ts.target_date DESC
        """), {"org": org_id, "nms": list(all_nm_ids)})
        for r in ts_rows:
            if r[0] in product_info:
                product_info[r[0]]["price"] = sf(r[1])
                product_info[r[0]]["price_spp"] = sf(r[2])
                product_info[r[0]]["price_discount"] = sf(r[3])

        # Дополняем из reference_book (статус, класс)
        rb_rows = await db.execute(text("""
            SELECT DISTINCT ON (rb.nm_id) rb.nm_id,
                rb.product_status as status,
                rb.product_class as abc_class
            FROM reference_book rb
            WHERE rb.organization_id = :org AND rb.nm_id = ANY(:nms)
            ORDER BY rb.nm_id, rb.valid_from DESC
        """), {"org": org_id, "nms": list(all_nm_ids)})
        for r in rb_rows:
            if r[0] in product_info:
                product_info[r[0]]["status"] = r[1] or ""
                product_info[r[0]]["abc_class"] = r[2] or ""

    # 3) Собираем список товаров
    products = []
    for nm_id in sorted(all_nm_ids):
        info = product_info.get(nm_id, {"nm_id": nm_id})
        camps = nm_to_campaigns.get(nm_id, [])

        # Суммарные метрики по всем РК товара
        total_views = sum(c["views"] for c in camps)
        total_clicks = sum(c["clicks"] for c in camps)
        total_spent = sum(c["spent"] for c in camps)
        total_orders = sum(c["orders"] for c in camps)

        active_camps = [c for c in camps if c["status"] in active_statuses]

        products.append({
            "nm_id": nm_id,
            "vendor_code": info.get("vendor_code", ""),
            "brand": info.get("brand", ""),
            "category": info.get("category", ""),
            "photo": info.get("photo", ""),
            "status": info.get("status", ""),
            "abc_class": info.get("abc_class", ""),
            "price": info.get("price", 0),
            "price_spp": info.get("price_spp", 0),
            "price_discount": info.get("price_discount", 0),
            "campaign_count": len(camps),
            "active_campaign_count": len(active_camps),
            "total_views": total_views,
            "total_clicks": total_clicks,
            "total_spent": round(total_spent, 2),
            "total_orders": total_orders,
            "ctr": round(total_clicks / total_views * 100, 2) if total_views else 0,
            "drr": round(total_spent / sum(c.get("sum_price", 0) for c in camps) * 100, 1) if sum(c.get("sum_price", 0) for c in camps) else 0,
            "campaigns": camps,
            "plan_orders": 0,  # TODO: из плана
            "fact_orders": total_orders,
            "plan_pct": 0,  # TODO: расчёт
        })

    # Фильтрация
    if search:
        s = search.lower()
        products = [p for p in products if s in str(p.get("vendor_code","")).lower() or s in str(p.get("nm_id","")) or s in str(p.get("brand","")).lower()]
    if status:
        products = [p for p in products if p.get("status") == status]
    if abc_class:
        products = [p for p in products if p.get("abc_class") == abc_class]
    if brand:
        products = [p for p in products if p.get("brand","").lower() == brand.lower()]

    # Список уникальных брендов для фильтра
    brands = sorted(set(p.get("brand","") for p in products if p.get("brand")))

    return {
        "products": products,
        "brands": brands,
        "total_products": len(products),
        "total_campaigns": len(all_campaigns),
    }


@router.get("/api/v1/nl/marketer/product/{nm_id}")
async def get_marketer_product_detail(
    nm_id: int,
    org_id: str,
    days: str = "30",
    db: AsyncSession = Depends(get_db)
):
    """Стол маркетолога — детальная карточка товара с РК по дням"""
    import decimal as _dec, json as _json

    try:
        days_int = int(days)
    except:
        days_int = 30

    def sf(v):
        if v is None: return 0
        return float(v) if isinstance(v, (_dec.Decimal, int)) else (float(v) if v else 0)

    date_from_sql = f"CURRENT_DATE - make_interval(days => {days_int})"

    # Инфо о товаре
    pe_rows = await db.execute(text("""
        SELECT DISTINCT ON (pe.nm_id) pe.nm_id, pe.brand, pe.subject_name, pe.photo_main,
            pe.vendor_code, pe.color, pe.weight, pe.chrt_id
        FROM product_entities pe
        WHERE pe.organization_id = :org AND pe.nm_id = :nm
        ORDER BY pe.nm_id, pe.created_at DESC
    """), {"org": org_id, "nm": nm_id})
    pe = pe_rows.first()
    if not pe:
        return {"error": "Товар не найден"}

    product = {
        "nm_id": nm_id,
        "brand": pe[1] or "",
        "category": pe[2] or "",
        "photo": pe[3] or "",
        "vendor_code": pe[4] or "",
    }

    # Цены
    ts_rows = await db.execute(text(f"""
        SELECT DISTINCT ON (ts.target_date) ts.target_date,
            ts.price, ts.price_spp, ts.price_discount, ts.impressions, ts.clicks, ts.ad_cost
        FROM tech_status ts
        WHERE ts.organization_id = :org AND ts.nm_id = :nm
          AND ts.target_date >= {date_from_sql}
        ORDER BY ts.target_date DESC, ts.created_at DESC
    """), {"org": org_id, "nm": nm_id})
    prices_by_date = {}
    organic_by_date = {}
    for r in ts_rows:
        prices_by_date[str(r[0])] = {
            "price": sf(r[1]), "price_spp": sf(r[2]), "price_discount": sf(r[3]),
            "organic_impressions": int(r[4] or 0), "organic_clicks": int(r[5] or 0), "organic_ad_cost": sf(r[6]),
        }

    # Статус/класс из справочника
    rb_rows = await db.execute(text("""
        SELECT rb.product_status as status,
               rb.product_class as abc_class
        FROM reference_book rb
        WHERE rb.organization_id = :org AND rb.nm_id = :nm
        ORDER BY rb.valid_from DESC LIMIT 1
    """), {"org": org_id, "nm": nm_id})
    rb = rb_rows.first()
    if rb:
        product["status"] = rb[0] or ""
        product["abc_class"] = rb[1] or ""
    # Акции - TODO: promo_products table not yet created
    product["in_promo"] = False
    product["promo_name"] = ""


    # РК, которые рекламируют этот nm_id
    camp_rows = await db.execute(text(f"""
        SELECT c.wb_campaign_id, c.name, c.type, c.status, c.nm_ids, c.budget,
               c.daily_budget, c.payment_type, c.bid_type
        FROM ad_campaigns c
        WHERE c.organization_id = :org AND c.nm_ids @> CAST(:nm_arr AS jsonb)
        ORDER BY c.status ASC, c.name
    """), {"org": org_id, "nm_arr": _json.dumps([nm_id])})

    active_statuses = ('7', '9')
    campaigns = []

    for r in camp_rows:
        camp_id = r[0]
        # Статистика по дням для этой РК — только по конкретному nm_id
        stat_rows = await db.execute(text(f"""
            SELECT s.stat_date,
                   SUM(s.views) as views, SUM(s.clicks) as clicks, SUM(s.spent) as spent,
                   CASE WHEN SUM(s.views) > 0 THEN ROUND(SUM(s.clicks)::numeric / SUM(s.views) * 100, 2) ELSE 0 END as ctr,
                   CASE WHEN SUM(s.clicks) > 0 THEN ROUND(SUM(s.spent) / SUM(s.clicks), 2) ELSE 0 END as cpc,
                   SUM(s.orders) as orders, SUM(s.atbs) as atbs,
                   CASE WHEN SUM(s.clicks) > 0 THEN ROUND(SUM(s.orders)::numeric / SUM(s.clicks) * 100, 2) ELSE 0 END as cr,
                   SUM(s.sum_price) as sum_price
            FROM ad_stats_nm s
            WHERE s.organization_id = :org AND s.wb_campaign_id = :cid AND s.nm_id = :nm
              AND s.stat_date >= {date_from_sql}
            GROUP BY s.stat_date
            ORDER BY s.stat_date
        """), {"org": org_id, "cid": camp_id, "nm": nm_id})

        daily_stats = []
        for sr in stat_rows:
            daily_stats.append({
                "date": str(sr[0]),
                "views": int(sr[1] or 0),
                "clicks": int(sr[2] or 0),
                "spent": round(sf(sr[3]), 2),
                "ctr": round(sf(sr[4]), 2),
                "cpc": round(sf(sr[5]), 2),
                "orders": int(sr[6] or 0),
                "atbs": int(sr[7] or 0),
                "cr": round(sf(sr[8]), 2),
                "sum_price": round(sf(sr[9]), 2),
            })

        # Итого по РК
        camp_total = {
            "views": sum(d["views"] for d in daily_stats),
            "clicks": sum(d["clicks"] for d in daily_stats),
            "spent": sum(d["spent"] for d in daily_stats),
            "orders": sum(d["orders"] for d in daily_stats),
            "atbs": sum(d["atbs"] for d in daily_stats),
            "sum_price": sum(d["sum_price"] for d in daily_stats),
        }
        camp_total["ctr"] = round(camp_total["clicks"] / camp_total["views"] * 100, 2) if camp_total["views"] else 0
        camp_total["cpc"] = round(camp_total["spent"] / camp_total["clicks"], 2) if camp_total["clicks"] else 0
        camp_total["cr"] = round(camp_total["orders"] / camp_total["clicks"] * 100, 2) if camp_total["clicks"] else 0

        campaigns.append({
            "campaign_id": camp_id,
            "name": r[1] or "Без названия",
            "type": str(r[2]) if r[2] else "",
            "status": str(r[3]) if r[3] else "",
            "nm_ids": [int(n) for n in (_json.loads(r[4]) if isinstance(r[4], str) else (r[4] or [])) if n],
            "budget": sf(r[5]),
            "daily_budget": sf(r[6]),
            "is_active": str(r[3]) in active_statuses,
            "daily": daily_stats,
            "totals": camp_total,
        })

    # Сводка «РК В ОБЩЕМ»
    all_daily = {}  # date -> aggregated
    for camp in campaigns:
        for d in camp["daily"]:
            dt = d["date"]
            if dt not in all_daily:
                all_daily[dt] = {"date": dt, "views": 0, "clicks": 0, "spent": 0, "orders": 0, "atbs": 0, "sum_price": 0}
            all_daily[dt]["views"] += d["views"]
            all_daily[dt]["clicks"] += d["clicks"]
            all_daily[dt]["spent"] += d["spent"]
            all_daily[dt]["orders"] += d["orders"]
            all_daily[dt]["atbs"] += d["atbs"]
            all_daily[dt]["sum_price"] += d["sum_price"]

    # Добавляем органику (из tech_status)
    for dt, info in organic_by_date.items():
        if dt in all_daily:
            all_daily[dt]["organic_impressions"] = info["organic_impressions"]
            all_daily[dt]["organic_clicks"] = info["organic_clicks"]

    summary_daily = sorted(all_daily.values(), key=lambda x: x["date"])

    grand_total = {
        "views": sum(d["views"] for d in summary_daily),
        "clicks": sum(d["clicks"] for d in summary_daily),
        "spent": sum(d["spent"] for d in summary_daily),
        "orders": sum(d["orders"] for d in summary_daily),
    }
    grand_total["ctr"] = round(grand_total["clicks"] / grand_total["views"] * 100, 2) if grand_total["views"] else 0
    grand_total["drr"] = round(grand_total["spent"] / grand_total["sum_price"] * 100, 1) if grand_total.get("sum_price") else 0

    # Лучший период — ищем день с макс profit (orders * price - spent)
    best_day = None
    best_profit = -float('inf')
    for d in summary_daily:
        price_day = prices_by_date.get(d["date"], {}).get("price", 0)
        profit = d["orders"] * price_day - d["spent"]
        if profit > best_profit and d["orders"] > 0:
            best_profit = profit
            best_day = {
                "date": d["date"],
                "orders": d["orders"],
                "spent": d["spent"],
                "views": d["views"],
                "price": price_day,
                "price_spp": prices_by_date.get(d["date"], {}).get("price_spp", 0),
                "profit": round(profit, 2),
                "in_promo": product.get("in_promo", False),
            }

    return {
        "product": product,
        "campaigns": campaigns,
        "summary_daily": summary_daily,
        "grand_total": grand_total,
        "best_period": best_day,
        "prices_by_date": prices_by_date,
    }



# ==================== UNIT ECONOMICS APIs ====================

async def build_unit_economics(
    org_id: str,
    db: AsyncSession,
    search: Optional[str] = None,
    limit: Optional[int] = None,
):
    """Юнит Экономика — сборка всех данных по SKU"""
    import asyncio
    from models.reference_book import ReferenceBook
    from sqlalchemy import text as sql_text
    from core.database import async_session as _make_session

    # Хелпер для параллельного выполнения SQL через отдельные сессии
    async def _run_query(query_text, params):
        async with _make_session() as s:
            result = await s.execute(sql_text(query_text), params)
            return result.all()

    # ── Redis-кэш: отдаём готовый результат если есть ──
    import redis as _redis_lib, json as _json
    _redis = _redis_lib.from_url("redis://redis:6379/0")
    _cache_key = f"ue_cache:{org_id}"
    if not search:
        _cached = _redis.get(_cache_key)
        if _cached:
            try:
                _cached_data = _json.loads(_cached)
                if limit and limit > 0:
                    _cached_data["items"] = _cached_data["items"][:limit]
                return _cached_data
            except Exception:
                pass

    # 1) Получаем список товаров из tech_status (последняя дата)
    dates_result = await db.execute(
        sql_text("SELECT DISTINCT target_date FROM tech_status WHERE organization_id = :org ORDER BY target_date DESC LIMIT 1"),
        {"org": org_id}
    )
    latest_date_row = dates_result.first()
    if not latest_date_row:
        return {"items": [], "total": 0}
    latest_date = latest_date_row[0]

    # 2) Продукты из product_entities (одна строка на nm_id + размер, как в Справочнике)
    prods_result = await db.execute(
        sql_text("""
            SELECT pe.id as entity_id, pe.nm_id, pe.vendor_code, COALESCE(ts.product_name, pe.product_name) as product_name, COALESCE(ts.photo_main, pe.photo_main) as photo_main,
                   (SELECT string_agg(eb.barcode, ', ') FROM entity_barcodes eb WHERE eb.entity_id = pe.id AND eb.is_active = true) as barcode,
                   ts.price, ts.price_discount, ts.tariff, ts.ad_cost,
                   pe.size_name, pe.subject_name,
                   pe.width, pe.height, pe.length
            FROM product_entities pe
            LEFT JOIN LATERAL (
                SELECT product_name, photo_main, price, price_discount, tariff, ad_cost
                FROM tech_status ts
                WHERE ts.organization_id = :org AND ts.nm_id = pe.nm_id AND ts.target_date = :dt
                LIMIT 1
            ) ts ON true
            WHERE pe.organization_id = :org
            ORDER BY pe.nm_id, pe.size_name
        """),
        {"org": org_id, "dt": latest_date}
    )
    products = prods_result.all()

    # ═══════════════════════════════════════════════════════════
    # [OPTIMIZED] Параллельное выполнение 5 SQL-запросов (asyncio.gather)
    # Запросы 3, 4, 5, 5b, 5c не зависят друг от друга → выполняются одновременно
    # ═══════════════════════════════════════════════════════════
    # ═══════════════════════════════════════════════════════════
    # [OPTIMIZED v2] 3 запроса вместо 5:
    #   - 1 широкий SELECT из reference_book (вместо 3 отдельных)
    #   - wb_tariff_snapshot
    #   - wb_box_tariffs
    # ═══════════════════════════════════════════════════════════
    _rb_sql = """SELECT entity_id, nm_id,
            mp_correction_pct, buyout_niche_pct, extra_costs, ad_plan_rub,
            price_before_spp_plan, price_before_spp_change, change_date,
            fulfillment_model, wb_club_discount_pct, storage_pct, product_status,
            mp_base_pct, wb_price_fact, wb_price_retail, wb_discount_pct, wb_prices_updated_at,
            cost_price, purchase_cost, logistics_cost, packaging_cost, other_costs,
            vat, product_class, brand, tax_system, tax_rate, vat_rate,
            fbs_warehouse
        FROM reference_book
        WHERE organization_id = :org AND (valid_to IS NULL OR valid_to >= CURRENT_DATE)
        ORDER BY entity_id NULLS LAST, valid_from DESC"""
    _q5_sql = """SELECT nm_id, logistics_tariff, storage_tariff, ad_cost_fact, buyout_pct_fact, commission_pct, price_retail, price_with_spp, spp_pct, commission_fbs_pct FROM wb_tariff_snapshot WHERE organization_id = :org ORDER BY target_date DESC"""
    _q5b_sql = """SELECT warehouse_name, box_delivery_base, box_delivery_liter, box_delivery_marketplace_base, box_delivery_marketplace_liter, box_delivery_coef, box_delivery_marketplace_coef FROM wb_box_tariffs WHERE organization_id = :org AND snapshot_date = (SELECT MAX(snapshot_date) FROM wb_box_tariffs WHERE organization_id = :org)"""
    _p = {"org": org_id}

    rb_rows, tsnap_rows, box_rows = await asyncio.gather(
        _run_query(_rb_sql, _p),
        _run_query(_q5_sql, _p),
        _run_query(_q5b_sql, _p),
    )

    # ── Обработка единого запроса reference_book ──────────
    # Колонки rb_rows:
    #  0: entity_id, 1: nm_id
    #  2-9: UE fields (mp_correction...fulfillment_model)
    # 10-12: wb_club_discount_pct, storage_pct, product_status
    # 13-17: mp_base_pct, wb_price_fact, wb_price_retail, wb_discount_pct, wb_prices_updated_at
    # 18-24: cost fields (cost_price...vat)
    # 25-28: product_class, brand, tax_system, tax_rate, vat_rate
    # 29: fbs_warehouse

    ue_by_entity = {}
    ue_by_nm_bc = {}
    cost_by_entity = {}
    cost_by_nm = {}
    ff_by_entity = {}
    ff_by_nm = {}

    for r in rb_rows:
        eid = str(r[0]) if r[0] else None
        nm = r[1]

        # ─ UE fields ─
        ue_fields = {
            "mp_correction_pct": r[2], "buyout_niche_pct": r[3],
            "extra_costs": r[4], "ad_plan_rub": r[5],
            "price_before_spp_plan": r[6], "price_before_spp_change": r[7],
            "change_date": r[8], "tariff_type": r[9],
            "wb_club_discount_pct": r[10],
            "product_status": r[12],
            "mp_base_pct": r[13],
            "wb_price_fact": r[14],
            "wb_price_retail": r[15],
            "wb_discount_pct": r[16],
            "wb_prices_updated_at": r[17],
            "fulfillment_model": r[9],
            "fbs_warehouse": r[29],
        }
        if eid:
            if eid not in ue_by_entity:
                ue_by_entity[eid] = ue_fields
        else:
            key = (nm, "")
            if key not in ue_by_nm_bc:
                ue_by_nm_bc[key] = ue_fields

        # ─ Cost fields ─
        cost_fields = {
            "cost_price": r[18], "purchase_cost": r[19], "logistics_cost": r[20],
            "packaging_cost": r[21], "other_costs": r[22], "vat": r[23],
            "product_class": r[24], "brand": r[25], "tax_system": r[26],
            "tax_rate": r[27], "vat_rate": r[28],
            "product_status": r[12],
        }
        if eid:
            if eid not in cost_by_entity:
                cost_by_entity[eid] = cost_fields
            else:
                for k, v in cost_fields.items():
                    if v is not None and v != 0:
                        cost_by_entity[eid][k] = v
        if nm:
            if nm not in cost_by_nm:
                cost_by_nm[nm] = cost_fields
            else:
                for k, v in cost_fields.items():
                    if v is not None and v != 0:
                        cost_by_nm[nm][k] = v

        # ─ FF fields ─
        fful = r[9]  # fulfillment_model
        fwh = r[29]  # fbs_warehouse
        if eid:
            if eid not in ff_by_entity:
                ff_by_entity[eid] = (fful, fwh)
        if nm:
            if nm not in ff_by_nm:
                ff_by_nm[nm] = (fful, fwh)

    # ── WB-данные из wb_tariff_snapshot ──────────
    snap_by_nm = {}
    for r in tsnap_rows:
        if r[0] not in snap_by_nm:
            snap_by_nm[r[0]] = {
                "logistics_tariff": float(r[1]) if r[1] else 0,
                "storage_tariff": float(r[2]) if r[2] else 0,
                "ad_cost_fact": float(r[3]) if r[3] else 0,
                "buyout_pct_fact": float(r[4]) if r[4] else 0,
                "commission_pct": float(r[5]) if r[5] else 0,
                "price_retail": float(r[6]) if r[6] else 0,
                "price_with_spp": float(r[7]) if r[7] else 0,
                "spp_pct": float(r[8]) if r[8] else 0,
                "commission_fbs_pct": float(r[9]) if r[9] else 0,
            }

    # ── Тарифы коробной логистики ──────────
    import math
    FBO_WH_NAMES = ["Коледино", "Краснодар", "Казань"]
    box_tariffs = {}
    fbo_delivery_sum = 0
    fbo_liter_sum = 0
    fbo_coef_sum = 0
    fbo_count = 0
    for r in box_rows:
        wh_name = r[0]
        box_tariffs[wh_name] = {
            "fbo_base": float(r[1]) if r[1] else None,
            "fbo_liter": float(r[2]) if r[2] else None,
            "fbs_base": float(r[3]) if r[3] else None,
            "fbs_liter": float(r[4]) if r[4] else None,
            "fbo_coef": float(r[5]) if r[5] else None,
            "fbs_coef": float(r[6]) if len(r) > 6 and r[6] else None,
        }
        if wh_name in FBO_WH_NAMES:
            if r[1] is not None and r[2] is not None:
                fbo_delivery_sum += float(r[1])
                fbo_liter_sum += float(r[2])
                fbo_coef_sum += float(r[5]) if len(r) > 5 and r[5] else 0
                fbo_count += 1

    fbo_avg_base = round(fbo_delivery_sum / fbo_count, 2) if fbo_count else 0
    fbo_avg_liter = round(fbo_liter_sum / fbo_count, 2) if fbo_count else 0
    fbo_avg_coef = round(fbo_coef_sum / fbo_count, 2) if fbo_count else 0

    # WB-сетка тарификации для товаров <= 1 литра (базовые ставки без коэфф.)
    _WB_TIER_RATES = [
        (0.001, 0.200, 23.0),
        (0.201, 0.400, 26.0),
        (0.401, 0.600, 29.0),
        (0.601, 0.800, 30.0),
        (0.801, 1.000, 32.0),
    ]
    _WB_BASE_FIRST_LITER = 46.0   # базовая ставка за 1-й литр (>1л)
    _WB_BASE_NEXT_LITER = 14.0    # базовая ставка за каждый доп. литр

    def _wb_rate_per_liter(vol):
        """Получить ставку ₽/л по WB-сетке для объёма <= 1л"""
        for lo, hi, rate in _WB_TIER_RATES:
            if lo <= vol <= hi:
                return rate
        return _WB_BASE_FIRST_LITER  # fallback

    def _calc_delivery(volume_liters, fulfillment_model, fbs_warehouse):
        """
        Расчёт логистики до клиента по WB-методике.
        
        WB тарификация:
        - <= 1 литр: ставка по сетке (23-32 ₽/л) × коэфф. склада
        - > 1 литр: 46 ₽ за первый + 14 ₽ за каждый доп. литр × коэфф. склада
        - box_delivery_base/liter в API = уже с коэфф. склада
        
        Для ФБО используем усреднённый коэфф. по 3 складам.
        """
        if not volume_liters or volume_liters <= 0:
            return 0, {}

        debug = {}
        vol_ceil = math.ceil(volume_liters)  # округление вверх до целых литров

        if fulfillment_model == "fbs" and fbs_warehouse:
            # ФБС — конкретный склад
            wh_tariffs = None
            wh_found = None
            for wh_name in box_tariffs:
                if fbs_warehouse in wh_name or wh_name in fbs_warehouse:
                    wh_tariffs = box_tariffs[wh_name]
                    wh_found = wh_name
                    break
            if wh_tariffs:
                base = wh_tariffs.get("fbs_base") or wh_tariffs.get("fbo_base") or 0
                liter = wh_tariffs.get("fbs_liter") or wh_tariffs.get("fbo_liter") or 0
                # Приоритет ФБС-коэфф., fallback на ФБО-коэфф., затем 100%
                coef = wh_tariffs.get("fbs_coef") or wh_tariffs.get("fbo_coef") or 100
                if base:
                    if volume_liters <= 1.0:
                        rate = _wb_rate_per_liter(volume_liters)
                        cost = round(rate * (coef / 100), 2)
                        debug = {"method": "ФБС (сетка <=1л)", "warehouse": wh_found, "vol": volume_liters, "tier_rate": rate, "coef": coef, "coef_source": "fbs" if wh_tariffs.get("fbs_coef") else "fbo", "formula": f"{rate} × {coef}%", "result": cost}
                    else:
                        cost = round(base + (vol_ceil - 1) * liter, 2)
                        debug = {"method": "ФБС (>1л)", "warehouse": wh_found, "vol_ceil": vol_ceil, "base": base, "liter": liter, "coef": coef, "coef_source": "fbs" if wh_tariffs.get("fbs_coef") else "fbo", "formula": f"{base} + {vol_ceil-1}×{liter}", "result": cost}
                    return cost, debug
            # Fallback на ФБС-тариф Коледино (План Б)
            _kd_tariffs = box_tariffs.get("\u041a\u043e\u043b\u0435\u0434\u0438\u043d\u043e")
            if _kd_tariffs and (_kd_tariffs.get("fbs_base") or _kd_tariffs.get("fbs_liter")):
                _kd_base = _kd_tariffs.get("fbs_base") or _kd_tariffs.get("fbo_base") or 0
                _kd_liter = _kd_tariffs.get("fbs_liter") or _kd_tariffs.get("fbo_liter") or 0
                _kd_coef = _kd_tariffs.get("fbs_coef") or _kd_tariffs.get("fbo_coef") or 100
                if volume_liters <= 1.0:
                    rate = _wb_rate_per_liter(volume_liters)
                    cost = round(rate * (_kd_coef / 100), 2)
                    debug = {"method": "ФБС Коледино (сетка <=1л, fallback)", "warehouse": "\u041a\u043e\u043b\u0435\u0434\u0438\u043d\u043e", "warehouse_requested": fbs_warehouse, "fallback_warehouse": "\u041a\u043e\u043b\u0435\u0434\u0438\u043d\u043e", "tier_rate": rate, "coef": _kd_coef, "formula": f"{rate} \u00d7 {_kd_coef}%", "result": cost}
                else:
                    cost = round(_kd_base + (vol_ceil - 1) * _kd_liter, 2)
                    debug = {"method": "ФБС Коледино (>1л, fallback)", "warehouse": "\u041a\u043e\u043b\u0435\u0434\u0438\u043d\u043e", "warehouse_requested": fbs_warehouse, "fallback_warehouse": "\u041a\u043e\u043b\u0435\u0434\u0438\u043d\u043e", "vol_ceil": vol_ceil, "base": _kd_base, "liter": _kd_liter, "coef": _kd_coef, "formula": f"{_kd_base} + {vol_ceil-1}\u00d7{_kd_liter}", "result": cost}
                return cost, debug
            return 0, {}

        # ФБО — усреднённое по 3 складам (Коледино/Краснодар/Казань)
        if fbo_avg_base:
            if volume_liters <= 1.0:
                rate = _wb_rate_per_liter(volume_liters)
                cost = round(rate * (fbo_avg_coef / 100), 2)
                debug = {"method": "ФБО-среднее (сетка <=1л)", "vol": volume_liters, "tier_rate": rate, "avg_coef": fbo_avg_coef,
                         "kd_coef": box_tariffs.get("Коледино", {}).get("fbo_coef", 0),
                         "kr_coef": box_tariffs.get("Краснодар", {}).get("fbo_coef", 0),
                         "kz_coef": box_tariffs.get("Казань", {}).get("fbo_coef", 0),
                         "formula": f"{rate} × {fbo_avg_coef}% (среднее 3 складов)", "result": cost}
            else:
                cost = round(fbo_avg_base + (vol_ceil - 1) * fbo_avg_liter, 2)
                debug = {"method": "ФБО-среднее (>1л)", "vol_ceil": vol_ceil, "avg_base": fbo_avg_base, "avg_liter": fbo_avg_liter,
                         "formula": f"{fbo_avg_base} + {vol_ceil-1}×{fbo_avg_liter}", "result": cost}
            return cost, debug
        return 0, {}


    def _calc_reverse_delivery(volume_liters):
        """
        Расчёт обратной логистики (возврат от покупателя на склад).
        Обратная логистика = базовый тариф за объём, БЕЗ коэффициента склада.
        Подтверждено API: прямая ~420₽, обратная ~205₽ (коэфф. склада ~2x не применяется).
        """
        if not volume_liters or volume_liters <= 0:
            return 0, {}
        
        debug = {}
        vol_ceil = math.ceil(volume_liters)
        
        if volume_liters <= 1.0:
            rate = _wb_rate_per_liter(volume_liters)
            cost = round(rate, 2)  # Без коэффициента склада!
            debug = {"method": "Обратная лог. (сетка <=1л)", "vol": volume_liters, "tier_rate": rate, "formula": f"{rate} ₽ (без коэфф. склада)", "result": cost}
        else:
            cost = round(_WB_BASE_FIRST_LITER + (vol_ceil - 1) * _WB_BASE_NEXT_LITER, 2)
            debug = {"method": "Обратная лог. (>1л)", "vol_ceil": vol_ceil, "base": _WB_BASE_FIRST_LITER, "liter": _WB_BASE_NEXT_LITER, "formula": f"{_WB_BASE_FIRST_LITER} + {vol_ceil-1}×{_WB_BASE_NEXT_LITER}", "result": cost}
        
        return cost, debug

    # 8) Собираем результат
    items = []
    search_q = search.lower() if search else ""
    
    # size_name и subject_name берутся напрямую из product_entities (индексы 10, 11)

    for p in products:
        entity_id = str(p[0]) if p[0] else None
        nm_id = p[1]
        vendor_code = p[2] or ""
        product_name = p[3] or ""
        photo = p[4] or ""
        main_barcode = p[5] or ""
        price = float(p[6]) if p[6] else 0
        price_discount = float(p[7]) if p[7] else 0
        # size_name и subject_name из product_entities (индексы 10, 11)
        _pe_size_name = p[10] or ""
        _pe_subject_name = p[11] or ""
        # Габариты из product_entities (индексы 12, 13, 14) — в см
        _pe_width = float(p[12]) if p[12] else 0
        _pe_height = float(p[13]) if p[13] else 0
        _pe_length = float(p[14]) if p[14] else 0
        # Объём в литрах (Д×Ш×В см / 1000)
        _volume_liters = round(_pe_width * _pe_height * _pe_length / 1000, 3) if (_pe_width and _pe_height and _pe_length) else 0

        # Фульфилмент модель и склад ФБС
        ff_info = ff_by_entity.get(entity_id, ff_by_nm.get(nm_id, (None, None)))
        _fulfillment_model = ff_info[0] or "fbo"
        _fbs_warehouse = ff_info[1] if ff_info[1] and ff_info[1] != "0" else None

        # Расчёт логистики до клиента
        _delivery_to_client, _delivery_debug = _calc_delivery(_volume_liters, _fulfillment_model, _fbs_warehouse)
        _reverse_logistics, _reverse_debug = _calc_reverse_delivery(_volume_liters)

        # Tooltip расшифровка логистики (детальная WB-методика)
        _logistics_tooltip_parts = []
        if _pe_width and _pe_height and _pe_length:
            _logistics_tooltip_parts.append(f"Габариты: {_pe_length}x{_pe_width}x{_pe_height} см")
            _logistics_tooltip_parts.append(f"Объём: {_volume_liters:.3f} л (окр. {math.ceil(_volume_liters)})")

        _meth = _delivery_debug.get("method", "")
        # Модель отгрузки
        if _fulfillment_model == "fbs":
            _model_label = f"Модель: ФБС (склад: {_fbs_warehouse or 'не указан'})"
        else:
            _model_label = "Модель: ФБО"
        _logistics_tooltip_parts.append(_model_label)
        _logistics_tooltip_parts.append(f"Методика: {_meth}")

        if "сетка" in _meth:
            # <= 1 литр: показываем сетку WB
            _logistics_tooltip_parts.append("")
            _logistics_tooltip_parts.append("WB-сетка (<= 1 л):")
            _logistics_tooltip_parts.append("  0.001-0.200 л → 23 ₽/л")
            _logistics_tooltip_parts.append("  0.201-0.400 л → 26 ₽/л")
            _logistics_tooltip_parts.append("  0.401-0.600 л → 29 ₽/л")
            _logistics_tooltip_parts.append("  0.601-0.800 л → 30 ₽/л")
            _logistics_tooltip_parts.append("  0.801-1.000 л → 32 ₽/л")
            _tr = _delivery_debug.get("tier_rate", 0)
            _logistics_tooltip_parts.append(f"  → Товар {_volume_liters:.3f} л = {_tr:.0f} ₽/л")
            _logistics_tooltip_parts.append("")
            if "ФБО-среднее" in _meth:
                _kd = box_tariffs.get("Коледино", {})
                _kr = box_tariffs.get("Краснодар", {})
                _kz = box_tariffs.get("Казань", {})
                _logistics_tooltip_parts.append(f"Коледино: коэфф. {_kd.get('fbo_coef', 0):.0f}%")
                _logistics_tooltip_parts.append(f"Краснодар: коэфф. {_kr.get('fbo_coef', 0):.0f}%")
                _logistics_tooltip_parts.append(f"Казань: коэфф. {_kz.get('fbo_coef', 0):.0f}%")
                _ac = _delivery_debug.get("avg_coef", 0)
                _logistics_tooltip_parts.append(f"Средний коэфф.: {_ac:.2f}%")
                _logistics_tooltip_parts.append(f"Формула: {_tr:.0f} × {_ac:.2f}% = {_delivery_to_client:.2f} ₽")
            else:
                _wh = _delivery_debug.get("warehouse", "?")
                _coef = _delivery_debug.get("coef", 0)
                _logistics_tooltip_parts.append(f"Склад: {_wh}, коэфф. {_coef:.0f}%")
                _logistics_tooltip_parts.append(f"Формула: {_tr:.0f} × {_coef:.0f}% = {_delivery_to_client:.2f} ₽")
        elif ">1л" in _meth:
            # > 1 литр: показываем формулу
            _logistics_tooltip_parts.append("")
            _vc = _delivery_debug.get("vol_ceil", 0)
            if "ФБС" in _meth:
                # ФБС: конкретный склад
                _b = _delivery_debug.get("base", 0)
                _l = _delivery_debug.get("liter", 0)
                _c = _delivery_debug.get("coef", 0)
                _wh = _delivery_debug.get("warehouse", "?")
                _fw = _delivery_debug.get("warehouse_requested", "")
                if _fw:
                    _logistics_tooltip_parts.append(f"Склад: {_fw} → тариф по {_wh}")
                else:
                    _logistics_tooltip_parts.append(f"Склад: {_wh}")
                _logistics_tooltip_parts.append(f"Базовый тариф: {_b:.2f} ₽ + {_l:.2f} ₽/л, коэфф. {_c:.0f}%")
                _logistics_tooltip_parts.append(f"Формула: {_b:.2f} + ({_vc}-1) × {_l:.2f} = {_delivery_to_client:.2f} ₽")
            elif "ФБО" in _meth:
                _kd = box_tariffs.get("Коледино", {})
                _kr = box_tariffs.get("Краснодар", {})
                _kz = box_tariffs.get("Казань", {})
                _ab = round((_kd.get('fbo_base') or 0) + (_kr.get('fbo_base') or 0) + (_kz.get('fbo_base') or 0), 2)
                _al = round((_kd.get('fbo_liter') or 0) + (_kr.get('fbo_liter') or 0) + (_kz.get('fbo_liter') or 0), 2)
                _logistics_tooltip_parts.append(f"Коледино: {_kd.get('fbo_base', 0):.2f} + {_kd.get('fbo_liter', 0):.2f}/л (коэфф. {_kd.get('fbo_coef', 0):.0f}%)")
                _logistics_tooltip_parts.append(f"Краснодар: {_kr.get('fbo_base', 0):.2f} + {_kr.get('fbo_liter', 0):.2f}/л (коэфф. {_kr.get('fbo_coef', 0):.0f}%)")
                _logistics_tooltip_parts.append(f"Казань: {_kz.get('fbo_base', 0):.2f} + {_kz.get('fbo_liter', 0):.2f}/л (коэфф. {_kz.get('fbo_coef', 0):.0f}%)")
                _logistics_tooltip_parts.append(f"Среднее: {_ab/3:.2f} + {_al/3:.2f}/л")

        _logistics_tooltip_parts.append("")
        _logistics_tooltip_parts.append(f"Итого логистика: {_delivery_to_client:.2f} ₽")
        _logistics_tooltip = chr(10).join(_logistics_tooltip_parts)

        # Фильтр поиска
        if search_q and search_q not in str(nm_id) and search_q not in product_name.lower() and search_q not in vendor_code.lower():
            continue

        cost = cost_by_nm.get(nm_id, cost_by_entity.get(entity_id, {}))
        ue = ue_by_entity.get(entity_id, ue_by_nm_bc.get((nm_id, main_barcode), ue_by_nm_bc.get((nm_id, ""), {})))

        item = {
            "entity_id": entity_id,
            "nm_id": nm_id,
            "vendor_code": vendor_code,
            "product_name": product_name,
            "photo": photo.replace("/hq/", "/c246x328/").replace("/big/", "/c246x328/").replace("/tm/", "/c246x328/") if photo else "",
            "barcode": main_barcode,
            "size_name": _pe_size_name,
            "subject_name": _pe_subject_name or cost.get("subject_name", ""),
            "sku": f"{vendor_code}_{main_barcode}" if vendor_code else str(nm_id),

            # Из справочника / себестоимости
            # Себестоимость в Юните = Итого из справочника (cost_price + extra_costs)
            "cost_price": (float(cost.get("cost_price") or 0)) + (float(ue.get("extra_costs") or 0)),
            "purchase_cost": float(cost.get("purchase_cost") or 0),
            "logistics_cost": float(cost.get("logistics_cost") or 0),
            "packaging_cost": float(cost.get("packaging_cost") or 0),
            "other_costs": float(cost.get("other_costs") or 0),
            "product_class": cost.get("product_class"),
            "brand": cost.get("brand"),
            "tax_system": cost.get("tax_system"),
            "tax_rate": float(cost.get("tax_rate") or 0),
            "vat_rate": float(cost.get("vat_rate") or 0),

            # Из wb_tariff_snapshot (автоподтяжка)
            "mp_base_pct": (lambda _s=snap_by_nm.get(nm_id,{}), _u=ue, _p=p: (float(_u.get("mp_base_pct") or 0) if _u.get("mp_base_pct") else ((_s.get("commission_fbs_pct") if (_u.get("tariff_type") or "fbo") == "fbs" else _s.get("commission_pct")) or float(_p[8] or 0))))(),
            "buyout_fact_pct": snap_by_nm.get(nm_id, {}).get("buyout_pct_fact", 0),
            "logistics_tariff": _delivery_to_client,
            "reverse_logistics": _reverse_logistics,
            "logistics_actual": 0,  # Будет из финотчётов
            "storage_tariff": snap_by_nm.get(nm_id, {}).get("storage_tariff", 0),
            "storage_actual": 0,  # Будет из финотчётов
            "acceptance_avg": 0,  # Будет из API приёмки
            "price_before_spp": float(ue.get("wb_price_fact")) if ue.get("wb_price_fact") else (snap_by_nm.get(nm_id, {}).get("price_retail") or price),
            "spp_pct": 0,  # заглушка, пока нет источника
            "price_with_spp": 0,  # заглушка, пока нет источника
            "ad_fact_pct": 0,  # заглушка, позже из финотчёта
            "ad_fact_rub": 0,  # заглушка, позже из финотчёта
            "wb_club_discount_pct_api": 0,

            # Из справочника
            "product_status": ue.get("product_status") or cost.get("product_status", ""),

            # Ручные вводы
            "mp_correction_pct": float(ue.get("mp_correction_pct") or 0),
            "buyout_niche_pct": float(ue.get("buyout_niche_pct") or 0),
            "extra_costs": 0,  # Уже включена в cost_price (Итого из справочника)
            "ad_plan_pct": min(99, max(0, float(ue.get("ad_plan_rub")) if ue.get("ad_plan_rub") not in (None, "", 0) else 5)),
            "ad_plan_rub": 0,  # рассчитывается ниже по цене
            "price_before_spp_plan": float(ue.get("price_before_spp_plan") or 0),
            "price_before_spp_change": float(ue.get("price_before_spp_change") or 0),
            "change_date": str(ue.get("change_date")) if ue.get("change_date") else None,
            "tariff_type": ue.get("tariff_type") or "box",
            "wb_club_discount_pct": float(ue.get("wb_club_discount_pct") or 0),

            # Логистика до клиента
            "volume_liters": _volume_liters,
            "volume_rounded": math.ceil(_volume_liters) if _volume_liters else 0,
            "fulfillment_model": _fulfillment_model,
            "fbs_warehouse": _fbs_warehouse or "",
            "delivery_to_client": _delivery_to_client,
            "logistics_tooltip": _logistics_tooltip,
        }

        # Расчётные формулы
        mp_total_pct = item["mp_base_pct"] + item["mp_correction_pct"]
        item["mp_total_pct"] = mp_total_pct

        # Реклама: расчёт ₽ из %
        item["ad_plan_rub"] = round(item["price_before_spp"] * item["ad_plan_pct"] / 100, 2)
        item["ad_fact_rub"] = round(item["price_with_spp"] * item["ad_fact_pct"] / 100, 2)  # Заглушка

        # Комиссия МП
        mp_commission = round(item["price_with_spp"] * mp_total_pct / 100, 2)
        item["mp_commission"] = mp_commission

        # Эквайринг 1.5%
        acquiring = round(item["price_with_spp"] * 0.015, 2)

        # Налог
        tax = 0
        ts = item["tax_system"]
        if ts == "usn":
            tax = round(item["price_with_spp"] * item["tax_rate"] / 100, 2)
        elif ts == "usn_dr":
            income = item["price_with_spp"] - mp_commission - item["cost_price"]  # extra_costs уже в cost_price
            tax = round(max(income, 0) * item["tax_rate"] / 100, 2)
        elif ts == "osn":
            nds = round(item["price_with_spp"] * item["vat_rate"] / 100, 2)
            input_nds = round(item["purchase_cost"] / 120 * item["vat_rate"] if item["purchase_cost"] else 0, 2)
            tax = round(nds - input_nds, 2)

        item["tax_total"] = tax

        # === БЛОК 7: Расчёт по ФАКТУ ===
        # Обратная логистика на 1 шт (с учётом % невозврата = 100% - %выкупа)
        # Обратная логистика = базовый тариф за объём, без коэфф. склада, без % выкупа
        _reverse_log_amount = item["reverse_logistics"] or 0

        # Логистика с % выкупа (затраты на 1 отправленный товар)
        _buyout_pct = item["buyout_fact_pct"] if item["buyout_fact_pct"] and item["buyout_fact_pct"] > 0 else item["buyout_niche_pct"]
        _buyout_ratio = float(_buyout_pct) / 100 if _buyout_pct else 1
        item["logistics_with_buyout"] = round(item["delivery_to_client"] + _reverse_log_amount * (1 - _buyout_ratio), 2)

        expenses_fact = (
            item["cost_price"] + item["logistics_cost"] + item["packaging_cost"] +
            item["other_costs"] +  # extra_costs уже включена в cost_price
            mp_commission + item["logistics_actual"] + item["storage_actual"] +
            item["acceptance_avg"] + acquiring + tax +
            item["ad_fact_rub"] +
            _reverse_log_amount
        )
        profit_fact = round(item["price_with_spp"] - expenses_fact, 2)
        margin_fact = round(profit_fact / item["price_with_spp"] * 100, 2) if item["price_with_spp"] else 0
        roi_fact = round(profit_fact / item["cost_price"] * 100, 2) if item["cost_price"] else 0
        to_account_fact = round(item["price_with_spp"] - mp_commission - tax, 2)

        item["expenses_fact"] = round(expenses_fact, 2)
        item["profit_fact"] = profit_fact
        item["margin_fact"] = margin_fact
        item["roi_fact"] = roi_fact
        item["to_account_fact"] = to_account_fact

        # === БЛОК 8: Расчёт по ПЛАНУ ===
        plan_price = float(item["price_before_spp_plan"] or item["price_before_spp"])
        plan_price_spp = round(plan_price * (1 - item["spp_pct"] / 100), 2) if item["spp_pct"] else plan_price
        plan_mp = round(plan_price_spp * mp_total_pct / 100, 2)
        plan_acquiring = round(plan_price_spp * 0.015, 2)
        
        # Пересчёт налога для плановой цены
        plan_tax = 0
        if ts == "usn":
            plan_tax = round(plan_price_spp * item["tax_rate"] / 100, 2)
        elif ts == "usn_dr":
            plan_income = plan_price_spp - plan_mp - item["cost_price"]  # extra_costs уже в cost_price
            plan_tax = round(max(plan_income, 0) * item["tax_rate"] / 100, 2)
        elif ts == "osn":
            plan_nds = round(plan_price_spp * item["vat_rate"] / 100, 2)
            plan_input_nds = round(item["purchase_cost"] / 120 * item["vat_rate"] if item["purchase_cost"] else 0, 2)
            plan_tax = round(plan_nds - plan_input_nds, 2)

        expenses_plan = (
            item["cost_price"] + item["logistics_cost"] + item["packaging_cost"] +
            item["other_costs"] +  # extra_costs уже включена в cost_price
            plan_mp + item["logistics_actual"] + item["storage_actual"] +
            item["acceptance_avg"] + plan_acquiring + plan_tax +
            round(plan_price_spp * item["ad_plan_pct"] / 100, 2) +
            _reverse_log_amount
        )
        profit_plan = round(plan_price_spp - expenses_plan, 2)
        margin_plan = round(profit_plan / plan_price_spp * 100, 2) if plan_price_spp else 0
        roi_plan = round(profit_plan / item["cost_price"] * 100, 2) if item["cost_price"] else 0
        to_account_plan = round(plan_price_spp - plan_mp - plan_tax, 2)

        item["plan_price_spp"] = plan_price_spp
        item["expenses_plan"] = round(expenses_plan, 2)
        item["profit_plan"] = profit_plan
        item["margin_plan"] = margin_plan
        item["roi_plan"] = roi_plan
        item["to_account_plan"] = to_account_plan

        # === БЛОК 9: После изменений ===
        change_price = float(item["price_before_spp_change"] or item["price_before_spp"])
        change_price_spp = round(change_price * (1 - item["spp_pct"] / 100), 2) if item["spp_pct"] else change_price
        change_mp = round(change_price_spp * mp_total_pct / 100, 2)
        
        change_tax = 0
        if ts == "usn":
            change_tax = round(change_price_spp * item["tax_rate"] / 100, 2)
        elif ts == "usn_dr":
            change_income = change_price_spp - change_mp - item["cost_price"]  # extra_costs уже в cost_price
            change_tax = round(max(change_income, 0) * item["tax_rate"] / 100, 2)
        elif ts == "osn":
            change_nds = round(change_price_spp * item["vat_rate"] / 100, 2)
            change_input_nds = round(item["purchase_cost"] / 120 * item["vat_rate"] if item["purchase_cost"] else 0, 2)
            change_tax = round(change_nds - change_input_nds, 2)

        expenses_change = (
            item["cost_price"] + item["logistics_cost"] + item["packaging_cost"] +
            item["other_costs"] +  # extra_costs уже включена в cost_price
            change_mp + item["logistics_actual"] + item["storage_actual"] +
            item["acceptance_avg"] + round(change_price_spp * 0.015, 2) + change_tax +
            item["ad_fact_rub"] +
            _reverse_log_amount
        )
        profit_change = round(change_price_spp - expenses_change, 2)
        roi_change = round(profit_change / item["cost_price"] * 100, 2) if item["cost_price"] else 0

        item["profit_change"] = profit_change
        item["margin_change"] = round(profit_change / change_price_spp * 100, 2) if change_price_spp else 0
        item["roi_change"] = roi_change

        items.append(item)

    # Сохраняем в Redis-кэш на 30 минут (ПОЛНЫЙ набор)
    _result_full = {"items": items, "total": len(items)}
    if not search:
        try:
            _redis.setex(_cache_key, 1800, _json.dumps(_result_full, ensure_ascii=False, default=str))
        except Exception:
            pass

    # Пагинация: если limit указан — обрезаем items
    if limit and limit > 0:
        return {"items": items[:limit], "total": len(items)}
    return _result_full


@router.get("/api/v1/nl/unit-economics")
async def get_unit_economics(
    org_id: str,
    search: Optional[str] = None,
    limit: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Юнит Экономика — только для участников организации."""
    org_id = await resolve_org_id(org_id, db)
    await require_organization_role(uuid.UUID(org_id), Role.VIEWER, current_user, db)
    return await build_unit_economics(org_id, db, search=search, limit=limit)


class UnitEconSave(BaseModel):
    nm_id: int
    barcode: Optional[str] = None
    entity_id: Optional[str] = None
    mp_correction_pct: Optional[float] = None
    buyout_niche_pct: Optional[float] = None
    extra_costs: Optional[float] = None
    ad_plan_rub: Optional[float] = None
    price_before_spp_plan: Optional[float] = None
    price_before_spp_change: Optional[float] = None
    change_date: Optional[str] = None
    tariff_type: Optional[str] = None
    wb_club_discount_pct: Optional[float] = None


@router.post("/api/v1/nl/unit-economics")
async def save_unit_economics(
    data: UnitEconSave,
    org_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Сохранить ручные вводы Юнит Экономики"""
    org_id = await resolve_org_id(org_id, db)
    await require_organization_role(uuid.UUID(org_id), Role.ADMIN, current_user, db)
    from models.reference_book import ReferenceBook
    from datetime import datetime as dt_mod

    change_date = date.today()

    # Определяем entity_id
    entity_id_ue = data.entity_id if hasattr(data, "entity_id") and data.entity_id else None
    if not entity_id_ue:
        from sqlalchemy import text as sql_text_sync
        ent_q = await db.execute(sql_text_sync(
            "SELECT pe.id FROM product_entities pe "
            "WHERE pe.organization_id = :org AND pe.nm_id = :nm "
            "ORDER BY CASE WHEN pe.size_name = :sz THEN 0 ELSE 1 END LIMIT 1"
        ), {"org": org_id, "nm": data.nm_id, "sz": data.barcode or ""})
        ent_row = ent_q.first()
        entity_id_ue = ent_row[0] if ent_row else None
    ins = pg_insert(ReferenceBook).values(
        organization_id=org_id,
        nm_id=data.nm_id,
        barcode=data.barcode,
        entity_id=entity_id_ue,
        valid_from=date.today(),
        mp_correction_pct=data.mp_correction_pct,
        buyout_niche_pct=data.buyout_niche_pct,
        extra_costs=data.extra_costs,
        ad_plan_rub=data.ad_plan_rub,
        price_before_spp_plan=data.price_before_spp_plan,
        price_before_spp_change=data.price_before_spp_change,
        change_date=date.today(),
        fulfillment_model=data.tariff_type or "fbo",
        wb_club_discount_pct=data.wb_club_discount_pct,
    )
    stmt = ins.on_conflict_do_update(
        constraint="reference_book_org_entity_vf_key",
        set_={
            "mp_correction_pct": ins.excluded.mp_correction_pct,
            "buyout_niche_pct": ins.excluded.buyout_niche_pct,
            "extra_costs": ins.excluded.extra_costs,
            "ad_plan_rub": ins.excluded.ad_plan_rub,
            "price_before_spp_plan": ins.excluded.price_before_spp_plan,
            "price_before_spp_change": ins.excluded.price_before_spp_change,
            "change_date": date.today(),
            "fulfillment_model": ins.excluded.fulfillment_model,
            "wb_club_discount_pct": ins.excluded.wb_club_discount_pct,
        }
    )
    await db.execute(stmt)
    await db.commit()
    try:
        import redis as _redis_lib
        _redis_lib.from_url("redis://redis:6379/0").delete(f"ue_cache:{org_id}")
    except Exception:
        pass
    return {"ok": True}


# ─── ОБНОВЛЕНИЕ ЦЕН ИЗ WB API ─────────────────────────────

# Кулдаун: минимальный интервал между обновлениями цен (секунды)
PRICES_REFRESH_COOLDOWN = 15 * 60  # 15 минут


@router.post("/api/v1/nl/prices/refresh")
async def refresh_prices_from_wb(org_id: str, db: AsyncSession = Depends(get_db)):
    """
    Обновить цены из WB Prices API и сохранить в reference_book.
    
    Тянет discountedPrice (цена со скидкой, реально на витрине),
    price (цена до скидки), discount (скидка %).
    
    Кулдаун 15 мин — защита от бана WB API.
    """
    from services.wb_api.keys import get_all_wb_keys as _get_keys
    from sqlalchemy.ext.asyncio import create_async_engine, async_sessionmaker
    from core.config import settings
    from datetime import datetime as _dt, timezone as _tz
    
    org_id = await resolve_org_id(org_id, db)
    
    # Проверяем кулдаун — когда последний раз обновляли цены
    cooldown_sql = "SELECT MAX(wb_prices_updated_at) FROM reference_book WHERE organization_id = :org AND wb_prices_updated_at IS NOT NULL"
    cooldown_result = await db.execute(text(cooldown_sql), {"org": org_id})
    last_update_row = cooldown_result.first()
    last_update = last_update_row[0] if last_update_row else None
    
    if last_update:
        now_utc = _dt.now(_tz.utc)
        if last_update.tzinfo is None:
            last_update = last_update.replace(tzinfo=_tz.utc)
        elapsed = (now_utc - last_update).total_seconds()
        remaining = PRICES_REFRESH_COOLDOWN - elapsed
        if remaining > 0:
            mins = int(remaining // 60)
            secs = int(remaining % 60)
            raise HTTPException(429, f"Кулдаун. Доступно через {mins}:{secs:02d}")
    
    # Получаем API ключи
    engine = create_async_engine(settings.DATABASE_URL, echo=False)
    sf = async_sessionmaker(engine, expire_on_commit=False)
    try:
        all_keys = await _get_keys(sf)
    finally:
        await engine.dispose()
    
    # Находим ключ для этой организации
    api_key = None
    for oid, key in all_keys:
        if oid == org_id:
            api_key = key
            break
    
    if not api_key:
        raise HTTPException(400, "Нет WB API ключа для этой организации")
    
    # Запрашиваем цены из WB API
    from services.wb_api.client import WBApiClient
    
    try:
        async with WBApiClient(api_key) as client:
            prices_data = await client.get_all_prices()
    except Exception as e:
        raise HTTPException(502, f"Ошибка WB API: {str(e)}")
    
    items = prices_data if isinstance(prices_data, list) else prices_data.get("items", [])
    if not items:
        raise HTTPException(404, "WB API вернул пустой список товаров")
    
    # Строим маппинг nm_id -> цены
    price_map = {}
    for item in items:
        nm_id = item.get("nmID") or item.get("nmId") or item.get("nm_id")
        if not nm_id:
            continue
        nm_id = int(nm_id)
        discount = item.get("discount", 0)
        sizes = item.get("sizes", [])
        if sizes:
            sz = sizes[0]
            price_retail = float(sz.get("price", 0))
            price_fact = float(sz.get("discountedPrice", 0))
            if price_retail > 0:
                price_map[nm_id] = {
                    "price_retail": price_retail,
                    "price_fact": price_fact,
                    "discount": discount,
                }
    
    # Обновляем reference_book
    now = _dt.now(_tz.utc)
    updated_count = 0
    
    for nm_id, prices in price_map.items():
        update_sql = (
            "UPDATE reference_book "
            "SET wb_price_fact = :pf, "
            "    wb_price_retail = :pr, "
            "    wb_discount_pct = :disc, "
            "    wb_prices_updated_at = :now "
            "WHERE organization_id = :org "
            "  AND nm_id = :nm "
            "  AND (valid_to IS NULL OR valid_to >= CURRENT_DATE)"
        )
        result = await db.execute(text(update_sql), {
            "pf": prices["price_fact"],
            "pr": prices["price_retail"],
            "disc": prices["discount"],
            "now": now,
            "org": org_id,
            "nm": nm_id,
        })
        updated_count += result.rowcount
    
    await db.commit()
    
    return {
        "ok": True,
        "updated": updated_count,
        "total_items": len(items),
        "total_with_prices": len(price_map),
        "updated_at": now.isoformat(),
        "cooldown_seconds": PRICES_REFRESH_COOLDOWN,
    }


@router.get("/api/v1/nl/prices/last-refresh")
async def get_last_prices_refresh(org_id: str, db: AsyncSession = Depends(get_db)):
    """Когда последний раз обновляли цены из WB API"""
    org_id = await resolve_org_id(org_id, db)
    last_sql = "SELECT MAX(wb_prices_updated_at) FROM reference_book WHERE organization_id = :org AND wb_prices_updated_at IS NOT NULL"
    result = await db.execute(text(last_sql), {"org": org_id})
    row = result.first()
    last_update = row[0] if row else None
    
    remaining = 0
    if last_update:
        from datetime import datetime as _dt2, timezone as _tz2
        now_utc = _dt2.now(_tz2.utc)
        if last_update.tzinfo is None:
            last_update = last_update.replace(tzinfo=_tz2.utc)
        elapsed = (now_utc - last_update).total_seconds()
        if elapsed < PRICES_REFRESH_COOLDOWN:
            remaining = int(PRICES_REFRESH_COOLDOWN - elapsed)
    
    return {
        "last_update": last_update.isoformat() if last_update else None,
        "cooldown_remaining_seconds": remaining,
        "can_refresh": remaining == 0,
    }


@router.get("/api/v1/nl/profile")
async def nl_profile(token: str = Query(""), db: AsyncSession = Depends(get_db)):
    """Профиль текущего пользователя + список магазинов (быстро, без подсчётов)"""
    from models.organization import Membership, Organization, WbApiKey
    payload = decode_token(token)
    if not payload:
        raise HTTPException(401, "Не авторизован")
    user_id = payload.get("sub")
    
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(401, "Пользователь не найден")
    
    result = await db.execute(
        select(Membership, Organization)
        .join(Organization, Membership.organization_id == Organization.id)
        .where(Membership.user_id == user_id)
    )
    rows = result.all()
    
    shops = []
    for m, o in rows:
        keys_result = await db.execute(
            select(WbApiKey).where(WbApiKey.organization_id == o.id)
        )
        keys = keys_result.scalars().all()
        
        keys_info = [{"id": str(k.id), "name": k.name, "created_at": str(k.created_at)[:10] if k.created_at else ""} for k in keys]
        
        shops.append({
            "id": str(o.id),
            "name": o.name,
            "wb_seller_id": o.wb_seller_id,
            "role": m.role.value if hasattr(m.role, "value") else str(m.role),
            "keys": keys_info,
            "keys_count": len(keys),
        })
    
    return {"email": user.email, "is_superuser": user.is_superuser or False, "shops": shops, "shops_count": len(shops)}


@router.post("/api/v1/nl/verify-wb-key")
async def nl_verify_wb_key(data: dict, token: str = Query(""), db: AsyncSession = Depends(get_db)):
    """Проверить работает ли WB API ключ магазина — реальный запрос к WB"""
    import httpx
    from core.security import decrypt_data
    
    payload = decode_token(token)
    if not payload:
        raise HTTPException(401, "Не авторизован")
    
    org_id = data.get("org_id", "").strip()
    if not org_id:
        raise HTTPException(400, "org_id обязателен")
    
    # Get first active key for this org
    from models.organization import WbApiKey
    result = await db.execute(
        select(WbApiKey).where(WbApiKey.organization_id == org_id)
    )
    keys = result.scalars().all()
    
    if not keys:
        return {"status": "error", "message": "Нет API ключей", "products_count": 0}
    
    # Decrypt and try WB API
    for key in keys:
        try:
            api_token = decrypt_data(key.personal_token)
            async with httpx.AsyncClient(timeout=10) as client:
                r = await client.post(
                    "https://content-api.wildberries.ru/content/v2/get/cards/list",
                    headers={"Authorization": api_token},
                    json={"settings": {"cursor": {"limit": 1}, "filter": {"withPhoto": -1}}}
                )
                if r.status_code == 200:
                    body = r.json()
                    cards = body.get("cards", [])
                    total = 0
                    try:
                        from models.product_entity import ProductEntity
                        from sqlalchemy import func as sqlfunc
                        result2 = await db.execute(
                            select(sqlfunc.count(ProductEntity.id)).where(ProductEntity.organization_id == org_id)
                        )
                        row = result2.first()
                        if row:
                            total = row[0]
                    except Exception:
                        pass
                    
                    return {
                        "status": "ok",
                        "message": "Ключ работает",
                        "key_name": key.name,
                        "products_count": total,
                        "wb_response": "200 OK"
                    }
                elif r.status_code == 401:
                    return {"status": "error", "message": "Ключ не авторизован (401)", "key_name": key.name, "products_count": 0}
                elif r.status_code == 429:
                    return {"status": "warn", "message": "Слишком много запросов (429), попробуйте позже", "key_name": key.name, "products_count": 0}
                else:
                    return {"status": "error", "message": f"Ошибка WB: {r.status_code}", "key_name": key.name, "products_count": 0}
        except Exception as e:
            return {"status": "error", "message": f"Ошибка: {str(e)[:100]}", "products_count": 0}
    
    return {"status": "error", "message": "Не удалось проверить", "products_count": 0}



@router.post("/api/v1/nl/rename-org")
async def nl_rename_org(data: dict, token: str = Query(""), db: AsyncSession = Depends(get_db)):
    """Переименовать магазин (организацию)"""
    from models.organization import Membership, Role
    payload = decode_token(token)
    if not payload:
        raise HTTPException(401, "Не авторизован")
    user_id = payload.get("sub")
    
    org_id = data.get("org_id", "").strip()
    new_name = data.get("name", "").strip()
    
    if not org_id or not new_name:
        raise HTTPException(400, "org_id и name обязательны")
    if len(new_name) > 100:
        raise HTTPException(400, "Название слишком длинное (макс 100 символов)")
    
    # Check user is OWNER or ADMIN in this org
    result = await db.execute(
        select(Membership).where(
            Membership.user_id == user_id,
            Membership.organization_id == org_id
        )
    )
    membership = result.scalar_one_or_none()
    if not membership or membership.role not in (Role.OWNER, Role.ADMIN):
        raise HTTPException(403, "Только OWNER или ADMIN может переименовывать")
    
    from models.organization import Organization
    result = await db.execute(select(Organization).where(Organization.id == org_id))
    org = result.scalar_one_or_none()
    if not org:
        raise HTTPException(404, "Организация не найдена")
    
    org.name = new_name
    await db.commit()
    
    return {"status": "ok", "name": new_name}

@router.post("/api/v1/nl/invite")
async def nl_invite(data: dict, token: str = Query(""), db: AsyncSession = Depends(get_db)):
    """Пригласить коллегу в организацию"""
    from models.organization import Membership, Role
    import secrets
    
    payload = decode_token(token)
    if not payload:
        raise HTTPException(401, "Не авторизован")
    user_id = payload.get("sub")
    
    org_id = data.get("org_id", "").strip()
    email = data.get("email", "").strip().lower()
    role_str = data.get("role", "VIEWER").upper()
    
    if not org_id or not email:
        raise HTTPException(400, "org_id и email обязательны")
    
    # Check inviter is ADMIN+ in this org
    result = await db.execute(
        select(Membership).where(
            Membership.user_id == user_id,
            Membership.organization_id == org_id
        )
    )
    membership = result.scalar_one_or_none()
    if not membership or membership.role not in (Role.OWNER, Role.ADMIN):
        raise HTTPException(403, "Только OWNER или ADMIN может приглашать")
    
    # Check role is valid
    if role_str not in ("ADMIN", "VIEWER"):
        role_str = "VIEWER"
    
    from models.organization import Invitation, InvitationStatus
    invite_token = secrets.token_urlsafe(32)
    from datetime import datetime, timedelta
    invitation = Invitation(
        email=email,
        organization_id=org_id,
        role=Role[role_str],
        token=invite_token,
        status=InvitationStatus.PENDING,
        expires_at=datetime.utcnow() + timedelta(days=7)
    )
    db.add(invitation)
    await db.commit()
    
    return {
        "status": "ok",
        "email": email,
        "role": role_str,
        "invite_token": invite_token,
        "expires_at": str(invitation.expires_at)[:19]
    }


@router.get("/nl/v2", response_class=HTMLResponse)
async def nl_page():
    """НЛ — главная страница"""
    html = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>НЛ — Аналитика v3</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f5f7fa;color:#1a1a2e}

.header{background:#fff;border-bottom:1px solid #e0e0e0;padding:12px 20px;display:flex;align-items:center;gap:12px}
.header h1{font-size:1.2em;color:#6c5ce7}
.header .logo{font-size:1.4em}
.header .logout{margin-left:auto;color:#999;cursor:pointer;font-size:0.85em}
.header .logout:hover{color:#e74c3c}
.org-switcher{display:flex;align-items:center;gap:4px}
.org-switcher select{border:1px solid #e0e0e0;border-radius:4px;padding:4px 8px;font-size:.85em;background:#fff;cursor:pointer}
.btn-sm{background:#6c5ce7;color:#fff;border:none;width:24px;height:24px;border-radius:50%;cursor:pointer;font-size:1em;display:flex;align-items:center;justify-content:center}
.btn-sm:hover{background:#5a4bd1}
.wb-key-item{display:flex;align-items:center;gap:10px;padding:8px 12px;background:#fff;border-radius:6px;margin-bottom:6px;box-shadow:0 1px 2px rgba(0,0,0,.06)}
.wb-key-item .name{font-weight:500}
.wb-key-item .date{color:#999;font-size:.8em}
.wb-key-item .del{margin-left:auto;color:#e74c3c;cursor:pointer;font-size:.8em}
.user-info{color:#666;font-size:0.85em}

.tabs{display:flex;background:#fff;border-bottom:2px solid #e0e0e0;padding:0 20px}
.tab{padding:12px 24px;cursor:pointer;color:#666;font-weight:500;border-bottom:3px solid transparent;transition:all .2s}
.tab:hover{color:#6c5ce7}
.tab.active{color:#6c5ce7;border-bottom-color:#6c5ce7}

.content{padding:20px}
.tab-content{display:none}
.tab-content.active{display:block}

table{width:100%;border-collapse:collapse;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.08)}
th{background:#f8f9fa;padding:10px 8px;text-align:left;font-size:.78em;color:#666;text-transform:uppercase;letter-spacing:.5px;border-bottom:2px solid #e0e0e0}
td{padding:8px;border-bottom:1px solid #f0f0f0;font-size:.85em}
tr:hover{background:#f8f9ff}
.r{text-align:right}
input[type="number"],input[type="text"],input[type="email"],input[type="password"]{width:100%;border:1px solid #e0e0e0;border-radius:4px;padding:6px 8px;font-size:.85em}
input:focus{outline:none;border-color:#6c5ce7;box-shadow:0 0 0 2px rgba(108,92,231,.15)}
.btn{background:#6c5ce7;color:#fff;border:none;padding:8px 16px;border-radius:6px;cursor:pointer;font-size:.9em;font-weight:500}
.btn:hover{background:#5a4bd1}
.btn-outline{background:#fff;color:#6c5ce7;border:1px solid #6c5ce7}
.btn-outline:hover{background:#f0edfc}
.save-btn{background:#6c5ce7;color:#fff;border:none;padding:4px 10px;border-radius:4px;cursor:pointer;font-size:.8em}
.save-btn:hover{background:#5a4bd1}
.save-btn.saved{background:#00b894}
.empty{text-align:center;padding:40px;color:#999}
.photo{width:36px;height:36px;border-radius:4px;object-fit:cover}

/* Auth */
.auth-container{max-width:400px;margin:80px auto;background:#fff;padding:32px;border-radius:12px;box-shadow:0 2px 12px rgba(0,0,0,.08)}
.auth-container h2{color:#6c5ce7;margin-bottom:8px;font-size:1.3em}
.auth-container p{color:#999;font-size:.85em;margin-bottom:20px}
.auth-container .field{margin-bottom:14px}
.auth-container label{display:block;font-size:.8em;color:#666;margin-bottom:4px}
.auth-container .toggle{color:#6c5ce7;cursor:pointer;font-size:.85em;text-align:center;margin-top:16px}
.auth-container .toggle:hover{text-decoration:underline}
//.auth-error{color:#e74c3c;font-size:.85em;margin-bottom:10px}
.metric-card{background:#fff;border-radius:8px;padding:12px 14px;box-shadow:0 1px 3px rgba(0,0,0,.06)}
.metric-card .mc-label{font-size:.75em;color:#999;text-transform:uppercase;letter-spacing:.3px;margin-bottom:4px}
.metric-card .mc-value{font-size:1.15em;font-weight:600;color:#1a1a2e}
.metric-card .mc-delta{font-size:.75em;margin-top:2px}
.metric-card .mc-delta.pos{color:#00b894}
.metric-card .mc-delta.neg{color:#e74c3c}
.metric-card .mc-sub{font-size:.75em;color:#999;margin-top:2px}
.metric-card.expense{border-left:3px solid #e17055}
.metric-card.expense .mc-value{color:#d63031}
.alert-card{background:#fff3cd;border:1px solid #ffc107;border-radius:8px;padding:10px 14px;margin-bottom:8px;font-size:.85em;display:flex;align-items:center;gap:8px}
.alert-card.red{background:#ffeaea;border-color:#e74c3c}
.alert-card.yellow{background:#fff8e1;border-color:#f39c12}

/* Sidebar layout */
.app-layout{display:flex;min-height:100vh}
.sidebar{width:220px;background:#1a1a2e;color:#fff;padding:0;flex-shrink:0;position:fixed;top:0;left:0;bottom:0;overflow-y:auto;z-index:50}
.sidebar .logo{padding:16px 20px;font-size:1.1em;font-weight:700;color:#fff;border-bottom:1px solid rgba(255,255,255,.1)}
.sidebar .nav-group{padding:8px 0}
.sidebar .nav-label{padding:6px 20px;font-size:.7em;text-transform:uppercase;color:rgba(255,255,255,.4);letter-spacing:1px}
.sidebar .nav-item{display:flex;align-items:center;gap:10px;padding:8px 20px;color:rgba(255,255,255,.7);cursor:pointer;font-size:.85em;transition:all .15s;text-decoration:none;border-left:3px solid transparent}
.sidebar .nav-item:hover{background:rgba(255,255,255,.08);color:#fff}
.sidebar .nav-item.active{background:rgba(108,92,231,.2);color:#fff;border-left-color:#6c5ce7}
.sidebar .nav-item .icon{width:20px;text-align:center}
.sidebar .user-block{position:absolute;bottom:0;left:0;right:0;padding:12px 20px;border-top:1px solid rgba(255,255,255,.1);font-size:.8em;color:rgba(255,255,255,.5)}
.sidebar .user-block .logout-btn{color:#e74c3c;cursor:pointer;display:block;margin-top:6px}
.sidebar .user-block .logout-btn:hover{color:#ff6b6b}
.sidebar select option{background:#2d2d44;color:#fff}
.main-area{margin-left:220px;flex:1;min-height:100vh;background:#f5f7fa}
.top-bar{background:#fff;border-bottom:1px solid #e0e0e0;padding:10px 24px;display:flex;align-items:center;gap:12px;position:sticky;top:0;z-index:10}
.top-bar .page-title{font-size:1.1em;font-weight:600;color:#1a1a2e}
.top-bar .filters{display:flex;align-items:center;gap:10px;margin-left:auto}
.top-bar select,.top-bar input{border:1px solid #e0e0e0;border-radius:4px;padding:4px 8px;font-size:.85em}
.page-content{padding:20px 24px}
.page-section{display:none}
.page-section.active{display:block}

th.sortable { cursor: pointer; user-select: none; white-space: nowrap; }
th.sortable:hover { background: rgba(255,255,255,0.15); }
th.sortable::after { content: ' ↕'; font-size: 0.7em; opacity: 0.5; }
th.sortable.asc::after { content: ' ↑'; opacity: 1; }
th.sortable.desc::after { content: ' ↓'; opacity: 1; }

.rnp-wrap{font-size:.78em;background:#fff;border-radius:10px;box-shadow:0 1px 4px rgba(0,0,0,.08);overflow:hidden;border:1px solid #eee;margin-bottom:14px}
.rnp-table-wrap{overflow-x:auto}
.rnp-table{border-collapse:collapse;width:max-content;min-width:100%}
.rnp-table th,.rnp-table td{border:1px solid #eee;padding:5px 8px;white-space:nowrap;vertical-align:top;font-size:.8em}
.rnp-table thead th{background:#f8f9fa;font-weight:600;position:sticky;top:0;z-index:2}
.rnp-table thead th.sticky-col{position:sticky;z-index:3}
.rnp-table tbody td.sticky-col{position:sticky;background:#fff;z-index:1}
.rnp-table .row-label{color:#888;font-weight:500;min-width:120px;max-width:160px;overflow:hidden;text-overflow:ellipsis}
.rnp-table .val-cell{text-align:right;min-width:95px}
.rnp-table .day-header{text-align:center;font-weight:700;color:#6c5ce7;font-size:.85em;background:#f0eeff;min-width:85px}
.rnp-table .day-header.today{background:#6c5ce7;color:#fff}
.rnp-table .val-cell.today{background:#faf8ff}
.rnp-table .section-title{font-weight:700;font-size:.85em;color:#333;background:#f8f9fa;padding:6px 8px}
.rnp-table .card-divider td{background:#f0f0f0;height:4px;padding:0;border:none}
.rnp-card-photo{width:50px;height:50px;border-radius:4px;overflow:hidden}
.rnp-card-photo img{width:100%;height:100%;object-fit:cover}
.rnp-card-info{min-width:140px;max-width:180px}
.rnp-card-nm{font-weight:700;color:#6c5ce7;font-size:.9em}
.rnp-card-name{font-weight:600;font-size:.82em;max-width:180px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.rnp-card-detail{color:#999;font-size:.78em}
.rnp-card-krrr{font-weight:700;font-size:.95em}
.rnp-header-table{border-collapse:collapse;width:max-content;min-width:100%}
.rnp-header-table th,.rnp-header-table td{border:1px solid #eee;padding:5px 8px;white-space:nowrap;font-size:.8em}
.rnp-header-table .row-label{color:#888;font-weight:500;min-width:120px}
.rnp-header-table .val-cell{text-align:right;min-width:95px}
.rnp-header-table .day-header{text-align:center;font-weight:700;color:#6c5ce7;font-size:.85em;background:#f0eeff;min-width:85px}
.rnp-header-table .day-header.today{background:#6c5ce7;color:#fff}
.rnp-header-table .section-title{font-weight:700;font-size:.85em;color:#333;background:#f8f9fa;padding:6px 8px}
.rnp-ctrl{display:flex;align-items:center;gap:12px;margin-bottom:16px;flex-wrap:wrap}
.rnp-ctrl select,.rnp-ctrl input{border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em}
.rnp-summary-bar{display:flex;gap:16px;margin-bottom:12px;font-size:.85em;flex-wrap:wrap;background:#fff;padding:12px;border-radius:8px;box-shadow:0 1px 3px rgba(0,0,0,.08)}
.rnp-pos{color:#00b894}.rnp-neg{color:#e74c3c}


/* === Tabulator overrides === */
.tabulator{border:1px solid #e0e0e0;border-radius:8px;overflow:hidden;font-size:.82em;background:#fff}
.tabulator .tabulator-header{background:#f8f9fa;border-bottom:2px solid #e0e0e0;font-size:.7em;text-transform:none;letter-spacing:0;color:#333}
.tabulator .tabulator-header .tabulator-col{border-right:1px solid #e0e0e0}
.tabulator .tabulator-header .tabulator-col .tabulator-col-content{padding:6px 8px}
.tabulator .tabulator-header .tabulator-col.tabulator-sortable .tabulator-col-title{padding-right:20px}
.tabulator .tabulator-header .tabulator-col-group{background:#f0f1f5;font-weight:600}
.tabulator .tabulator-tableholder .tabulator-table .tabulator-row .tabulator-cell{padding:4px 6px;border-right:1px solid #f0f0f0;border-bottom:1px solid #f0f0f0}
.tabulator .tabulator-tableholder .tabulator-table .tabulator-row:hover{background:#f8f9ff}
.tabulator .tabulator-tableholder .tabulator-table .tabulator-row .tabulator-cell input,
.tabulator .tabulator-tableholder .tabulator-table .tabulator-row .tabulator-cell select{width:100%;border:1px solid #e0e0e0;border-radius:3px;padding:2px 4px;font-size:.85em}
.tabulator .tabulator-tableholder .tabulator-table .tabulator-row .tabulator-cell input:focus,
.tabulator .tabulator-tableholder .tabulator-table .tabulator-row .tabulator-cell select:focus{outline:none;border-color:#6c5ce7;box-shadow:0 0 0 2px rgba(108,92,231,.15)}
.season-cell{background:#fffde7 !important}
.topquery-cell{background:#ede7f6 !important}
.tax-cell{text-align:center;color:#6c5ce7;font-weight:600;background:#f8f7ff !important}
.tabulator .tabulator-footer{background:#f8f9fa;border-top:1px solid #e0e0e0;padding:6px 12px}
.tabulator .tabulator-tableholder .tabulator-table .tabulator-group{background:#f0f1f5;border-bottom:1px solid #e0e0e0;font-size:11px;color:#333;font-weight:400;padding:4px 8px}
.tabulator .tabulator-tableholder .tabulator-table .tabulator-group .tabulator-arrow{color:#999;margin-right:4px}
</style>
<!-- Tabulator CSS -->
<link href="/static/lib/tabulator.min.css" rel="stylesheet">
<link href="/static/lib/tabulator_modern.min.css" rel="stylesheet">
<!-- Tabulator JS -->
<script type="text/javascript" src="/static/lib/tabulator.min.js"></script>
<!-- NL Grid Module -->
<!-- Chart.js -->
<script src="/static/lib/chart.min.js"></script>
<script type="text/javascript" src="/static/js/nl-grid.js?v=20260605"></script>
<!-- Cost Grid Module -->
<script type="text/javascript" src="/static/js/cost-grid.js?v=20260603f"></script>
<script type="text/javascript" src="/static/js/ue-grid.js?v=20260611-auth"></script>
<script type="text/javascript" src="/static/js/promo-grid.js?v=20260525a"></script>
<script type="text/javascript" src="/static/js/ads-grid.js?v=20260608j"></script>
<script type="text/javascript" src="/static/js/ads-arts-grid.js?v=20260609d"></script>
</head>
<body>

<!-- Auth -->
<div id="auth-section">
<div class="auth-container">
<div id="auth-login">
<h2>🔑 Вход</h2>
<p>Войдите в свой аккаунт НЛ</p>
<div id="login-error" class="auth-error" style="display:none"></div>
<div class="field"><label>Email</label><input type="email" id="login-email"></div>
<div class="field"><label>Пароль</label><input type="password" id="login-password"></div>
<button class="btn" onclick="doLogin()" style="width:100%">Войти</button>
<div style="text-align:center;margin-top:16px;font-size:.85em"><a href="#" onclick="showRegister();return false" style="color:#6c5ce7;text-decoration:none">Нет аккаунта? Зарегистрироваться</a></div>
</div>

<div id="auth-register" style="display:none">
<h2>📝 Регистрация</h2>
<p>Создайте аккаунт для доступа к аналитике</p>
<div id="reg-error" class="auth-error" style="display:none"></div>
<div class="field"><label>Email</label><input type="email" id="reg-email"></div>
<div class="field"><label>Пароль</label><input type="password" id="reg-password"></div>
<div class="field"><label>Название организации</label><input type="text" id="reg-org" value="Моя организация"></div>
<button class="btn" onclick="doRegister()" style="width:100%">Зарегистрироваться</button>
<a href="#" class="toggle" onclick="showLogin();return false">Уже есть аккаунт? Войти</a>
</div>
</div>
</div>

<!-- Main app -->
<div id="app-section" style="display:none">
<div class="app-layout">
<aside class="sidebar">
<div class="logo">📊 НЛ Аналитика</div>
<div class="nav-group">
<div class="nav-label">Финансы</div>
<a class="nav-item active" onclick="navTo('stats',this)"><span class="icon">📊</span>Основные показатели</a>
<a class="nav-item" onclick="navTo('rnp',this)"><span class="icon">🎯</span>РНП</a>
<a class="nav-item" onclick="navTo('opiu',this)"><span class="icon">📑</span>ОПиУ</a>
<a class="nav-item" onclick="navTo('analytics',this)"><span class="icon">📈</span>Аналитика по товарам</a>
<a class="nav-item" onclick="navTo('costprice',this)"><span class="icon">📖</span>Справочник</a>
<a class="nav-item" onclick="navTo('salesplan',this)"><span class="icon">🎯</span>План продаж</a>
<a class="nav-item" onclick="navTo('warehouses',this)"><span class="icon">📦</span>Склады</a>
<a class="nav-item" onclick="navTo('opexpenses',this)"><span class="icon">📝</span>Опер. расходы</a>
<a class="nav-item" onclick="navTo('ads',this)"><span class="icon">📢</span>Реклама</a>
<a class="nav-item" onclick="navTo('marketer',this)"><span class="icon">📊</span>Стол маркетолога</a>
<a class="nav-item" onclick="navTo('extads',this)"><span class="icon">🎯</span>Внешняя реклама</a>
<a class="nav-item" onclick="navTo('fboneeds',this)"><span class="icon">🚚</span>Потребность FBO</a>
<a class="nav-item" onclick="navTo('unitecon',this)"><span class="icon">🧮</span>Юнит Экономика</a>
</div>
<a class="nav-item" onclick="navTo('promo',this)"><span class="icon">🏷</span>Акции</a>
<div class="nav-group">
<div class="nav-label">Остальное</div>
<a class="nav-item" onclick="navTo('connectors',this)"><span class="icon">🔌</span>Подключения</a>
<a class="nav-item" onclick="navTo('subscription',this)"><span class="icon">💳</span>Подписка</a>
<a class="nav-item" onclick="navTo('settings',this)"><span class="icon">⚙️</span>Настройки</a>
<a class="nav-item" onclick="navTo('help',this)"><span class="icon">❓</span>Помощь</a>
</div>
<div class="user-block">
<div id="user-email"></div>
<div style="margin-top:4px">
<select id="org-select" onchange="switchOrg()" style="width:100%;background:#2d2d44;color:#fff;border:1px solid rgba(255,255,255,.25);border-radius:4px;padding:4px;font-size:.85em"></select>
</div>
<span class="logout-btn" onclick="doLogout()">Выйти</span>
</div>
</aside>
<div class="main-area">
<div class="top-bar">
<span class="page-title" id="page-title">Основные показатели</span>
<div class="filters" id="top-filters">
<select id="filter-store" onchange="switchTopStore()" style="min-width:160px;border:1px solid #e0e0e0;border-radius:4px;padding:4px 8px;font-size:.85em"></select>
<select id="filter-period"><option value="yesterday">Вчера</option><option value="week">Неделя</option><option value="month" selected>Месяц</option></select>
<input type="text" id="filter-article" placeholder="Артикул" style="width:120px">
</div>
</div>
<div class="page-content">
<div id="page-stats" class="page-section active">
<!-- Фильтр по дате -->
<div style="display:flex;align-items:center;gap:12px;margin-bottom:20px;flex-wrap:wrap">
<select id="stats-date" onchange="_statsLimit=50;loadStats()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em;cursor:pointer"></select>
<button class="btn" onclick="_statsLimit=50;loadStats()" style="padding:6px 14px;font-size:.85em">🔄 Обновить</button>
</div>

<!-- Прибыль -->
<div style="margin-bottom:24px">
<div style="font-size:.85em;color:#999;margin-bottom:8px">ПРИБЫЛЬ</div>
<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px">
<div class="metric-card"><div class="mc-label">Выручка</div><div class="mc-value" id="v-revenue">—</div></div>
<div class="metric-card"><div class="mc-label">Выкупов</div><div class="mc-value" id="v-buyouts-count">—</div></div>
<div class="metric-card" id="mc-profit"><div class="mc-label">Прибыль</div><div class="mc-value" id="v-profit">—</div><div class="mc-delta" id="d-profit"></div></div>
<div class="metric-card" id="mc-realization"><div class="mc-label">Реализация</div><div class="mc-value" id="v-realization">—</div><div class="mc-delta" id="d-realization"></div></div>
<div class="metric-card"><div class="mc-label">Продажи</div><div class="mc-value" id="v-sales">—</div><div class="mc-delta" id="d-sales"></div></div>
<div class="metric-card"><div class="mc-label">К выплате</div><div class="mc-value" id="v-topay">—</div><div class="mc-delta" id="d-topay"></div></div>
<div class="metric-card"><div class="mc-label">Продано штук</div><div class="mc-value" id="v-sold">—</div><div class="mc-delta" id="d-sold"></div></div>
<div class="metric-card"><div class="mc-label">Отмен штук</div><div class="mc-value" id="v-cancelled">—</div><div class="mc-delta" id="d-cancelled"></div></div>
<div class="metric-card"><div class="mc-label">Возвратов штук</div><div class="mc-value" id="v-returned">—</div><div class="mc-delta" id="d-returned"></div></div>
<div class="metric-card"><div class="mc-label">% возвратов</div><div class="mc-value" id="v-retpercent">—</div><div class="mc-delta" id="d-retpercent"></div></div>
<div class="metric-card"><div class="mc-label">ROI</div><div class="mc-value" id="v-roi">—</div><div class="mc-delta" id="d-roi"></div></div>
<div class="metric-card"><div class="mc-label">Рентабельность</div><div class="mc-value" id="v-rent">—</div><div class="mc-delta" id="d-rent"></div></div>
<div class="metric-card"><div class="mc-label">Процент выкупов</div><div class="mc-value" id="v-buyout">—</div><div class="mc-delta" id="d-buyout"></div></div>
<div class="metric-card"><div class="mc-label">Прибыль/ед.</div><div class="mc-value" id="v-profitunit">—</div><div class="mc-delta" id="d-profitunit"></div></div>
</div>
</div>

<!-- Расходы -->
<div style="margin-bottom:24px">
<div style="font-size:.85em;color:#999;margin-bottom:8px">РАСХОДЫ</div>
<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px">
<div class="metric-card expense"><div class="mc-label">Налог</div><div class="mc-value" id="v-tax">—</div></div>
<div class="metric-card expense"><div class="mc-label">Себестоимость</div><div class="mc-value" id="v-costprice">—</div></div>
<div class="metric-card expense"><div class="mc-label">Комиссия ВБ</div><div class="mc-value" id="v-commission">—</div></div>
<div class="metric-card expense"><div class="mc-label">Реклама (ДРР)</div><div class="mc-value" id="v-ads">—</div></div>
<div class="metric-card expense"><div class="mc-label">Логистика</div><div class="mc-value" id="v-logistics">—</div></div>
<div class="metric-card expense"><div class="mc-label">Штрафы</div><div class="mc-value" id="v-fines">—</div></div>
<div class="metric-card expense"><div class="mc-label">Хранение</div><div class="mc-value" id="v-storage">—</div></div>
<div class="metric-card expense"><div class="mc-label">Платная приемка</div><div class="mc-value" id="v-reception">—</div></div>
<div class="metric-card expense"><div class="mc-label">Прочие удержания</div><div class="mc-value" id="v-other">—</div></div>
</div>
</div>

<!-- Остатки -->
<div style="margin-bottom:24px">
<div style="font-size:.85em;color:#999;margin-bottom:8px">ОСТАТКИ</div>
<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:12px">
<div class="metric-card"><div class="mc-label">Остатки всего</div><div class="mc-value" id="v-stock-total">—</div><div class="mc-sub" id="v-stock-sub"></div></div>
<div class="metric-card"><div class="mc-label">На складах WB</div><div class="mc-value" id="v-stock-wb">—</div><div class="mc-sub" id="v-stock-wb-sub"></div></div>
<div class="metric-card"><div class="mc-label">На моих складах</div><div class="mc-value" id="v-stock-my">—</div><div class="mc-sub" id="v-stock-my-sub"></div></div>
</div>
</div>

<!-- Алерты -->
<div id="stats-alerts" style="margin-bottom:20px"></div>

<!-- Таблица товаров -->
<div style="font-size:.85em;color:#999;margin-bottom:8px">ТОВАРЫ</div>
<table id="stats-table">
<thead><tr><th>Фото</th><th>Арт WB</th><th>Название</th><th>Размер</th><th>ШК</th><th>Остаток</th><th>Заказы</th><th>Выкупы</th><th>Возвраты</th><th>Рейтинг</th><th>Показы</th><th>Клики</th><th>CTR</th><th>Реклама ₽</th><th>Цена</th></tr></thead>
<tbody id="stats-products"><tr><td colspan="15" class="empty">Выберите дату</td></tr></tbody>
</table>
<div id="stats-pagination" style="margin-top:8px;display:flex;align-items:center;gap:10px;font-size:.85em;color:#666"><span id="stats-shown"></span><button id="stats-more-btn" class="btn" style="padding:4px 12px;font-size:.85em;display:none" onclick="loadStatsMore()">Показать ещё</button></div>
</div>

<div id="page-analytics" class="page-section"></div>



<div id="page-rnp" class="page-section"></div>

<div id="page-opiu" class="page-section"></div>
<div id="page-costprice" class="page-section"></div>
<div id="page-salesplan" class="page-section"></div>
<div id="page-warehouses" class="page-section"></div>
<div id="page-opexpenses" class="page-section"></div>

<div id="page-ads" class="page-section"></div><div id="page-marketer" class="page-section"></div>
<div id="page-extads" class="page-section"></div>
<div id="page-fboneeds" class="page-section"></div>
<div id="page-unitecon" class="page-section"></div>
<div id="page-promo" class="page-section"></div>
<div id="page-connectors" class="page-section"></div>
<div id="page-subscription" class="page-section"></div>
<div id="page-help" class="page-section"></div>
<div id="page-settings" class="page-section"></div>


<script>
function esc(s) { return (s||"").replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;").replace(/"/g,"&quot;"); }

function toggleGroup(tr) {
    let el = tr.nextElementSibling;
    let show = el && el.style.display === 'none';
    while (el && el.classList.contains('group-child')) {
        el.style.display = show ? '' : 'none';
        el = el.nextElementSibling;
    }
    const badge = tr.querySelector('span');
    if (badge) badge.textContent = show ? badge.textContent.replace('▸', '▾') : badge.textContent.replace('▾', '▸');
}

let TOKEN = localStorage.getItem('nl_token');
let ORG_ID = new URL(location).searchParams.get('org') || localStorage.getItem('nl_org_id');
let FBS_WAREHOUSES = [];
let _taxSettings = {tax_system: '', tax_rate: null, vat_type: 'нет'};
let _costDirty = false;  // Флаг несохранённых изменений в справочнике
// Защита от потери данных при закрытии/перезагрузке
window.addEventListener('beforeunload', function(e) {
    if (_costDirty) {
        e.preventDefault();
        e.returnValue = '';
    }
});

async function loadFbsWarehouses() {
    try {
        const r = await fetch('/api/v1/nl/fbs-warehouses?org_id=' + ORG_ID);
        if (r.ok) FBS_WAREHOUSES = await r.json();
    } catch(e) { console.warn('FBS warehouses load failed', e); }
}

function showRegister() {
    document.getElementById('auth-login').style.display = 'none';
    document.getElementById('auth-register').style.display = '';
}
function showLogin() {
    document.getElementById('auth-register').style.display = 'none';
    document.getElementById('auth-login').style.display = '';
}

async function showApp() {
    document.getElementById('auth-section').style.display = 'none';
    document.getElementById('app-section').style.display = '';
    try {
        if (ORG_ID) {
            await loadDates();
            loadStats();
            loadOpiu();
            loadAnalytics();
            initRnpMonths();
            loadRnp();
            
            loadWarehouses();
            loadOpEx();
        }
    } catch(e) { console.error('init error:', e); }
}

var _statsLimit=50;function loadStatsMore(){_statsLimit+=50;loadStats()}

async function loadStats() {
    if (!ORG_ID) return;
    const sel = document.getElementById('stats-date') || document.getElementById('ref-date');
    let dateVal = sel ? sel.value : '';
    if (!dateVal || dateVal === 'Нет данных') dateVal = new Date().toISOString().split('T')[0];
    try {
        const res = await fetch('/api/v1/nl/control?org_id=' + ORG_ID + '&target_date=' + dateVal);
        if (!res.ok) return;
        const data = await res.json();
        const s = data.summary || {};
        // Заполняем карточки
        const fmt = (v, suffix) => { if (v == null) return '—'; return Number(v).toLocaleString('ru-RU', {maximumFractionDigits:2}) + (suffix || ''); };
        // Карточки прибыли
        const revenue = s.total_revenue || 0;
        const adCost = s.total_ad_cost || 0;
        const profit = revenue - adCost;
        document.getElementById('v-profit').textContent = fmt(profit, ' ₽');
        document.getElementById('v-profit').style.color = profit >= 0 ? '#00b894' : '#e74c3c';
        document.getElementById('v-sold').textContent = fmt(s.total_orders);
        document.getElementById('v-returned').textContent = fmt(s.total_returns);
        document.getElementById('v-stock-total').textContent = fmt(s.total_stock, ' шт');
        document.getElementById('v-stock-wb').textContent = fmt(s.total_stock_fbo, ' шт');
        document.getElementById('v-stock-my').textContent = fmt(s.total_stock_fbs, ' шт');
        document.getElementById('v-ads').textContent = fmt(adCost, ' ₽');
        document.getElementById('v-buyout').textContent = s.total_orders ? (s.total_buyouts / s.total_orders * 100).toFixed(1) + '%' : '—';
        // Доп. карточки
        const el1 = document.getElementById('v-revenue');
        if (el1) el1.textContent = fmt(revenue, ' ₽');
        const el2 = document.getElementById('v-buyouts-count');
        if (el2) el2.textContent = s.total_buyouts || 0;
        const el3 = document.getElementById('v-profitunit');
        if (el3) el3.textContent = s.total_buyouts ? fmt(profit / s.total_buyouts, ' ₽') : '—';
        // Алерты
        let alerts = '';
        if (s.zero_stock_count > 0) alerts += '<div class="alert-card red">🔴 Нет в наличии: ' + s.zero_stock_count + ' товаров</div>';
        if (s.low_stock_count > 0) alerts += '<div class="alert-card yellow">🟡 Низкий остаток (≤5): ' + s.low_stock_count + ' товаров</div>';
        document.getElementById('stats-alerts').innerHTML = alerts;
        // Таблица товаров — группировка по nm_id
        const tbody = document.getElementById('stats-products');
        const prods = data.products || [];
        if (!prods.length) { tbody.innerHTML = '<tr><td colspan="15" class="empty">Нет данных</td></tr>'; return; }
        
        // Группируем по nm_id
        const groups = {};
        const order = [];
        prods.forEach(p => {
            const key = p.nm_id;
            if (!groups[key]) { groups[key] = []; order.push(key); }
            groups[key].push(p);
        });
        
        let html = '';
        order.slice(0, _statsLimit).forEach(nmId => {
            const items = groups[nmId];
            const hasSizes = items.length > 1 || (items.length === 1 && items[0].size_name && items[0].size_name !== '0' && items[0].size_name !== 'ONE SIZE');
            
            // Агрегация для родительской строки
            let totalStock = 0, totalStockFbo = 0, totalOrders = 0, totalBuyouts = 0, totalReturns = 0;
            let totalImpressions = 0, totalClicks = 0, totalAd = 0;
            items.forEach(p => {
                totalStock += (p.stock_qty || 0) + (p.stock_fbo_qty || 0);
                totalStockFbo += p.stock_fbo_qty || 0;
                totalOrders += p.orders_count || 0;
                totalBuyouts += p.buyouts_count || 0;
                totalReturns += p.returns_count || 0;
                totalImpressions += p.impressions || 0;
                totalClicks += p.clicks || 0;
                totalAd += p.ad_cost || 0;
            });
            const avgRating = items.reduce((s,p) => s + (p.rating||0), 0) / items.length;
            const avgPrice = items[0].wb_price_fact || items[0].price || 0;
            const ctr = totalImpressions > 0 ? (totalClicks / totalImpressions * 100).toFixed(1) + '%' : '—';
            const thumb = (items[0].photo_main || '').replace('/hq/', '/c246x328/').replace('/big/', '/c246x328/').replace('/tm/', '/c246x328/');
            const totalStockFbs = totalStock - totalStockFbo;
            const stockColor = totalStock <= 0 ? '#e74c3c' : totalStock <= 5 ? '#e17055' : '';
            
            if (hasSizes) {
                // Родительская строка (кликабельная)
                html += '<tr class="group-parent" onclick="toggleGroup(this)" style="cursor:pointer;background:#f8f9ff">' +
                '<td>' + (thumb ? '<img src="' + thumb + '" style="width:36px;height:36px;border-radius:4px;object-fit:cover" loading="lazy">' : '') + '</td>' +
                '<td><b>' + nmId + '</b> <span style="font-size:.7em;color:#6c5ce7">▸ ' + items.length + ' разм.</span></td>' +
                '<td style="max-width:150px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="' + esc(items[0].product_name) + '">' + esc(items[0].product_name) + '</td>' +
                '<td></td><td></td>' +
                '<td style="font-weight:600;line-height:1.1"><div style="color:#0984e3;font-size:.9em">' + totalStockFbo + '</div><div style="color:#6c5ce7;font-size:.75em">' + totalStockFbs + '</div></td>' +
                '<td>' + totalOrders + '</td><td>' + totalBuyouts + '</td><td>' + totalReturns + '</td>' +
                '<td>' + (avgRating ? avgRating.toFixed(1) : '—') + '</td>' +
                '<td>' + totalImpressions + '</td><td>' + totalClicks + '</td><td>' + ctr + '</td>' +
                '<td>' + fmt(totalAd) + '</td><td>' + fmt(avgPrice) + '</td></tr>';
                
                // Строки размеров (скрыты по умолчанию)
                items.forEach(p => {
                    const sCtr = p.impressions > 0 ? (p.clicks / p.impressions * 100).toFixed(1) + '%' : '—';
                    const sizeLabel = p.size_name && p.size_name !== '0' && p.size_name !== 'ONE SIZE' ? p.size_name : '—';
                    html += '<tr class="group-child" style="display:none;font-size:.85em">' +
                    '<td></td><td></td><td></td>' +
                    '<td style="color:#6c5ce7;font-weight:500">' + sizeLabel + '</td>' +
                    '<td style="font-size:.7em;color:#999">' + (p.barcode || '') + '</td>' +
                    '<td style="font-weight:600;line-height:1.1"><div style="color:#0984e3;font-size:.9em">' + (p.stock_fbo_qty ?? '—') + '</div><div style="color:#6c5ce7;font-size:.75em">' + (p.stock_qty ?? '—') + '</div></td>' +
                    '<td>' + (p.orders_count ?? '—') + '</td><td>' + (p.buyouts_count ?? '—') + '</td><td>' + (p.returns_count ?? '—') + '</td>' +
                    '<td>' + (p.rating ?? '—') + '</td><td>' + (p.impressions ?? '—') + '</td><td>' + (p.clicks ?? '—') + '</td><td>' + sCtr + '</td>' +
                    '<td>' + fmt(p.ad_cost) + '</td><td>' + fmt(p.wb_price_fact || p.price) + '</td></tr>';
                });
            } else {
                // Один размер — обычная строка
                const p = items[0];
                const sizeLabel = p.size_name && p.size_name !== '0' && p.size_name !== 'ONE SIZE' ? p.size_name : '';
                html += '<tr data-entity="' + (p.entity_id||'') + '">' +
                '<td>' + (thumb ? '<img src="' + thumb + '" style="width:36px;height:36px;border-radius:4px;object-fit:cover" loading="lazy">' : '') + '</td>' +
                '<td>' + (p.nm_id || '') + '</td><td style="max-width:150px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="' + esc(p.product_name) + '">' + esc(p.product_name) + '</td>' +
                '<td style="font-size:.8em;color:#636e72">' + sizeLabel + '</td>' +
                '<td style="font-size:.7em;color:#999">' + (p.barcode || '') + '</td>' +
                '<td style="font-weight:600;line-height:1.1"><div style="color:#0984e3;font-size:.9em">' + (p.stock_fbo_qty ?? '—') + '</div><div style="color:#6c5ce7;font-size:.75em">' + (p.stock_qty ?? '—') + '</div></td>' +
                '<td>' + (p.orders_count ?? '—') + '</td>' +
                '<td>' + (p.buyouts_count ?? '—') + '</td><td>' + (p.returns_count ?? '—') + '</td>' +
                '<td>' + (p.rating ?? '—') + '</td><td>' + (p.impressions ?? '—') + '</td>' +
                '<td>' + (p.clicks ?? '—') + '</td><td>' + ctr + '</td>' +
                '<td>' + fmt(p.ad_cost) + '</td><td>' + fmt(p.wb_price_fact || p.price) + '</td></tr>';
            }
        });
        tbody.innerHTML = html;
        var _pagEl=document.getElementById('stats-shown');var _morBtn=document.getElementById('stats-more-btn');
        if(_pagEl)_pagEl.textContent='Показано '+Math.min(_statsLimit,order.length)+' из '+order.length+' товаров';
        if(_morBtn)_morBtn.style.display=_statsLimit>=order.length?'none':'';
        // === ITOGO ===
        let gStock=0, gStockFbo=0, gOrders=0, gBuyouts=0, gReturns=0, gImpressions=0, gClicks=0, gAd=0, gRatingSum=0, gRatingCount=0;
        prods.forEach(p => {
            gStock += (p.stock_qty || 0) + (p.stock_fbo_qty || 0);
            gStockFbo += p.stock_fbo_qty || 0;
            gOrders += p.orders_count || 0;
            gBuyouts += p.buyouts_count || 0;
            gReturns += p.returns_count || 0;
            gImpressions += p.impressions || 0;
            gClicks += p.clicks || 0;
            gAd += p.ad_cost || 0;
            if (p.rating && p.rating > 0) { gRatingSum += p.rating; gRatingCount++; }
        });
        const gCtr = gImpressions > 0 ? (gClicks / gImpressions * 100).toFixed(1) + '%' : '-';
        const gRating = gRatingCount > 0 ? (gRatingSum / gRatingCount).toFixed(1) : '-';
        const gBuyoutPct = gOrders > 0 ? (gBuyouts / gOrders * 100).toFixed(1) + '%' : '-';
        const gActiveCount = prods.filter(p => ((p.stock_qty || 0) + (p.stock_fbo_qty || 0)) > 0).length;
        const totRow = document.createElement('tr');
        totRow.id = 'stats-total-row';
        totRow.style.cssText = 'background:linear-gradient(135deg,#6c5ce7,#a29bfe);color:#fff;font-weight:700;position:sticky;top:0;z-index:5';
        var totLabel = '\u0418\u0422\u041E\u0413\u041E (' + order.length + ' \u0442\u043e\u0432., ' + gActiveCount + ' \u0430\u043a\u0442.)';
        var rbl = ' \u20bd';
        totRow.innerHTML = '<td colspan="5" style="text-align:left;padding-left:12px;font-size:.95em">' + totLabel + '</td>' +
        '<td style="font-weight:600;line-height:1.1"><div style="color:#0984e3;font-size:.9em">' + gStockFbo + '</div><div style="color:#6c5ce7;font-size:.75em">' + (gStock - gStockFbo) + '</div></td>' +
        '<td>' + gOrders + '</td>' +
        '<td>' + gBuyouts + '</td>' +
        '<td>' + gReturns + '</td>' +
        '<td>' + gRating + '</td>' +
        '<td>' + gImpressions + '</td>' +
        '<td>' + gClicks + '</td>' +
        '<td>' + gCtr + '</td>' +
        '<td>' + fmt(gAd) + rbl + '</td>' +
        '<td>' + gBuyoutPct + '</td>';
        tbody.insertBefore(totRow, tbody.firstChild);
    } catch(e) { console.error('loadStats error:', e); }
}

function showAuth() {
    document.getElementById('app-section').style.display = 'none';
    document.getElementById('auth-section').style.display = '';
    document.getElementById('auth-login').style.display = '';
    document.getElementById('auth-register').style.display = 'none';
    TOKEN = null; ORG_ID = null;
    localStorage.removeItem('nl_token');
    localStorage.removeItem('nl_org_id');
}

async function doLogin() {
    const email = document.getElementById('login-email').value;
    const password = document.getElementById('login-password').value;
    const err = document.getElementById('login-error');
    try {
        const res = await fetch('/api/v1/nl/login', {
            method: 'POST', headers: {'Content-Type': 'application/json'},
            body: JSON.stringify({email, password})
        });
        if (!res.ok) { const d = await res.json(); throw new Error(d.detail || 'Ошибка'); }
        const data = await res.json();
        TOKEN = data.access_token; ORG_ID = data.org_id;
        localStorage.setItem('nl_token', TOKEN);
        localStorage.setItem('nl_org_id', ORG_ID);
        document.getElementById('user-email').textContent = email;
        showApp();
        loadOrgs();
        loadWbKeys();
        loadProfile();
    } catch(e) { err.textContent = e.message; err.style.display = ''; }
}

async function doRegister() {
    const email = document.getElementById('reg-email').value;
    const password = document.getElementById('reg-password').value;
    const org_name = document.getElementById('reg-org').value;
    const err = document.getElementById('reg-error');
    try {
        const res = await fetch('/api/v1/nl/register', {
            method: 'POST', headers: {'Content-Type': 'application/json'},
            body: JSON.stringify({email, password, org_name})
        });
        if (!res.ok) { const d = await res.json(); throw new Error(d.detail || 'Ошибка'); }
        const data = await res.json();
        TOKEN = data.access_token; ORG_ID = data.org_id;
        localStorage.setItem('nl_token', TOKEN);
        localStorage.setItem('nl_org_id', ORG_ID);
        document.getElementById('user-email').textContent = email;
        showApp();
        loadOrgs();
        loadWbKeys();
        loadProfile();
    } catch(e) { err.textContent = e.message; err.style.display = ''; }
}

function doLogout() { showAuth(); }

function switchTab(name, el) { navTo(name, el); }

async function confirmDirty() {
    if (!_costDirty) return true;
    const answer = confirm('В справочнике есть несохранённые изменения. Нажмите ОК — сохранить, Отмена — перейти без сохранения.');
    if (answer) {
        _costDirty = false;
        await saveAllCostPrices();
    } else {
        _costDirty = false;
    }
    return true;
}

// === Реестр секций для lazy loading ===
var _sectionRegistry = {
    stats:      { title:'Основные показатели', topFilters:true },
    rnp:        { title:'РНП', topFilters:true },
    opiu:       { title:'ОПиУ', topFilters:true },
    analytics:  { title:'Аналитика по товарам', topFilters:true },
    costprice:  { title:'Справочник' },
    salesplan:  { title:'План продаж' },
    warehouses: { title:'Склады' },
    opexpenses: { title:'Опер. расходы' },
    ads:        { title:'Реклама', topFilters:true },
    marketer:   { title:'Стол маркетолога' },
    extads:     { title:'Внешняя реклама' },
    fboneeds:   { title:'Потребность FBO' },
    unitecon:   { title:'Юнит Экономика' },
    promo:      { title:'Акции' },
    connectors: { title:'Подключения' },
    subscription:{ title:'Подписка' },
    settings:   { title:'Настройки' },
    help:       { title:'Помощь' }
};

// Загрузчик данных для каждой секции (lazy init + data fetch)

// === Lazy Loading Section Store ===
var _sectionHTML = {
    'opiu': `
<div style="display:flex;align-items:center;gap:12px;margin-bottom:20px;flex-wrap:wrap">
<select id="opiu-period" onchange="loadOpiu()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em">
<option value="4">Последние 4 недели</option><option value="8">8 недель</option><option value="12">12 недель</option>
</select>
<button class="btn" onclick="loadOpiu()" style="padding:6px 14px;font-size:.85em">🔄</button>
<button class="btn btn-outline" onclick="exportOpiu()" style="padding:6px 14px;font-size:.85em">📥 Excel</button>
</div>
<table id="opiu-table"><thead><tr><th>Статья</th><th>Итого</th><th>%</th></tr></thead>
<tbody id="opiu-body"><tr><td colspan="3" class="empty">Нажмите обновить</td></tr></tbody></table>
`,
    'costprice': `
<!-- Верхняя панель: текущий магазин -->
<div style="display:flex;align-items:center;gap:12px;margin-bottom:16px;padding:10px 16px;background:#f8f9fb;border-radius:8px;flex-wrap:wrap">
<span style="font-size:.9em;color:#666">🏪 Магазин:</span>
<select id="cp-store" onchange="switchCostStore()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em;min-width:200px"></select>
</div>


<!-- Налоговые настройки кабинета -->
<div id="tax-settings-panel" style="display:flex;align-items:center;gap:16px;margin-bottom:16px;padding:12px 16px;background:#fff;border:1px solid #e0e0e0;border-radius:8px;flex-wrap:wrap">
<div style="display:flex;align-items:center;gap:8px">
<label style="font-size:.85em;color:#333;font-weight:600;white-space:nowrap">Вид налогообложения</label>
<select id="tax-system-select" onchange="onTaxSystemChange()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 10px;font-size:.9em;min-width:180px">
<option value="">— не выбрано —</option>
<option value="УСН Доходы">УСН "Доходы"</option>
<option value="УСН Доходы-Расходы">УСН "Доходы-Расходы"</option>
<option value="ОСНО">ОСНО</option>
<option value="АУСН Доходы">АУСН "Доходы"</option>
<option value="АУСН Доходы-Расходы">АУСН "Доходы-Расходы"</option>
</select>
<input type="number" id="tax-rate-input" placeholder="% ставки" step="0.01" min="0" max="100" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 10px;font-size:.9em;width:80px">
</div>
<div style="display:flex;align-items:center;gap:8px;margin-left:24px">
<label style="font-size:.85em;color:#333;font-weight:600;white-space:nowrap">НДС от дохода</label>
<select id="vat-type-select" onchange="onTaxSettingsChange()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 10px;font-size:.9em;min-width:100px">
<option value="нет">нет</option>
<option value="5%">5%</option>
<option value="7%">7%</option>
</select>
</div>
<button id="tax-apply-btn" onclick="applyTaxToAll()" title="Применить налоги ко всем строкам" style="padding:6px 10px;font-size:.85em;background:none;border:1px solid #ddd;border-radius:6px;cursor:pointer;margin-left:8px">📋 Применить ко всему</button>
<button id="tax-lock-btn" onclick="toggleTaxLock()" title="Заблокировать/разблокировать" style="padding:6px 10px;font-size:1.2em;background:none;border:1px solid #ddd;border-radius:6px;cursor:pointer;margin-left:8px;transition:all .2s">🔒</button>
<button class="btn" onclick="document.getElementById('cost-file-input').click()" style="padding:6px 14px;font-size:.85em">📤 Загрузить Excel</button>
<input type="file" id="cost-file-input" accept=".xlsx,.csv" style="display:none" onchange="uploadCostExcel(this)">
<button class="btn btn-outline" onclick="downloadEmptyTemplate()" style="padding:6px 14px;font-size:.85em">📥 Скачать шаблон</button>
<button class="btn btn-outline" onclick="exportCostTemplate()" style="padding:6px 14px;font-size:.85em">📤 Экспорт данных</button>
<span style="font-size:.85em;color:#999" id="cost-count"></span>
<button class="btn" onclick="autoFillReference()" style="padding:5px 12px;font-size:.82em;background:#6c5ce7;color:#fff;border-radius:4px">🔄 Автозаполнение</button>
<button class="btn" onclick="saveAllCostPrices()" style="padding:7px 16px;font-size:.88em;background:#00b894;color:#fff;font-weight:600;border-radius:6px">СОХРАНИТЬ<br><span style="font-size:.65em;font-weight:400;opacity:.7">отправить на сервер</span></button>
<span style="flex:1"></span>
<span id="cost-selected-info" style="font-size:.85em;color:#6c5ce7;font-weight:600;display:none">☑ Выделено: <span id="cost-selected-count">0</span></span>
</div>

<!-- Фильтры по столбцам -->
<div id="cost-filters" style="display:flex;align-items:center;gap:8px;margin-bottom:12px;flex-wrap:wrap;padding:8px 12px;background:#f0f1f5;border-radius:8px;font-size:.82em">
<label style="color:#666;font-weight:600">Фильтры:</label>
<select id="flt-fulfillment" onchange="applyCostFilters()" style="border:1px solid #ddd;border-radius:4px;padding:4px 8px;font-size:.9em"><option value="">Отгрузка: все</option><option value="fbo">ФБО</option><option value="fbs">ФБС</option></select>
<select id="flt-tax-system" onchange="applyCostFilters()" style="border:1px solid #ddd;border-radius:4px;padding:4px 8px;font-size:.9em"><option value="">Налог. система: все</option><option value="usn">УСН</option><option value="usn_dr">Доходы-Расходы</option><option value="osn">ОСН</option></select>
<select id="flt-product-class" onchange="applyCostFilters()" style="border:1px solid #ddd;border-radius:4px;padding:4px 8px;font-size:.9em"><option value="">Класс: все</option></select>
<select id="flt-brand" onchange="applyCostFilters()" style="border:1px solid #ddd;border-radius:4px;padding:4px 8px;font-size:.9em"><option value="">Бренд: все</option></select>
<select id="flt-product-status" onchange="applyCostFilters()" style="border:1px solid #ddd;border-radius:4px;padding:4px 8px;font-size:.9em"><option value="">Статус: все</option><option value="Новинка">🟢 Новинка</option><option value="Выводим">🔴 Выводим</option><option value="ТОП (А)">🔵 ТОП (А)</option><option value="Двигаем (В)">🟡 Двигаем (В)</option><option value="Категория С">⚪ Категория С</option><option value="Планируется к запуску">🟣 Планируется к запуску</option></select>
<select id="flt-has-cost" onchange="applyCostFilters()" style="border:1px solid #ddd;border-radius:4px;padding:4px 8px;font-size:.9em"><option value="">Себестоимость: все</option><option value="yes">Заполнена</option><option value="no">Не заполнена</option></select>
<input type="text" id="cost-search" placeholder="🔍 Поиск по названию / артикулу / ШК" oninput="applyCostFilters()" style="border:1px solid #ddd;border-radius:4px;padding:4px 8px;font-size:.9em;width:200px">
<button onclick="clearCostFilters()" style="border:none;background:none;color:#e17055;cursor:pointer;font-size:.9em;padding:4px 8px">✕ Сбросить</button>
</div>

<!-- Таблица -->
<div id="cost-tabulator-host"></div>

<!-- Плавающая панель массовых действий -->
<div id="cost-bulk-bar" style="display:none;position:sticky;bottom:0;left:0;right:0;background:#6c5ce7;color:#fff;padding:10px 16px;border-radius:8px 8px 0 0;display:none;align-items:center;gap:12px;flex-wrap:wrap;z-index:10;box-shadow:0 -2px 10px rgba(0,0,0,.15)">
<span style="font-weight:600" id="bulk-bar-count">Выделено: 0</span>
<select id="bulk-field" onchange="onBulkFieldChange()" style="border:1px solid rgba(255,255,255,.3);border-radius:4px;padding:4px 8px;font-size:.9em;background:#fff;color:#333">
<option value="">Выберите поле...</option>
<optgroup label="📌 Основное">
<option value="product_status">Статус товара</option>
<option value="product_class">Класс товара</option>
<option value="brand">Бренд</option>
</optgroup>
<optgroup label="🚚 Логистика">
<option value="fulfillment_model">Отгрузка</option>
<option value="fbs_warehouse">Склад FBS</option>
</optgroup>
<optgroup label="💰 Себестоимость">
<option value="cost_price">Себестоимость ₽</option>
<option value="extra_costs">Доп расходы ₽</option>
<option value="vat_rate">НДС от дохода</option>
</optgroup>
<optgroup label="📐 Габариты ПЛАН">
<option value="plan_length">Длина</option><option value="plan_width">Ширина</option><option value="plan_height">Высота</option><option value="plan_volume">Объём, л</option><option value="plan_weight">Вес, гр</option>
</optgroup>
<optgroup label="📊 Сезонность">
<option value="season_jan">янв</option><option value="season_feb">фев</option><option value="season_mar">мар</option><option value="season_apr">апр</option><option value="season_may">май</option><option value="season_jun">июн</option><option value="season_jul">июл</option><option value="season_aug">авг</option><option value="season_sep">сен</option><option value="season_oct">окт</option><option value="season_nov">ноя</option><option value="season_dec">дек</option>
</optgroup>
<optgroup label="🔍 ТОП запросы">
<option value="top_query_1">1</option><option value="top_query_2">2</option><option value="top_query_3">3</option>
</optgroup>
<optgroup label="🎯 Расчёты">
<option value="buyout_niche_pct">% выкупа по кат.</option>
<option value="mp_correction_pct">Корр. комиссии %</option>
<option value="ad_plan_rub">Рекл. расходы %</option>
<option value="supply_days">Скорость достав., дн</option>
<option value="min_batch_fbo">Мин партия</option>
<option value="rrc_price">РРЦ</option>
<option value="min_price">Мин. цена</option>

</optgroup>
</select>
<input type="text" id="bulk-value" placeholder="Значение" style="border:1px solid rgba(255,255,255,.3);border-radius:4px;padding:4px 8px;font-size:.9em;width:120px;background:#fff;color:#333">
<button onclick="applyBulkEdit()" style="background:#00b894;color:#fff;border:none;border-radius:4px;padding:6px 14px;font-size:.85em;cursor:pointer;font-weight:600">✅ Применить</button>
<button onclick="clearBulkSelection()" style="background:rgba(255,255,255,.15);color:#fff;border:none;border-radius:4px;padding:6px 14px;font-size:.85em;cursor:pointer">Снять выделение</button>
</div>

<div style="margin-top:12px;display:flex;gap:16px;font-size:.85em" id="cost-summary"></div>
`,
    'salesplan': `
<!-- Фильтры -->
<div style="display:flex;align-items:center;gap:12px;margin-bottom:16px;padding:10px 16px;background:#f8f9fb;border-radius:8px;flex-wrap:wrap">
<select id="sp-period" onchange="loadSalesPlans()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em"></select>
<select id="sp-type" onchange="loadSalesPlans()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em"><option value="">Тип: все</option><option value="quantity">Штуки</option><option value="revenue">Сумма</option></select>
<select id="sp-season" onchange="applySpFilters()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em"><option value="">Сезонность: все</option><option value="low">Низкая</option><option value="medium">Средняя</option><option value="high">Высокая</option><option value="peak">Пик</option></select>
<select id="sp-status" onchange="applySpFilters()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em"><option value="">Статус: все</option><option value="green">🟢 ≥90%</option><option value="yellow">🟡 70-89%</option><option value="red">🔴 <70%</option></select>
<input type="text" id="sp-search" placeholder="🔍 Артикул / название" oninput="applySpFilters()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em;width:180px">
<button class="btn" onclick="loadSalesPlans()" style="padding:6px 14px;font-size:.85em">🔄 Обновить</button>
</div>

<!-- Сводка -->
<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px;margin-bottom:16px" id="sp-summary-cards">
<div class="metric-card"><div class="mc-label">План (всего)</div><div class="mc-value" id="sp-total-plan">—</div></div>
<div class="metric-card"><div class="mc-label">Факт (всего)</div><div class="mc-value" id="sp-total-actual">—</div></div>
<div class="metric-card"><div class="mc-label">% выполнения</div><div class="mc-value" id="sp-total-pct">—</div></div>
<div class="metric-card"><div class="mc-label">🟢 Выполняют</div><div class="mc-value" id="sp-green-count">—</div></div>
<div class="metric-card"><div class="mc-label">🟡 Отстают</div><div class="mc-value" id="sp-yellow-count">—</div></div>
<div class="metric-card"><div class="mc-label">🔴 Не выполняют</div><div class="mc-value" id="sp-red-count">—</div></div>
</div>

<!-- Панель действий -->
<div style="display:flex;align-items:center;gap:12px;margin-bottom:12px;flex-wrap:wrap">
<button class="btn" onclick="openSpAddModal()" style="padding:6px 14px;font-size:.85em;background:#6c5ce7;color:#fff">➕ Добавить план</button>
<button class="btn" onclick="openSpBatchModal()" style="padding:6px 14px;font-size:.85em;background:#00b894;color:#fff">📋 Массовое назначение</button>
<span style="font-size:.85em;color:#999" id="sp-count"></span>
<span style="flex:1"></span>
<span id="sp-selected-info" style="font-size:.85em;color:#6c5ce7;font-weight:600;display:none">☑ Выделено: <span id="sp-selected-count">0</span></span>
<button class="btn" onclick="saveAllSpChanges()" style="padding:6px 14px;font-size:.85em;background:#00b894;color:#fff;display:none" id="sp-save-all-btn">💾 Сохранить всё</button>
</div>

<!-- Таблица -->
<div style="overflow-x:auto;position:relative">
<table id="sp-table" style="font-size:.82em"><thead><tr>
<th style="position:sticky;left:0;z-index:2;background:#fff"><input type="checkbox" id="sp-check-all" onchange="toggleAllSpRows(this.checked)" style="cursor:pointer"></th>
<th>Фото</th><th>Арт ВБ</th><th>Арт поставщика</th><th>Товар</th><th>Размер</th>
<th>Период</th><th>Тип плана</th><th>План</th><th>Факт</th><th>% выполн.</th><th>Статус</th>
<th>Темп продаж</th><th>Сезонность</th><th>Действия</th>
</tr></thead>
<tbody id="sp-body"><tr><td colspan="15" class="empty">Загрузка...</td></tr></tbody></table>
</div>

<!-- Плавающая панель массовых действий -->
<div id="sp-bulk-bar" style="display:none;position:sticky;bottom:0;left:0;right:0;background:#6c5ce7;color:#fff;padding:10px 16px;border-radius:8px 8px 0 0;align-items:center;gap:12px;flex-wrap:wrap;z-index:10;box-shadow:0 -2px 10px rgba(0,0,0,.15)">
<span style="font-weight:600" id="sp-bulk-count">Выделено: 0</span>
<select id="sp-bulk-field" style="border:1px solid rgba(255,255,255,.3);border-radius:4px;padding:4px 8px;font-size:.9em;background:#fff;color:#333">
<option value="">Выберите поле...</option>
<option value="plan_value">План</option>
<option value="plan_type">Тип плана</option>
<option value="seasonality">Сезонность</option>
<option value="actual_value">Факт</option>
<option value="sales_temp">Темп продаж</option>
</select>
<input type="text" id="sp-bulk-value" placeholder="Значение" style="border:1px solid rgba(255,255,255,.3);border-radius:4px;padding:4px 8px;font-size:.9em;width:120px;background:#fff;color:#333">
<button onclick="applySpBulkEdit()" style="background:#00b894;color:#fff;border:none;border-radius:4px;padding:6px 14px;font-size:.85em;cursor:pointer;font-weight:600">✅ Применить</button>
<button onclick="clearSpBulkSelection()" style="background:rgba(255,255,255,.15);color:#fff;border:none;border-radius:4px;padding:6px 14px;font-size:.85em;cursor:pointer">Снять выделение</button>
</div>
`,
    'warehouses': `
<div style="display:flex;align-items:center;gap:12px;margin-bottom:16px;flex-wrap:wrap">
<select id="wh-date" onchange="loadWarehouses()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em"></select>
<button class="btn" onclick="loadWarehouses()" style="padding:6px 14px;font-size:.85em">🔄 Обновить</button>
<label style="font-size:.85em;display:flex;align-items:center;gap:6px"><input type="checkbox" id="wh-sizes" onchange="loadWarehouses()"> Размеры</label>
</div>
<div style="overflow-x:auto">
<table id="wh-table"><thead><tr><th>Арт WB</th><th>Арт продавца</th><th>Товар</th><th>Склад</th><th>Кол-во, шт</th><th>Общая себестоимость</th></tr></thead>
<tbody id="wh-body"><tr><td colspan="6" class="empty">Выберите дату</td></tr></tbody></table>
</div>
<div style="margin-top:12px;font-size:.85em;color:#999" id="wh-count"></div>
`,
    'opexpenses': `
<div style="display:flex;align-items:center;gap:12px;margin-bottom:16px;flex-wrap:wrap">
<button class="btn" onclick="showOpExDialog()" style="padding:6px 14px;font-size:.85em">➕ Добавить</button>
<div style="font-size:.85em;color:#999;margin-left:auto" id="opex-count"></div>
</div>
<table id="opex-table"><thead><tr><th>Дата</th><th>Статья расходов</th><th>Описание</th><th>Сумма ₽</th><th>НДС</th><th>Магазин</th><th></th></tr></thead>
<tbody id="opex-body"><tr><td colspan="7" class="empty">Нет записей</td></tr></tbody></table>

<div id="opex-dialog" style="display:none;position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,.4);z-index:100;align-items:center;justify-content:center">
<div style="background:#fff;padding:24px;border-radius:12px;width:400px;box-shadow:0 4px 20px rgba(0,0,0,.15)">
<h3 style="color:#6c5ce7;margin-bottom:16px">➕ Операционный расход</h3>
<div style="margin-bottom:10px"><label style="font-size:.8em;color:#666">Дата</label><input type="date" id="opex-date" style="width:100%"></div>
<div style="margin-bottom:10px"><label style="font-size:.8em;color:#666">Статья</label><select id="opex-category" style="width:100%"><option>Аренда</option><option>Зарплата</option><option>Маркетинг</option><option>Логистика</option><option>Упаковка</option><option>Прочее</option></select></div>
<div style="margin-bottom:10px"><label style="font-size:.8em;color:#666">Описание</label><input type="text" id="opex-desc" style="width:100%"></div>
<div style="margin-bottom:10px"><label style="font-size:.8em;color:#666">Сумма ₽</label><input type="number" id="opex-amount" style="width:100%"></div>
<div style="margin-bottom:10px"><label style="font-size:.8em;color:#666">НДС %</label><input type="number" id="opex-vat" placeholder="0" style="width:100%"></div>
<div style="display:flex;gap:8px;margin-top:16px">
<button class="btn" onclick="saveOpEx()" style="flex:1">Сохранить</button>
<button class="btn btn-outline" onclick="hideOpExDialog()" style="flex:1">Отмена</button>
</div>
</div>
</div>
`,
    'marketer': `
<!-- Верхняя панель: магазин + период + фильтры -->
<div style="display:flex;align-items:center;gap:12px;margin-bottom:16px;padding:10px 16px;background:#f8f9fb;border-radius:8px;flex-wrap:wrap">
<span style="font-size:.9em;color:#666">🏪 Магазин:</span>
<select id="mkt-store" onchange="switchMktStore()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em;min-width:200px"></select>
<span style="font-size:.9em;color:#666;margin-left:8px">📅 Период:</span>
<select id="mkt-period" onchange="loadMarketer()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em">
<option value="7">7 дней</option>
<option value="14">14 дней</option>
<option value="30" selected>30 дней</option>
<option value="60">60 дней</option>
</select>
<button class="btn" onclick="loadMarketer()" style="padding:6px 14px;font-size:.85em">🔄 Обновить</button>
</div>

<!-- Фильтры -->
<div id="mkt-filters" style="display:flex;align-items:center;gap:8px;margin-bottom:12px;flex-wrap:wrap;padding:8px 12px;background:#f0f1f5;border-radius:8px;font-size:.82em">
<label style="color:#666;font-weight:600">Фильтры:</label>
<select id="mkt-flt-status" onchange="filterMarketer()" style="border:1px solid #ddd;border-radius:4px;padding:4px 8px;font-size:.9em"><option value="">Статус: все</option><option value="Новинка">🟢 Новинка</option><option value="Выводим">🔴 Выводим</option><option value="ТОП (А)">🔵 ТОП (А)</option><option value="Двигаем (В)">🟡 Двигаем (В)</option><option value="Категория С">⚪ Категория С</option><option value="Планируется к запуску">🟣 Планируется</option></select>
<select id="mkt-flt-class" onchange="filterMarketer()" style="border:1px solid #ddd;border-radius:4px;padding:4px 8px;font-size:.9em"><option value="">Класс: все</option><option value="A">A</option><option value="B">B</option><option value="C">C</option></select>
<select id="mkt-flt-brand" onchange="filterMarketer()" style="border:1px solid #ddd;border-radius:4px;padding:4px 8px;font-size:.9em"><option value="">Бренд: все</option></select>
<input type="text" id="mkt-flt-search" placeholder="🔍 Поиск по артикулу / названию" oninput="filterMarketer()" style="border:1px solid #ddd;border-radius:4px;padding:4px 8px;font-size:.9em;width:200px">
<button onclick="resetMktFilters()" style="border:1px solid #ddd;border-radius:4px;padding:4px 10px;font-size:.9em;background:#fff;cursor:pointer">✕ Сбросить</button>
<span style="margin-left:auto;font-size:.85em;color:#999" id="mkt-count"></span>
</div>

<!-- Список товаров -->
<div id="mkt-products-list"></div>

<!-- Карточка товара (скрыта по умолчанию) -->
<div id="mkt-product-detail" style="display:none">
<button onclick="closeMktDetail()" style="margin-bottom:12px;padding:6px 14px;font-size:.85em;background:#fff;border:1px solid #ddd;border-radius:6px;cursor:pointer">← Назад к списку</button>
<div id="mkt-detail-content"></div>
</div>
`,
    'extads': `
<div style="display:flex;align-items:center;gap:12px;margin-bottom:16px;padding:10px 16px;background:#f8f9fb;border-radius:8px;flex-wrap:wrap">
<select id="ext-type" onchange="loadExtAds()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em">
<option value="">Все типы</option><option value="ad">Реклама</option><option value="buyout">Самовыкуп</option></select>
<select id="ext-source" onchange="loadExtAds()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em;min-width:140px"><option value="">Все источники</option></select>
<input type="date" id="ext-date-from" onchange="loadExtAds()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em">
<input type="date" id="ext-date-to" onchange="loadExtAds()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em">
<input type="text" id="ext-search" placeholder="🔍 Поиск..." oninput="loadExtAds()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em;width:160px">
<span style="flex:1"></span>
<button class="btn" onclick="showExtAdModal()" style="padding:6px 14px;font-size:.85em;background:#6c5ce7;color:#fff">➕ Добавить</button>
<button class="btn" onclick="loadExtAds()" style="padding:6px 14px;font-size:.85em">🔄 Обновить</button>
</div>
<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px;margin-bottom:16px">
<div style="background:#fff;border-radius:8px;padding:12px;text-align:center;border:1px solid #e0e0e0"><div style="font-size:1.4em;font-weight:700;color:#6c5ce7" id="ext-total-count">0</div><div style="font-size:.75em;color:#999">Всего записей</div></div>
<div style="background:#fff;border-radius:8px;padding:12px;text-align:center;border:1px solid #e0e0e0"><div style="font-size:1.4em;font-weight:700;color:#e17055" id="ext-total-amount">0 ₽</div><div style="font-size:.75em;color:#999">Общая сумма</div></div>
<div style="background:#fff;border-radius:8px;padding:12px;text-align:center;border:1px solid #e0e0e0"><div style="font-size:1.4em;font-weight:700;color:#00b894" id="ext-total-orders">0</div><div style="font-size:.75em;color:#999">Заказов</div></div>
<div style="background:#fff;border-radius:8px;padding:12px;text-align:center;border:1px solid #e0e0e0"><div style="font-size:1.4em;font-weight:700;color:#0984e3" id="ext-total-reach">0</div><div style="font-size:.75em;color:#999">Охват</div></div>
</div>
<div style="overflow-x:auto;position:relative">
<table id="ext-table" style="font-size:.82em"><thead><tr>
<th style="position:sticky;left:0;z-index:2;background:#fff"><input type="checkbox" id="ext-check-all" onchange="toggleAllExtRows(this.checked)"></th>
<th>Фото</th><th>Артикул</th><th>Арт WB</th><th>Арт продавца</th><th>Товар</th>
<th>Карточка</th><th>Подменная</th><th>UTM</th>
<th>Источник</th><th>Запрос</th><th>Дата</th>
<th>Охват</th><th>Сумма ₽</th><th>Заказы</th><th>Заказов/нед</th>
<th>Тип</th><th>Заметки</th><th>⚡</th>
</tr></thead>
<tbody id="ext-body"><tr><td colspan="20" class="empty">Нажмите "Добавить" для создания записи</td></tr></tbody></table>
</div>
<div id="ext-bulk-bar" style="display:none;position:sticky;bottom:0;left:0;right:0;background:#6c5ce7;color:#fff;padding:10px 16px;border-radius:8px 8px 0 0;align-items:center;gap:12px;flex-wrap:wrap;z-index:10;box-shadow:0 -2px 10px rgba(0,0,0,.15)">
<span style="font-weight:600" id="ext-bulk-count">Выделено: 0</span>
<select id="ext-bulk-field" style="border:1px solid rgba(255,255,255,.3);border-radius:4px;padding:4px 8px;font-size:.9em;background:#fff;color:#333">
<option value="">Поле...</option><option value="source">Источник</option><option value="ad_type">Тип</option><option value="amount">Сумма</option><option value="notes">Заметки</option></select>
<input type="text" id="ext-bulk-value" placeholder="Значение" style="border:1px solid rgba(255,255,255,.3);border-radius:4px;padding:4px 8px;font-size:.9em;width:120px;background:#fff;color:#333">
<button onclick="applyExtBulkEdit()" style="background:#00b894;color:#fff;border:none;border-radius:4px;padding:6px 14px;font-size:.85em;cursor:pointer;font-weight:600">✅ Применить</button>
<button onclick="clearExtBulkSelection()" style="background:rgba(255,255,255,.15);color:#fff;border:none;border-radius:4px;padding:6px 14px;font-size:.85em;cursor:pointer">Снять</button>
</div>
</div>

<div id="ext-modal" style="display:none;position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,.5);z-index:100;align-items:center;justify-content:center">
<div style="background:#fff;border-radius:12px;padding:24px;width:600px;max-width:90vw;max-height:85vh;overflow-y:auto">
<h3 id="ext-modal-title" style="margin-bottom:16px;color:#1a1a2e">Добавить запись</h3>
<input type="hidden" id="ext-edit-id">
<div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
<div><label style="font-size:.85em;color:#666">Артикул WB (nm_id)</label><input type="number" id="ext-nm-id" style="width:100%;border:1px solid #e0e0e0;border-radius:6px;padding:6px 8px;margin-top:4px" oninput="extAutoFill()"></div>
<div><label style="font-size:.85em;color:#666">Арт продавца</label><input type="text" id="ext-vendor-code" style="width:100%;border:1px solid #e0e0e0;border-radius:6px;padding:6px 8px;margin-top:4px"></div>
<div><label style="font-size:.85em;color:#666">Артикул (свой)</label><input type="text" id="ext-article" style="width:100%;border:1px solid #e0e0e0;border-radius:6px;padding:6px 8px;margin-top:4px"></div>
<div><label style="font-size:.85em;color:#666">Тип</label><select id="ext-ad-type" style="width:100%;border:1px solid #e0e0e0;border-radius:6px;padding:6px 8px;margin-top:4px"><option value="ad">Реклама</option><option value="buyout">Самовыкуп</option></select></div>
<div><label style="font-size:.85em;color:#666">Источник</label><input type="text" id="ext-source-input" style="width:100%;border:1px solid #e0e0e0;border-radius:6px;padding:6px 8px;margin-top:4px" placeholder="Telegram-канал, блогер..."></div>
<div><label style="font-size:.85em;color:#666">Запрос</label><input type="text" id="ext-query" style="width:100%;border:1px solid #e0e0e0;border-radius:6px;padding:6px 8px;margin-top:4px"></div>
<div><label style="font-size:.85em;color:#666">Дата</label><input type="date" id="ext-ad-date" style="width:100%;border:1px solid #e0e0e0;border-radius:6px;padding:6px 8px;margin-top:4px"></div>
<div><label style="font-size:.85em;color:#666">Охват</label><input type="number" id="ext-reach" style="width:100%;border:1px solid #e0e0e0;border-radius:6px;padding:6px 8px;margin-top:4px"></div>
<div><label style="font-size:.85em;color:#666">Сумма ₽</label><input type="number" step="0.01" id="ext-amount" style="width:100%;border:1px solid #e0e0e0;border-radius:6px;padding:6px 8px;margin-top:4px"></div>
<div><label style="font-size:.85em;color:#666">Заказов</label><input type="number" id="ext-orders-count" style="width:100%;border:1px solid #e0e0e0;border-radius:6px;padding:6px 8px;margin-top:4px"></div>
<div><label style="font-size:.85em;color:#666">Заказов/нед (ср.)</label><input type="number" step="0.01" id="ext-orders-avg" style="width:100%;border:1px solid #e0e0e0;border-radius:6px;padding:6px 8px;margin-top:4px"></div>
<div><label style="font-size:.85em;color:#666">Фото (URL)</label><input type="text" id="ext-photo-url" style="width:100%;border:1px solid #e0e0e0;border-radius:6px;padding:6px 8px;margin-top:4px"></div>
<div style="grid-column:span 2"><label style="font-size:.85em;color:#666">Подменная ссылка</label><input type="text" id="ext-sub-url" style="width:100%;border:1px solid #e0e0e0;border-radius:6px;padding:6px 8px;margin-top:4px"></div>
<div style="grid-column:span 2"><label style="font-size:.85em;color:#666">UTM ссылка</label><input type="text" id="ext-utm-url" style="width:100%;border:1px solid #e0e0e0;border-radius:6px;padding:6px 8px;margin-top:4px"></div>
<div style="grid-column:span 2"><label style="font-size:.85em;color:#666">Заметки</label><textarea id="ext-notes" rows="2" style="width:100%;border:1px solid #e0e0e0;border-radius:6px;padding:6px 8px;margin-top:4px;resize:vertical"></textarea></div>
</div>
<div style="display:flex;gap:12px;margin-top:16px;justify-content:flex-end">
<button onclick="closeExtAdModal()" style="padding:8px 16px;border:1px solid #e0e0e0;border-radius:6px;cursor:pointer;font-size:.9em">Отмена</button>
<button onclick="saveExtAd()" style="padding:8px 16px;background:#6c5ce7;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:.9em;font-weight:600">💾 Сохранить</button>
</div>
</div>
`,
    'fboneeds': `
<div style="display:flex;align-items:center;gap:12px;margin-bottom:16px;padding-bottom:12px;border-bottom:1px solid #e0e0e0;flex-wrap:wrap;background:#f8f9fb;padding:10px 16px;border-radius:8px">
<select id="fbo-warehouse-filter" onchange="filterFboTable()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em;min-width:160px"><option value="">Все склады</option></select>
<select id="fbo-period" onchange="loadFboNeeds()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em">
<option value="7">7 дней</option><option value="14" selected>14 дней</option><option value="21">21 день</option><option value="30">30 дней</option></select>
<label style="font-size:.85em;color:#666;display:flex;align-items:center;gap:4px"><input type="checkbox" id="fbo-only-needs" onchange="filterFboTable()" checked> Только с потребностью</label>
</div>
<div style="display:flex;align-items:center;gap:12px;margin-bottom:16px;flex-wrap:wrap">
<button class="btn" onclick="loadFboNeeds()" style="padding:6px 14px;font-size:.85em">🔄 Рассчитать</button>
<input type="text" id="fbo-search" placeholder="🔍 Поиск по артикулу/названию" oninput="filterFboTable()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em;width:240px">
<button class="btn btn-outline" onclick="exportFboExcel()" style="padding:6px 14px;font-size:.85em">📥 Excel</button>
<span style="font-size:.85em;color:#999;margin-left:auto" id="fbo-count"></span>
<button class="btn" onclick="saveFboEdits()" style="padding:6px 14px;font-size:.85em;background:#00b894;color:#fff">💾 Сохранить</button>
</div>
<div style="overflow-x:auto;max-height:65vh;position:relative">
<table id="fbo-table" style="font-size:.82em"><thead><tr>
<th>Фото</th><th>Арт WB</th><th>Товар</th><th>Размер</th><th>Склад</th>
<th class="r">Остаток</th><th class="r">Заказов</th><th class="r">Темп/день</th>
<th class="r">Дней до 0</th><th>Срок поставки</th><th>Мин. партия</th>
<th class="r" style="background:#ffeaa7">Потребность</th>
<th class="r" style="background:#fdcb6e;color:#d63031;font-weight:700">✏ К отправке</th>
</tr></thead>
<tbody id="fbo-body"><tr><td colspan="13" class="empty">Нажмите «Рассчитать» для загрузки данных</td></tr></tbody>
</table>
</div>
`,
    'unitecon': `
<!-- Верхняя панель: магазин (как в Справочнике) -->
<div style="display:flex;align-items:center;gap:12px;margin-bottom:16px;padding:10px 16px;background:#f8f9fb;border-radius:8px;flex-wrap:wrap">
<span style="font-size:.9em;color:#666">🏪 Магазин:</span>
<select id="ue-store" onchange="switchUEStore()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em;min-width:200px"></select>
</div>

<!-- Фильтры по столбцам (как в Справочнике) -->
<div id="ue-filters" style="display:flex;align-items:center;gap:8px;margin-bottom:12px;flex-wrap:wrap;padding:8px 12px;background:#f0f1f5;border-radius:8px;font-size:.82em">
<label style="color:#666;font-weight:600">Фильтры:</label>
<select id="ue-flt-status" onchange="applyUEFilters()" style="border:1px solid #ddd;border-radius:4px;padding:4px 8px;font-size:.9em"><option value="">Статус: все</option><option value="Новинка">🟢 Новинка</option><option value="Выводим">🔴 Выводим</option><option value="ТОП (А)">🔵 ТОП (А)</option><option value="Двигаем (В)">🟡 Двигаем (В)</option><option value="Категория С">⚪ Категория С</option><option value="Планируется к запуску">🟣 Планируется</option></select>
<select id="ue-flt-class" onchange="applyUEFilters()" style="border:1px solid #ddd;border-radius:4px;padding:4px 8px;font-size:.9em"><option value="">Класс: все</option><option value="A">A</option><option value="B">B</option><option value="C">C</option></select>
<select id="ue-flt-brand" onchange="applyUEFilters()" style="border:1px solid #ddd;border-radius:4px;padding:4px 8px;font-size:.9em"><option value="">Бренд: все</option></select>
<select id="ue-flt-ff" onchange="applyUEFilters()" style="border:1px solid #ddd;border-radius:4px;padding:4px 8px;font-size:.9em"><option value="">Отгрузка: все</option><option value="fbo">ФБО</option><option value="fbs">ФБС</option></select>
<input type="text" id="ue-flt-search" placeholder="🔍 Поиск по названию / артикулу" oninput="applyUEFilters()" style="border:1px solid #ddd;border-radius:4px;padding:4px 8px;font-size:.9em;width:200px">
<button onclick="resetUEFilters()" style="border:1px solid #ddd;border-radius:4px;padding:4px 10px;font-size:.9em;background:#fff;cursor:pointer">✕ Сбросить</button>
</div>

<div style="display:flex;align-items:center;gap:12px;margin-bottom:16px;flex-wrap:wrap">
<button class="btn" onclick="loadUEData()" style="padding:6px 14px;font-size:.85em">🔄 Обновить</button>
<button class="btn" id="btn-refresh-prices" onclick="refreshPricesFromWB()" style="padding:6px 14px;font-size:.85em;background:#6c5ce7;color:#fff" disabled>💱 Цены из WB...</button>
<button class="btn btn-outline" onclick="exportUEExcel()" style="padding:6px 14px;font-size:.85em">📥 Excel</button>
<span style="font-size:.85em;color:#999;margin-left:auto" id="ue-count"></span>
<button class="btn" onclick="saveUEData()" style="padding:6px 14px;font-size:.85em;background:#00b894;color:#fff">💾 Сохранить</button>
</div>

<div id="ue-tabulator" style="overflow-x:auto;max-height:70vh"></div>

<div style="margin-top:12px;display:flex;gap:16px;font-size:.85em;flex-wrap:wrap" id="ue-summary"></div>
`,
    'promo': `
<div style="display:flex;align-items:center;gap:12px;margin-bottom:16px;padding:10px 16px;background:#f8f9fb;border-radius:8px;flex-wrap:wrap">
<span style="font-size:.9em;color:#666">🏪 Магазин:</span>
<select id="promo-store" onchange="switchPromoStore()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em;min-width:200px"></select>
</div>
<div id="promo-filters" style="display:flex;align-items:center;gap:8px;margin-bottom:12px;flex-wrap:wrap;padding:8px 12px;background:#f0f1f5;border-radius:8px;font-size:.82em">
<label style="color:#666;font-weight:600">Фильтры:</label>
<select id="promo-flt-action" onchange="applyPromoFilters()" style="border:1px solid #ddd;border-radius:4px;padding:4px 8px;font-size:.9em"><option value="">Акция: все</option></select>
<select id="promo-flt-status" onchange="applyPromoFilters()" style="border:1px solid #ddd;border-radius:4px;padding:4px 8px;font-size:.9em"><option value="">Участие: все</option><option value="in">В акции</option><option value="plan">План</option><option value="out">Не участвует</option></select>
<select id="promo-flt-brand" onchange="applyPromoFilters()" style="border:1px solid #ddd;border-radius:4px;padding:4px 8px;font-size:.9em"><option value="">Бренд: все</option></select>
<input type="text" id="promo-flt-search" placeholder="🔍 Поиск по названию / артикулу" oninput="applyPromoFilters()" style="border:1px solid #ddd;border-radius:4px;padding:4px 8px;font-size:.9em;width:200px">
<button onclick="resetPromoFilters()" style="border:1px solid #ddd;border-radius:4px;padding:4px 10px;font-size:.9em;background:#fff;cursor:pointer">✕ Сбросить</button>
</div>
<div style="display:flex;align-items:center;gap:12px;margin-bottom:16px;flex-wrap:wrap">
<button class="btn" onclick="loadPromoData()" style="padding:6px 14px;font-size:.85em">🔄 Обновить</button>
<button class="btn btn-outline" onclick="exportPromoExcel()" style="padding:6px 14px;font-size:.85em">📥 Excel</button>
<button class="btn btn-outline" onclick="document.getElementById('promo-upload-input').click()" style="padding:6px 14px;font-size:.85em">📤 Загрузить шаблон</button>
<input type="file" id="promo-upload-input" accept=".xlsx,.xls" style="display:none" onchange="uploadPromoTemplate(this)">
<button class="btn" onclick="savePromoData()" style="padding:6px 14px;font-size:.85em;background:#00b894;color:#fff">💾 Сохранить</button>
<span style="font-size:.85em;color:#999;margin-left:auto" id="promo-count"></span>
</div>
<div id="promo-tabulator" style="overflow-x:auto;max-height:70vh"></div>
<div style="margin-top:12px;display:flex;gap:16px;font-size:.85em;flex-wrap:wrap" id="promo-summary"></div>
`,
    'connectors': `
<div style="max-width:700px;margin:0 auto;padding:20px">
<h3 style="color:#6c5ce7;margin-bottom:16px">🔌 Подключения</h3>

<!-- Profile -->
<div style="background:#fff;border-radius:8px;padding:16px 20px;box-shadow:0 1px 3px rgba(0,0,0,.08);margin-bottom:16px;display:flex;align-items:center;gap:12px">
<div style="width:40px;height:40px;border-radius:50%;background:#6c5ce7;color:#fff;display:flex;align-items:center;justify-content:center;font-size:1.2em;font-weight:700" id="profile-avatar">?</div>
<div>
<div style="font-weight:600" id="profile-email">Загрузка...</div>
<div style="font-size:.85em;color:#666" id="profile-summary"></div>
</div>
</div>

<!-- My shops -->
<div style="margin-bottom:16px">
<div style="font-weight:600;margin-bottom:10px;font-size:1.05em">🏪 Мои магазины</div>
<div id="shops-list">Загрузка...</div>
</div>

<!-- Add new shop -->
<div style="background:#fff;border-radius:8px;padding:20px;box-shadow:0 1px 3px rgba(0,0,0,.08);margin-bottom:16px;border-left:4px solid #6c5ce7">
<div style="font-weight:600;margin-bottom:12px;font-size:1.1em">➕ Подключить новый магазин</div>
<div style="margin-bottom:10px">
<input type="text" id="new-shop-name" placeholder="Название магазина" style="width:100%;padding:8px 12px;border:1px solid #ddd;border-radius:6px;box-sizing:border-box">
</div>
<div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:8px">
<input type="text" id="new-shop-key" placeholder="API ключ WB (JWT токен)" style="flex:1;min-width:300px;padding:8px 12px;border:1px solid #ddd;border-radius:6px">
<button class="btn" onclick="connectNewShop()" id="btn-connect-shop" style="background:#6c5ce7;color:#fff;padding:8px 20px;border-radius:6px;border:none;cursor:pointer;font-weight:600">Подключить</button>
</div>
<p style="color:#999;font-size:.8em;margin-top:4px">Кабинет WB → Настройки → Доступ к API → Создать токен</p>
<div id="connect-status" style="display:none;margin-top:10px;padding:10px;border-radius:6px;font-size:.9em"></div>
</div>

<!-- Invite colleague -->
<div style="background:#fff;border-radius:8px;padding:20px;box-shadow:0 1px 3px rgba(0,0,0,.08);margin-bottom:16px;border-left:4px solid #00b894">
<div style="font-weight:600;margin-bottom:12px;font-size:1.1em">👥 Пригласить коллегу</div>
<div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:8px">
<input type="email" id="invite-email" placeholder="Email коллеги" style="flex:1;min-width:200px;padding:8px 12px;border:1px solid #ddd;border-radius:6px">
<select id="invite-role" style="padding:8px 12px;border:1px solid #ddd;border-radius:6px">
<option value="VIEWER">Просмотр (VIEWER)</option>
<option value="ADMIN">Администратор (ADMIN)</option>
</select>
<select id="invite-org" style="padding:8px 12px;border:1px solid #ddd;border-radius:6px">
<option value="">Выберите магазин</option>
</select>
<button class="btn" onclick="inviteColleague()" style="background:#00b894;color:#fff;padding:8px 20px;border-radius:6px;border:none;cursor:pointer;font-weight:600">Пригласить</button>
</div>
<div id="invite-status" style="display:none;margin-top:10px;padding:10px;border-radius:6px;font-size:.9em"></div>
</div>

</div>
`,
    'subscription': `
<div style="max-width:900px;margin:0 auto;padding:20px;text-align:center">
<h3 style="color:#6c5ce7;margin-bottom:24px">Тарифные планы</h3>
<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:16px">
<div style="background:#fff;border-radius:12px;padding:20px;box-shadow:0 1px 3px rgba(0,0,0,.08)">
<h4>Новичок</h4><div style="font-size:1.5em;font-weight:700;color:#6c5ce7;margin:12px 0">990 ₽/мес</div>
<ul style="text-align:left;font-size:.85em;color:#666;list-style:none;padding:0"><li>✅ 1 магазин</li><li>✅ Дашборд + ОПиУ</li><li>✅ Аналитика по товарам</li></ul>
<button class="btn" style="width:100%;margin-top:12px">Выбрать</button>
</div>
<div style="background:#fff;border-radius:12px;padding:20px;box-shadow:0 1px 3px rgba(0,0,0,.08);border:2px solid #6c5ce7">
<h4>Старт ⭐</h4><div style="font-size:1.5em;font-weight:700;color:#6c5ce7;margin:12px 0">1490 ₽/мес</div>
<ul style="text-align:left;font-size:.85em;color:#666;list-style:none;padding:0"><li>✅ 2 магазина</li><li>✅ Всё из Новичок</li><li>✅ РНП + Склады</li></ul>
<button class="btn" style="width:100%;margin-top:12px">Выбрать</button>
</div>
<div style="background:#fff;border-radius:12px;padding:20px;box-shadow:0 1px 3px rgba(0,0,0,.08)">
<h4>Бизнес</h4><div style="font-size:1.5em;font-weight:700;color:#6c5ce7;margin:12px 0">3990 ₽/мес</div>
<ul style="text-align:left;font-size:.85em;color:#666;list-style:none;padding:0"><li>✅ 5 магазинов</li><li>✅ Всё из Старт</li><li>✅ Приоритетная поддержка</li></ul>
<button class="btn" style="width:100%;margin-top:12px">Выбрать</button>
</div>
</div>
</div>
`,
    'help': `
<div style="max-width:600px;margin:0 auto;padding:40px;text-align:center;color:#999"><div style="font-size:3em;margin-bottom:16px">❓</div><h3>Помощь</h3><p>support@nl-table.ru</p></div>
`,
    'settings': `
<h3 style="margin-bottom:12px;color:#6c5ce7">🔑 WB API ключи</h3>
<div id="seller-id-display" style="margin-bottom:12px;font-size:.9em;color:#666"></div>
<div id="wb-keys-list"></div>
<div style="margin-top:16px;padding:16px;background:#fff;border-radius:8px;box-shadow:0 1px 3px rgba(0,0,0,.08)">
<h4 style="margin-bottom:10px">Добавить ключ</h4>
<div style="display:flex;gap:8px;flex-wrap:wrap">
<input type="text" id="wb-key-name" placeholder="Название" style="width:150px">
<input type="text" id="wb-key-value" placeholder="API ключ WB" style="flex:1;min-width:200px">
<button class="btn" onclick="addWbKey()">Добавить</button>
</div>
<p style="color:#999;font-size:.8em;margin-top:8px">Получить ключ: Кабинет WB → Настройки → Доступ к API</p>
</div>
</div>
</div>
</div>

<!-- New org dialog -->
<div id="new-org-dialog" style="display:none;position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,.4);z-index:100;display:none;align-items:center;justify-content:center">
<div style="background:#fff;padding:24px;border-radius:12px;width:360px;box-shadow:0 4px 20px rgba(0,0,0,.15)">
<h3 style="color:#6c5ce7;margin-bottom:12px">Новый магазин</h3>
<input type="text" id="new-org-name" placeholder="Название магазина" style="width:100%;margin-bottom:12px">
<div style="display:flex;gap:8px">
<button class="btn" onclick="createNewOrg()" style="flex:1">Создать</button>
<button class="btn btn-outline" onclick="hideNewOrgDialog()" style="flex:1">Отмена</button>
</div>
</div>
</div>

</div>
`,
    'analytics': `
<div style="display:flex;align-items:center;gap:12px;margin-bottom:16px;flex-wrap:wrap">
<input type="text" id="analytics-search" placeholder="🔍 Поиск по артикулу/названию" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em;width:250px" oninput="loadAnalytics()">
<select id="analytics-date" onchange="loadAnalytics()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em"></select>
<button class="btn" onclick="loadAnalytics()" style="padding:6px 14px;font-size:.85em">🔄 Обновить</button>
<button class="btn btn-outline" onclick="exportAnalytics()" style="padding:6px 14px;font-size:.85em">📥 Скачать</button>
</div>
<div style="overflow-x:auto">
<table id="analytics-table" style="font-size:.8em">
<thead><tr>
<th>Фото</th>
<th>Арт продавца</th>
<th>Арт WB</th>
<th>Товар</th>
<th>Остаток</th>
<th>Заказы</th>
<th>Выкупы</th>
<th>Отмены</th>
<th>Возвраты</th>
<th>% выкупа</th>
<th>Цена</th>
<th>Цена со скидкой</th>
<th>Комиссия ВБ %</th>
<th>Комиссия ВБ ₽</th>
<th>Логистика</th>
<th>Реклама ₽</th>
<th>ДРР %</th>
<th>Штрафы</th>
<th>Хранение</th>
<th>Приёмка</th>
<th>Прочие удерж.</th>
<th>Средний чек</th>
<th>Сумма реализации</th>
<th>К выплате</th>
<th>Себестоимость</th>
<th>Маржа</th>
<th>Маржа/ед.</th>
<th>Рентабельность</th>
<th>ROI</th>
<th>Рейтинг</th>
<th>Показы</th>
<th>Клики</th>
<th>CTR</th>
<th>Оборачив.</th>
<th>В пути к клиенту</th>
</tr></thead>
<tbody id="analytics-body"><tr><td colspan="50" class="empty">Загрузка...</td></tr></tbody>
</table>
</div>
<div style="margin-top:12px;display:flex;align-items:center;gap:12px;font-size:.85em;color:#999">
<span id="analytics-count"></span>
<select id="analytics-pagesize" onchange="loadAnalytics()" style="border:1px solid #e0e0e0;border-radius:4px;padding:4px 8px;font-size:.85em">
<option value="10">10</option><option value="25">25</option><option value="50">50</option><option value="100">100</option>
</select>
</div>
`,
    'rnp': `<div class="rnp-ctrl">
<select id="rnp-month" onchange="loadRnp()" style="min-width:130px"></select>
<label style="font-size:.85em;display:flex;align-items:center;gap:4px;cursor:pointer"><input type="checkbox" id="rnp-buyout-pct" onchange="loadRnp()"> Учесть % выкупа</label>
<select id="rnp-sort" onchange="loadRnp()">
<option value="orders_revenue">Сортировка: Заказы, руб</option><option value="roi">ROI</option><option value="buyout_pct">% выкупа</option>
</select>
<input type="text" id="rnp-search" placeholder="🔍 Поиск" style="width:200px" oninput="loadRnp()">
<button class="btn" onclick="loadRnp()" style="padding:6px 14px;font-size:.85em">🔄</button>
</div>
<div id="rnp-summary" class="rnp-summary-bar"></div>
<div id="rnp-header-wrap" style="margin-bottom:8px"></div>
<div id="rnp-cards" style="font-size:.82em"></div>
<div style="margin-top:12px;font-size:.85em;color:#999" id="rnp-count"></div>`,
    'ads': `
<div style="display:flex;align-items:center;gap:12px;margin-bottom:16px;padding:10px 16px;background:#f8f9fb;border-radius:8px;flex-wrap:wrap">
<span style="font-size:.9em;color:#666">🏪 Магазин:</span>
<select id="ads-store" onchange="switchAdsStore()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em;min-width:200px"></select>
<span style="font-size:.9em;color:#666;margin-left:8px">📅 Период:</span>
<select id="ads-period" onchange="adsPeriodPreset()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em">
<option value="1">Сегодня</option>
<option value="2">Вчера</option>
<option value="7">7 дней</option>
<option value="14">14 дней</option>
<option value="30" selected>30 дней</option>
<option value="60">60 дней</option>
<option value="custom">📅 Свой</option>
</select>
<input type="date" id="ads-date-from" onchange="adsCustomDateChange()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em">
<span style="color:#999;font-size:.85em">—</span>
<input type="date" id="ads-date-to" onchange="adsCustomDateChange()" style="border:1px solid #e0e0e0;border-radius:6px;padding:6px 12px;font-size:.9em">
<button class="btn" onclick="refreshAds()" style="padding:6px 14px;font-size:.85em">ð Обновить</button>
<div style="margin-left:auto;font-size:.85em;color:#999" id="ads-updated"></div>
</div>

<!-- Карточки метрик (сгруппированные) -->
<div style="display:flex;flex-wrap:wrap;gap:6px;margin-bottom:8px;align-items:stretch" id="ads-metrics">
<!-- Финансы -->
<div style="display:flex;gap:6px;align-items:stretch">
<div style="background:#fff;border-radius:6px;padding:6px 10px;text-align:center;min-width:78px;border:1px solid #eee"><div style="font-size:.68em;color:#999">Расход</div><div style="font-size:.95em;font-weight:700;color:#e17055" id="ad-spent">—</div></div>
<div style="background:#fff;border-radius:6px;padding:6px 10px;text-align:center;min-width:78px;border:1px solid #eee"><div style="font-size:.68em;color:#999" title="Доля рекламных расходов от оборота">ДРР %</div><div style="font-size:.95em;font-weight:700;color:#e17055" id="ad-drr">—</div></div>
<div style="background:#fff;border-radius:6px;padding:6px 10px;text-align:center;min-width:78px;border:1px solid #eee"><div style="font-size:.68em;color:#999">Баланс</div><div style="font-size:.95em;font-weight:700;color:#2d3436" id="ad-balance">—</div></div>
</div>
<div style="width:1px;background:#e0e0e0;margin:0 4px;align-self:stretch"></div>
<!-- Трафик -->
<div style="display:flex;gap:6px;align-items:stretch">
<div style="background:#fff;border-radius:6px;padding:6px 10px;text-align:center;min-width:78px;border:1px solid #eee"><div style="font-size:.68em;color:#999">Показы</div><div style="font-size:.95em;font-weight:700;color:#6c5ce7" id="ad-views">—</div></div>
<div style="background:#fff;border-radius:6px;padding:6px 10px;text-align:center;min-width:78px;border:1px solid #eee"><div style="font-size:.68em;color:#999">Клики</div><div style="font-size:.95em;font-weight:700;color:#0984e3" id="ad-clicks">—</div></div>
<div style="background:#fff;border-radius:6px;padding:6px 10px;text-align:center;min-width:78px;border:1px solid #eee"><div style="font-size:.68em;color:#999">CTR</div><div style="font-size:.95em;font-weight:700;color:#00b894" id="ad-ctr">—</div></div>
<div style="background:#fff;border-radius:6px;padding:6px 10px;text-align:center;min-width:78px;border:1px solid #eee"><div style="font-size:.68em;color:#999">CPC</div><div style="font-size:.95em;font-weight:700;color:#fdcb6e" id="ad-cpc">—</div></div>
</div>
<div style="width:1px;background:#e0e0e0;margin:0 4px;align-self:stretch"></div>
<!-- Конверсии -->
<div style="display:flex;gap:6px;align-items:stretch">
<div style="background:#fff;border-radius:6px;padding:6px 10px;text-align:center;min-width:78px;border:1px solid #eee"><div style="font-size:.68em;color:#999">Заказы</div><div style="font-size:.95em;font-weight:700;color:#00cec9" id="ad-orders">—</div></div>
<div style="background:#fff;border-radius:6px;padding:6px 10px;text-align:center;min-width:78px;border:1px solid #eee"><div style="font-size:.68em;color:#999">CR</div><div style="font-size:.95em;font-weight:700;color:#e84393" id="ad-cr">—</div></div>
<div style="background:#fff;border-radius:6px;padding:6px 10px;text-align:center;min-width:78px;border:1px solid #eee"><div style="font-size:.68em;color:#999">В корзину</div><div style="font-size:.95em;font-weight:700;color:#636e72" id="ad-arts-count">—</div></div>
</div>
</div>
<!-- Статистика по дням (на всю ширину, под метриками) -->
<div style="background:#fff;border-radius:8px;border:1px solid #eee;padding:8px 12px;margin-bottom:8px">
<div style="display:flex;align-items:center;gap:8px;margin-bottom:4px;cursor:pointer" onclick="toggleDailyTable()"><span style="font-size:.78em;font-weight:600;color:#6c5ce7">📅 По дням</span><span style="font-size:.7em;color:#999" id="ads-daily-count"></span><span style="font-size:.7em;color:#999" id="ads-daily-toggle">▼</span></div>
<div id="ads-daily-wrapper" style="display:none;max-height:180px;overflow-y:auto;font-size:.78em">
<table id="ads-daily-table" style="width:100%">
<thead><tr><th style="padding:2px 6px;text-align:left">Дата</th><th style="padding:2px 6px;text-align:right">Расход ₽</th><th style="padding:2px 6px;text-align:right">Показы</th><th style="padding:2px 6px;text-align:right">Клики</th><th style="padding:2px 6px;text-align:right">CTR</th><th style="padding:2px 6px;text-align:right">CPC ₽</th><th style="padding:2px 6px;text-align:right">Заказы</th><th style="padding:2px 6px;text-align:right">CR</th><th style="padding:2px 6px;text-align:right">В корзину</th><th style="padding:2px 6px;text-align:right">ДРР %</th></tr></thead>
<tbody id="ads-daily-body"><tr><td colspan="9" class="empty" style="padding:4px">Загрузка...</td></tr></tbody>
</table>
</div>
</div>

<!-- Переключалка вида + фильтры статусов в одну строку -->
<div style="display:flex;align-items:center;gap:8px;margin-bottom:8px;flex-wrap:wrap">
<button id="ads-view-rk" class="ads-view-btn active" onclick="switchAdsView('rk')" style="padding:4px 12px;border:1px solid #6c5ce7;background:#6c5ce7;color:#fff;border-radius:6px;font-size:.8em;cursor:pointer;font-weight:600">📊 По РК</button>
<button id="ads-view-art" class="ads-view-btn" onclick="switchAdsView('art')" style="padding:4px 12px;border:1px solid #ddd;background:#fff;color:#333;border-radius:6px;font-size:.8em;cursor:pointer;font-weight:600">📦 По артикулам</button>
<span style="width:1px;height:20px;background:#ddd;margin:0 4px"></span>
<button class="ads-status-btn" data-status="7" onclick="toggleAdsStatusFilter(this)" style="padding:4px 10px;border:1px solid #00b894;background:#00b894;color:#fff;border-radius:6px;font-size:.78em;cursor:pointer;font-weight:600">🟢 Активные</button>
<button class="ads-status-btn" data-status="9" onclick="toggleAdsStatusFilter(this)" style="padding:4px 10px;border:1px solid #fdcb6e;background:#fdcb6e;color:#fff;border-radius:6px;font-size:.78em;cursor:pointer;font-weight:600">⏸ Приостановл.</button>
<button class="ads-status-btn" data-status="11" onclick="toggleAdsStatusFilter(this)" style="padding:4px 10px;border:1px solid #dfe6e9;background:#fff;color:#636e72;border-radius:6px;font-size:.78em;cursor:pointer;font-weight:600">✅ Завершённые</button>
<span style="width:1px;height:20px;background:#ddd;margin:0 4px"></span>
<select id="ads-flt-status" onchange="applyAdsColumnFilters()" style="border:1px solid #ddd;border-radius:4px;padding:3px 6px;font-size:.82em"><option value="">Статус: все</option><option value="Новинка">🟢 Новинка</option><option value="Выводим">🔴 Выводим</option><option value="ТОП (А)">🔵 ТОП (А)</option><option value="Двигаем (В)">🟡 Двигаем (В)</option><option value="Категория С">⚪ Категория С</option><option value="Планируется к запуску">🟣 Планируется</option></select>
<select id="ads-flt-class" onchange="applyAdsColumnFilters()" style="border:1px solid #ddd;border-radius:4px;padding:3px 6px;font-size:.82em"><option value="">Класс: все</option><option value="A">A</option><option value="B">B</option><option value="C">C</option></select>
<select id="ads-flt-brand" onchange="applyAdsColumnFilters()" style="border:1px solid #ddd;border-radius:4px;padding:3px 6px;font-size:.82em"><option value="">Бренд: все</option></select>
<input type="text" id="ads-flt-search" placeholder="🔍 Поиск" oninput="applyAdsColumnFilters()" style="border:1px solid #ddd;border-radius:4px;padding:3px 6px;font-size:.82em;width:140px">
<button onclick="resetAdsColumnFilters()" style="border:1px solid #ddd;border-radius:4px;padding:3px 8px;font-size:.82em;background:#fff;cursor:pointer">✕</button>
<span style="font-size:.8em;color:#999" id="ads-filter-count"></span>
</div>

<!-- Таблицы: артикулы / кампании -->
<div id="ads-arts-container" style="display:none">
<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px">
<button class="btn btn-outline" onclick="exportAdsArtsExcel()" style="padding:3px 10px;font-size:.78em">📥 Excel</button>
<span style="font-size:.8em;color:#999" id="ads-arts-count"></span>
</div>
<div id="ads-arts-tabulator"></div>
</div>
<div id="ads-rk-container">
<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px">
<button class="btn btn-outline" onclick="exportAdsExcel()" style="padding:3px 10px;font-size:.78em">📥 Excel</button>
<span style="font-size:.8em;color:#999" id="ads-camp-count"></span>
</div>
<div id="ads-campaigns-tabulator"></div>
</div><!-- end ads-rk-container -->

<!-- Модал детализации РК -->
<div id="ads-detail-modal" style="display:none;position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,.5);z-index:100;align-items:center;justify-content:center" onclick="if(event.target===this)closeAdsDetailModal()">
<div style="background:#fff;border-radius:12px;padding:24px;width:700px;max-width:90vw;max-height:85vh;overflow-y:auto">
<button onclick="closeAdsDetailModal()" style="float:right;border:none;background:none;font-size:1.2em;cursor:pointer;color:#999">✕</button>
<div id="ads-detail-content"></div>
</div>
</div>

<!-- По дням перенесён выше в блок метрик -->

</div>`,
};

function _lazyInit(name) {
    var html = _sectionHTML[name];
    if (html) document.getElementById('page-' + name).innerHTML = html;
}

function _clearSection(id) {
    var el = document.getElementById(id);
    if (el) el.innerHTML = '';
}

async function _sectionEnter(name) {
    switch(name) {
        case 'stats': loadStats(); break;
        case 'analytics':
            if (!_analyticsInited) { _lazyInit('analytics'); await loadDates(); _analyticsInited = true; }
            loadAnalytics(); break;
        case 'rnp':
            if (!_rnpInited) { _lazyInit('rnp'); if (typeof initRnpMonths === 'function') initRnpMonths(); _rnpInited = true; }
            if (_rnpState) {
                var m = document.getElementById('rnp-month'); if (m) m.value = _rnpState.month;
                var s = document.getElementById('rnp-sort'); if (s) s.value = _rnpState.sort;
                var q = document.getElementById('rnp-search'); if (q) q.value = _rnpState.search;
                var b = document.getElementById('rnp-buyout-pct'); if (b) b.checked = _rnpState.buyoutPct;
            }
            loadRnp(); break;
        case 'opiu':
            if (!_opiuInited) { _lazyInit('opiu'); _opiuInited = true; }
            loadOpiu(); break;
        case 'costprice':
            if (!_costpriceInited) { _lazyInit('costprice'); _costpriceInited = true; }
            loadTaxSettings(); loadCostPrices(); break;
        case 'salesplan':
            if (!_salesplanInited) { _lazyInit('salesplan'); _salesplanInited = true; }
            loadSalesPlans(); break;
        case 'warehouses':
            if (!_warehousesInited) { _lazyInit('warehouses'); _warehousesInited = true; }
            loadWarehouses(); break;
        case 'opexpenses':
            if (!_opexpensesInited) { _lazyInit('opexpenses'); _opexpensesInited = true; }
            loadOpEx(); break;
        case 'ads':
            if (!_adsInited) { _lazyInit('ads'); loadOrgs(); _adsInited = true; }
            if (!adsTabulator) initAdsGrid(); loadAds(); break;
        case 'marketer':
            if (!_marketerInited) { _lazyInit('marketer'); loadOrgs(); _marketerInited = true; }
            loadMarketer(); break;
        case 'extads':
            if (!_extadsInited) { _lazyInit('extads'); loadOrgs(); _extadsInited = true; }
            loadExtAds(); break;
        case 'fboneeds':
            if (!_fboneedsInited) { _lazyInit('fboneeds'); _fboneedsInited = true; }
            loadFboNeeds(); break;
        case 'unitecon':
            if (!_uniteconInited) { _lazyInit('unitecon'); loadOrgs(); _uniteconInited = true; }
            if (!ueTabulator) initUEGrid(); loadUEData(); break;
        case 'promo':
            if (!_promoInited) { _lazyInit('promo'); _promoInited = true; }
            if (typeof promoTabulator === 'undefined' || !promoTabulator) initPromoGrid(); loadPromoData(); break;
        case 'connectors':
            if (!_connectorsInited) { _lazyInit('connectors'); _connectorsInited = true; }
            loadProfile(); loadWbKeys(); break;
        case 'subscription':
            if (!_subscriptionInited) { _lazyInit('subscription'); _subscriptionInited = true; }
            break;
        case 'settings':
            if (!_settingsInited) { _lazyInit('settings'); _settingsInited = true; }
            break;
        case 'help':
            if (!_helpInited) { _lazyInit('help'); _helpInited = true; }
            break;
    }
}

function _sectionLeave(name) {
    switch(name) {
        case 'rnp':
            _rnpState = {
                month: document.getElementById('rnp-month')?.value || '',
                sort: document.getElementById('rnp-sort')?.value || 'orders_revenue',
                search: document.getElementById('rnp-search')?.value || '',
                buyoutPct: document.getElementById('rnp-buyout-pct')?.checked || false
            };
            _clearSection('page-rnp'); _rnpInited = false; break;
        case 'analytics': _clearSection('page-analytics'); _analyticsInited = false; break;
        case 'opiu': _clearSection('page-opiu'); _opiuInited = false; break;
        case 'costprice':
            if (typeof costTabulator !== 'undefined' && costTabulator) { costTabulator.destroy(); costTabulator = null; }
            _clearSection('page-costprice'); _costpriceInited = false; break;
        case 'salesplan': _clearSection('page-salesplan'); _salesplanInited = false; break;
        case 'warehouses': _clearSection('page-warehouses'); _warehousesInited = false; break;
        case 'opexpenses': _clearSection('page-opexpenses'); _opexpensesInited = false; break;
        case 'ads':
            if (typeof adsTabulator !== 'undefined' && adsTabulator) { adsTabulator.destroy(); adsTabulator = null; }
            if (typeof adsArtsTabulator !== 'undefined' && adsArtsTabulator) { adsArtsTabulator.destroy(); adsArtsTabulator = null; }
            _clearSection('page-ads'); _adsInited = false; break;
        case 'marketer': _clearSection('page-marketer'); _marketerInited = false; break;
        case 'extads': _clearSection('page-extads'); _extadsInited = false; break;
        case 'fboneeds': _clearSection('page-fboneeds'); _fboneedsInited = false; break;
        case 'unitecon':
            if (typeof ueTabulator !== 'undefined' && ueTabulator) { ueTabulator.destroy(); ueTabulator = null; }
            _clearSection('page-unitecon'); _uniteconInited = false; break;
        case 'promo':
            if (typeof promoTabulator !== 'undefined' && promoTabulator) { promoTabulator.destroy(); promoTabulator = null; }
            _clearSection('page-promo'); _promoInited = false; break;
        case 'connectors': _clearSection('page-connectors'); _connectorsInited = false; break;
        case 'subscription': _clearSection('page-subscription'); _subscriptionInited = false; break;
        case 'settings': _clearSection('page-settings'); _settingsInited = false; break;
        case 'help': _clearSection('page-help'); _helpInited = false; break;
    }
}

var _rnpState = null, _rnpInited = false, _adsInited = false, _analyticsInited = false;
var _opiuInited = false, _costpriceInited = false, _salesplanInited = false, _warehousesInited = false;
var _opexpensesInited = false, _marketerInited = false, _extadsInited = false, _fboneedsInited = false;
var _uniteconInited = false, _promoInited = false, _connectorsInited = false, _subscriptionInited = false;
var _settingsInited = false, _helpInited = false;
var _currentSection = 'stats';


async function navTo(name, el) {
    if (_costDirty) await confirmDirty();
    // Leave current section
    if (_currentSection && _currentSection !== name) _sectionLeave(_currentSection);
    // Nav UI
    document.querySelectorAll('.nav-item').forEach(n => n.classList.remove('active'));
    if (el) el.classList.add('active');
    document.querySelectorAll('.page-section').forEach(t => t.classList.remove('active'));
    var target = document.getElementById('page-' + name);
    if (target) target.classList.add('active');
    // Title
    var reg = _sectionRegistry[name] || {};
    document.getElementById('page-title').textContent = reg.title || name;
    // Top filters
    var topFilters = document.getElementById('top-filters');
    if (topFilters) topFilters.style.display = reg.topFilters ? 'flex' : 'none';
    // Cleanup popups
    if (typeof cleanupCostPopups === 'function') cleanupCostPopups();
    // Track current & enter
    _currentSection = name;
    _sectionEnter(name);
}

async function loadAds() {
    if (!ORG_ID || ORG_ID === 'null') { console.warn('loadAds: no ORG_ID'); return; }
    const range = getAdsDateRange();
    let url;
    if (range.days) {
        url = '/api/v1/nl/ad-stats?org_id=' + ORG_ID + '&days=' + range.days;
    } else if (range.date_from && range.date_to) {
        url = '/api/v1/nl/ad-stats?org_id=' + ORG_ID + '&date_from=' + range.date_from + '&date_to=' + range.date_to;
    } else {
        url = '/api/v1/nl/ad-stats?org_id=' + ORG_ID + '&days=30';
    }
    try {
        const r = await fetch(url, {headers:{'Authorization':'Bearer '+TOKEN}});
        const d = await r.json();
        // Totals
        const t = d.totals || {};
        const fmt = (v, s='') => v != null ? Number(v).toLocaleString('ru-RU', {maximumFractionDigits: v >= 1000 ? 0 : 2}) + s : '—';
        document.getElementById('ad-views').textContent = fmt(t.views);
        document.getElementById('ad-clicks').textContent = fmt(t.clicks);
        document.getElementById('ad-spent').textContent = fmt(t.spent, ' ₽');
        document.getElementById('ad-ctr').textContent = t.ctr ? t.ctr + '%' : '—';
        document.getElementById('ad-cpc').textContent = fmt(t.cpc, ' ₽');
        document.getElementById('ad-orders').textContent = fmt(t.orders);
        document.getElementById('ad-cr').textContent = t.cr ? t.cr + '%' : '—';
        document.getElementById('ad-drr').textContent = t.drr ? t.drr + '%' : '—';
        // Подсчёт уникальных артикулов из кампаний
        var _nmSet = new Set(); (d.campaigns||[]).forEach(function(c){ (c.products||[]).forEach(function(p){ if(p.nm_id) _nmSet.add(p.nm_id); }); });
        document.getElementById('ad-arts-count').textContent = _nmSet.size || '—';
        // Balance
        if (d.balance) {
            var balNum = d.balance.net || d.balance.balance || d.balance.balanceXdiscount || 0;
            if (typeof balNum !== 'number') balNum = 0;
            document.getElementById('ad-balance').textContent = fmt(balNum, ' ₽');
        } else {
            document.getElementById('ad-balance').textContent = '—';
        }
        // Daily table
        const daily = d.daily || [];
        const db = document.getElementById('ads-daily-body');
        if (!daily.length) {
            db.innerHTML = '<tr><td colspan="5" class="empty" style="padding:4px">Нет данных</td></tr>';
        } else {
            db.innerHTML = daily.map(r => '<tr><td style="padding:2px 6px">'+r.date+'</td><td style="padding:2px 6px;text-align:right">'+r.spent.toLocaleString('ru-RU',{maximumFractionDigits:0})+'₽</td><td style="padding:2px 6px;text-align:right">'+r.views.toLocaleString('ru-RU')+'</td><td style="padding:2px 6px;text-align:right">'+r.clicks.toLocaleString('ru-RU')+'</td><td style="padding:2px 6px;text-align:right">'+r.ctr+'%</td><td style="padding:2px 6px;text-align:right">'+r.cpc.toFixed(2)+'₽</td><td style="padding:2px 6px;text-align:right">'+r.orders+'</td><td style="padding:2px 6px;text-align:right">'+r.cr+'%</td><td style="padding:2px 6px;text-align:right">'+(r.atbs||0)+'</td><td style="padding:2px 6px;text-align:right">'+(r.drr > 100 ? '<span style="color:#b2bec3">н/д</span>' : r.drr ? '<span style="color:'+(r.drr>50?'#e74c3c':r.drr>25?'#e17055':'#00b894')+';font-weight:600">'+r.drr+'%</span>' : '\u2014')+'</td></tr>').join('');
            document.getElementById('ads-daily-count').textContent = daily.length + ' дней';
        }
        // --- Рендер таблицы кампаний (Tabulator) ---
        window._adsAllCampaigns = d.campaigns || [];
        updateAdsTabulator(d.campaigns || []);

        document.getElementById('ads-updated').textContent = 'Обновлено: ' + new Date().toLocaleTimeString('ru-RU');
    } catch(e) {
        console.error('loadAds error:', e);
        document.getElementById('ads-daily-body').innerHTML = '<tr><td colspan="5" class="empty" style="padding:4px">Ошибка: '+e.message+'</td></tr>';
    }
}




function toggleDailyTable() {
    var w = document.getElementById('ads-daily-wrapper');
    var t = document.getElementById('ads-daily-toggle');
    if (w.style.display === 'none') { w.style.display = 'block'; t.textContent = '▲'; }
    else { w.style.display = 'none'; t.textContent = '▼'; }
}

// ===== ADS PERIOD & CALENDAR =====
function adsPeriodPreset() {
    if(typeof _adsCurrentView !== "undefined" && _adsCurrentView === "art") { loadAdsArts(); return; }
    const sel = document.getElementById('ads-period');
    const val = sel.value;
    const df = document.getElementById('ads-date-from');
    const dt = document.getElementById('ads-date-to');
    if (val === 'custom') {
        // Дефолт: последние 30 дней если пусто
        if (!df.value || !dt.value) {
            const today = new Date();
            const from = new Date(today);
            from.setDate(from.getDate() - 30);
            df.value = from.toISOString().split('T')[0];
            dt.value = today.toISOString().split('T')[0];
        }
        return; // ждём ручной ввод + кнопку Обновить
    }
    // Пресет — обновляем сразу
    loadAds();
}

function adsCustomDateChange() {
    const df = document.getElementById('ads-date-from').value;
    const dt = document.getElementById('ads-date-to').value;
    if (df && dt) {
        document.getElementById('ads-period').value = 'custom';
        if(typeof _adsCurrentView !== "undefined" && _adsCurrentView === "art") { loadAdsArts(); return; }
        loadAds();
    }
}

function getAdsDateRange() {
    const val = document.getElementById('ads-period').value;
    let days = parseInt(val);
    if (!isNaN(days)) {
        return { days: days };
    }
    const df = document.getElementById('ads-date-from').value;
    const dt = document.getElementById('ads-date-to').value;
    return { date_from: df, date_to: dt };
}

function filterMarketer() {
    // Re-fetch with filters
    loadMarketer();
}

function resetMktFilters() {
    document.getElementById('mkt-flt-status').value = '';
    document.getElementById('mkt-flt-class').value = '';
    document.getElementById('mkt-flt-brand').value = '';
    document.getElementById('mkt-flt-search').value = '';
    loadMarketer();
}

function switchMktStore() {
    var sel = document.getElementById('mkt-store');
    if (sel && sel.value) {
        ORG_ID = sel.value;
        loadMarketer();
    }
}

function initMktStores() {
    // Populate store selector from existing store selectors
    var ueStore = document.getElementById('ue-store');
    var mktStore = document.getElementById('mkt-store');
    if (ueStore && mktStore) {
        mktStore.innerHTML = ueStore.innerHTML;
        mktStore.value = ORG_ID;
    }
}

function renderMktProducts() {
    var el = document.getElementById('mkt-products-list');
    if (!mktAllProducts.length) {
        el.innerHTML = '<div class="empty">Нет товаров с рекламными данными за выбранный период</div>';
        return;
    }

    var fmt = function(v) {
        if (v == null) return '—';
        if (v >= 1000) return v.toLocaleString('ru-RU', {maximumFractionDigits: 0});
        return typeof v === 'number' ? v.toFixed(2) : v;
    };

    var html = '<table style="width:100%;border-collapse:collapse;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.08);font-size:.82em">';
    html += '<thead><tr style="background:#f8f9fa">';
    html += '<th style="padding:8px;text-align:left">📷</th>';
    html += '<th style="padding:8px;text-align:left">Артикул</th>';
    html += '<th style="padding:8px;text-align:left">Бренд</th>';
    html += '<th style="padding:8px;text-align:center">Статус</th>';
    html += '<th style="padding:8px;text-align:center">Класс</th>';
    html += '<th style="padding:8px;text-align:right">Цена</th>';
    html += '<th style="padding:8px;text-align:center">РК</th>';
    html += '<th style="padding:8px;text-align:right">Показы</th>';
    html += '<th style="padding:8px;text-align:right">Клики</th>';
    html += '<th style="padding:8px;text-align:right">CTR %</th>';
    html += '<th style="padding:8px;text-align:right">Расход ₽</th>';
    html += '<th style="padding:8px;text-align:right">Заказы</th>';
    html += '<th style="padding:8px;text-align:right">ДРР %</th>';
    html += '</tr></thead><tbody>';

    mktAllProducts.forEach(function(p) {
        var photo = p.photo ? '<img src="' + p.photo.replace('/hq/', '/c246x328/') + '" style="width:40px;height:52px;object-fit:cover;border-radius:4px" onerror="this.style.opacity=0">' : '—';
        var statusBadge = p.status ? '<span style="font-size:.75em;padding:2px 6px;border-radius:3px;background:' + getStatusColor(p.status) + ';color:#fff">' + p.status + '</span>' : '—';
        var classBadge = p.abc_class ? '<span style="font-size:.8em;font-weight:700;color:' + getClassColor(p.abc_class) + '">' + p.abc_class + '</span>' : '—';

        html += '<tr style="cursor:pointer;border-bottom:1px solid #f0f0f0" onclick="openMktDetail(' + p.nm_id + ')">';
        html += '<td style="padding:6px">' + photo + '</td>';
        html += '<td style="padding:6px"><div style="font-weight:600;font-size:.85em">' + (p.vendor_code || p.nm_id) + '</div><div style="font-size:.7em;color:#999">nm ' + p.nm_id + '</div></td>';
        html += '<td style="padding:6px;font-size:.85em">' + (p.brand || '—') + '</td>';
        html += '<td style="padding:6px;text-align:center">' + statusBadge + '</td>';
        html += '<td style="padding:6px;text-align:center">' + classBadge + '</td>';
        html += '<td style="padding:6px;text-align:right">' + fmt(p.price) + '</td>';
        html += '<td style="padding:6px;text-align:center"><span style="background:#6c5ce7;color:#fff;padding:2px 8px;border-radius:10px;font-size:.8em">' + p.active_campaign_count + '/' + p.campaign_count + '</span></td>';
        html += '<td style="padding:6px;text-align:right">' + fmt(p.total_views) + '</td>';
        html += '<td style="padding:6px;text-align:right">' + fmt(p.total_clicks) + '</td>';
        html += '<td style="padding:6px;text-align:right">' + fmt(p.ctr) + '</td>';
        html += '<td style="padding:6px;text-align:right;font-weight:600">' + fmt(p.total_spent) + '</td>';
        html += '<td style="padding:6px;text-align:right">' + p.total_orders + '</td>';
        html += '<td style="padding:6px;text-align:right;color:' + (p.drr > 15 ? '#e74c3c' : '#27ae60') + '">' + fmt(p.drr) + '</td>';
        html += '</tr>';
    });

    html += '</tbody></table>';
    el.innerHTML = html;
}

function getStatusColor(s) {
    var map = {'Новинка':'#27ae60','Выводим':'#e74c3c','ТОП (А)':'#2980b9','Двигаем (В)':'#f39c12','Категория С':'#95a5a6','Планируется к запуску':'#8e44ad'};
    return map[s] || '#999';
}

function getClassColor(c) {
    var map = {'A':'#2980b9','B':'#f39c12','C':'#95a5a6'};
    return map[c] || '#999';
}

async function openMktDetail(nmId) {
    var period = document.getElementById('mkt-period').value;
    try {
        var r = await fetch('/api/v1/nl/marketer/product/' + nmId + '?org_id=' + ORG_ID + '&days=' + period, {
            headers: {'Authorization': 'Bearer ' + TOKEN}
        });
        var d = await r.json();

        if (d.error) {
            alert(d.error);
            return;
        }

        document.getElementById('mkt-products-list').style.display = 'none';
        document.getElementById('mkt-filters').style.display = 'none';
        document.getElementById('mkt-product-detail').style.display = 'block';

        renderMktDetail(d);
    } catch(e) {
        console.error('openMktDetail error:', e);
        alert('Ошибка: ' + e.message);
    }
}

function closeMktDetail() {
    document.getElementById('mkt-products-list').style.display = 'block';
    document.getElementById('mkt-filters').style.display = 'flex';
    document.getElementById('mkt-product-detail').style.display = 'none';
}

function renderMktDetail(d) {
    var p = d.product;
    var el = document.getElementById('mkt-detail-content');
    var fmt = function(v) { return v != null ? (v >= 1000 ? v.toLocaleString('ru-RU', {maximumFractionDigits:0}) : (typeof v === 'number' ? v.toFixed(2) : v)) : '—'; };

    var html = '';

    // Шапка товара
    html += '<div style="display:flex;gap:20px;margin-bottom:20px;padding:16px;background:#fff;border-radius:8px;box-shadow:0 1px 3px rgba(0,0,0,.08);flex-wrap:wrap">';

    // Фото
    if (p.photo) {
        html += '<img src="' + p.photo.replace('/hq/', '/c246x328/') + '" style="width:100px;height:130px;object-fit:cover;border-radius:8px" onerror="this.style.opacity=0">';
    }

    html += '<div style="flex:1;min-width:200px">';
    html += '<h3 style="color:#6c5ce7;margin-bottom:8px">' + (p.vendor_code || 'Арт ' + p.nm_id) + '</h3>';
    html += '<div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:8px">';
    if (p.status) html += '<span style="font-size:.8em;padding:3px 8px;border-radius:4px;background:' + getStatusColor(p.status) + ';color:#fff">' + p.status + '</span>';
    if (p.abc_class) html += '<span style="font-size:.8em;padding:3px 8px;border-radius:4px;background:' + getClassColor(p.abc_class) + ';color:#fff">Класс ' + p.abc_class + '</span>';
    if (p.in_promo) html += '<span style="font-size:.8em;padding:3px 8px;border-radius:4px;background:#e67e22;color:#fff">🏷 В акции</span>';
    html += '</div>';
    html += '<div style="font-size:.85em;color:#666">';
    html += '<div><strong>Арт WB:</strong> ' + p.nm_id + '</div>';
    html += '<div><strong>Бренд:</strong> ' + (p.brand || '—') + '</div>';
    html += '<div><strong>Категория:</strong> ' + (p.category || '—') + '</div>';
    html += '</div>';
    html += '</div>';

    // Цены
    html += '<div style="text-align:right;min-width:150px">';
    html += '<div style="font-size:.85em;color:#666">Цена до СПП</div>';
    html += '<div style="font-size:1.3em;font-weight:700;color:#333">' + fmt(p.price || 0) + ' ₽</div>';
    html += '<div style="font-size:.85em;color:#666;margin-top:6px">Цена с СПП</div>';
    html += '<div style="font-size:1.1em;color:#6c5ce7">' + fmt(p.price_spp || 0) + ' ₽</div>';
    html += '</div>';

    html += '</div>';

    // Лучший период
    if (d.best_period) {
        var bp = d.best_period;
        html += '<div style="padding:12px 16px;background:#f0fff4;border:1px solid #27ae60;border-radius:8px;margin-bottom:20px">';
        html += '<div style="font-weight:600;color:#27ae60;margin-bottom:6px">🏆 Лучший период</div>';
        html += '<div style="display:flex;gap:16px;flex-wrap:wrap;font-size:.85em">';
        html += '<span><strong>Дата:</strong> ' + bp.date + '</span>';
        html += '<span><strong>Заказы:</strong> ' + bp.orders + '</span>';
        html += '<span><strong>Цена:</strong> ' + fmt(bp.price) + ' ₽</span>';
        html += '<span><strong>Цена с СПП:</strong> ' + fmt(bp.price_spp) + ' ₽</span>';
        html += '<span><strong>Прибыль:</strong> ' + fmt(bp.profit) + ' ₽</span>';
        html += '<span><strong>Расход:</strong> ' + fmt(bp.spent) + ' ₽</span>';
        html += '</div></div>';
    }

    // Сводка «РК В ОБЩЕМ»
    var gt = d.grand_total || {};
    html += '<div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(140px,1fr));gap:10px;margin-bottom:20px">';
    html += '<div class="metric-card"><div class="mc-label">Показы</div><div class="mc-value">' + fmt(gt.views || 0) + '</div></div>';
    html += '<div class="metric-card"><div class="mc-label">Клики</div><div class="mc-value">' + fmt(gt.clicks || 0) + '</div></div>';
    html += '<div class="metric-card"><div class="mc-label">CTR</div><div class="mc-value">' + fmt(gt.ctr || 0) + '%</div></div>';
    html += '<div class="metric-card"><div class="mc-label">Расход</div><div class="mc-value">' + fmt(gt.spent || 0) + ' ₽</div></div>';
    html += '<div class="metric-card"><div class="mc-label">Заказы</div><div class="mc-value">' + (gt.orders || 0) + '</div></div>';
    html += '</div>';

    // Графики по РК
    var campaigns = d.campaigns || [];
    var activeCamps = campaigns.filter(function(c) { return c.is_active; });
    var inactiveCamps = campaigns.filter(function(c) { return !c.is_active; });
    
    // График: РК В ОБЩЕМ + Органика
    html += '<div style="background:#fff;border-radius:8px;margin-bottom:20px;padding:16px;box-shadow:0 1px 3px rgba(0,0,0,.08)">';
    html += '<h4 style="color:#6c5ce7;margin-bottom:12px;font-size:.95em">📊 РК В ОБЩЕМ — тренды по дням</h4>';
    html += '<canvas id="mkt-chart-summary" height="110"></canvas>';
    html += '</div>';
    
    // Графики по каждой активной РК
    campaigns.forEach(function(camp, idx) {
        var statusLabel = camp.is_active ? '🟢' : '🔴';
        html += '<div style="background:#fff;border-radius:8px;margin-bottom:16px;padding:16px;box-shadow:0 1px 3px rgba(0,0,0,.08)">';
        html += '<h4 style="color:#6c5ce7;margin-bottom:12px;font-size:.95em">' + statusLabel + ' ' + camp.name + ' (ID: ' + camp.campaign_id + ')</h4>';
        html += '<canvas id="mkt-chart-rk-' + idx + '" height="100"></canvas>';
        html += '</div>';
    });

    // По каждой РК
    var campaigns = d.campaigns || [];
    if (campaigns.length) {
        html += '<h3 style="color:#6c5ce7;margin-bottom:12px;font-size:1em">📢 Рекламные кампании (' + campaigns.length + ')</h3>';

        campaigns.forEach(function(camp, idx) {
            var isActive = camp.is_active;
            var borderColor = isActive ? '#6c5ce7' : '#ccc';
            var statusText = isActive ? '🟢 Активна' : '🔴 Приостановлена';

            html += '<div style="border:1px solid ' + borderColor + ';border-radius:8px;margin-bottom:16px;overflow:hidden">';

            // Шапка РК
            html += '<div style="padding:10px 16px;background:#f8f9fa;border-bottom:1px solid ' + borderColor + ';display:flex;align-items:center;gap:12px;flex-wrap:wrap">';
            html += '<span style="font-weight:600;font-size:.95em">РК ' + (idx+1) + ': ' + camp.name + '</span>';
            html += '<span style="font-size:.8em;padding:2px 8px;border-radius:10px;background:' + (isActive ? '#27ae60' : '#e74c3c') + ';color:#fff">' + statusText + '</span>';
            html += '<span style="font-size:.8em;color:#999">ID: ' + camp.campaign_id + '</span>';
            html += '<span style="font-size:.8em;color:#999">Тип: ' + camp.type + '</span>';
            html += '</div>';

            // Метрики РК
            var ct = camp.totals || {};
            html += '<div style="padding:12px 16px;display:grid;grid-template-columns:repeat(auto-fill,minmax(120px,1fr));gap:8px;font-size:.85em">';
            html += '<div><span style="color:#999">Показы:</span> <strong>' + fmt(ct.views) + '</strong></div>';
            html += '<div><span style="color:#999">Клики:</span> <strong>' + fmt(ct.clicks) + '</strong></div>';
            html += '<div><span style="color:#999">CTR:</span> <strong>' + fmt(ct.ctr) + '%</strong></div>';
            html += '<div><span style="color:#999">CPC:</span> <strong>' + fmt(ct.cpc) + ' ₽</strong></div>';
            html += '<div><span style="color:#999">Расход:</span> <strong>' + fmt(ct.spent) + ' ₽</strong></div>';
            html += '<div><span style="color:#999">Заказы:</span> <strong>' + (ct.orders || 0) + '</strong></div>';
            html += '<div><span style="color:#999">CR:</span> <strong>' + fmt(ct.cr) + '%</strong></div>';
            html += '</div>';

            // Таблица по дням
            var daily = camp.daily || [];
            if (daily.length) {
                html += '<div style="padding:0 16px 12px;overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-size:.78em">';
                html += '<thead><tr style="background:#f0f1f5">';
                html += '<th style="padding:4px 6px;text-align:left">Дата</th>';
                html += '<th style="padding:4px 6px;text-align:right">Показы</th>';
                html += '<th style="padding:4px 6px;text-align:right">Клики</th>';
                html += '<th style="padding:4px 6px;text-align:right">CTR</th>';
                html += '<th style="padding:4px 6px;text-align:right">CPC ₽</th>';
                html += '<th style="padding:4px 6px;text-align:right">Расход ₽</th>';
                html += '<th style="padding:4px 6px;text-align:right">Заказы</th>';
                html += '<th style="padding:4px 6px;text-align:right">В корзину</th>';
                html += '</tr></thead><tbody>';

                daily.reverse().forEach(function(d) {
                    html += '<tr style="border-bottom:1px solid #f0f0f0">';
                    html += '<td style="padding:3px 6px">' + d.date + '</td>';
                    html += '<td style="padding:3px 6px;text-align:right">' + d.views + '</td>';
                    html += '<td style="padding:3px 6px;text-align:right">' + d.clicks + '</td>';
                    html += '<td style="padding:3px 6px;text-align:right">' + d.ctr + '</td>';
                    html += '<td style="padding:3px 6px;text-align:right">' + d.cpc + '</td>';
                    html += '<td style="padding:3px 6px;text-align:right;font-weight:600">' + d.spent + '</td>';
                    html += '<td style="padding:3px 6px;text-align:right">' + d.orders + '</td>';
                    html += '<td style="padding:3px 6px;text-align:right">' + d.atbs + '</td>';
                    html += '</tr>';
                });

                html += '</tbody></table></div>';
            }

            html += '</div>';
        });
    } else {
        html += '<div class="empty" style="padding:20px;text-align:center">Нет рекламных кампаний для этого товара за выбранный период</div>';
    }

    // Ссылки на неработающие разделы
    html += '<div style="margin-top:20px;padding:12px 16px;background:#f8f9fb;border-radius:8px;font-size:.85em;color:#999">';
    html += '<div>🔗 <strong>В разработке:</strong></div>';
    html += '<div>• Ссылка на ОПиУ товара</div>';
    html += '<div>• План заказов (шт) и % выполнения</div>';
    html += '<div>• Стратегия РК (ручной ввод)</div>';
    html += '<div>• Ссылка на РК в кабинете WB</div>';
    html += '<div>• Графики Chart.js (по дням)</div>';
    html += '</div>';

    el.innerHTML = html;
    renderMktCharts(d);
}

function renderMktCharts(d) {
    if (window._mktCharts) { window._mktCharts.forEach(function(c) { c.destroy(); }); }
    window._mktCharts = [];

    var summaryDaily = (d.summary_daily || []).sort(function(a,b) { return a.date < b.date ? -1 : 1; });
    var labels = summaryDaily.map(function(dd) { return dd.date.substring(5); });

    // Summary chart
    var ctxS = document.getElementById('mkt-chart-summary');
    if (ctxS && summaryDaily.length > 1) {
        var cS = new Chart(ctxS.getContext('2d'), {
            type: 'line',
            data: {
                labels: labels,
                datasets: [
                    { label: '\u041f\u043e\u043a\u0430\u0437\u044b', data: summaryDaily.map(function(dd) { return dd.views; }), borderColor: '#6c5ce7', backgroundColor: 'rgba(108,92,231,0.1)', yAxisID: 'y', tension: 0.2, pointRadius: 3 },
                    { label: '\u041a\u043b\u0438\u043a\u0438', data: summaryDaily.map(function(dd) { return dd.clicks; }), borderColor: '#00b894', backgroundColor: 'rgba(0,184,148,0.1)', yAxisID: 'y', tension: 0.2, pointRadius: 3 },
                    { label: '\u0417\u0430\u043a\u0430\u0437\u044b', data: summaryDaily.map(function(dd) { return dd.orders; }), borderColor: '#fdcb6e', backgroundColor: 'rgba(253,203,110,0.1)', yAxisID: 'y', tension: 0.2, pointRadius: 3 },
                    { label: '\u0420\u0430\u0441\u0445\u043e\u0434 \u20bd', data: summaryDaily.map(function(dd) { return dd.spent; }), borderColor: '#e17055', backgroundColor: 'rgba(225,112,85,0.1)', yAxisID: 'y1', tension: 0.2, pointRadius: 3 }
                ]
            },
            options: {
                responsive: true, interaction: { mode: 'index', intersect: false },
                plugins: { legend: { position: 'top', labels: { font: { size: 11 } } } },
                scales: {
                    y: { type: 'linear', position: 'left', title: { display: true, text: '\u041f\u043e\u043a\u0430\u0437\u044b / \u041a\u043b\u0438\u043a\u0438 / \u0417\u0430\u043a\u0430\u0437\u044b' } },
                    y1: { type: 'linear', position: 'right', title: { display: true, text: '\u0420\u0430\u0441\u0445\u043e\u0434 \u20bd' }, grid: { drawOnChartArea: false } }
                }
            }
        });
        window._mktCharts.push(cS);
    }

    // Per-campaign charts
    var campaigns = d.campaigns || [];
    campaigns.forEach(function(camp, idx) {
        var ctx = document.getElementById('mkt-chart-rk-' + idx);
        if (!ctx) return;
        var daily = (camp.daily || []).sort(function(a,b) { return a.date < b.date ? -1 : 1; });
        if (daily.length < 2) return;
        var campLabels = daily.map(function(dd) { return dd.date.substring(5); });
        var c = new Chart(ctx.getContext('2d'), {
            type: 'line',
            data: {
                labels: campLabels,
                datasets: [
                    { label: '\u041f\u043e\u043a\u0430\u0437\u044b', data: daily.map(function(dd) { return dd.views; }), borderColor: '#6c5ce7', backgroundColor: 'rgba(108,92,231,0.1)', yAxisID: 'y', tension: 0.2, pointRadius: 3 },
                    { label: '\u041a\u043b\u0438\u043a\u0438', data: daily.map(function(dd) { return dd.clicks; }), borderColor: '#00b894', backgroundColor: 'rgba(0,184,148,0.1)', yAxisID: 'y', tension: 0.2, pointRadius: 3 },
                    { label: '\u0417\u0430\u043a\u0430\u0437\u044b', data: daily.map(function(dd) { return dd.orders; }), borderColor: '#fdcb6e', backgroundColor: 'rgba(253,203,110,0.1)', yAxisID: 'y', tension: 0.2, pointRadius: 3 },
                    { label: '\u0420\u0430\u0441\u0445\u043e\u0434 \u20bd', data: daily.map(function(dd) { return dd.spent; }), borderColor: '#e17055', backgroundColor: 'rgba(225,112,85,0.1)', yAxisID: 'y1', tension: 0.2, pointRadius: 3 },
                    { label: 'CTR %', data: daily.map(function(dd) { return dd.ctr; }), borderColor: '#e84393', backgroundColor: 'rgba(232,67,147,0.1)', yAxisID: 'y2', tension: 0.2, pointRadius: 3, borderDash: [5,5] }
                ]
            },
            options: {
                responsive: true, interaction: { mode: 'index', intersect: false },
                plugins: { legend: { position: 'top', labels: { font: { size: 11 } } } },
                scales: {
                    y: { type: 'linear', position: 'left', title: { display: true, text: '\u041f\u043e\u043a\u0430\u0437\u044b / \u041a\u043b\u0438\u043a\u0438' } },
                    y1: { type: 'linear', position: 'right', title: { display: true, text: '\u0420\u0430\u0441\u0445\u043e\u0434 \u20bd' }, grid: { drawOnChartArea: false } },
                    y2: { type: 'linear', position: 'right', display: false }
                }
            }
        });
        window._mktCharts.push(c);
    });
}


// ===== EXTERNAL ADS =====
var extAdData = [];
var extSelectedIds = new Set();

async function loadExtAds() {
    try {
        var params = new URLSearchParams({org_id: ORG_ID});
        var tp = document.getElementById('ext-type').value;
        var src = document.getElementById('ext-source').value;
        var df = document.getElementById('ext-date-from').value;
        var dt = document.getElementById('ext-date-to').value;
        var sr = document.getElementById('ext-search').value;
        if (tp) params.set('ad_type', tp);
        if (src) params.set('source', src);
        if (df) params.set('date_from', df);
        if (dt) params.set('date_to', dt);
        if (sr) params.set('search', sr);
        var r = await fetch('/api/v1/nl/external-ads?' + params, {headers:{'Authorization':'Bearer '+TOKEN}});
        extAdData = await r.json();
        renderExtAds();
        loadExtSources();
    } catch(e) { console.error('loadExtAds error:', e); }
}

async function loadExtSources() {
    try {
        var r = await fetch('/api/v1/nl/external-ads/sources/list?org_id=' + ORG_ID, {headers:{'Authorization':'Bearer '+TOKEN}});
        var sources = await r.json();
        var sel = document.getElementById('ext-source');
        var cur = sel.value;
        sel.innerHTML = '<option value="">Все источники</option>' + sources.map(s => '<option value="'+s+'"'+(s===cur?' selected':'')+'>'+s+'</option>').join('');
    } catch(e) {}
}

function renderExtAds() {
    var body = document.getElementById('ext-body');
    if (!extAdData.length) {
        body.innerHTML = '<tr><td colspan="20" class="empty">Нет записей. Нажмите "Добавить"</td></tr>';
    } else {
        body.innerHTML = extAdData.map(function(r) {
            var photo = r.photo_url ? '<img src="'+r.photo_url+'" style="width:40px;height:40px;object-fit:cover;border-radius:4px">' : '—';
            var cardLink = r.card_url ? '<a href="'+r.card_url+'" target="_blank" style="color:#6c5ce7;font-size:.9em">🔗</a>' : '—';
            var subLink = r.substitution_url ? '<a href="'+r.substitution_url+'" target="_blank" style="color:#0984e3;font-size:.9em" title="'+r.substitution_url+'">🔗</a>' : '—';
            var utmLink = r.utm_url ? '<a href="'+r.utm_url+'" target="_blank" style="color:#00b894;font-size:.9em" title="'+r.utm_url+'">🔗</a>' : '—';
            var typeLabel = r.ad_type === 'buyout' ? '<span style="background:#fdcb6e;color:#333;padding:2px 8px;border-radius:10px;font-size:.75em">Самовыкуп</span>' : '<span style="background:#dfe6e9;color:#333;padding:2px 8px;border-radius:10px;font-size:.75em">Реклама</span>';
            var checked = extSelectedIds.has(r.id) ? 'checked' : '';
            return '<tr data-id="'+r.id+'">'
                + '<td style="position:sticky;left:0;background:#fff"><input type="checkbox" '+checked+' onclick="toggleExtRow(r.id,this.checked)"></td>'
                + '<td>'+photo+'</td>'
                + '<td>'+(r.article||'—')+'</td>'
                + '<td>'+(r.nm_id||'—')+'</td>'
                + '<td>'+(r.vendor_code||'—')+'</td>'
                + '<td style="max-width:150px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="'+(r.product_name||'')+'">'+(r.product_name||'—')+'</td>'
                + '<td>'+cardLink+'</td>'
                + '<td>'+subLink+'</td>'
                + '<td>'+utmLink+'</td>'
                + '<td>'+(r.source||'—')+'</td>'
                + '<td>'+(r.query||'—')+'</td>'
                + '<td>'+(r.ad_date||'—')+'</td>'
                + '<td class="r">'+(r.reach!=null?r.reach.toLocaleString('ru-RU'):'—')+'</td>'
                + '<td class="r">'+(r.amount!=null?Number(r.amount).toLocaleString('ru-RU',{maximumFractionDigits:2})+' ₽':'—')+'</td>'
                + '<td class="r">'+(r.orders_count||'—')+'</td>'
                + '<td class="r">'+(r.orders_avg_weekly!=null?Number(r.orders_avg_weekly).toFixed(1):'—')+'</td>'
                + '<td>'+typeLabel+'</td>'
                + '<td style="max-width:120px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="'+(r.notes||'')+'">'+(r.notes||'—')+'</td>'
                + '<td><button onclick="editExtAd(r.id)" style="background:none;border:none;cursor:pointer;font-size:1em" title="Редактировать">✏️</button>'
                + '<button onclick="deleteExtAd(r.id)" style="background:none;border:none;cursor:pointer;font-size:1em" title="Удалить">🗑️</button></td>'
                + '</tr>';
        }).join('');
    }
    // Stats
    var total = extAdData.length;
    var totalAmount = extAdData.reduce(function(s,r){return s+(r.amount?Number(r.amount):0)},0);
    var totalOrders = extAdData.reduce(function(s,r){return s+(r.orders_count||0)},0);
    var totalReach = extAdData.reduce(function(s,r){return s+(r.reach||0)},0);
    document.getElementById('ext-total-count').textContent = total;
    document.getElementById('ext-total-amount').textContent = totalAmount.toLocaleString('ru-RU',{maximumFractionDigits:0}) + ' ₽';
    document.getElementById('ext-total-orders').textContent = totalOrders.toLocaleString('ru-RU');
    document.getElementById('ext-total-reach').textContent = totalReach.toLocaleString('ru-RU');
}

function showExtAdModal(id) {
    document.getElementById('ext-edit-id').value = '';
    document.getElementById('ext-modal-title').textContent = 'Добавить запись';
    ['ext-nm-id','ext-vendor-code','ext-article','ext-source-input','ext-query','ext-ad-date','ext-reach','ext-amount','ext-orders-count','ext-orders-avg','ext-photo-url','ext-sub-url','ext-utm-url','ext-notes'].forEach(function(id){document.getElementById(id).value=''});
    document.getElementById('ext-ad-type').value = 'ad';
    document.getElementById('ext-modal').style.display = 'flex';
}

function closeExtAdModal() {
    document.getElementById('ext-modal').style.display = 'none';
}

function editExtAd(id) {
    var item = extAdData.find(function(r){return r.id===id});
    if (!item) return;
    document.getElementById('ext-edit-id').value = id;
    document.getElementById('ext-modal-title').textContent = 'Редактировать';
    document.getElementById('ext-nm-id').value = item.nm_id || '';
    document.getElementById('ext-vendor-code').value = item.vendor_code || '';
    document.getElementById('ext-article').value = item.article || '';
    document.getElementById('ext-ad-type').value = item.ad_type || 'ad';
    document.getElementById('ext-source-input').value = item.source || '';
    document.getElementById('ext-query').value = item.query || '';
    document.getElementById('ext-ad-date').value = item.ad_date || '';
    document.getElementById('ext-reach').value = item.reach || '';
    document.getElementById('ext-amount').value = item.amount || '';
    document.getElementById('ext-orders-count').value = item.orders_count || '';
    document.getElementById('ext-orders-avg').value = item.orders_avg_weekly || '';
    document.getElementById('ext-photo-url').value = item.photo_url || '';
    document.getElementById('ext-sub-url').value = item.substitution_url || '';
    document.getElementById('ext-utm-url').value = item.utm_url || '';
    document.getElementById('ext-notes').value = item.notes || '';
    document.getElementById('ext-modal').style.display = 'flex';
}

async function saveExtAd() {
    var editId = document.getElementById('ext-edit-id').value;
    var payload = {
        nm_id: document.getElementById('ext-nm-id').value || null,
        vendor_code: document.getElementById('ext-vendor-code').value || null,
        article: document.getElementById('ext-article').value || null,
        ad_type: document.getElementById('ext-ad-type').value || 'ad',
        source: document.getElementById('ext-source-input').value || null,
        query: document.getElementById('ext-query').value || null,
        ad_date: document.getElementById('ext-ad-date').value || null,
        reach: document.getElementById('ext-reach').value || null,
        amount: document.getElementById('ext-amount').value || null,
        orders_count: document.getElementById('ext-orders-count').value || null,
        orders_avg_weekly: document.getElementById('ext-orders-avg').value || null,
        photo_url: document.getElementById('ext-photo-url').value || null,
        substitution_url: document.getElementById('ext-sub-url').value || null,
        utm_url: document.getElementById('ext-utm-url').value || null,
        notes: document.getElementById('ext-notes').value || null,
    };
    if (payload.nm_id) payload.nm_id = parseInt(payload.nm_id);
    if (payload.reach) payload.reach = parseInt(payload.reach);
    if (payload.orders_count) payload.orders_count = parseInt(payload.orders_count);
    if (payload.amount) payload.amount = parseFloat(payload.amount);
    if (payload.orders_avg_weekly) payload.orders_avg_weekly = parseFloat(payload.orders_avg_weekly);

    try {
        var url, method;
        if (editId) {
            url = '/api/v1/nl/external-ads/' + editId + '?org_id=' + ORG_ID;
            method = 'PUT';
        } else {
            url = '/api/v1/nl/external-ads?org_id=' + ORG_ID;
            method = 'POST';
        }
        var r = await fetch(url, {
            method: method,
            headers: {'Content-Type':'application/json','Authorization':'Bearer '+TOKEN},
            body: JSON.stringify(payload)
        });
        if (!r.ok) { var err = await r.json(); throw new Error(err.detail || 'Ошибка сохранения'); }
        closeExtAdModal();
        loadExtAds();
    } catch(e) { showToast('❌ Ошибка: ' + e.message, 'error'); }
}

async function deleteExtAd(id) {
    if (!confirm('Удалить запись?')) return;
    try {
        await fetch('/api/v1/nl/external-ads/' + id + '?org_id=' + ORG_ID, {
            method: 'DELETE',
            headers: {'Authorization':'Bearer '+TOKEN}
        });
        loadExtAds();
    } catch(e) { alert('Ошибка удаления: ' + e.message); }
}

function extAutoFill() {
    // Автозаполнение фото при вводе nm_id
    var nmId = document.getElementById('ext-nm-id').value;
    if (nmId && extAdData.length) {
        var found = extAdData.find(function(r){ return r.nm_id == nmId; });
        if (found && found.photo_url && !document.getElementById('ext-photo-url').value) {
            document.getElementById('ext-photo-url').value = found.photo_url;
        }
    }
}

function toggleExtRow(id, checked) {
    if (checked) extSelectedIds.add(id); else extSelectedIds.delete(id);
    updateExtBulkBar();
}

function toggleAllExtRows(checked) {
    extSelectedIds.clear();
    if (checked) extAdData.forEach(function(r){ extSelectedIds.add(r.id); });
    renderExtAds();
    updateExtBulkBar();
}

function updateExtBulkBar() {
    var bar = document.getElementById('ext-bulk-bar');
    var cnt = document.getElementById('ext-bulk-count');
    if (extSelectedIds.size > 0) {
        bar.style.display = 'flex';
        cnt.textContent = 'Выделено: ' + extSelectedIds.size;
    } else {
        bar.style.display = 'none';
    }
}

function clearExtBulkSelection() {
    extSelectedIds.clear();
    document.getElementById('ext-check-all').checked = false;
    renderExtAds();
    updateExtBulkBar();
}

async function applyExtBulkEdit() {
    var field = document.getElementById('ext-bulk-field').value;
    var value = document.getElementById('ext-bulk-value').value;
    if (!field || extSelectedIds.size === 0) return;
    try {
        await fetch('/api/v1/nl/external-ads/bulk-update?org_id=' + ORG_ID, {
            method: 'POST',
            headers: {'Content-Type':'application/json','Authorization':'Bearer '+TOKEN},
            body: JSON.stringify({ids: Array.from(extSelectedIds), updates: {[field]: value}})
        });
        clearExtBulkSelection();
        loadExtAds();
    } catch(e) { showToast('❌ Ошибка: ' + e.message, 'error'); }
}


async function loadDates() {
    if (!ORG_ID) return [];
    const res = await fetch('/api/v1/nl/dates?org_id=' + ORG_ID);
    if (!res.ok) return [];
    const dates = await res.json();
    // Fill both date selectors
    // Заполнить периоды для плана продаж
    var spPeriodSel = document.getElementById('sp-period');
    if (spPeriodSel && spPeriodSel.tagName === 'SELECT') {
        spPeriodSel.innerHTML = '';
        // Генерируем последние 12 месяцев
        var now = new Date();
        for (var m = 0; m < 12; m++) {
            var d = new Date(now.getFullYear(), now.getMonth() - m, 1);
            var val = d.toISOString().substring(0,10);
            var label = d.toLocaleDateString('ru-RU', {month:'long', year:'numeric'});
            var opt = document.createElement('option');
            opt.value = val;
            opt.textContent = label.charAt(0).toUpperCase() + label.slice(1);
            spPeriodSel.appendChild(opt);
        }
    }
    ['ref-date', 'stats-date', 'analytics-date', 'wh-date', 'opiu-period'].forEach(id => {
        const sel = document.getElementById(id);
        if (!sel || sel.tagName !== 'SELECT') return;
        // Keep period selects
        if (id === 'opiu-period') return;
        sel.innerHTML = '';
        if (!dates.length) { sel.innerHTML = '<option>Нет данных</option>'; return; }
        dates.forEach(d => {
            const opt = document.createElement('option');
            opt.value = d;
            const dt = new Date(d + 'T00:00:00');
            opt.textContent = dt.toLocaleDateString('ru-RU', {day:'numeric', month:'short', year:'numeric'});
            sel.appendChild(opt);
        });
    });
    return dates[0];
}

async function loadAnalytics() {
    if (!ORG_ID) return;
    const sel = document.getElementById('analytics-date') || document.getElementById('ref-date');
    const dateVal = sel ? sel.value : '';
    if (!dateVal || dateVal === 'Нет данных') return;
    const search = document.getElementById('analytics-search')?.value || '';
    try {
        const res = await fetch('/api/v1/nl/analytics?org_id=' + ORG_ID + '&target_date=' + dateVal + (search ? '&search=' + encodeURIComponent(search) : ''));
        if (!res.ok) { document.getElementById('analytics-body').innerHTML = '<tr><td colspan="37" class="empty">Ошибка загрузки</td></tr>'; return; }
        const data = await res.json();
        const prods = data.products || [];
        document.getElementById('analytics-count').textContent = prods.length + ' товаров';
        if (!prods.length) { document.getElementById('analytics-body').innerHTML = '<tr><td colspan="37" class="empty">Нет данных</td></tr>'; return; }
        const fmt = (v, s) => { if (v == null) return '—'; return Number(v).toLocaleString('ru-RU', {maximumFractionDigits:2}) + (s || ''); };
        const size = parseInt(document.getElementById('analytics-pagesize')?.value || '25');
        document.getElementById('analytics-body').innerHTML = prods.slice(0, size).map(p => {
            const thumb = (p.photo_main || '').replace('/hq/', '/c246x328/').replace('/big/', '/c246x328/').replace('/tm/', '/c246x328/');
            return '<tr>' +
                '<td>' + (thumb ? '<img src="' + thumb + '" style="width:32px;height:32px;border-radius:4px;object-fit:cover">' : '') + '</td>' +
            '<td>' + esc(p.vendor_code || '') + '</td>' +
            '<td>' + (p.nm_id || '') + '</td>' +
            '<td style="max-width:150px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">' + esc(p.product_name || '') + '</td>' +
            '<td>' + (p.stock_qty ?? '—') + '</td>' +
            '<td>' + (p.orders_count ?? '—') + '</td>' +
            '<td>' + (p.buyouts_count ?? '—') + '</td>' +
            '<td>—</td>' +
            '<td>' + (p.returns_count ?? '—') + '</td>' +
            '<td>' + (p.buyout_percent || '—') + '%</td>' +
            '<td>' + fmt(p.price) + '</td>' +
            '<td>' + fmt(p.price_discount) + '</td>' +
            '<td>' + (p.tariff_percent || '—') + '%</td>' +
            '<td>' + fmt(p.commission) + '</td>' +
            '<td>' + fmt(p.logistics) + '</td>' +
            '<td>' + fmt(p.ad_cost) + '</td>' +
            '<td>' + (p.drr || '—') + '%</td>' +
            '<td>' + fmt(p.fines) + '</td>' +
            '<td>' + fmt(p.storage) + '</td>' +
            '<td>' + fmt(p.reception) + '</td>' +
            '<td>' + fmt(p.other_deductions) + '</td>' +
            '<td>' + fmt(p.avg_check) + '</td>' +
            '<td>' + fmt(p.revenue) + '</td>' +
            '<td>' + fmt(p.payout) + '</td>' +
            '<td>' + fmt(p.cost_price) + '</td>' +
            '<td>' + fmt(p.margin) + '</td>' +
            '<td>' + fmt(p.margin_per_unit) + '</td>' +
            '<td>' + (p.profitability || '—') + '%</td>' +
            '<td>' + (p.roi || '—') + '%</td>' +
            '<td>' + (p.rating || '—') + '</td>' +
            '<td>' + (p.impressions ?? '—') + '</td>' +
            '<td>' + (p.clicks ?? '—') + '</td>' +
            '<td>' + (p.ctr || '—') + '%</td>' +
            '<td>' + (p.turnover || '—') + '</td>' +
            '<td>' + (p.in_transit || '—') + '</td>' +
            '</tr>';
        }).join('');
    } catch(e) { console.error('loadAnalytics error:', e); }
}

async function loadOpiu() {
    if (!ORG_ID) return;
    const period = document.getElementById('opiu-period')?.value || '4';
    try {
        const res = await fetch('/api/v1/nl/opiu?org_id=' + ORG_ID + '&period=' + period);
        if (!res.ok) return;
        const data = await res.json();
        const weeks = data.weeks || [];
        const total = data.total || {};
        const fmt = (v) => v != null ? Number(v).toLocaleString('ru-RU', {maximumFractionDigits:0}) : '—';
        
        // Строим таблицу
        const thead = document.querySelector('#opiu-table thead tr');
        thead.innerHTML = '<th>Статья</th><th>Итого ₽</th><th>Итого %</th>';
        weeks.forEach(w => { thead.innerHTML += '<th>' + w.label + ' ₽</th><th>%</th>'; });
        
        const tbody = document.getElementById('opiu-body');
        const rows = [
            {name: 'Реализация', key: 'revenue'},
            {name: 'Продажи', key: 'buyouts'},
            {name: 'Возвраты', key: 'returns'},
            {name: 'Расходы', key: 'ad_cost'},
            {name: 'Комиссия ВБ', key: null},
            {name: 'Реклама', key: 'ad_cost'},
            {name: 'Логистика', key: null},
        ];
        
        const totalRev = total.revenue || 1;
        let html = '';
        rows.forEach(r => {
            const t = r.key ? total[r.key] : 0;
            const tp = totalRev ? (t / totalRev * 100).toFixed(1) : '—';
            html += '<tr><td><strong>' + r.name + '</strong></td><td>' + fmt(t) + '</td><td>' + tp + '%</td>';
            weeks.forEach(w => {
                const v = r.key ? w[r.key] : 0;
                const p = totalRev ? (v / totalRev * 100).toFixed(1) : '—';
                html += '<td>' + fmt(v) + '</td><td>' + p + '%</td>';
            });
            html += '</tr>';
        });
        
        // Чистая прибыль
        const profit = (total.revenue || 0) - (total.ad_cost || 0);
        const profitP = totalRev ? (profit / totalRev * 100).toFixed(1) : '—';
        html += '<tr style="font-weight:700;background:#f0edfc"><td>Чистая прибыль</td><td>' + fmt(profit) + '</td><td>' + profitP + '%</td>';
        weeks.forEach(w => {
            const p = (w.revenue || 0) - (w.ad_cost || 0);
            const pp = totalRev ? (p / totalRev * 100).toFixed(1) : '—';
            html += '<td>' + fmt(p) + '</td><td>' + pp + '%</td>';
        });
        html += '</tr>';
        
        tbody.innerHTML = html || '<tr><td colspan="3" class="empty">Нет данных</td></tr>';
    } catch(e) { console.error('loadOpiu error:', e); }
}

async function exportOpiu() { alert('Экспорт ОПиУ в разработке'); }

async function loadOpiu() {
    if (!ORG_ID) return;
    const period = document.getElementById('opiu-period')?.value || '4';
    try {
        const res = await fetch('/api/v1/nl/opiu?org_id=' + ORG_ID + '&period=' + period);
        if (!res.ok) return;
        const data = await res.json();
        const weeks = data.weeks || [];
        const total = data.total || {};
        const fmt = (v) => v != null ? Number(v).toLocaleString('ru-RU', {maximumFractionDigits:0}) : '—';
        const thead = document.querySelector('#opiu-table thead tr');
        thead.innerHTML = '<th>Статья</th><th>Итого</th><th>%</th>';
        weeks.forEach(w => { thead.innerHTML += '<th>' + w.label + '</th>'; });
        const tbody = document.getElementById('opiu-body');
        const tr = total.revenue || 1;
        const rows = [
            ['Реализация', total.revenue], ['Продажи (шт)', total.buyouts],
            ['Возвраты (шт)', total.returns], ['Реклама', total.ad_cost],
        ];
        let html = '';
        rows.forEach(([name, val]) => {
            html += '<tr><td><strong>' + name + '</strong></td><td>' + fmt(val) + '</td><td>' + (tr ? (val/tr*100).toFixed(1) : '—') + '%</td>';
            weeks.forEach(w => { const v = name.includes('Реал') ? w.revenue : name.includes('Прод') ? w.buyouts : name.includes('Воз') ? w.returns : w.ad_cost; html += '<td>' + fmt(v) + '</td>'; });
            html += '</tr>';
        });
        const profit = (total.revenue || 0) - (total.ad_cost || 0);
        html += '<tr style="font-weight:700;background:#f0edfc"><td>Чистая прибыль</td><td>' + fmt(profit) + '</td><td>' + (tr ? (profit/tr*100).toFixed(1) : '—') + '%</td>';
        weeks.forEach(w => { const p = (w.revenue||0)-(w.ad_cost||0); html += '<td>' + fmt(p) + '</td>'; });
        html += '</tr>';
        tbody.innerHTML = html || '<tr><td colspan="3" class="empty">Нет данных</td></tr>';
    } catch(e) { console.error('loadOpiu', e); }
}
async function exportOpiu() { alert('В разработке'); }

// Глобальный кэш для фильтрации
let _costProducts = [];
let _costMap = {};

function calcPlanVol(el) {
  var row = el.closest('tr');
  if (!row) return;
  var l = parseFloat(row.querySelector('[data-field="plan_length"]')?.value) || 0;
  var w = parseFloat(row.querySelector('[data-field="plan_width"]')?.value) || 0;
  var h = parseFloat(row.querySelector('[data-field="plan_height"]')?.value) || 0;
  var volCell = row.querySelector('.plan-vol-cell');
  if (volCell) {
    if (l > 0 && w > 0 && h > 0) {
      volCell.textContent = ((l * w * h) / 1000);
    } else {
      volCell.innerHTML = String.fromCharCode(8212);
    }
  }
}


// === Налоговые настройки кабинета ===
async function loadTaxSettings() {
    if (!ORG_ID) return;
    try {
        const r = await fetch('/api/v1/nl/tax-settings?org_id=' + ORG_ID);
        if (r.ok) {
            _taxSettings = await r.json();
            document.getElementById('tax-system-select').value = _taxSettings.tax_system || '';
            document.getElementById('tax-rate-input').value = _taxSettings.tax_rate != null ? _taxSettings.tax_rate : '';
            document.getElementById('vat-type-select').value = _taxSettings.vat_type || 'нет';
            updateTaxColHeader();
            // Restore lock state from localStorage
            const lockKey = 'tax_locked_' + ORG_ID;
            const isLocked = localStorage.getItem(lockKey);
            if (isLocked === 'true') {
                setTaxLock(true);
            }
        }
    } catch(e) { console.error('loadTaxSettings', e); }
}

function onTaxSystemChange() {
    updateTaxColHeader();
}

function onTaxSettingsChange() {
    updateTaxColHeader();
}

function updateTaxColHeader() {
    const ts = document.getElementById('tax-system-select').value;
    const tr = document.getElementById('tax-rate-input').value;
    const sub = document.getElementById('tax-col-subheader');
    if (sub) sub.textContent = ts ? ts + (tr ? ' ' + tr + '%' : '') : '—';
}

async function saveTaxSettings() {
    if (!ORG_ID) return;
    const data = {
        tax_system: document.getElementById('tax-system-select').value,
        tax_rate: document.getElementById('tax-rate-input').value || null,
        vat_type: document.getElementById('vat-type-select').value || 'нет'
    };
    try {
        const r = await fetch('/api/v1/nl/tax-settings?org_id=' + ORG_ID, {
            method: 'POST',
            headers: {'Content-Type': 'application/json'},
            body: JSON.stringify(data)
        });
        if (r.ok) {
            _taxSettings = await r.json();
            updateTaxColHeader();
            // Save to localStorage
            localStorage.setItem('tax_settings_' + ORG_ID, JSON.stringify(_taxSettings));
            // Auto-lock after save
            setTaxLock(true);
            localStorage.setItem('tax_locked_' + ORG_ID, 'true');
            // Update grid with new tax values (only rows without manual override)
            if (costTabulator) {
                const rows = costTabulator.getData();
                costTabulator.updateData(rows.filter(r => !r._tax_rate_override).map(r => ({
                    _id: r._id,
                    _tax_rate_override: ''
                })));
                costTabulator.redraw(true);
            }
            // showToast handled by caller
        } else {
            const err = await r.json();
            showToast('Ошибка: ' + (err.detail || r.status), 'error');
        }
    } catch(e) {
        console.error('saveTaxSettings', e);
        showToast('Ошибка сохранения', 'error');
    }
}


function toggleTaxLock() {
    const lockKey = 'tax_locked_' + ORG_ID;
    const isLocked = localStorage.getItem(lockKey) === 'true';
    if (isLocked) {
        setTaxLock(false);
        localStorage.setItem(lockKey, 'false');
    } else {
        setTaxLock(true);
        localStorage.setItem(lockKey, 'true');
    }
}

function setTaxLock(locked) {
    var btn = document.getElementById('tax-lock-btn');
    var sel = document.getElementById('tax-system-select');
    var inp = document.getElementById('tax-rate-input');
    var vat = document.getElementById('vat-type-select');
    var excelBtn = document.getElementById('cost-file-input');
    if (locked) {
        btn.textContent = '\U0001f512';
        btn.style.borderColor = '#6c5ce7';
        btn.style.background = '#f0edff';
        sel.disabled = true;
        inp.disabled = true;
        vat.disabled = true;
        _taxLocked = true;
        if (costTabulator) {
            costTabulator.updateColumnDefinition('_tax_rate_override', {editable:false});
            costTabulator.updateColumnDefinition('vat_rate', {editable:false});
        }
        if (excelBtn) excelBtn.disabled = true;
    } else {
        btn.textContent = '\U0001f513';
        btn.style.borderColor = '#ddd';
        btn.style.background = 'none';
        sel.disabled = false;
        inp.disabled = false;
        vat.disabled = false;
        _taxLocked = false;
        if (costTabulator) {
            costTabulator.updateColumnDefinition('_tax_rate_override', {editable:true});
            costTabulator.updateColumnDefinition('vat_rate', {editable:true});
        }
        if (excelBtn) excelBtn.disabled = false;
    }
}

var _taxLocked = false;

function applyTaxToAll() {
    if (!costTabulator) return;
    var taxRate = parseFloat(document.getElementById('tax-rate-input').value) || 0;
    var vatSel = document.getElementById('vat-type-select').value;
    var vatNum = vatSel === '5%' ? 5 : vatSel === '7%' ? 7 : 0;
    var rows = costTabulator.getData();
    costTabulator.updateData(rows.map(function(r) {
        return { _id: r._id, _tax_rate_override: taxRate, vat_rate: vatNum };
    }));
    showToast('\u041d\u0430\u043b\u043e\u0433 ' + taxRate + '% \u0438 \u041d\u0414\u0421 ' + (vatNum ? vatNum + '%' : '\u043d\u0435\u0442') + ' \u043f\u0440\u0438\u043c\u0435\u043d\u0435\u043d\u044b \u043a\u043e \u0432\u0441\u0435\u043c');
}

function showToast(msg, type) {
    var el = document.createElement('div');
    el.textContent = msg;
    el.style.cssText = 'position:fixed;bottom:20px;right:20px;padding:12px 20px;border-radius:8px;color:#fff;font-size:.9em;z-index:9999;animation:fadeInUp .3s;' + (type==='error'?'background:#e74c3c':'background:#00b894');
    document.body.appendChild(el);
    setTimeout(function(){ el.style.opacity='0'; el.style.transition='opacity .3s'; setTimeout(function(){ el.remove(); }, 300); }, 2500);
}

async function loadCostPrices() {
    console.log("[CP] ORG_ID=", ORG_ID);
    loadFbsWarehouses();
    if (!ORG_ID) { console.warn("[CP] No ORG_ID"); return; }
    try {
        const datesRes = await fetch('/api/v1/nl/dates?org_id=' + ORG_ID);
        const dates = datesRes.ok ? await datesRes.json() : [];
        if (!dates.length) { document.getElementById('cost-count').textContent = 'Нет данных'; return; }
        
        const prodsRes = await fetch('/api/v1/nl/control?org_id=' + ORG_ID + '&target_date=' + dates[0]);
        if (!prodsRes.ok) return;
        const prodsData = await prodsRes.json();
        _costProducts = prodsData.products || [];
        console.log("[CP] products:", _costProducts.length, "dates:", dates.length);
        
        const costRes = await fetch('/api/v1/nl/cost-prices?org_id=' + ORG_ID);
        _costMap = {};  // entity_id -> cost data
        _costSizes = {};
        if (costRes.ok) {
            const costs = await costRes.json();
            costs.forEach(c => {
                _costMap[c.entity_id] = c;  // ключ = entity_id
            });
        }
        
        // Заполнить фильтры-дропдауны уникальными значениями
        populateCostFilterOptions();
        console.log("[CP] costMap:", Object.keys(_costMap).length);
        applyCostFilters();
    } catch(e) { console.error('loadCostPrices', e); }
}

function populateCostFilterOptions() {
    const classes = new Set(), brands = new Set(), statuses = new Set();
    _costProducts.forEach(p => {
        const c = _costMap[p.entity_id] || {};
        if (c.product_class) classes.add(c.product_class);
        if (c.brand) brands.add(c.brand);
        if (c.product_status) statuses.add(c.product_status);
    });
    const fillSel = (id, vals) => {
        const el = document.getElementById(id);
        const cur = el.value;
        const opts = ['<option value="">' + el.options[0].text + '</option>'];
        [...vals].sort().forEach(v => opts.push('<option value="' + esc(v) + '">' + esc(v) + '</option>'));
        el.innerHTML = opts.join('');
        el.value = cur;
    };
    fillSel('flt-product-class', classes);
    fillSel('flt-brand', brands);
    // product_status filter is now static with predefined values
}

function clearCostFilters() {
    document.getElementById('flt-fulfillment').value = '';
    document.getElementById('flt-tax-system').value = '';
    document.getElementById('flt-product-class').value = '';
    document.getElementById('flt-brand').value = '';
    document.getElementById('flt-product-status').value = '';
    document.getElementById('flt-has-cost').value = '';
    document.getElementById('cost-search').value = '';
    applyCostFilters();
}

function applyCostFilters() {
    const search = (document.getElementById('cost-search')?.value || '').toLowerCase();
    const fltFF = document.getElementById('flt-fulfillment')?.value || '';
    const fltTax = document.getElementById('flt-tax-system')?.value || '';
    const fltClass = document.getElementById('flt-product-class')?.value || '';
    const fltBrand = document.getElementById('flt-brand')?.value || '';
    const fltStatus = document.getElementById('flt-product-status')?.value || '';
    const fltCost = document.getElementById('flt-has-cost')?.value || '';
    
    let products = _costProducts;
    
    if (search) {
        products = products.filter(p => 
            (p.product_name||'').toLowerCase().includes(search) || 
            String(p.nm_id).includes(search) || 
            (p.vendor_code||'').toLowerCase().includes(search) ||
            (p.barcodes||'').includes(search)
        );
    }
    
    const checkMatch = (c, field, filterVal) => {
        if (!filterVal) return true;
        return (c[field] || '') === filterVal;
    };
    
    products = products.filter(p => {
        const c = _costMap[p.entity_id] || {};
        if (fltFF && !checkMatch(c, 'fulfillment_model', fltFF)) return false;
        if (fltTax && !checkMatch(c, 'tax_system', fltTax)) return false;
        if (fltClass && !checkMatch(c, 'product_class', fltClass)) return false;
        if (fltBrand && !checkMatch(c, 'brand', fltBrand)) return false;
        if (fltStatus && !checkMatch(c, 'product_status', fltStatus)) return false;
        if (fltCost === 'yes' && !c.cost_price) return false;
        if (fltCost === 'no' && c.cost_price) return false;
        return true;
    });
        
        document.getElementById('cost-count').textContent = products.length + ' товаров';
        
        // Группируем по nm_id
        const groups = {};
        const order = [];
        products.forEach(p => {
            if (!groups[p.nm_id]) { groups[p.nm_id] = []; order.push(p.nm_id); }
            groups[p.nm_id].push(p);
        });
        
        // === TABULATOR: если загружен, рендерим через него ===
        if (typeof Tabulator !== "undefined" && typeof updateCostTabulator === "function") {
            updateCostTabulator(products);
            // Итоги
            let tc = 0, fc = 0;
            products.forEach(p => {
                const c = _costMap[p.entity_id] || {};
                if (c.cost_price) { tc += parseFloat(c.cost_price); fc++; }
            });
            document.getElementById("cost-summary").innerHTML =
                "<span>💰 Заполнено: <strong>" + fc + "/" + products.length + "</strong></span>" +
                "<span>📊 Сумма себестоимости: <strong>" + tc.toLocaleString("ru-RU") + " ₽</strong></span>" +
                (fc > 0 ? "<span>📄 Средняя: <strong>" + Math.round(tc/fc).toLocaleString("ru-RU") + " ₽</strong></span>" : "");
            return;
        }
        // === ОШИБКА: Tabulator не загружен ===
        var reasons = [];
        if (typeof Tabulator === "undefined") reasons.push("Tabulator CDN не загружен");
        if (typeof updateCostTabulator === "undefined") reasons.push("cost-grid.js не загружен");
        var errMsg = "⚠️ Ошибка загрузки таблицы: " + (reasons.length ? reasons.join(", ") : "неизвестная ошибка") + ". Обновите страницу (Ctrl+F5).";
        var errDiv = document.getElementById('cost-error-msg');
        if (!errDiv) {
            errDiv = document.createElement('div');
            errDiv.id = 'cost-error-msg';
            errDiv.style.cssText = 'padding:20px;text-align:center;color:#e17055;font-size:1.1em;background:#fff5f5;border:1px solid #e17055;border-radius:8px;margin:10px 0';
            var wrapper = document.getElementById('cost-table-wrapper');
            if (wrapper) { wrapper.style.display = 'none'; wrapper.parentNode.appendChild(errDiv); }
            else { document.querySelector('.main-content')?.appendChild(errDiv); }
        }
        errDiv.textContent = errMsg;
        document.getElementById('cost-count').textContent = '0 товаров';
        document.getElementById('cost-summary').innerHTML = '';
}

// Автоудаление зависших попапов
function cleanupCostPopups() {
    ['cost-info-popup','extra-costs-info-popup','total-cost-info-popup'].forEach(function(id) {
        var el = document.getElementById(id);
        if (el) el.remove();
    });
}
document.addEventListener('DOMContentLoaded', cleanupCostPopups);

function showCostInfo(el) {
    var existing = document.getElementById('cost-info-popup');
    if (existing) { existing.remove(); return; }
    var popup = document.createElement('div');
    popup.id = 'cost-info-popup';
    popup.style.cssText = 'position:fixed;top:50%;left:50%;transform:translate(-50%,-50%);background:#fff;border:1px solid #ddd;border-radius:8px;padding:16px 20px;box-shadow:0 4px 24px rgba(0,0,0,.18);z-index:9999;max-width:400px;font:6px Arial;line-height:1.6;color:#333';
    popup.innerHTML = '<div style="font-size:7px;font-weight:600;margin-bottom:6px;color:#6c5ce7">ℹ️ Себестоимость</div>' +
        '<div style="font-size:6px">В колонке себестоимости указывается закупочная цена + все затраты на данный товар, напр, доставка до склада, упаковка, маркировка, фулфилмент.</div>' +
        '<div style="font-size:5px;margin-top:8px;color:#999">Ввод: ручной или через загрузочный лист. Считается в рублях.</div>' +
        '<div style="margin-top:8px;text-align:right"><button onclick="this.parentNode.parentNode.remove()" style="font:6px Arial;background:#6c5ce7;color:#fff;border:none;border-radius:4px;padding:3px 10px;cursor:pointer">Закрыть</button></div>';
    document.body.appendChild(popup);
    popup.addEventListener('click', function(e) { if (e.target === popup) popup.remove(); });
}


function showExtraCostsInfo(el) {
    var existing = document.getElementById('extra-costs-info-popup');
    if (existing) { existing.remove(); return; }
    var popup = document.createElement('div');
    popup.id = 'extra-costs-info-popup';
    popup.style.cssText = 'position:fixed;top:50%;left:50%;transform:translate(-50%,-50%);background:#fff;border:1px solid #ddd;border-radius:8px;padding:16px 20px;box-shadow:0 4px 24px rgba(0,0,0,.18);z-index:9999;max-width:400px;font:6px Arial;line-height:1.6;color:#333';
    popup.innerHTML = '<div style="font-size:7px;font-weight:600;margin-bottom:6px;color:#6c5ce7">ℹ️ Доп расходы</div>' +
        '<div style="font-size:6px">Все дополнительные расходы, которые не учтены в Себестоимости. Считается в рублях.</div>' +
        '<div style="margin-top:8px;text-align:right"><button onclick="this.parentNode.parentNode.remove()" style="font:6px Arial;background:#6c5ce7;color:#fff;border:none;border-radius:4px;padding:3px 10px;cursor:pointer">Закрыть</button></div>';
    document.body.appendChild(popup);
    popup.addEventListener('click', function(e) { if (e.target === popup) popup.remove(); });
}

function showTotalCostInfo(el) {
    var existing = document.getElementById('total-cost-info-popup');
    if (existing) { existing.remove(); return; }
    var popup = document.createElement('div');
function showCommCorrInfo(el) {
    var existing = document.getElementById('comm-corr-popup');
    if (existing) { existing.remove(); return; }
    cleanupCostPopups();
    var div = document.createElement('div');
    div.id = 'comm-corr-popup';
    div.style.cssText = 'position:fixed;top:50%;left:50%;transform:translate(-50%,-50%);background:#fff;border:1px solid #6c5ce7;border-radius:8px;padding:16px;z-index:9999;max-width:350px;box-shadow:0 4px 20px rgba(0,0,0,.2);font-family:Arial;font-size:6px;line-height:1.5';
    div.innerHTML = '<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px"><b style="font-size:7px">Коррекция к комиссии МП</b><span onclick="this.parentNode.parentNode.remove()" style="cursor:pointer;font-size:9px">✕</span></div><div>Скорректируйте базовый % МП на величину всех опций из Конструктора тарифов. Значения могут быть как со знаком +, так и -.</div>';
    document.body.appendChild(div);
}function showDeliveryInfo(el) {
    var existing = document.getElementById('delivery-popup');
    if (existing) { existing.remove(); return; }
    cleanupCostPopups();
    var div = document.createElement('div');
    div.id = 'delivery-popup';
    div.style.cssText = 'position:fixed;top:50%;left:50%;transform:translate(-50%,-50%);background:#fff;border:1px solid #6c5ce7;border-radius:8px;padding:16px;z-index:9999;max-width:350px;box-shadow:0 4px 20px rgba(0,0,0,.2);font-family:Arial;font-size:6px;line-height:1.5';
    div.innerHTML = '<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px"><b style="font-size:7px">Скорость доставаемости</b><span onclick="this.parentNode.parentNode.remove()" style="cursor:pointer;font-size:9px">✕</span></div><div>От закупа до склада поставщика</div>';
    document.body.appendChild(div);
}



    popup.id = 'total-cost-info-popup';
    popup.style.cssText = 'position:fixed;top:50%;left:50%;transform:translate(-50%,-50%);background:#fff;border:1px solid #ddd;border-radius:8px;padding:16px 20px;box-shadow:0 4px 24px rgba(0,0,0,.18);z-index:9999;max-width:400px;font:6px Arial;line-height:1.6;color:#333';
    popup.innerHTML = '<div style="font-size:7px;font-weight:600;margin-bottom:6px;color:#6c5ce7">ℹ️ Себестоимость итого</div>' +
        '<div style="font-size:6px">Сумма колонок «Себестоимость» и «Доп расходы». Автоматический расчёт. Считается в рублях.</div>' +
        '<div style="margin-top:8px;text-align:right"><button onclick="this.parentNode.parentNode.remove()" style="font:6px Arial;background:#6c5ce7;color:#fff;border:none;border-radius:4px;padding:3px 10px;cursor:pointer">Закрыть</button></div>';
    document.body.appendChild(popup);
    popup.addEventListener('click', function(e) { if (e.target === popup) popup.remove(); });
}


// Автопересчёт Себестоимость итого
document.addEventListener('input', function(e) {
    if (e.target.matches('[data-field="cost_price"], [data-field="extra_costs"]')) {
        var row = e.target.closest('tr');
        if (!row) return;
        var cp = parseFloat(row.querySelector('[data-field="cost_price"]')?.value) || 0;
        var ec = parseFloat(row.querySelector('[data-field="extra_costs"]')?.value) || 0;
        var tc = row.querySelector('[data-field="total_cost"]');
        if (tc) tc.value = (cp + ec).toFixed(2);
    }
});

function onShipmentChange(sel) {
    var row = sel.closest('tr');
    var nmId = row.getAttribute('data-nm');
    console.log('[shipment] change nmId=', nmId, 'value=', sel.value);
    if (!nmId) return;
    
    // Try _costMap first (number key), then string key
    var c = _costMap[nmId] || _costMap[parseInt(nmId)];
    console.log('[shipment] _costMap entry=', c ? 'found (subject_id=' + c.subject_id + ')' : 'NOT FOUND');
    
    // Fallback: find subject_id from _costProducts
    var subjectId = null;
    if (c && c.subject_id) {
        subjectId = c.subject_id;
    } else {
        for (var i = 0; i < _costProducts.length; i++) {
            if (String(_costProducts[i].nm_id) === String(nmId)) {
                subjectId = _costProducts[i].subject_id;
                break;
            }
        }
    }
    
    if (!subjectId) {
        console.log('[shipment] No subject_id for nm_id', nmId);
        return;
    }
    
    var model = sel.value;
    var mpInput = row.querySelector('input[data-field="mp_base_pct"]');
    var url = '/api/v1/nl/commission-rate?org_id=' + encodeURIComponent(ORG_ID) + '&subject_id=' + subjectId + '&model=' + model;
    console.log('[shipment] fetching', url);
    
    fetch(url)
        .then(function(r) { return r.json(); })
        .then(function(data) {
            console.log('[shipment] response', data);
            if (data.commission_pct !== null && data.commission_pct !== undefined && mpInput) {
                mpInput.value = data.commission_pct;
                mpInput.style.background = '#d4edda';
                setTimeout(function() { mpInput.style.background = ''; }, 1500);
            }
        })
        .catch(function(e) { console.error('[shipment] error', e); });
}

function toggleCostGroup(row) {
    let next = row.nextElementSibling;
    while (next && next.classList.contains('cost-group-child')) {
        next.style.display = next.style.display === 'none' ? '' : 'none';
        next = next.nextElementSibling;
    }
}

function updateBulkBar() {
    // Считаем через Tabulator данные (а не DOM чекбоксы)
    let count = 0;
    if (typeof costTabulator !== 'undefined' && costTabulator) {
        count = costTabulator.getData().filter(r => r._selected).length;
    } else {
        count = document.querySelectorAll('.cost-row-check:checked').length;
    }
    const bar = document.getElementById('cost-bulk-bar');
    const info = document.getElementById('cost-selected-info');
    const countEl = document.getElementById('cost-selected-count');
    const barCount = document.getElementById('bulk-bar-count');
    
    if (count > 0) {
        bar.style.display = 'flex';
        info.style.display = 'inline';
    } else {
        bar.style.display = 'none';
        info.style.display = 'none';
    }
    if (countEl) countEl.textContent = count;
    if (barCount) barCount.textContent = '\u0412\u044b\u0434\u0435\u043b\u0435\u043d\u043e: ' + count;
    
    const allChecks = document.querySelectorAll('.cost-row-check');
    const checkAll = document.getElementById('cost-check-all');
    if (checkAll) checkAll.checked = allChecks.length > 0 && count === allChecks.length;
}

function clearBulkSelection() {
    if (typeof costTabulator !== "undefined" && costTabulator) {
        costTabulator.getData().forEach(function(r) { r._selected = false; });
        costTabulator.replaceData(costTabulator.getData());
    }
    updateBulkBar();
}

function onBulkFieldChange() {
    const field = document.getElementById('bulk-field').value;
    const container = document.getElementById('bulk-value');
    const parentEl = container.parentNode;

    // Values for list-type fields (must match cost-grid.js editorParams)
    const listValues = {
        'product_class': [['','—'],['A','A'],['B','B'],['C','C']],
        'product_status': [['','—'],['Новинка','🟢 Новинка'],['Выводим','🔴 Выводим'],['ТОП (А)','🔵 ТОП (А)'],['Двигаем (В)','🟡 Двигаем (В)'],['Категория С','⚪ Категория С'],['Планируется к запуску','🟣 Планируется к запуску']],
        'fulfillment_model': [['','—'],['fbo','ФБО'],['fbs','ФБС']],
        'vat_rate': [['','—'],['0','нет'],['5','5%'],['7','7%']],
    };

    // Replace input with appropriate element
    if (listValues[field]) {
        const sel = document.createElement('select');
        sel.id = 'bulk-value';
        sel.style.cssText = 'border:1px solid rgba(255,255,255,.3);border-radius:4px;padding:4px 8px;font-size:.9em;width:160px;background:#fff;color:#333';
        listValues[field].forEach(function(item) {
            const opt = document.createElement('option');
            opt.value = item[0];
            opt.textContent = item[1];
            sel.appendChild(opt);
        });
        parentEl.replaceChild(sel, container);
    } else {
        // Restore text input if not a list field
        if (container.tagName === 'SELECT') {
            const inp = document.createElement('input');
            inp.type = 'text';
            inp.id = 'bulk-value';
            inp.placeholder = 'Значение';
            inp.style.cssText = 'border:1px solid rgba(255,255,255,.3);border-radius:4px;padding:4px 8px;font-size:.9em;width:120px;background:#fff;color:#333';
            parentEl.replaceChild(inp, container);
        }
    }
}

function applyBulkEdit() {
    _costDirty = true;
    const field = document.getElementById('bulk-field').value;
    const value = document.getElementById('bulk-value').value;
    if (!field) { alert('\u0412\u044b\u0431\u0435\u0440\u0438\u0442\u0435 \u043f\u043e\u043b\u0435 \u0434\u043b\u044f \u0438\u0437\u043c\u0435\u043d\u0435\u043d\u0438\u044f'); return; }

    // Берём выделенные строки через Tabulator API (а не DOM — Tabulator виртуализирует строки)
    let selectedRows = [];
    if (typeof costTabulator !== 'undefined' && costTabulator) {
        selectedRows = costTabulator.getRows().filter(r => r.getData()._selected);
    }

    if (!selectedRows.length) { alert('\u0412\u044b\u0434\u0435\u043b\u0438\u0442\u0435 \u0445\u043e\u0442\u044f \u0431\u044b \u043e\u0434\u043d\u0443 \u0441\u0442\u0440\u043e\u043a\u0443'); return; }

    const count = selectedRows.length;
    if (!confirm('\u0417\u0430\u043f\u0438\u0441\u0430\u0442\u044c \u00ab' + value + '\u00bb \u0432 \u043f\u043e\u043b\u0435 \u00ab' + field + '\u00bb \u0434\u043b\u044f ' + count + ' \u0442\u043e\u0432\u0430\u0440\u043e\u0432?')) return;

    // Применяем значение через Tabulator row.update — триггерит cellEdited и синхронизацию
    const update = {};
    update[field] = value;
    selectedRows.forEach(row => {
        row.update(update);
    });

    document.getElementById('bulk-field').value = '';
    document.getElementById('bulk-value').value = '';
}


async function autoFillReference() {
    if (!ORG_ID) return;
    const btn = event.target.closest('button');
    const origText = btn.innerHTML;
    btn.innerHTML = '⏳ Заполнение...';
    btn.disabled = true;
    try {
        const res = await fetch('/api/v1/nl/cost-prices/auto-fill?org_id=' + ORG_ID, {method: 'POST'});
        const data = await res.json();
        if (data.ok) {
            const s = data.stats;
            const filled = s.fields_filled;
            let msg = 'Обновлено: ' + s.updated + ' записей';
            if (s.skipped) msg += ', пропущено: ' + s.skipped;
            const details = Object.entries(filled).map(([k,v]) => {
                const names = {
                    'mp_base_pct': 'Комиссия МП',
                    'logistics_cost': 'Логистика',
                    'storage_pct': 'Хранение',
                    'price_before_spp_plan': 'Цена',
                    'buyout_niche_pct': '% выкупа',
                    'ad_plan_rub': 'Реклама'
                };
                return (names[k] || k) + ': ' + v;
            }).join(', ');
            if (details) msg += String.fromCharCode(10) + 'Поля: ' + details;
            alert(msg);
            loadCostPrices();
        } else {
            alert('Ошибка: ' + (data.error || 'неизвестная'));
        }
    } catch(e) {
        alert('Ошибка автозаполнения: ' + e.message);
    } finally {
        btn.innerHTML = origText;
        btn.disabled = false;
    }
}

async function saveAllCostPrices() {
    // Сначала сохраняем налоговые настройки (тихо, без alert)
    try { await saveTaxSettings(); } catch(e) { console.warn("Tax save failed", e); }

    // === TABULATOR: batch-сохранение ===
    if (typeof Tabulator !== "undefined" && costTabulator && typeof getCostDataForSave === "function") {
        const saveData = getCostDataForSave();
        try {
            const resp = await fetch("/api/v1/nl/cost-prices/batch?org_id=" + ORG_ID, {
                method: "POST", headers: {"Content-Type": "application/json"},
                body: JSON.stringify(saveData)
            });
            const result = await resp.json();
            if (resp.ok) {
                _costDirty = false;
                showToast("✅ Сохранено: " + (result.saved || 0) + " из " + saveData.length + (result.errors ? " (ошибок: " + result.errors + ")" : ""));
            } else {
                showToast("❌ Ошибка сохранения: " + (result.detail || resp.status), "error");
            }
        } catch(e) { showToast("❌ Ошибка: " + e.message, "error"); }
        loadCostPrices();
        return;
    }
    // === FALLBACK: диагностика ===
    var reasons = [];
    if (typeof Tabulator === "undefined") reasons.push("Tabulator CDN не загружен");
    if (typeof costTabulator === "undefined" || !costTabulator) reasons.push("costTabulator не инициализирован");
    if (typeof getCostDataForSave === "undefined") reasons.push("cost-grid.js не загружен");
    showToast("⚠️ Сохранение недоступно: " + reasons.join(", ") + ". Обновите страницу (Ctrl+F5).", "error");
}

async function uploadCostExcel(input) {
    if (!input.files.length) return;
    const file = input.files[0];
    const body = await file.arrayBuffer();
    try {
        const res = await fetch('/api/v1/nl/cost-prices/upload?org_id=' + ORG_ID, {
            method: 'POST', headers: {'Content-Type': 'application/octet-stream', 'x-filename': file.name},
            body: body
        });
        const data = await res.json();
        if (!res.ok) throw new Error(data.detail || 'Ошибка');
        showToast('✅ Загружено: ' + data.updated + ' из ' + data.total);
        loadCostPrices();
    } catch(e) { showToast('❌ Ошибка: ' + e.message, 'error'); }
    input.value = '';
}
function downloadEmptyTemplate() {
    var hdr = "Арт WB;Арт продавца;Баркод;Размер;Категория;" +
        "Себестоимость;Доп расходы;Закупка;Логистика;Упаковка;Прочее;Мин. цена;НДС руб;" +
        "ФБО/ФБС;Склад отгрузки FBS;" +
        "Баз. % МП;Корр. % МП;% хранения;% выкупа по категории;" +
        "Цена до СПП план;Цена до СПП к изм.;Дата правок;Скидка WB Клуб %;РРЦ;" +
        "Рекл. расходы %;" +
        "Класс товара;Бренд;Статус товара;" +
        "Налог. система;" +
        "Сезон янв;Сезон фев;Сезон мар;Сезон апр;Сезон май;Сезон июн;" +
        "Сезон июл;Сезон авг;Сезон сен;Сезон окт;Сезон ноя;Сезон дек;" +
        "План длина;План ширина;План высота;План объём;План вес;" +
        "Доставка до склада (дни);Доставка до МП (дни);" +
        "ТОП запрос 1;ТОП запрос 2;ТОП запрос 3;" +
        "Заметки";
    var blob = new Blob(["\ufeff" + hdr], {type: "text/csv;charset=utf-8"});
    var a = document.createElement("a");
    a.href = URL.createObjectURL(blob);
    a.download = "template_spravochnik.csv";
    a.click();
    URL.revokeObjectURL(a.href);
    showToast("📥 Шаблон скачан");
}

function exportCostTemplate() {
    var hdr = "Арт WB;Арт продавца;Баркод;Размер;Категория;" +
        "Себестоимость;Доп расходы;Закупка;Логистика;Упаковка;Прочее;Мин. цена;НДС руб;" +
        "ФБО/ФБС;Склад отгрузки FBS;" +
        "Баз. % МП;Корр. % МП;% хранения;% выкупа по категории;" +
        "Цена до СПП план;Цена до СПП к изм.;Дата правок;Скидка WB Клуб %;РРЦ;" +
        "Рекл. расходы %;" +
        "Класс товара;Бренд;Статус товара;" +
        "Налог. система;" +
        "Сезон янв;Сезон фев;Сезон мар;Сезон апр;Сезон май;Сезон июн;" +
        "Сезон июл;Сезон авг;Сезон сен;Сезон окт;Сезон ноя;Сезон дек;" +
        "План длина;План ширина;План высота;План объём;План вес;" +
        "Доставка до склада (дни);Доставка до МП (дни);" +
        "ТОП запрос 1;ТОП запрос 2;ТОП запрос 3;" +
        "Заметки";
    var csv = hdr;

    // Tabulator: читаем из costTabulator
    if (typeof costTabulator !== "undefined" && costTabulator) {
        var data = costTabulator.getData();
        if (data.length) {
            csv += String.fromCharCode(10);
            data.forEach(function(d) {
                var vol = "";
                var l = parseFloat(d.plan_length) || 0;
                var w = parseFloat(d.plan_width) || 0;
                var h = parseFloat(d.plan_height) || 0;
                if (l > 0 && w > 0 && h > 0) vol = (l * w * h / 1000);
                var cols = [
                    d.nm_id || "", d.vendor_code || "", d._barcodes || "", d.size_name || "", d.subject_name || "",
                    d.cost_price || "", d.extra_costs || "", "", "", "", "", d.min_price || "", "",
                    d.fulfillment_model === "fbs" ? "fbs" : "fbo", d.fbs_warehouse || "",
                    "", d.mp_correction_pct || "", "", d.buyout_niche_pct || "",
                    "", "", d.change_date || "", "", d.rrc_price || "",
                    d.ad_plan_rub || "",
                    d.product_class || "", d.brand || "", d.product_status || "",
                    "",
                    d.season_jan || "", d.season_feb || "", d.season_mar || "", d.season_apr || "",
                    d.season_may || "", d.season_jun || "", d.season_jul || "", d.season_aug || "",
                    d.season_sep || "", d.season_oct || "", d.season_nov || "", d.season_dec || "",
                    d.plan_length || "", d.plan_width || "", d.plan_height || "", vol, d.plan_weight || "",
                    "", "",
                    d.top_query_1 || "", d.top_query_2 || "", d.top_query_3 || "",
                    ""
                ];
                csv += cols.join(";") + String.fromCharCode(10);
            });
        }
    }

    var blob = new Blob(["\ufeff" + csv], {type: "text/csv;charset=utf-8"});
    var a = document.createElement("a");
    a.href = URL.createObjectURL(blob);
    a.download = "export_spravochnik.csv";
    a.click();
    URL.revokeObjectURL(a.href);
}

async function loadWarehouses() {
    if (!ORG_ID) return;
    const sel = document.getElementById('wh-date') || document.getElementById('ref-date');
    const d = sel?.value;
    if (!d || d === 'Нет данных') return;
    try {
        const res = await fetch('/api/v1/nl/warehouses?org_id=' + ORG_ID + '&target_date=' + d);
        if (!res.ok) return;
        const items = await res.json();
        document.getElementById('wh-count').textContent = items.length + ' записей';
        if (!items.length) { document.getElementById('wh-body').innerHTML = '<tr><td colspan="6" class="empty">Нет данных</td></tr>'; return; }
        document.getElementById('wh-body').innerHTML = items.map(i =>
            '<tr><td>' + (i.nm_id||'') + '</td><td>' + esc(i.vendor_code||'') + '</td><td>' + esc(i.product_name||'') +
            '</td><td>' + (i.warehouse||'—') + '</td><td>' + (i.qty||0) + '</td><td>—</td></tr>'
        ).join('');
    } catch(e) { console.error('loadWarehouses', e); }
}

function showOpExDialog() { document.getElementById('opex-dialog').style.display = 'flex'; document.getElementById('opex-date').value = new Date().toISOString().split('T')[0]; }
function hideOpExDialog() { document.getElementById('opex-dialog').style.display = 'none'; }
async function saveOpEx() {
    const data = {date: document.getElementById('opex-date').value, category: document.getElementById('opex-category').value, description: document.getElementById('opex-desc').value, amount: parseFloat(document.getElementById('opex-amount').value), vat: parseFloat(document.getElementById('opex-vat').value||'0')};
    if (!data.date || !data.amount) { alert('Заполните дату и сумму'); return; }
    try {
        await fetch('/api/v1/nl/operating-expenses?org_id=' + ORG_ID, {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify(data)});
        hideOpExDialog(); loadOpEx();
    } catch(e) { showToast('❌ Ошибка: ' + e.message, 'error'); }
}
async function loadOpEx() {
    if (!ORG_ID) return;
    try {
        const res = await fetch('/api/v1/nl/operating-expenses?org_id=' + ORG_ID);
        if (!res.ok) return;
        const items = await res.json();
        document.getElementById('opex-count').textContent = items.length + ' записей';
        if (!items.length) { document.getElementById('opex-body').innerHTML = '<tr><td colspan="7" class="empty">Нет записей. Нажмите добавить.</td></tr>'; return; }
        document.getElementById('opex-body').innerHTML = items.map(i =>
            '<tr><td>' + (i.date||'') + '</td><td>' + esc(i.category||'') + '</td><td>' + esc(i.description||'') +
            '</td><td>' + (i.amount||0) + '</td><td>' + (i.vat||0) + '%</td><td>—</td><td style="color:#e74c3c;cursor:pointer">🗑</td></tr>'
        ).join('');
    } catch(e) { console.error('loadOpEx', e); }
}

function initRnpMonths() {
    // Period selector now: dropdown with presets
    var sel = document.getElementById("rnp-month");
    if (!sel) return;
    if (sel.options.length > 0) return;
    var now = new Date();
    var presets = [
        {label: "90 дней", value: ""},
        {label: "60 дней", value: "60"},
        {label: "30 дней", value: "30"},
    ];
    // Add last 6 months as options
    for (var i = 0; i < 6; i++) {
        var d = new Date(now.getFullYear(), now.getMonth() - i, 1);
        var val = d.getFullYear() + "-" + ("0" + (d.getMonth()+1)).slice(-2);
        var label = d.toLocaleDateString("ru-RU", {month:"long", year:"numeric"});
        presets.push({label: label.charAt(0).toUpperCase() + label.slice(1), value: val});
    }
    for (var j = 0; j < presets.length; j++) {
        var opt = document.createElement("option");
        opt.value = presets[j].value;
        opt.textContent = presets[j].label;
        if (j === 0) opt.selected = true;
        sel.appendChild(opt);
    }
}
async function loadRnp() {
    if (!ORG_ID) return;
    var monthSel = document.getElementById('rnp-month');
    var selVal = monthSel ? monthSel.value : '';
    var sort = document.getElementById('rnp-sort')?.value || 'orders_revenue';
    var search = document.getElementById('rnp-search')?.value || '';
    var ubp = document.getElementById('rnp-buyout-pct')?.checked ? '1' : '0';
    try {
        var url = '/api/v1/nl/rnp?org_id=' + ORG_ID + '&sort_by=' + sort + '&use_buyout_pct=' + ubp;
        // Если выбран месяц (YYYY-MM), передаём month; если дни — передаём days
        if (selVal && selVal.indexOf('-') > 0 && !/^\d+$/.test(selVal)) {
            url += '&month=' + selVal;
        } else if (selVal && /^\d+$/.test(selVal)) {
            url += '&days=' + selVal;
        }
        if (search) url += '&search=' + encodeURIComponent(search);
        var res = await fetch(url);
        if (!res.ok) { document.getElementById('rnp-cards').innerHTML = '<div class="empty">Ошибка ' + res.status + '</div>'; return; }
        var data = await res.json();
        var prods = data.products || [];
        var days = data.day_list || [];
        var s = data.summary || {};

        // Сводка
        document.getElementById('rnp-summary').innerHTML =
            '<div><b>Товаров:</b> ' + s.total_products + '</div>' +
            '<div><b>Заказы:</b> ' + s.total_orders + ' шт / ' + fmtR(s.total_orders_revenue) + ' ₽</div>' +
            '<div><b>Выкупы:</b> ' + s.total_buyouts + ' шт / ' + fmtR(s.total_buyouts_revenue) + ' ₽</div>' +
            '<div><b>Реклама:</b> ' + fmtR(s.total_ad_cost) + ' ₽</div>' +
            '<div><b>ДРР:</b> ' + s.total_drr + '%</div>' +
            '<div><b>Маржа до ДРР:</b> ' + fmtR(s.total_margin_before_drr) + ' ₽</div>' +
            '<div><b>Прибыль расч:</b> ' + fmtR(s.total_profit_calc) + ' ₽</div>' +
            '<div><b>Маржа с ДРР:</b> ' + fmtR(s.total_margin_with_drr) + ' ₽</div>' +
            '<div><b>Остатки:</b> ' + s.total_stock + ' шт</div>';
        document.getElementById('rnp-count').textContent = prods.length + ' товаров | ' + (data.month || '') + ' | ' + days.length + ' дней';
        if (!prods.length) { document.getElementById('rnp-cards').innerHTML = '<div class="empty">Нет данных</div>'; return; }

        // Считаем summary по дням
        var todayStr = new Date().toISOString().substring(0, 10);
        var sumDays = {};
        for (var di = 0; di < days.length; di++) {
            sumDays[days[di]] = {orders_count:0,orders_revenue:0,buyouts_count:0,buyouts_revenue:0,ad_cost:0,profit_calc:0,margin_with_drr:0,margin_before_drr:0,drr:0};
        }
        for (var pi = 0; pi < prods.length; pi++) {
            var pd = prods[pi].days || [];
            for (var dj = 0; dj < pd.length; dj++) {
                var dk = pd[dj].date;
                if (sumDays[dk]) {
                    sumDays[dk].orders_count += pd[dj].orders_count || 0;
                    sumDays[dk].orders_revenue += pd[dj].orders_revenue || 0;
                    sumDays[dk].buyouts_count += pd[dj].buyouts_count || 0;
                    sumDays[dk].buyouts_revenue += pd[dj].buyouts_revenue || 0;
                    sumDays[dk].ad_cost += pd[dj].ad_cost || 0;
                    sumDays[dk].profit_calc += pd[dj].profit_calc || 0;
                    sumDays[dk].margin_with_drr += pd[dj].margin_with_drr || 0;
                    sumDays[dk].margin_before_drr += pd[dj].margin_before_drr || 0;
                }
            }
        }

        // Рисуем шапку-сводку (таблица 1)
        var hhtml = '<div class="rnp-wrap" style="overflow-x:auto"><table class="rnp-header-table"><thead><tr>';
        hhtml += '<th class="row-label" style="min-width:140px">Показатель</th>';
        hhtml += '<th class="val-cell" style="min-width:90px">Итого</th>';
        for (var hi = 0; hi < days.length; hi++) {
            var dShort = days[hi].substring(5);
            hhtml += '<th class="day-header' + (days[hi] === todayStr ? ' today' : '') + '">' + dShort + '</th>';
        }
        hhtml += '</tr></thead><tbody>';

        // Строки шапки
        var hrows = [
            {label:'Заказы, Σ ₽', key:'orders_revenue', fmt:1},
            {label:'Заказы, кол', key:'orders_count', fmt:0},
            {label:'Выкупы, Σ ₽', key:'buyouts_revenue', fmt:1},
            {label:'Выкупы, кол', key:'buyouts_count', fmt:0},
            {label:'Рекл. бюджет, ₽', key:'ad_cost', fmt:1},
            {label:'Маржа до ДРР, ₽', key:'margin_before_drr', fmt:1},
            {label:'Маржа с ДРР, ₽', key:'margin_with_drr', fmt:1},
            {label:'Прибыль расчёт, ₽', key:'profit_calc', fmt:1},
        ];
        for (var ri = 0; ri < hrows.length; ri++) {
            var r = hrows[ri];
            hhtml += '<tr><td class="row-label">' + r.label + '</td>';
            var totalV = 0;
            for (var tdi = 0; tdi < days.length; tdi++) { totalV += (sumDays[days[tdi]][r.key] || 0); }
            hhtml += '<td class="val-cell"><b>' + (r.fmt ? fmtR(totalV) : totalV) + '</b></td>';
            for (var ci = 0; ci < days.length; ci++) {
                var cv = sumDays[days[ci]][r.key] || 0;
                var cls = 'val-cell' + (days[ci] === todayStr ? ' today' : '');
                var color = r.key === 'margin_with_drr' || r.key === 'profit_calc' || r.key === 'margin_before_drr' ? (cv >= 0 ? ' class="' + cls + '" style="color:#00b894"' : ' class="' + cls + '" style="color:#e74c3c"') : ' class="' + cls + '"';
                hhtml += '<td' + color + '>' + (r.fmt ? fmtR(cv) : cv) + '</td>';
            }
            hhtml += '</tr>';
        }
        hhtml += '</tbody></table></div>';
        document.getElementById('rnp-header-wrap').innerHTML = hhtml;

        // Рисуем карточки (таблица 2)
        var html = '<div class="rnp-wrap"><div class="rnp-table-wrap"><table class="rnp-table"><thead><tr>';
        html += '<th class="sticky-col" style="min-width:160px;left:0">Карточка</th>';
        html += '<th class="sticky-col" style="min-width:100px;left:160px">За период</th>';
        html += '<th class="sticky-col" style="min-width:100px;left:260px">План/Факт</th>';
        for (var dhi = 0; dhi < days.length; dhi++) {
            var dS = days[dhi].substring(5);
            html += '<th class="day-header' + (days[dhi] === todayStr ? ' today' : '') + '">' + dS + '</th>';
        }
        html += '</tr></thead><tbody>';

        for (var i = 0; i < prods.length; i++) {
            var p = prods[i];
            var mc = p.margin_with_drr >= 0 ? '#00b894' : '#e74c3c';
            var pDays = p.days || [];
            var daysMap = {};
            for (var j = 0; j < pDays.length; j++) { daysMap[pDays[j].date] = pDays[j]; }

            // Фотография + информация
            var photoHtml = '';
            var thumbUrl = (p.photo_main || '').replace('/hq/', '/c246x328/').replace('/big/', '/c246x328/').replace('/tm/', '/c246x328/');
            if (thumbUrl) photoHtml = '<img src="' + esc(thumbUrl) + '">'; 
            else photoHtml = '📦';
            var cardInfo = '<div class="rnp-card-photo">' + photoHtml + '</div>';
            cardInfo += '<div class="rnp-card-nm">' + (p.nm_id||'') + '</div>';
            cardInfo += '<div class="rnp-card-name" title="' + esc(p.product_name||'') + '">' + esc((p.product_name||'').substring(0,35)) + '</div>';
            cardInfo += '<div class="rnp-card-detail">' + esc(p.product_class||'') + '</div>';
            cardInfo += '<div class="rnp-card-detail">' + esc(p.brand||'') + '</div>';
            cardInfo += '<div class="rnp-card-detail">' + esc(p.vendor_code||'') + (p.size_name ? ' / ' + esc(p.size_name) : '') + '</div>';
            cardInfo += '<div class="rnp-card-detail">' + esc(p.tags||'') + '</div>';
            cardInfo += '<div class="rnp-card-detail">Ост: ' + (p.current_stock||0) + ' (хватит ' + (p.enough_days||0) + 'д)</div>';
            cardInfo += '<div class="rnp-card-krrr" style="color:' + mc + '">КРРР: ' + (p.krrr||0) + '%</div>';

            // За 30 дней
            var m30 = '';
            m30 += '<div>Зак: <b>' + fmtR(p.total_orders_revenue) + '₽</b></div>';
            m30 += '<div>     ' + (p.total_orders||0) + ' шт</div>';
            m30 += '<div>Вык: <b>' + fmtR(p.total_buyouts_revenue) + '₽</b></div>';
            m30 += '<div>     ' + (p.total_buyouts||0) + ' шт</div>';
            m30 += '<div>%Вык: ' + (p.buyout_pct||0) + '%</div>';
            m30 += '<div>М.ДРР: <span style="color:' + mc + '">' + fmtR(p.margin_with_drr) + '₽</span></div>';
            m30 += '<div>М.до: ' + fmtR(p.margin_before_drr) + '₽</div>';
            m30 += '<div>Рекл: ' + fmtR(p.total_ad_cost) + '₽</div>';
            m30 += '<div>ДРР: ' + (p.drr||0) + '%</div>';
            m30 += '<div>CTR: ' + (p.ctr||0) + '% / CPL: ' + fmtR(p.cpl) + '</div>';
            m30 += '<div>Рейтинг: ' + (p.rating_reviews != null ? p.rating_reviews : '-') + '</div>';

            // План/Факт
            var pf = '';
            pf += '<div>План: ' + fmtR(p.plan_value) + '</div>';
            pf += '<div>%Вып: <span style="color:' + (p.pct_complete >= 100 ? '#00b894' : p.pct_complete >= 70 ? '#fdcb6e' : '#e74c3c') + '">' + (p.pct_complete||0) + '%</span></div>';
            pf += '<div>Ф/д: ' + fmtR(p.daily_norm) + '</div>';
            pf += '<div>Ц.СПП: ' + fmtR(p.price_retail) + '₽</div>';
            pf += '<div>Ц+СПП: ' + fmtR(p.price_with_spp) + '₽</div>';
            pf += '<div>СПП%: ' + (p.spp_pct||0) + '%</div>';
            pf += '<div>Акция: ' + (p.in_promo ? '✅' : '-') + '</div>';
            pf += '<div>ROI: ' + (p.roi||0) + '%</div>';
            pf += '<div>Сб.ост: ' + fmtR(p.cost_of_stock) + '₽</div>';

            html += '<tr>';
            html += '<td class="sticky-col" style="left:0">' + cardInfo + '</td>';
            html += '<td class="sticky-col" style="left:160px">' + m30 + '</td>';
            html += '<td class="sticky-col" style="left:260px">' + pf + '</td>';

            // Дни
            for (var di = 0; di < days.length; di++) {
                var dd = daysMap[days[di]] || {};
                var dmc = (dd.margin_with_drr||0) >= 0 ? '#00b894' : '#e74c3c';
                var cls = 'val-cell' + (days[di] === todayStr ? ' today' : '');
                html += '<td class="' + cls + '">';
                html += '<div>Зак <b>' + fmtR(dd.orders_revenue) + '</b></div>';
                html += '<div style="color:#aaa">' + (dd.orders_count||0) + ' шт</div>';
                html += '<div>Вык <b>' + fmtR(dd.buyouts_revenue) + '</b></div>';
                html += '<div style="color:#aaa">' + (dd.buyouts_count||0) + ' шт</div>';
                html += '<div>М.ДРР <span style="color:' + dmc + '">' + fmtR(dd.margin_with_drr) + '</span></div>';
                html += '<div>ДРР ' + (dd.drr||0) + '%</div>';
                html += '</td>';
            }
            html += '</tr>';

            // Разделитель между карточками
            html += '<tr class="card-divider"><td colspan="' + (3 + days.length) + '"></td></tr>';
        }

        html += '</tbody></table></div></div>';
        document.getElementById('rnp-cards').innerHTML = html;

        // Синхронизация горизонтального скролла шапки и карточек
        var headerWrap = document.getElementById('rnp-header-wrap');
        var cardsWrap = document.querySelector('.rnp-table-wrap');
        if (headerWrap && cardsWrap) {
            cardsWrap.addEventListener('scroll', function() { headerWrap.scrollLeft = cardsWrap.scrollLeft; });
            headerWrap.addEventListener('scroll', function() { cardsWrap.scrollLeft = headerWrap.scrollLeft; });
        }

    } catch(e) { console.error('loadRnp', e); document.getElementById('rnp-cards').innerHTML = '<div class="empty">Ошибка: '+esc(e.message)+'</div>'; }
}
function fmtR(v) { if (v==null||v===0) return '—'; return Number(v).toLocaleString('ru-RU',{maximumFractionDigits:0}); }

// ─── ПЛАН ПРОДАЖ ─────────────────────────────────────────
var spData = [];
var spChanged = new Set();
var spSelected = new Set();

async function loadSalesPlans() {
    if (!ORG_ID) return;
    var periodSel = document.getElementById('sp-period');
    var period = periodSel ? periodSel.value : '';
    var type = document.getElementById('sp-type')?.value || '';
    var url = '/api/v1/nl/sales-plans?org_id=' + ORG_ID;
    if (period) url += '&period=' + period;
    try {
        var r = await fetch(url, {headers:{'Authorization':'Bearer '+TOKEN}});
        if (!r.ok) throw new Error('HTTP ' + r.status);
        var data = await r.json();
        // Фильтр по типу на клиенте
        if (type) data = data.filter(function(d) { return d.plan_type === type; });
        spData = data;
        spChanged.clear();
        document.getElementById('sp-save-all-btn').style.display = 'none';
        renderSpTable(data);
        // Загрузить сводку
        loadSpSummary(period);
    } catch(e) {
        console.error('loadSalesPlans error:', e);
        document.getElementById('sp-body').innerHTML = '<tr><td colspan="15" class="empty">Ошибка: ' + esc(e.message) + '</td></tr>';
    }
}

async function loadSpSummary(period) {
    if (!ORG_ID) return;
    var url = '/api/v1/nl/sales-plans/summary?org_id=' + ORG_ID;
    if (period) url += '&period=' + period;
    try {
        var r = await fetch(url, {headers:{'Authorization':'Bearer '+TOKEN}});
        var data = await r.json();
        var fmt = function(v) { return v != null ? Number(v).toLocaleString('ru-RU', {maximumFractionDigits:0}) : '—'; };
        var totalPlan = 0, totalActual = 0, greenCount = 0, yellowCount = 0, redCount = 0;
        data.forEach(function(d) {
            totalPlan += d.total_plan;
            totalActual += d.total_actual;
            greenCount += d.green_count;
            yellowCount += d.yellow_count;
            redCount += d.red_count;
        });
        var pct = totalPlan > 0 ? (totalActual / totalPlan * 100).toFixed(1) : '—';
        document.getElementById('sp-total-plan').textContent = fmt(totalPlan);
        document.getElementById('sp-total-actual').textContent = fmt(totalActual);
        document.getElementById('sp-total-pct').textContent = pct + '%';
        document.getElementById('sp-green-count').textContent = greenCount;
        document.getElementById('sp-yellow-count').textContent = yellowCount;
        document.getElementById('sp-red-count').textContent = redCount;
    } catch(e) { console.error('loadSpSummary error:', e); }
}

function renderSpTable(data) {
    var body = document.getElementById('sp-body');
    if (!data || !data.length) {
        body.innerHTML = '<tr><td colspan="15" class="empty">Нет планов продаж. Нажмите «➕ Добавить план» или «📋 Массовое назначение»</td></tr>';
        document.getElementById('sp-count').textContent = '';
        return;
    }
    // Фильтры
    var season = document.getElementById('sp-season')?.value || '';
    var status = document.getElementById('sp-status')?.value || '';
    var search = (document.getElementById('sp-search')?.value || '').toLowerCase();
    var filtered = data.filter(function(d) {
        if (season && d.seasonality !== season) return false;
        if (status) {
            var pct = d.pct_complete || 0;
            if (status === 'green' && pct < 90) return false;
            if (status === 'yellow' && (pct < 70 || pct >= 90)) return false;
            if (status === 'red' && pct >= 70) return false;
        }
        if (search) {
            var s = (d.nm_id + ' ' + (d.vendor_code||'') + ' ' + (d.product_name||'') + ' ' + (d.size_name||'')).toLowerCase();
            if (s.indexOf(search) < 0) return false;
        }
        return true;
    });
    document.getElementById('sp-count').textContent = filtered.length + ' из ' + data.length + ' записей';
    var fmt = function(v) { return v != null ? Number(v).toLocaleString('ru-RU', {maximumFractionDigits:2}) : '—'; };
    var seasonLabel = {low:'Низкая',medium:'Средняя',high:'Высокая',peak:'Пик'};
    var typeLabel = {quantity:'шт',revenue:'₽'};
    body.innerHTML = filtered.map(function(d, i) {
        var thumb = (d.photo_main||'').replace('/hq/','/c246x328/').replace('/big/','/c246x328/').replace('/tm/','/c246x328/');
        var pct = d.pct_complete || 0;
        var statusColor = pct >= 90 ? '#00b894' : (pct >= 70 ? '#fdcb6e' : '#e17055');
        var statusIcon = pct >= 90 ? '🟢' : (pct >= 70 ? '🟡' : '🔴');
        var origIdx = spData.indexOf(d);
        return '<tr data-idx="' + origIdx + '">' +
        '<td style="position:sticky;left:0;background:#fff;z-index:1"><input type="checkbox" class="sp-row-check" data-idx="' + origIdx + '" onchange="onSpRowCheck(this)" style="cursor:pointer"></td>' +
        '<td>' + (thumb ? '<img src="' + esc(thumb) + '" style="width:36px;height:36px;border-radius:4px;object-fit:cover" loading="lazy">' : '') + '</td>' +
        '<td>' + d.nm_id + '</td>' +
        '<td>' + esc(d.vendor_code||'') + '</td>' +
        '<td style="max-width:150px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="' + esc(d.product_name||'') + '">' + esc(d.product_name||'') + '</td>' +
        '<td>' + esc(d.size_name||'') + '</td>' +
        '<td>' + (d.period||'').substring(0,7) + '</td>' +
        '<td>' + (typeLabel[d.plan_type]||d.plan_type) + '</td>' +
        '<td><input type="number" value="' + (d.plan_value||0) + '" data-field="plan_value" data-idx="' + origIdx + '" onchange="onSpFieldChange(this)" style="width:80px;border:1px solid #ddd;border-radius:4px;padding:2px 4px;font-size:.9em"></td>' +
        '<td><input type="number" value="' + (d.actual_value||0) + '" data-field="actual_value" data-idx="' + origIdx + '" onchange="onSpFieldChange(this)" style="width:80px;border:1px solid #ddd;border-radius:4px;padding:2px 4px;font-size:.9em"></td>' +
        '<td style="font-weight:600;color:' + statusColor + '">' + pct.toFixed(1) + '%</td>' +
        '<td style="text-align:center">' + statusIcon + '</td>' +
        '<td><input type="number" step="0.1" value="' + (d.sales_temp||'') + '" data-field="sales_temp" data-idx="' + origIdx + '" onchange="onSpFieldChange(this)" style="width:70px;border:1px solid #ddd;border-radius:4px;padding:2px 4px;font-size:.9em" placeholder="—"></td>' +
        '<td><select data-field="seasonality" data-idx="' + origIdx + '" onchange="onSpFieldChange(this)" style="border:1px solid #ddd;border-radius:4px;padding:2px 4px;font-size:.9em">' +
            '<option value="low"' + (d.seasonality==='low'?' selected':'') + '>Низкая</option>' +
            '<option value="medium"' + (d.seasonality==='medium'?' selected':'') + '>Средняя</option>' +
            '<option value="high"' + (d.seasonality==='high'?' selected':'') + '>Высокая</option>' +
            '<option value="peak"' + (d.seasonality==='peak'?' selected':'') + '>Пик</option>' +
        '</select></td>' +
        '<td><button onclick="deleteSpRow(' + origIdx + ')" style="border:none;background:none;color:#e17055;cursor:pointer;font-size:1.1em" title="Удалить">🗑</button></td>' +
        '</tr>';
    }).join('');
}

function applySpFilters() { renderSpTable(spData); }

function onSpFieldChange(el) {
    var idx = parseInt(el.dataset.idx);
    var field = el.dataset.field;
    spData[idx][field] = el.type === 'number' ? parseFloat(el.value) || 0 : el.value;
    spChanged.add(idx);
    document.getElementById('sp-save-all-btn').style.display = spChanged.size ? 'inline-block' : 'none';
    // Пересчитать %
    var d = spData[idx];
    d.pct_complete = d.plan_value > 0 ? (d.actual_value / d.plan_value * 100) : 0;
    renderSpTable(spData);
}

function onSpRowCheck(el) {
    var idx = parseInt(el.dataset.idx);
    if (el.checked) spSelected.add(idx); else spSelected.delete(idx);
    updateSpBulkBar();
}

function toggleAllSpRows(checked) {
    spSelected.clear();
    if (checked) spData.forEach(function(_, i) { spSelected.add(i); });
    document.querySelectorAll('.sp-row-check').forEach(function(cb) { cb.checked = checked; });
    updateSpBulkBar();
}

function updateSpBulkBar() {
    var bar = document.getElementById('sp-bulk-bar');
    var count = spSelected.size;
    bar.style.display = count > 0 ? 'flex' : 'none';
    document.getElementById('sp-bulk-count').textContent = 'Выделено: ' + count;
    document.getElementById('sp-selected-info').style.display = count > 0 ? 'inline' : 'none';
    document.getElementById('sp-selected-count').textContent = count;
}

function applySpBulkEdit() {
    var field = document.getElementById('sp-bulk-field').value;
    var value = document.getElementById('sp-bulk-value').value;
    if (!field || !value) return;
    spSelected.forEach(function(idx) {
        spData[idx][field] = field === 'seasonality' || field === 'plan_type' ? value : parseFloat(value) || 0;
        spData[idx].pct_complete = spData[idx].plan_value > 0 ? (spData[idx].actual_value / spData[idx].plan_value * 100) : 0;
        spChanged.add(idx);
    });
    document.getElementById('sp-save-all-btn').style.display = 'inline-block';
    renderSpTable(spData);
}

function clearSpBulkSelection() {
    spSelected.clear();
    document.getElementById('sp-check-all').checked = false;
    document.querySelectorAll('.sp-row-check').forEach(function(cb) { cb.checked = false; });
    updateSpBulkBar();
}

async function saveAllSpChanges() {
    if (!ORG_ID || !spChanged.size) return;
    var updates = [];
    spChanged.forEach(function(idx) {
        var d = spData[idx];
        if (d.id) {
            updates.push(fetch('/api/v1/nl/sales-plans/' + d.id + '?org_id=' + ORG_ID, {
                method:'PUT', headers:{'Content-Type':'application/json','Authorization':'Bearer '+TOKEN},
                body: JSON.stringify({plan_value:d.plan_value, actual_value:d.actual_value, sales_temp:d.sales_temp, seasonality:d.seasonality, plan_type:d.plan_type})
            }));
        }
    });
    try {
        await Promise.all(updates);
        spChanged.clear();
        document.getElementById('sp-save-all-btn').style.display = 'none';
        loadSalesPlans();
        alert('✅ Сохранено ' + updates.length + ' записей');
    } catch(e) { showToast('❌ Ошибка: ' + e.message, 'error'); }
}

async function deleteSpRow(idx) {
    var d = spData[idx];
    if (!d.id || !confirm('Удалить план для ' + d.nm_id + '?')) return;
    try {
        await fetch('/api/v1/nl/sales-plans/' + d.id + '?org_id=' + ORG_ID, {
            method:'DELETE', headers:{'Authorization':'Bearer '+TOKEN}
        });
        spData.splice(idx, 1);
        renderSpTable(spData);
    } catch(e) { showToast('❌ Ошибка: ' + e.message, 'error'); }
}

function openSpAddModal() {
    var nmId = prompt('Введите артикул WB:');
    if (!nmId) return;
    var period = document.getElementById('sp-period')?.value || new Date().toISOString().substring(0,7) + '-01';
    var planValue = prompt('План (значение):', '100');
    if (!planValue) return;
    var type = confirm('ОК = штуки, Отмена = сумма') ? 'quantity' : 'revenue';
    fetch('/api/v1/nl/sales-plans?org_id=' + ORG_ID, {
        method:'POST',
        headers:{'Content-Type':'application/json','Authorization':'Bearer '+TOKEN},
        body: JSON.stringify({nm_id:parseInt(nmId), period:period, plan_type:type, plan_value:parseFloat(planValue), actual_value:0, seasonality:'medium'})
    }).then(function(r) {
        if (r.ok) { loadSalesPlans(); } else { r.text().then(function(t) { alert('Ошибка: ' + t); }); }
    }).catch(function(e) { alert('Ошибка: ' + e.message); });
}

function openSpBatchModal() {
    var nmIds = prompt('Введите артикулы WB через запятую:');
    if (!nmIds) return;
    var period = document.getElementById('sp-period')?.value || new Date().toISOString().substring(0,7) + '-01';
    var planValue = prompt('План на каждый товар:', '100');
    if (!planValue) return;
    var type = confirm('ОК = штуки, Отмена = сумма') ? 'quantity' : 'revenue';
    var items = nmIds.split(',').map(function(nm) {
        return {nm_id:parseInt(nm.trim()), period:period, plan_type:type, plan_value:parseFloat(planValue), actual_value:0, seasonality:'medium'};
    }).filter(function(it) { return !isNaN(it.nm_id); });
    if (!items.length) { alert('Нет валидных артикулов'); return; }
    fetch('/api/v1/nl/sales-plans/batch?org_id=' + ORG_ID, {
        method:'POST',
        headers:{'Content-Type':'application/json','Authorization':'Bearer '+TOKEN},
        body: JSON.stringify(items)
    }).then(function(r) {
        if (r.ok) { loadSalesPlans(); alert('✅ Добавлено ' + items.length + ' планов'); }
        else { r.text().then(function(t) { alert('Ошибка: ' + t); }); }
    }).catch(function(e) { alert('Ошибка: ' + e.message); });
}


async function loadRefData() {
    if (!ORG_ID) { document.getElementById('ref-body').innerHTML = '<tr><td colspan="13" class="empty">Нет данных. Добавьте WB API ключ в настройках.</td></tr>'; return; }
    const dateVal = document.getElementById('ref-date').value;
    const target_date = dateVal && dateVal !== 'Нет данных' ? '&target_date=' + dateVal : '';
    const [prodRes, refRes] = await Promise.all([
        fetch('/api/v1/nl/products?org_id=' + ORG_ID + target_date),
        fetch('/api/v1/nl/reference?org_id=' + ORG_ID + target_date)
    ]);
    const products = await prodRes.json();
    const refData = await refRes.json();
    const refMap = {};
    refData.forEach(r => refMap[r.nm_id] = r);
    const tbody = document.getElementById('ref-body');
    tbody.innerHTML = '';
    document.getElementById('ref-stats').textContent = products.length + ' товаров';
    if (!products.length) { tbody.innerHTML = '<tr><td colspan="13" class="empty">Нет товаров на эту дату. Подключите WB API ключ и дождитесь синхронизации.</td></tr>'; return; }
    products.forEach(p => {
        const ref = refMap[p.nm_id] || {};
        const thumb = (p.photo_main || '').replace('/hq/', '/c246x328/').replace('/big/', '/c246x328/').replace('/tm/', '/c246x328/');
        const img = thumb ? '<img class="photo" src="' + thumb + '" loading="lazy">' : '📦';
        const tr = document.createElement('tr');
        tr.innerHTML =
            '<td>' + img + '</td>' +
            '<td><b>' + p.nm_id + '</b></td>' +
            '<td>' + (p.vendor_code||'') + '</td>' +
            '<td>' + (p.product_name||'').substring(0,30) + '</td>' +
            '<td style="font-size:.78em;color:#666">' + (p.barcode||'—') + '</td>' +
            '<td style="font-size:.78em;color:#666">' + (p.sku||'—') + '</td>' +
            '<td><input type="number" data-field="cost_price" value="' + (ref.cost_price||'') + '" step="0.01" style="width:80px;font-weight:600;background:' + (ref.cost_price ? '#f0fff0' : '#fff8f0') + '"></td>' +
            '<td><input type="number" data-field="purchase_price" value="' + (ref.purchase_price||'') + '" step="0.01" style="width:75px"></td>' +
            '<td><input type="number" data-field="packaging_cost" value="' + (ref.packaging_cost||'') + '" step="0.01" style="width:70px"></td>' +
            '<td><input type="number" data-field="logistics_cost" value="' + (ref.logistics_cost||'') + '" step="0.01" style="width:70px"></td>' +
            '<td><input type="number" data-field="other_costs" value="' + (ref.other_costs||'') + '" step="0.01" style="width:65px"></td>' +
            '<td><input type="text" data-field="notes" value="' + esc(ref.notes) + '" style="width:100px"></td>' +
            '<td><button class="save-btn" data-nm="' + p.nm_id + '" data-vc="' + (p.vendor_code || '') + '" data-pn="' + (p.product_name || '').replace(/"/g, '&quot;') + '" onclick="saveRowBtn(this)">💾</button></td>';
        tbody.appendChild(tr);
    });
}

async function saveAllRows() {
    const btns = document.querySelectorAll('#ref-body .save-btn');
    for (const btn of btns) { await saveRowBtn(btn); }
}

async function saveRowBtn(btn) {
    const nmId = parseInt(btn.dataset.nm);
    const vc = btn.dataset.vc || '';
    const pn = btn.dataset.pn || '';
    const dateVal = document.getElementById('ref-date').value;
    const row = btn.closest('tr');
    const inputs = row.querySelectorAll('input');
    const data = {nm_id: nmId, vendor_code: vc, product_name: pn, target_date: dateVal};
    inputs.forEach(inp => {
        const f = inp.dataset.field;
        if (f) data[f] = inp.type === 'number' ? (parseFloat(inp.value)||null) : inp.value;
    });
    const res = await fetch('/api/v1/nl/reference?org_id=' + ORG_ID, {
        method: 'POST', headers: {'Content-Type': 'application/json'}, body: JSON.stringify(data)
    });
    if (res.ok) { btn.textContent='✅'; btn.classList.add('saved'); setTimeout(()=>{btn.textContent='💾';btn.classList.remove('saved');},1500); }
}

// Org & WB key management
async function loadOrgs() {
    const res = await fetch('/api/v1/nl/organizations?token=' + TOKEN);
    if (!res.ok) return;
    const orgs = await res.json();
    const sel = document.getElementById('org-select');
    sel.innerHTML = '';
    orgs.forEach(o => {
        const opt = document.createElement('option');
        opt.value = o.id;
        opt.textContent = o.name + (o.wb_seller_id ? ' (ID ' + o.wb_seller_id + ')' : '') + (o.wb_keys_count ? ' ' + o.wb_keys_count + '🔑' : '');
        if (o.id === ORG_ID) opt.selected = true;
        sel.appendChild(opt);
    });
    // Show seller ID in settings
    const current = orgs.find(o => o.id === ORG_ID);
    const sid = document.getElementById('seller-id-display');
    if (sid && current && current.wb_seller_id) sid.innerHTML = '<strong>ID ' + current.wb_seller_id + '</strong>';
    else if (sid) sid.innerHTML = '';
    if (!ORG_ID && orgs.length) {
        ORG_ID = orgs[0].id;
        localStorage.setItem('nl_org_id', ORG_ID);
        sel.value = ORG_ID;
    }
    // Sync cp-store dropdown in Справочник
    const cpSel = document.getElementById('cp-store');
    if (cpSel) {
        cpSel.innerHTML = '';
        orgs.forEach(o => {
            const opt = document.createElement('option');
            opt.value = o.id;
            opt.textContent = o.name + (o.wb_seller_id ? ' (ID ' + o.wb_seller_id + ')' : '');
            if (o.id === ORG_ID) opt.selected = true;
            cpSel.appendChild(opt);
        });
    }
    // Sync ue-store dropdown in Юнит-экономика
    const ueSel = document.getElementById("ue-store");
    if (ueSel) {
        ueSel.innerHTML = "";
        orgs.forEach(o => {
            const opt = document.createElement("option");
            opt.value = o.id;
            opt.textContent = o.name + (o.wb_seller_id ? " (ID " + o.wb_seller_id + ")" : "");
            if (o.id === ORG_ID) opt.selected = true;
            ueSel.appendChild(opt);
        });
    }
    // Sync promo-store dropdown in Акции
    const promoSel = document.getElementById("promo-store");
    if (promoSel) {
        promoSel.innerHTML = "";
        orgs.forEach(o => {
            const opt = document.createElement("option");
            opt.value = o.id;
            opt.textContent = o.name + (o.wb_seller_id ? " (ID " + o.wb_seller_id + ")" : "");
            if (o.id === ORG_ID) opt.selected = true;
            promoSel.appendChild(opt);
        });
    }
    // Sync analytics-store dropdown in Основные показатели
    const anSel = document.getElementById("analytics-store");
    if (anSel) {
        anSel.innerHTML = "";
        orgs.forEach(o => {
            const opt = document.createElement("option");
            opt.value = o.id;
            opt.textContent = o.name + (o.wb_seller_id ? " (ID " + o.wb_seller_id + ")" : "");
            if (o.id === ORG_ID) opt.selected = true;
            anSel.appendChild(opt);
        });
    }
    // Sync filter-store in top-bar
    const fSel = document.getElementById("filter-store");
    if (fSel) {
        fSel.innerHTML = "";
        orgs.forEach(o => {
            const opt = document.createElement("option");
            opt.value = o.id;
            opt.textContent = o.name + (o.wb_seller_id ? " (ID " + o.wb_seller_id + ")" : "");
            if (o.id === ORG_ID) opt.selected = true;
            fSel.appendChild(opt);
        });
    }
    // Sync mkt-store dropdown in Стол маркетолога
    const mktSel = document.getElementById("mkt-store");
    if (mktSel) {
        mktSel.innerHTML = "";
        orgs.forEach(o => {
            const opt = document.createElement("option");
            opt.value = o.id;
            opt.textContent = o.name + (o.wb_seller_id ? " (ID " + o.wb_seller_id + ")" : "");
            if (o.id === ORG_ID) opt.selected = true;
            mktSel.appendChild(opt);
        });
    }
    // Sync ads-store dropdown in Реклама
    const adsSel = document.getElementById("ads-store");
    if (adsSel) {
        adsSel.innerHTML = "";
        orgs.forEach(o => {
            const opt = document.createElement("option");
            opt.value = o.id;
            opt.textContent = o.name + (o.wb_seller_id ? " (ID " + o.wb_seller_id + ")" : "");
            if (o.id === ORG_ID) opt.selected = true;
            adsSel.appendChild(opt);
        });
    }
}

async function switchOrg() {
    if (_costDirty && !await confirmDirty()) return;
    // Update URL so refresh keeps the right org
    ORG_ID = document.getElementById('org-select').value;
    localStorage.setItem('nl_org_id', ORG_ID);
    history.replaceState(null, '', '/nl/v2?org=' + ORG_ID);
    // Sync all store dropdowns
    const cpSel = document.getElementById('cp-store');
    if (cpSel) cpSel.value = ORG_ID;
    const ueSel2 = document.getElementById('ue-store');
    if (ueSel2) ueSel2.value = ORG_ID;
    const promoSel2 = document.getElementById('promo-store');
    if (promoSel2) promoSel2.value = ORG_ID;
    const anSel2 = document.getElementById('analytics-store');
    if (anSel2) anSel2.value = ORG_ID;
    const fSel2 = document.getElementById('filter-store');
    if (fSel2) fSel2.value = ORG_ID;
    const mktSel2 = document.getElementById('mkt-store');
    if (mktSel2) mktSel2.value = ORG_ID;
    const adsSel2 = document.getElementById('ads-store');
    if (adsSel2) adsSel2.value = ORG_ID;
    showApp();
    // Reload current active tab data for new org
    var activeTab = document.querySelector('.page-section.active');
    if (activeTab) {
        var tabName = activeTab.id.replace('page-', '');
        if (tabName === 'costprice') { loadTaxSettings(); loadCostPrices(); }
        else if (tabName === 'unitecon') { if (!ueTabulator) initUEGrid(); loadUEData(); }
        else if (tabName === 'promo') { if (typeof promoTabulator === 'undefined' || !promoTabulator) initPromoGrid(); loadPromoData(); }
        else if (tabName === 'salesplan') loadSalesPlans();
        else if (tabName === 'fboneeds') loadFboNeeds();
        else if (tabName === 'ads') { if (!adsTabulator) initAdsGrid(); loadAds(); }
        else if (tabName === 'extads') loadExtAds();
        else if (tabName === 'analytics') loadAnalytics();
        else if (tabName === 'marketer') loadMarketer();
    }
}

async function switchCostStore() {
    if (_costDirty && !await confirmDirty()) return;
    const cpSel = document.getElementById('cp-store');
    const newOrgId = cpSel.value;
    if (newOrgId === ORG_ID) return;
    ORG_ID = newOrgId;
    localStorage.setItem('nl_org_id', ORG_ID);
    // Sync sidebar org-select + ue-store + promo-store
    const sideSel = document.getElementById('org-select');
    if (sideSel) sideSel.value = ORG_ID;
    const ueSel3 = document.getElementById('ue-store');
    if (ueSel3) ueSel3.value = ORG_ID;
    const promoSel4 = document.getElementById('promo-store');
    if (promoSel4) promoSel4.value = ORG_ID;
    const mktSel4 = document.getElementById('mkt-store');
    if (mktSel4) mktSel4.value = ORG_ID;
    history.replaceState(null, '', '/nl/v2?org=' + ORG_ID);
    showApp();
    // Reload cost prices for new org
    loadTaxSettings();
    loadCostPrices();
}

async function switchUEStore() {
    const ueSel = document.getElementById("ue-store");
    const newOrgId = ueSel.value;
    if (newOrgId === ORG_ID) return;
    ORG_ID = newOrgId;
    localStorage.setItem("nl_org_id", ORG_ID);
    // Sync sidebar + cp-store + promo-store
    const sideSel = document.getElementById("org-select");
    if (sideSel) sideSel.value = ORG_ID;
    const cpSel = document.getElementById("cp-store");
    if (cpSel) cpSel.value = ORG_ID;
    const promoSel3 = document.getElementById("promo-store");
    if (promoSel3) promoSel3.value = ORG_ID;
    const mktSel3 = document.getElementById("mkt-store");
    if (mktSel3) mktSel3.value = ORG_ID;
    const adsSel4 = document.getElementById("ads-store");
    if (adsSel4) adsSel4.value = ORG_ID;
    history.replaceState(null, "", "/nl/v2?org=" + ORG_ID);
    loadUEData();
}

async function switchAdsStore() {
    const adsSel = document.getElementById("ads-store");
    const newOrgId = adsSel.value;
    if (newOrgId === ORG_ID) return;
    ORG_ID = newOrgId;
    localStorage.setItem("nl_org_id", ORG_ID);
    // Sync sidebar + cp-store + ue-store + promo-store + mkt-store
    const sideSel = document.getElementById("org-select");
    if (sideSel) sideSel.value = ORG_ID;
    const cpSel = document.getElementById("cp-store");
    if (cpSel) cpSel.value = ORG_ID;
    const ueSel4 = document.getElementById("ue-store");
    if (ueSel4) ueSel4.value = ORG_ID;
    const promoSel5 = document.getElementById("promo-store");
    if (promoSel5) promoSel5.value = ORG_ID;
    const mktSel5 = document.getElementById("mkt-store");
    if (mktSel5) mktSel5.value = ORG_ID;
    history.replaceState(null, "", "/nl/v2?org=" + ORG_ID);
    loadAds();
}

async function switchAnalyticsStore() {
    const anSel = document.getElementById("analytics-store");
    const newOrgId = anSel.value;
    if (newOrgId === ORG_ID) return;
    ORG_ID = newOrgId;
    localStorage.setItem("nl_org_id", ORG_ID);
    // Sync sidebar + cp-store + ue-store + promo-store
    const sideSel = document.getElementById("org-select");
    if (sideSel) sideSel.value = ORG_ID;
    const cpSel = document.getElementById("cp-store");
    if (cpSel) cpSel.value = ORG_ID;
    const ueSel = document.getElementById("ue-store");
    if (ueSel) ueSel.value = ORG_ID;
    const promoSel = document.getElementById("promo-store");
    if (promoSel) promoSel.value = ORG_ID;
    const anSel2 = document.getElementById("analytics-store");
    if (anSel2) anSel2.value = ORG_ID;
    history.replaceState(null, "", "/nl/v2?org=" + ORG_ID);
    loadAnalytics();
}

async function switchTopStore() {
    const fSel = document.getElementById('filter-store');
    const newOrgId = fSel.value;
    if (newOrgId === ORG_ID) return;
    ORG_ID = newOrgId;
    localStorage.setItem('nl_org_id', ORG_ID);
    // Sync all other selectors
    const sideSel = document.getElementById('org-select');
    if (sideSel) sideSel.value = ORG_ID;
    const cpSel = document.getElementById('cp-store');
    if (cpSel) cpSel.value = ORG_ID;
    const ueSel = document.getElementById('ue-store');
    if (ueSel) ueSel.value = ORG_ID;
    const promoSel = document.getElementById('promo-store');
    if (promoSel) promoSel.value = ORG_ID;
    const anSel = document.getElementById('analytics-store');
    if (anSel) anSel.value = ORG_ID;
    const adsSel = document.getElementById("ads-store");
    if (adsSel) adsSel.value = ORG_ID;
    const mktSel = document.getElementById("mkt-store");
    if (mktSel) mktSel.value = ORG_ID;
    const extAdsSel = document.getElementById("ext-ad-store");
    if (extAdsSel) extAdsSel.value = ORG_ID;
    history.replaceState(null, '', '/nl/v2?org=' + ORG_ID);
    // Reload current tab data
    var activeTab = document.querySelector('.page-section.active');
    if (activeTab) {
        var tabName = activeTab.id.replace('page-', '');
        if (tabName === 'stats') loadStats();
        else if (tabName === 'analytics') loadAnalytics();
        else if (tabName === 'rnp') loadRnp();
        else if (tabName === 'opiu') loadOpiu();
        else if (tabName === 'ads') loadAds();
        else if (tabName === 'marketer') loadMarketer();
        else if (tabName === 'extads') loadExtAds();
    }
}

function showNewOrgDialog() {
    const d = document.getElementById('new-org-dialog');
    d.style.display = 'flex';
}
function hideNewOrgDialog() {
    document.getElementById('new-org-dialog').style.display = 'none';
}

async function createNewOrg() {
    const name = document.getElementById('new-org-name').value || 'Новый магазин';
    const res = await fetch('/api/v1/nl/organizations?token=' + TOKEN, {
        method: 'POST', headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({name})
    });
    if (res.ok) {
        const data = await res.json();
        ORG_ID = data.id;
        localStorage.setItem('nl_org_id', ORG_ID);
        hideNewOrgDialog();
        loadOrgs();
        loadRefData();
        loadWbKeys();
    }
}

async function loadProfile() {
    if (!TOKEN) return;
    try {
        const res = await fetch('/api/v1/nl/profile?token=' + TOKEN);
        if (!res.ok) return;
        const data = await res.json();
        
        const avatar = document.getElementById('profile-avatar');
        const emailEl = document.getElementById('profile-email');
        const summaryEl = document.getElementById('profile-summary');
        
        if (avatar) avatar.textContent = data.email.charAt(0).toUpperCase();
        if (emailEl) emailEl.textContent = data.email;
        if (summaryEl) summaryEl.textContent = data.shops_count + ' ' + declShop(data.shops_count);
        
        const shopsEl = document.getElementById('shops-list');
        if (!shopsEl) return;
        
        if (!data.shops.length) {
            shopsEl.innerHTML = '<div style="color:#999;padding:12px">Магазины не подключены</div>';
            return;
        }
        
        shopsEl.innerHTML = data.shops.map(s => {
            const keysHtml = s.keys.map(k => 
                '<div style="display:flex;align-items:center;gap:6px;font-size:.85em;padding:3px 0">' +
                '<span style="color:#00b894">🔑</span>' +
                '<span>' + k.name + '</span>' +
                '<span style="color:#999">(' + k.created_at + ')</span>' +
                '<span style="cursor:pointer;color:#e74c3c;margin-left:4px" onclick="deleteWbKeyBtn(this)" data-keyid="' + k.id + '" data-orgid="' + s.id + '">✕</span>' +
                '</div>'
            ).join('');
            
            return '<div style="background:#fff;border-radius:8px;padding:14px 16px;box-shadow:0 1px 3px rgba(0,0,0,.06);margin-bottom:8px" id="shop-card-' + s.id + '">' +
                '<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px">' +
                '<span style="font-size:1.1em">🏪</span>' +
                '<span style="font-weight:600" id="shop-name-' + s.id + '">' + s.name + '</span>' +
                '<button onclick="startRenameShop(this)" data-orgid="' + s.id + '" style="background:none;border:none;cursor:pointer;font-size:.85em;color:#999;margin-left:4px" title="Rename">&#9998;</button>' +
                (s.wb_seller_id ? '<span style="font-size:.8em;color:#999;background:#f0f0f0;padding:2px 6px;border-radius:4px">ID ' + s.wb_seller_id + '</span>' : '') +
                '<span style="margin-left:auto;font-size:.8em;color:#999" id="shop-status-' + s.id + '">не проверен</span>' +
                '<button onclick="verifyShopKey(this)" data-orgid="' + s.id + '" style="background:none;border:1px solid #ddd;border-radius:4px;padding:3px 8px;cursor:pointer;font-size:.85em;color:#6c5ce7" title="Проверить ключ">🔍 Проверить</button>' +
                '</div>' +
                '<div style="margin-left:28px">' +
                (keysHtml || '<span style="color:#e74c3c;font-size:.85em">❌ Нет API ключей</span>') +
                '</div>' +
                '</div>';
        }).join('');
        
        const inviteOrgSel = document.getElementById('invite-org');
        if (inviteOrgSel) {
            inviteOrgSel.innerHTML = '<option value="">Выберите магазин</option>' + 
                data.shops.map(s => '<option value="' + s.id + '">' + s.name + '</option>').join('');
        }
        
    } catch(e) { console.error('loadProfile error:', e); }
}

async function verifyShopKey(btn) {
    const orgId = btn.dataset.orgid;
    const statusEl = document.getElementById('shop-status-' + orgId);
    if (!statusEl) return;
    
    btn.disabled = true;
    btn.textContent = '⏳ Проверка...';
    statusEl.textContent = 'проверяем...';
    statusEl.style.color = '#999';
    
    try {
        const res = await fetch('/api/v1/nl/verify-wb-key?token=' + TOKEN, {
            method: 'POST',
            headers: {'Content-Type': 'application/json'},
            body: JSON.stringify({org_id: orgId})
        });
        
        if (!res.ok) {
            const err = await res.json();
            throw new Error(err.detail || 'Ошибка');
        }
        
        const data = await res.json();
        
        if (data.status === 'ok') {
            statusEl.style.color = '#00b894';
            statusEl.innerHTML = '✅ Работает' + (data.products_count ? ' (' + data.products_count + ' тов.)' : '');
            btn.textContent = '🔍 Проверить';
        } else if (data.status === 'warn') {
            statusEl.style.color = '#e17055';
            statusEl.innerHTML = '⚠️ ' + data.message;
            btn.textContent = '🔍 Проверить';
        } else {
            statusEl.style.color = '#e74c3c';
            statusEl.innerHTML = '❌ ' + data.message;
            btn.textContent = '🔍 Проверить';
        }
    } catch(e) {
        statusEl.style.color = '#e74c3c';
        statusEl.innerHTML = '❌ ' + e.message;
        btn.textContent = '🔍 Проверить';
    } finally {
        btn.disabled = false;
    }
}

function declShop(n) {
    if (n === 1) return 'магазин';
    if (n >= 2 && n <= 4) return 'магазина';
    return 'магазинов';
}



async function deleteWbKeyBtn(span) {
    if (!confirm('Удалить ключ?')) return;
    const keyId = span.dataset.keyid;
    const orgId = span.dataset.orgid;
    await fetch('/api/v1/nl/wb-keys/' + keyId + '?org_id=' + orgId, {method: 'DELETE'});
    loadProfile();
    loadOrgs();
}

async function startRenameShop(btn) {
    const orgId = btn.dataset.orgid;
    const nameEl = document.getElementById('shop-name-' + orgId);
    if (!nameEl) return;
    const currentName = nameEl.textContent;
    const newName = prompt('Новое название:', currentName);
    if (!newName || newName.trim() === '' || newName.trim() === currentName) return;
    try {
        const res = await fetch('/api/v1/nl/rename-org?token=' + TOKEN, {
            method: 'POST',
            headers: {'Content-Type': 'application/json'},
            body: JSON.stringify({org_id: orgId, name: newName.trim()})
        });
        if (!res.ok) { const err = await res.json(); alert('Ошибка: ' + (err.detail || 'не удалось')); return; }
        const data = await res.json();
        nameEl.textContent = data.name;
        const inviteSel = document.getElementById('invite-org');
        if (inviteSel) { for (let i = 0; i < inviteSel.options.length; i++) { if (inviteSel.options[i].value === orgId) { inviteSel.options[i].textContent = data.name; break; } } }
        showToast("Сохранено: " + data.name);
        const orgSel = document.getElementById('org-selector');
        if (orgSel) { for (let i = 0; i < orgSel.options.length; i++) { if (orgSel.options[i].value === orgId) { orgSel.options[i].textContent = data.name; break; } } }
    } catch(e) { alert('Ошибка сети: ' + e.message); }
}

async function connectNewShop() {
    const name = document.getElementById('new-shop-name').value.trim();
    const api_key = document.getElementById('new-shop-key').value.trim();
    const statusEl = document.getElementById('connect-status');
    const btn = document.getElementById('btn-connect-shop');
    
    if (!api_key) {
        statusEl.style.display = 'block';
        statusEl.style.background = '#fff3f3';
        statusEl.style.color = '#c62828';
        statusEl.textContent = 'Введите API ключ';
        return;
    }
    
    btn.disabled = true;
    btn.textContent = 'Подключение...';
    statusEl.style.display = 'block';
    statusEl.style.background = '#e8f5e9';
    statusEl.style.color = '#2e7d32';
    statusEl.textContent = '⏳ Подключаем магазин...';
    
    try {
        const res = await fetch('/api/v1/nl/connect-wb?token=' + TOKEN, {
            method: 'POST',
            headers: {'Content-Type': 'application/json'},
            body: JSON.stringify({name: name || 'Новый магазин', api_key})
        });
        
        if (!res.ok) {
            const err = await res.json();
            throw new Error(err.detail || 'Ошибка подключения');
        }
        
        const data = await res.json();
        
        statusEl.style.background = '#e8f5e9';
        statusEl.style.color = '#2e7d32';
        statusEl.innerHTML = '✅ Магазин <strong>' + data.name + '</strong> подключен!' +
            (data.wb_seller_id ? ' (seller ID: ' + data.wb_seller_id + ')' : '') +
            '<br><span style="font-size:.8em;color:#666">Запускаем синхронизацию данных...</span>';
        
        document.getElementById('new-shop-name').value = '';
        document.getElementById('new-shop-key').value = '';
        
        ORG_ID = data.org_id;
        localStorage.setItem('nl_org_id', ORG_ID);
        
        await loadOrgs();
        loadProfile();
        
    } catch(e) {
        statusEl.style.background = '#fff3f3';
        statusEl.style.color = '#c62828';
        statusEl.textContent = '❌ ' + e.message;
    } finally {
        btn.disabled = false;
        btn.textContent = 'Подключить';
    }
}

async function inviteColleague() {
    const email = document.getElementById('invite-email').value.trim();
    const role = document.getElementById('invite-role').value;
    const orgId = document.getElementById('invite-org').value;
    const statusEl = document.getElementById('invite-status');
    
    if (!email || !orgId) {
        statusEl.style.display = 'block';
        statusEl.style.background = '#fff3f3';
        statusEl.style.color = '#c62828';
        statusEl.textContent = 'Укажите email и выберите магазин';
        return;
    }
    
    try {
        const res = await fetch('/api/v1/nl/invite?token=' + TOKEN, {
            method: 'POST',
            headers: {'Content-Type': 'application/json'},
            body: JSON.stringify({email, role, org_id: orgId})
        });
        
        if (!res.ok) {
            const err = await res.json();
            throw new Error(err.detail || 'Ошибка');
        }
        
        const data = await res.json();
        statusEl.style.display = 'block';
        statusEl.style.background = '#e8f5e9';
        statusEl.style.color = '#2e7d32';
        statusEl.innerHTML = '✅ Приглашение отправлено: <strong>' + data.email + '</strong> (' + data.role + ')';
        document.getElementById('invite-email').value = '';
        
    } catch(e) {
        statusEl.style.display = 'block';
        statusEl.style.background = '#fff3f3';
        statusEl.style.color = '#c62828';
        statusEl.textContent = '❌ ' + e.message;
    }
}

async function loadWbKeys() {
    if (!ORG_ID) return;
    const res = await fetch('/api/v1/nl/wb-keys?org_id=' + ORG_ID);
    const keys = await res.json();
    const el = document.getElementById('wb-keys-list');
    if (!keys.length) {
        el.innerHTML = '<div style="color:#999;font-size:.9em;padding:8px">Ключи не добавлены</div>';
        return;
    }
    el.innerHTML = keys.map(k =>
        '<div class="wb-key-item"><span class="name">🔑 ' + k.name + '</span><span class="date">' + (k.created_at||'').substring(0,10) + '</span><span class="del" data-keyid="' + k.id + '" onclick="deleteWbKeyBtn(this)">✕</span></div>'
    ).join('');
}

async function addWbKey() {
    const name = document.getElementById('wb-key-name').value || 'WB Key';
    const api_key = document.getElementById('wb-key-value').value;
    if (!api_key) return;
    const res = await fetch('/api/v1/nl/wb-keys?org_id=' + ORG_ID, {
        method: 'POST', headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({name, api_key})
    });
    if (res.ok) {
        document.getElementById('wb-key-name').value = '';
        document.getElementById('wb-key-value').value = '';
        loadWbKeys();
        loadOrgs(); // update key count
    } else {
        const err = await res.json();
        alert(err.detail || 'Ошибка');
    }
}

async function deleteWbKeyBtn(btn) { await deleteWbKey(btn.dataset.keyid); }
async function deleteWbKey(id) {
    if (!confirm('Удалить ключ?')) return;
    await fetch('/api/v1/nl/wb-keys/' + id + '?org_id=' + ORG_ID, {method: 'DELETE'});
    loadWbKeys();
    loadOrgs();
}



// ─── OPERATIONAL CONTROL ──────────────────────────────────

async function loadControl() {
    const dateVal = document.getElementById('ctrl-date').value;
    if (!dateVal || dateVal === 'Нет данных') return;
    const res = await fetch('/api/v1/nl/control?org_id=' + ORG_ID + '&target_date=' + dateVal);
    const data = await res.json();
    const s = data.summary;

    // Alerts
    const alerts = document.getElementById('ctrl-alerts');
    let alertHtml = '';
    if (s.zero_stock_count > 0) alertHtml += '<div style="background:#fff3f3;border:1px solid #ffcdd2;border-radius:6px;padding:8px 12px;margin-bottom:6px;color:#c62828">🔴 Нет в наличии: ' + s.zero_stock_count + ' товаров</div>';
    if (s.low_stock_count > 0) alertHtml += '<div style="background:#fff8e1;border:1px solid #ffe082;border-radius:6px;padding:8px 12px;margin-bottom:6px;color:#e65100">🟡 Низкий остаток (≤5): ' + s.low_stock_count + ' товаров</div>';
    if (s.low_rating_count > 0) alertHtml += '<div style="background:#fff8e1;border:1px solid #ffe082;border-radius:6px;padding:8px 12px;margin-bottom:6px;color:#e65100">⚠️ Низкий рейтинг (<4): ' + s.low_rating_count + ' товаров</div>';
    alerts.innerHTML = alertHtml;

    // Cards
    const cards = document.getElementById('ctrl-cards');
    const metrics = [
        {label: '📦 Товаров', value: s.total_products, color: '#6c5ce7'},
        {label: '🏪 Остаток', value: s.total_stock, color: s.total_stock > 0 ? '#00b894' : '#e74c3c'},
        {label: '📋 Заказы', value: s.total_orders, color: '#0984e3'},
        {label: '✅ Выкупы', value: s.total_buyouts, color: '#00b894'},
        {label: '↩️ Возвраты', value: s.total_returns, color: '#d63031'},
        {label: '👁 Показы', value: s.total_impressions, color: '#636e72'},
        {label: '👆 Клики', value: s.total_clicks, color: '#0984e3'},
        {label: '📊 CTR', value: s.ctr + '%', color: s.ctr > 5 ? '#00b894' : s.ctr > 2 ? '#fdcb6e' : '#e74c3c'},
        {label: '💰 Реклама ₽', value: s.total_ad_cost ? s.total_ad_cost.toFixed(0) : '0', color: '#e17055'},
        {label: '⭐ Ср. рейтинг', value: s.avg_rating || '—', color: '#fdcb6e'},
    ];
    cards.innerHTML = metrics.map(m => '<div style="background:#fff;border-radius:8px;padding:12px;box-shadow:0 1px 3px rgba(0,0,0,.08);text-align:center"><div style="font-size:.75em;color:#999;margin-bottom:4px">' + m.label + '</div><div style="font-size:1.3em;font-weight:700;color:' + m.color + '">' + m.value + '</div></div>').join('');

    // Products table
    const tbody = document.getElementById('ctrl-body');
    if (!data.products.length) { tbody.innerHTML = '<tr><td colspan="15" class="empty">Нет данных</td></tr>'; return; }
    tbody.innerHTML = data.products.map(p => {
        const thumb = (p.photo_main || '').replace('/hq/', '/c246x328/').replace('/big/', '/c246x328/').replace('/tm/', '/c246x328/');
        const img = thumb ? '<img class="photo" src="' + thumb + '" loading="lazy">' : '📦';
        const ctr = p.impressions > 0 ? (p.clicks / p.impressions * 100).toFixed(1) + '%' : '—';
        const stockColor = p.stock_qty <= 0 ? '#e74c3c' : p.stock_qty <= 5 ? '#e17055' : '#00b894';
        const sizeLabel = p.size_name && p.size_name !== '0' && p.size_name !== 'ONE SIZE' ? p.size_name : '';
        return '<tr data-entity="' + (p.entity_id||'') + '">' +
            '<td>' + img + '</td>' +
            '<td><b>' + p.nm_id + '</b></td>' +
            '<td>' + (p.product_name || '').substring(0, 25) + '</td>' +
            '<td style="font-size:.8em;color:#636e72">' + sizeLabel + '</td>' +
            '<td style="font-size:.7em;color:#999">' + (p.barcode || '') + '</td>' +
            '<td style="color:' + stockColor + ';font-weight:600">' + (p.stock_qty ?? '—') + '</td>' +
            '<td>' + (p.orders_count ?? '—') + '</td>' +
            '<td>' + (p.buyouts_count ?? '—') + '</td>' +
            '<td>' + (p.returns_count ?? '—') + '</td>' +
            '<td>' + (p.rating ? p.rating.toFixed(1) : '—') + '</td>' +
            '<td>' + (p.impressions ?? '—') + '</td>' +
            '<td>' + (p.clicks ?? '—') + '</td>' +
            '<td>' + ctr + '</td>' +
            '<td>' + (p.ad_cost ? p.ad_cost.toFixed(0) : '—') + '</td>' +
            '<td>' + (p.price ? p.price.toFixed(0) : '—') + '</td>' +
            '<td>' + (p.price_discount ? p.price_discount.toFixed(0) : '—') + '</td>' +
            '<td>' + (p.tariff ? p.tariff.toFixed(0) + '%' : '—') + '</td>' +
            '</tr>';
    }).join('');
}

// Sync ctrl dates with ref dates
function syncCtrlDates() {
    const refSel = document.getElementById('ref-date');
    const ctrlSel = document.getElementById('ctrl-date');
    if (!refSel || !ctrlSel) return;
    ctrlSel.innerHTML = refSel.innerHTML;
    ctrlSel.value = refSel.value;
}

// Override loadDates to also sync
const origLoadDates = loadDates;
loadDates = async function() {
    const result = await origLoadDates();
    syncCtrlDates();
    return result;
};

// Auto-login
try {
    if (TOKEN && ORG_ID) {
        fetch('/api/v1/nl/me?token=' + TOKEN).then(r => {
            if (r.ok) return r.json(); throw '';
        }).then(d => {
            document.getElementById('user-email').textContent = d.email;
            showApp();
            loadOrgs();
            loadWbKeys();
        }).catch(() => {
            TOKEN = null; ORG_ID = null;
            localStorage.removeItem('nl_token');
            localStorage.removeItem('nl_org_id');
            showAuth(); // Показываем логин при невалидном токене
        });
    }
} catch(e) { console.error('Auto-login error:', e); }

// === Сортировка таблиц ===
function parseVal(v) {
    if (v == null || v === '—' || v === '') return null;
    var s = String(v).replace(/[\s₽%]/g,'').replace(',','.');
    var n = parseFloat(s);
    return isNaN(n) ? v : n;
}

function addSorting(tableId) {
    var table = document.getElementById(tableId);
    if (!table) return;
    var headers = table.querySelectorAll('thead th');
    headers.forEach(function(th, ci) {
        th.classList.add('sortable');
        th.addEventListener('click', function() {
            var tbody = table.querySelector('tbody');
            var rows = Array.from(tbody.querySelectorAll('tr'));
            if (!rows.length) return;
            if (rows[0].querySelector('.empty')) return;
            var asc = th.classList.contains('asc') ? false : (th.classList.contains('desc') ? true : true);
            headers.forEach(h => h.classList.remove('asc','desc'));
            th.classList.add(asc ? 'asc' : 'desc');
            rows.sort(function(a, b) {
                // Get cell value: check for input/select first, then textContent
                function cellVal(row) {
                    var cell = row.children[ci];
                    if (!cell) return '';
                    var inp = cell.querySelector('input, select');
                    if (inp) return inp.value || '';
                    return cell.textContent.trim();
                }
                var va = parseVal(cellVal(a));
                var vb = parseVal(cellVal(b));
                if (va == null && vb == null) return 0;
                if (va == null) return 1;
                if (vb == null) return -1;
                if (typeof va === 'number' && typeof vb === 'number') return asc ? va - vb : vb - va;
                return asc ? String(va).localeCompare(String(vb),'ru') : String(vb).localeCompare(String(va),'ru');
            });
            rows.forEach(r => tbody.appendChild(r));
        });
    });
}

function initSorting() {
    // Stats products table
    var sp = document.getElementById('stats-products');
    if (sp && sp.closest('table')) addSorting(sp.closest('table').id || undefined);
    // All tables with thead
    document.querySelectorAll('table').forEach(function(t) {
        if (t.querySelector('thead') && t.id) addSorting(t.id);
    });
}

// Call initSorting after DOM is ready (deferred)
setTimeout(initSorting, 500);

// ==================== FBO NEEDS ====================
var fboAllData = [];
var fboEdits = {};  // key -> edited value

async function loadFboNeeds() {
    const days = document.getElementById('fbo-period').value;
    const tbody = document.getElementById('fbo-body');
    tbody.innerHTML = '<tr><td colspan="13" class="empty">Загрузка...</td></tr>';
    try {
        const r = await fetch('/api/v1/nl/fbo-needs?org_id=' + ORG_ID + '&days=' + days, {headers:{'Authorization':'Bearer '+TOKEN}});
        const d = await r.json();
        fboAllData = d.rows || [];
        fboEdits = {};
        // Fill warehouse filter
        const sel = document.getElementById('fbo-warehouse-filter');
        const curVal = sel.value;
        sel.innerHTML = '<option value="">Все склады</option>';
        (d.warehouses || []).forEach(w => {
            const opt = document.createElement('option');
            opt.value = w; opt.textContent = w;
            sel.appendChild(opt);
        });
        sel.value = curVal;
        filterFboTable();
    } catch(e) {
        tbody.innerHTML = '<tr><td colspan="13" class="empty">Ошибка: ' + e.message + '</td></tr>';
    }
}

function filterFboTable() {
    const wh = document.getElementById('fbo-warehouse-filter').value;
    const search = (document.getElementById('fbo-search').value || '').toLowerCase();
    const onlyNeeds = document.getElementById('fbo-only-needs').checked;
    let rows = fboAllData;
    if (wh) rows = rows.filter(r => r.warehouse_name === wh);
    if (search) rows = rows.filter(r =>
        String(r.nm_id).includes(search) ||
        (r.product_name || '').toLowerCase().includes(search)
    );
    if (onlyNeeds) rows = rows.filter(r => (fboEdits[r.entity_id + '_' + r.warehouse_name] ?? r.need) > 0);
    const tbody = document.getElementById('fbo-body');
    document.getElementById('fbo-count').textContent = rows.length + ' строк';
    if (!rows.length) { tbody.innerHTML = '<tr><td colspan="13" class="empty">Нет данных</td></tr>'; return; }
    var html = '';
    rows.forEach(r => {
        var key = r.entity_id + '_' + r.warehouse_name;
        var toSend = fboEdits[key] ?? r.need;
        var d2z = r.days_to_zero;
        var d2zColor = d2z <= 2 ? '#d63031' : d2z <= 5 ? '#e17055' : d2z <= 10 ? '#fdcb6e' : '#00b894';
        var toSendColor = toSend > 0 ? '#fdcb6e' : '';
        html += '<tr style="border-bottom:1px solid #f0f0f0">';
        html += '<td><img src="' + (r.photo_main || '') + '" style="width:40px;height:40px;object-fit:cover;border-radius:4px"></td>';
        html += '<td style="font-weight:600">' + r.nm_id + '</td>';
        html += '<td style="max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">' + esc(r.product_name || '') + '</td>';
        html += '<td>' + esc(r.size_name || '-') + '</td>';
        html += '<td style="font-size:.9em">' + esc(r.warehouse_name) + '</td>';
        html += '<td class="r">' + r.stock_qty + '</td>';
        html += '<td class="r">' + r.orders_total + '</td>';
        html += '<td class="r">' + r.order_rate.toFixed(1) + '</td>';
        html += '<td class="r" style="color:' + d2zColor + ';font-weight:600">' + (d2z >= 999 ? '\u221e' : d2z) + '</td>';
        html += '<td>' + r.supply_days + '</td>';
        html += '<td>' + r.min_batch + '</td>';
        html += '<td class="r" style="background:#ffeaa7">' + r.need + '</td>';
        html += '<td class="r" style="background:' + toSendColor + '"><input type="number" class="cost-input" data-fbo-key="' + key + '" value="' + toSend + '" min="0" style="width:60px;font-weight:700;font-size:1em" onchange="fboEditChange(this)"></td>';
        html += '</tr>';
    });
    tbody.innerHTML = html;
}

function fboEditChange(el) {
    var key = el.getAttribute('data-fbo-key');
    fboEdits[key] = parseInt(el.value) || 0;
}

function saveFboEdits() {
    // FBO edits are local only (for Excel export), no backend save needed yet
    alert('Изменения сохранены локально. Скачайте Excel для экспорта.');
}

function exportFboExcel() {
    var wh = document.getElementById('fbo-warehouse-filter').value;
    var search = (document.getElementById('fbo-search').value || '').toLowerCase();
    var onlyNeeds = document.getElementById('fbo-only-needs').checked;
    var rows = fboAllData;
    if (wh) rows = rows.filter(r => r.warehouse_name === wh);
    if (search) rows = rows.filter(r =>
        String(r.nm_id).includes(search) ||
        (r.product_name || '').toLowerCase().includes(search)
    );
    if (onlyNeeds) rows = rows.filter(r => (fboEdits[r.entity_id + '_' + r.warehouse_name] ?? r.need) > 0);
    if (!rows.length) { alert('Нет данных для экспорта'); return; }

    var lines = ['Арт WB\tТовар\tРазмер\tСклад\tОстаток\tЗаказов\tТемп/день\tДней до 0\tСрок поставки\tМин. партия\tПотребность\tК отправке'];
    rows.forEach(r => {
        var key = r.entity_id + '_' + r.warehouse_name;
        var toSend = fboEdits[key] ?? r.need;
        lines.push([r.nm_id, r.product_name, r.size_name, r.warehouse_name,
            r.stock_qty, r.orders_total, r.order_rate.toFixed(1),
            r.days_to_zero >= 999 ? 'inf' : r.days_to_zero,
            r.supply_days, r.min_batch, r.need, toSend
        ].join('\t'));
    });
    var blob = new Blob(['\ufeff' + lines.join(String.fromCharCode(10))], {type: 'text/tab-separated-values;charset=utf-8'});
    var a = document.createElement('a');
    a.href = URL.createObjectURL(blob);
    a.download = 'fbo_needs_' + new Date().toISOString().slice(0,10) + '.xls';
    a.click();
    URL.revokeObjectURL(a.href);
}

</script>
</body>
</html>
"""
    response = HTMLResponse(html)
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    return response
